"""
Bonus Module Routes — Payment of Bonus Act, 1965
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, abort
from app import db
from app.models.bonus import BonusRun, BonusEntry
from app.models.establishment import Establishment
from app.models.employee import Employee
from app.models.payroll import (MonthlyPayroll, PayrollEntry, PayrollEntryHead,
                                 SalaryHead, PayrollConfig)
from app.user_context import (user_establishments, verify_est_ownership, current_user_id,
                               capture_est_from_url)
from datetime import datetime, date
import json
import io
import calendar

bonus_bp = Blueprint('bonus', __name__)


# Role-agnostic hook: let ?establishment=X in URL restore session when lost.
# Works identically for admin and user — no role branches.
@bonus_bp.before_request
def _capture_url_establishment():
    if request.path and '/api/' in request.path:
        return None
    capture_est_from_url()
    return None


# =====================================================
# HELPER: Get Basic + DA amounts for all heads lookup
# =====================================================

def _get_basic_da_heads(establishment_id):
    """Return (basic_head, spl_basic_head, da_head) for an establishment."""
    heads = SalaryHead.query.filter_by(
        establishment_id=establishment_id,
        is_active=True,
        head_type='earning'
    ).all()
    basic_head = None
    spl_basic_head = None
    da_head = None
    for h in heads:
        code = (h.short_code or '').upper().strip()
        name = (h.head_name or '').upper().strip()
        if code == 'BASIC' or name == 'BASIC':
            basic_head = h
        elif code in ('DA', 'DEARNESS ALLOWANCE') or 'DEARNESS' in name:
            da_head = h
        elif code in ('SPECIAL BASIC', 'SPL BASIC', 'SPLBASIC') or 'SPECIAL BASIC' in name:
            spl_basic_head = h
    return basic_head, spl_basic_head, da_head


def _get_entry_basic_da(entry, basic_head, spl_basic_head, da_head):
    """Return (basic+spl_basic, da) earned amounts for a single PayrollEntry."""
    heads_rows = PayrollEntryHead.query.filter_by(payroll_entry_id=entry.id).all()
    amounts = {r.salary_head_id: r.earned_amount or 0 for r in heads_rows}
    basic = 0.0
    da = 0.0
    if basic_head:
        basic += amounts.get(basic_head.id, 0) or 0
    if spl_basic_head:
        basic += amounts.get(spl_basic_head.id, 0) or 0
    if da_head:
        da += amounts.get(da_head.id, 0) or 0
    return basic, da


# =====================================================
# ENGINE: Calculate bonus for a run
# =====================================================

def _calculate_bonus_run(run):
    """(Re)calculate all BonusEntry rows for the given BonusRun.

    BONUS BASIS: Attendance × Daily Rate (per-month) × Bonus %, summed for
    the FY. This matches how Vaishnavi Consultant manually calculates bonus
    for client-delivery sheets (see SM_Bonus_2025-26.xlsx sample).

    NO WAGE CEILING. NO ELIGIBILITY CAP. The old Basic+DA + ₹7,000 Sec.12
    cap logic was the wrong fit for daily-wage establishments where the
    whole gross sits in a single BASIC head and PayrollEntryHead.earned_amount
    isn't reliably populated — it produced 0/0 results and "calculation
    not coming" on the summary page.

    Only Sec.8 (min days worked) still gates eligibility — it doesn't change
    the math, just marks the row eligible / not eligible for client display.

    Field mapping into the existing BonusEntry schema (kept stable so
    Form C + the legacy Statement Excel continue to work without changes):
      total_days_worked  = total attendance (days_present + paid_holidays)
      total_basic_da     = total wage (Attendance × Daily Rate, summed)
      total_capped_wage  = total wage (no cap applied)
      bonus_at_ceiling   = total bonus (no cap = same as actual)
      bonus_at_actual    = total bonus
      monthly_data       = JSON with both new (attendance/daily_rate/
                           monthly_wage/monthly_bonus) AND legacy keys
                           (basic_da/capped/days/eligible) for back-compat.
    """
    bonus_pct = (run.bonus_percentage or 8.33) / 100.0
    min_days  = run.min_days_worked or 30

    # ── Section 1 — Attendance ──────────────────────────────────────────
    inc_nph_att     = bool(getattr(run, 'include_holiday_attendance', True))
    inc_ot_days_att = bool(getattr(run, 'att_include_ot_days',        False))
    skip_zero       = bool(getattr(run, 'att_skip_zero',              True))

    # ── Section 2 — Wage ────────────────────────────────────────────────
    use_full_gross  = bool(getattr(run, 'wage_use_full_gross',       False))
    wage_add_nph    = bool(getattr(run, 'wage_add_nph_wages',        False))
    wage_add_ot     = bool(getattr(run, 'include_overtime_in_wage',  False))
    wage_add_other  = bool(getattr(run, 'wage_add_other_allowance',  False))

    # ── Section 3 — Ceiling / Cap (None = not applicable) ───────────────
    raw_ceil = getattr(run, 'wage_ceiling_per_month', None)
    ceiling  = float(raw_ceil) if raw_ceil and float(raw_ceil) > 0 else None
    raw_cap  = getattr(run, 'bonus_cap_per_employee', None)
    bonus_cap = float(raw_cap) if raw_cap and float(raw_cap) > 0 else None

    # Pull all payrolls for this est across the FY months
    payrolls = MonthlyPayroll.query.filter(
        MonthlyPayroll.establishment_id == run.establishment_id,
        db.or_(
            db.and_(MonthlyPayroll.year == run.start_year, MonthlyPayroll.month >= 4),
            db.and_(MonthlyPayroll.year == run.end_year, MonthlyPayroll.month <= 3),
        )
    ).all()
    payroll_by_id = {p.id: p for p in payrolls}
    payroll_ids = list(payroll_by_id.keys())

    entries = PayrollEntry.query.filter(
        PayrollEntry.monthly_payroll_id.in_(payroll_ids)
    ).all() if payroll_ids else []

    emp_data = {}  # emp_id -> { monthly, total_attendance, total_wage, total_bonus }

    for entry in entries:
        payroll = payroll_by_id.get(entry.monthly_payroll_id)
        if not payroll:
            continue
        emp_id = entry.employee_id
        if not emp_id:
            continue

        present  = float(entry.days_present or 0)
        ph       = float(entry.paid_holidays or 0)
        ot_hours = float(entry.ot_hours or 0)
        ot_amt   = float(entry.ot_amount or 0)
        gross    = float(entry.gross_salary or 0)
        earned   = float(entry.earned_gross or 0)
        tot_earn = float(entry.total_earnings or 0)

        # ── Attendance ─────────────────────────────────────────────────
        attendance = present
        if inc_nph_att:
            attendance += ph
        if inc_ot_days_att and ot_hours > 0:
            # Convert OT hours → days assuming a standard 8-hour day
            attendance += round(ot_hours / 8.0, 2)

        # ── Daily rate (gross ≤ ₹2,000 → already daily; else / working days)
        wd = payroll.working_days or 26
        daily_rate = 0.0
        if gross > 0:
            daily_rate = gross if gross <= 2000 else (gross / wd)

        # ── Monthly wage ───────────────────────────────────────────────
        if use_full_gross:
            # Take everything the employee actually earned for the month
            monthly_wage = tot_earn
        else:
            # Base: attendance × daily rate
            base = attendance * daily_rate
            monthly_wage = base
            # NPH wages — only add if not already in attendance×rate
            if wage_add_nph and not inc_nph_att:
                monthly_wage += ph * daily_rate
            # OT wages
            if wage_add_ot:
                monthly_wage += ot_amt
            # Other allowance: anything in earned_gross beyond the base
            if wage_add_other:
                monthly_wage += max(0, earned - base)

        # ── Section 3 — per-month ceiling ──────────────────────────────
        if ceiling and monthly_wage > ceiling:
            monthly_wage = ceiling

        monthly_wage  = round(monthly_wage)
        monthly_bonus = round(monthly_wage * bonus_pct)
        month_key     = f"{payroll.year}-{payroll.month:02d}"

        bucket = emp_data.setdefault(emp_id, {
            'monthly': {},
            'total_attendance': 0.0,
            'total_wage': 0,
            'total_bonus': 0,
        })
        bucket['monthly'][month_key] = {
            # ── Vaishnavi-format fields (used by the Vaishnavi Excel) ──
            'attendance':    round(attendance, 1),
            'daily_rate':    round(daily_rate),
            'monthly_wage':  monthly_wage,
            'monthly_bonus': monthly_bonus,
            # ── Legacy fields (back-compat for Statement Excel + Form C) ──
            'basic_da':      monthly_wage,
            'capped':        monthly_wage,
            'days':          round(attendance, 1),
            'eligible':      True,           # no per-month gate
        }
        bucket['total_attendance'] += attendance
        bucket['total_wage']       += monthly_wage
        bucket['total_bonus']      += monthly_bonus

    # Wipe existing entries and rebuild
    BonusEntry.query.filter_by(bonus_run_id=run.id).delete()
    db.session.flush()

    total_emp = 0
    eligible_emp = 0
    total_bonus_sum = 0.0

    for emp_id, d in emp_data.items():
        total_attendance = d['total_attendance']

        # Section 1 toggle — skip zero-attendance employees entirely
        if skip_zero and total_attendance <= 0:
            continue

        total_emp += 1

        # Sec. 8 eligibility (min days)
        is_eligible = total_attendance >= min_days
        reason = None
        if not is_eligible:
            reason = f"Worked only {int(total_attendance)} days (need {min_days})"

        # Bonus is zero for ineligible employees (Sec. 8 — under min days)
        final_bonus = d['total_bonus'] if is_eligible else 0

        # Section 3 — per-employee bonus cap
        if bonus_cap and final_bonus > bonus_cap:
            final_bonus = bonus_cap

        be = BonusEntry(
            bonus_run_id=run.id,
            employee_id=emp_id,
            monthly_data=json.dumps(d['monthly']),
            months_eligible=len(d['monthly']),
            total_days_worked=round(total_attendance, 2),
            total_basic_da=d['total_wage'],         # repurposed: total wage
            total_capped_wage=d['total_wage'],      # display: same value
            bonus_at_ceiling=final_bonus,
            bonus_at_actual=final_bonus,
            is_eligible=is_eligible,
            ineligibility_reason=reason,
        )
        db.session.add(be)

        if is_eligible:
            eligible_emp += 1
            total_bonus_sum += final_bonus

    run.total_employees = total_emp
    run.eligible_employees = eligible_emp
    run.total_bonus_ceiling = round(total_bonus_sum, 2)
    run.total_bonus_actual  = round(total_bonus_sum, 2)
    db.session.commit()


# =====================================================
# ROUTES: list / create / view / recalc / finalize / delete
# =====================================================

@bonus_bp.route('/bonus')
def bonus_list():
    """List all bonus runs across user's establishments."""
    runs = BonusRun.query.join(Establishment).filter(
        Establishment.id.in_([e.id for e in user_establishments().all()])
    ).order_by(BonusRun.start_year.desc(), BonusRun.id.desc()).all()
    return render_template('bonus/list.html', runs=runs)


@bonus_bp.route('/bonus/new', methods=['GET', 'POST'])
def bonus_new():
    ests = user_establishments().order_by(Establishment.company_name).all()

    if request.method == 'POST':
        est_id = request.form.get('establishment_id', type=int)
        start_year = request.form.get('start_year', type=int)
        bonus_pct = request.form.get('bonus_percentage', type=float) or 8.33
        min_days = request.form.get('min_days_worked', type=int) or 30

        # Section 1 — Attendance
        att_nph    = 'include_holiday_attendance' in request.form
        att_ot_d   = 'att_include_ot_days'        in request.form
        skip_zero  = 'att_skip_zero'              in request.form
        # Section 2 — Wage
        full_gross = 'wage_use_full_gross'        in request.form
        w_nph      = 'wage_add_nph_wages'         in request.form
        w_ot       = 'include_overtime_in_wage'   in request.form
        w_other    = 'wage_add_other_allowance'   in request.form
        # Section 3 — Ceiling / Cap (blank → NULL = not applicable)
        ceil_raw = request.form.get('wage_ceiling_per_month', '').strip()
        cap_raw  = request.form.get('bonus_cap_per_employee', '').strip()
        try:
            wage_ceil = float(ceil_raw) if ceil_raw else None
        except ValueError:
            wage_ceil = None
        try:
            bonus_cap = float(cap_raw) if cap_raw else None
        except ValueError:
            bonus_cap = None

        est = Establishment.query.get_or_404(est_id)
        verify_est_ownership(est)

        # Block annual bonus for establishments that pay bonus MONTHLY —
        # their bonus is already disbursed each month in payroll, so an
        # annual run would double-pay.
        from app.models.payroll import PayrollConfig as _PC
        _cfg = _PC.query.filter_by(establishment_id=est_id).first()
        if _cfg and getattr(_cfg, 'monthly_bonus_applicable', False):
            flash(f'{est.company_name} is set to pay bonus MONTHLY (in Payroll Config), '
                  f'so it is excluded from the annual bonus run. Turn off '
                  f'"Pay Bonus Every Month" in Payroll Config to run annual bonus.', 'warning')
            return redirect(url_for('bonus.bonus_list'))

        # Prevent duplicate run for same est + FY
        existing = BonusRun.query.filter_by(
            establishment_id=est_id, start_year=start_year
        ).first()
        if existing:
            flash(f'A bonus run for {est.company_name} FY {start_year}-{str(start_year+1)[-2:]} already exists. Open it or delete it first.', 'warning')
            return redirect(url_for('bonus.bonus_view', run_id=existing.id))

        run = BonusRun(
            establishment_id=est_id,
            start_year=start_year,
            end_year=start_year + 1,
            bonus_percentage=bonus_pct,
            min_days_worked=min_days,
            include_holiday_attendance=att_nph,
            att_include_ot_days=att_ot_d,
            att_skip_zero=skip_zero,
            wage_use_full_gross=full_gross,
            wage_add_nph_wages=w_nph,
            include_overtime_in_wage=w_ot,
            wage_add_other_allowance=w_other,
            wage_ceiling_per_month=wage_ceil,
            bonus_cap_per_employee=bonus_cap,
            status='draft',
        )
        db.session.add(run)
        db.session.commit()

        _calculate_bonus_run(run)
        flash(f'Bonus run created and calculated for {est.company_name} — {run.fy_label}', 'success')
        return redirect(url_for('bonus.bonus_view', run_id=run.id))

    current_year = datetime.now().year
    years = list(range(current_year - 5, current_year + 1))
    return render_template('bonus/create.html', establishments=ests, years=years,
                           current_year=current_year)


@bonus_bp.route('/bonus/<int:run_id>')
def bonus_view(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = BonusEntry.query.filter_by(bonus_run_id=run_id).join(Employee).order_by(Employee.emp_code).all()
    # Decode monthly data
    for e in entries:
        try:
            e.monthly = json.loads(e.monthly_data) if e.monthly_data else {}
        except Exception:
            e.monthly = {}
    return render_template('bonus/view.html', run=run, entries=entries)


@bonus_bp.route('/bonus/<int:run_id>/recalculate', methods=['POST'])
def bonus_recalculate(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    if run.status == 'finalized':
        flash('Cannot recalculate a finalized run. Unfinalize first.', 'warning')
        return redirect(url_for('bonus.bonus_view', run_id=run_id))
    # Update the engine settings — bonus %, eligibility, and all three
    # composition sections in one go.
    run.bonus_percentage = request.form.get('bonus_percentage', type=float) or run.bonus_percentage
    run.min_days_worked  = request.form.get('min_days_worked',  type=int)   or run.min_days_worked
    # Section 1 — Attendance
    run.include_holiday_attendance = 'include_holiday_attendance' in request.form
    run.att_include_ot_days        = 'att_include_ot_days'        in request.form
    run.att_skip_zero              = 'att_skip_zero'              in request.form
    # Section 2 — Wage
    run.wage_use_full_gross        = 'wage_use_full_gross'        in request.form
    run.wage_add_nph_wages         = 'wage_add_nph_wages'         in request.form
    run.include_overtime_in_wage   = 'include_overtime_in_wage'   in request.form
    run.wage_add_other_allowance   = 'wage_add_other_allowance'   in request.form
    # Section 3 — Ceiling / Cap (blank → NULL)
    ceil_raw = (request.form.get('wage_ceiling_per_month') or '').strip()
    cap_raw  = (request.form.get('bonus_cap_per_employee') or '').strip()
    try:
        run.wage_ceiling_per_month = float(ceil_raw) if ceil_raw else None
    except ValueError:
        run.wage_ceiling_per_month = None
    try:
        run.bonus_cap_per_employee = float(cap_raw) if cap_raw else None
    except ValueError:
        run.bonus_cap_per_employee = None
    db.session.commit()
    _calculate_bonus_run(run)
    flash('Bonus recalculated with updated configuration.', 'success')
    return redirect(url_for('bonus.bonus_view', run_id=run_id))


@bonus_bp.route('/bonus/<int:run_id>/entry/<int:entry_id>/override', methods=['POST'])
def bonus_override(run_id, entry_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entry = BonusEntry.query.get_or_404(entry_id)
    if entry.bonus_run_id != run_id:
        abort(404)
    override = request.form.get('override_amount', type=float)
    remarks = request.form.get('remarks', '').strip()
    entry.override_amount = override if override is not None and override >= 0 else None
    entry.remarks = remarks or None
    db.session.commit()
    flash(f'Override saved for {entry.employee.name}', 'success')
    return redirect(url_for('bonus.bonus_view', run_id=run_id))


@bonus_bp.route('/bonus/<int:run_id>/finalize', methods=['POST'])
def bonus_finalize(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    pay_date_str = request.form.get('payment_date')
    if pay_date_str:
        try:
            run.payment_date = datetime.strptime(pay_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    run.status = 'finalized'
    run.finalized_at = datetime.utcnow()
    db.session.commit()
    flash('Bonus run finalized.', 'success')
    return redirect(url_for('bonus.bonus_view', run_id=run_id))


@bonus_bp.route('/bonus/<int:run_id>/unfinalize', methods=['POST'])
def bonus_unfinalize(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    run.status = 'draft'
    run.finalized_at = None
    db.session.commit()
    flash('Bonus run reopened for editing.', 'info')
    return redirect(url_for('bonus.bonus_view', run_id=run_id))


@bonus_bp.route('/bonus/<int:run_id>/delete', methods=['POST'])
def bonus_delete(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    db.session.delete(run)
    db.session.commit()
    flash('Bonus run deleted.', 'info')
    return redirect(url_for('bonus.bonus_list'))


# =====================================================
# REPORTS: Statement (month-wise) + Form C
# =====================================================

def _fy_month_keys(run):
    """Return ordered list of (year, month, month_label, key) for the FY."""
    out = []
    for m in range(4, 13):
        out.append((run.start_year, m, calendar.month_abbr[m],
                    f"{run.start_year}-{m:02d}"))
    for m in range(1, 4):
        out.append((run.end_year, m, calendar.month_abbr[m],
                    f"{run.end_year}-{m:02d}"))
    return out


@bonus_bp.route('/bonus/<int:run_id>/statement')
def bonus_statement(run_id):
    """Month-wise bonus statement — employee rows × month columns."""
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = BonusEntry.query.filter_by(bonus_run_id=run_id).join(Employee).order_by(Employee.emp_code).all()
    for e in entries:
        try:
            e.monthly = json.loads(e.monthly_data) if e.monthly_data else {}
        except Exception:
            e.monthly = {}
    months = _fy_month_keys(run)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('bonus/statement.html', run=run, entries=entries,
                           months=months, generated_on=generated_on)


@bonus_bp.route('/bonus/<int:run_id>/form-c/excel')
def bonus_form_c_excel(run_id):
    """Form C (Bonus Paid Register) — Excel download in Legal Landscape.
    Mirrors the HTML form_c view column-for-column so the statutory layout
    stays identical between print and Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)

    entries = (BonusEntry.query
               .filter_by(bonus_run_id=run_id)
               .join(Employee)
               .order_by(Employee.emp_code)
               .all())
    # Only eligible employees on Form C — matches the HTML view.
    entries = [e for e in entries if e.is_eligible]

    wb = Workbook()
    ws = wb.active
    ws.title = "Form C"

    # ── Styles (Calibri throughout — Form C is a statutory format and
    #    Calibri at the small body size prints cleanly on Legal sheets) ──
    title_font   = Font(bold=True, size=14, name='Calibri')
    subtitle_fnt = Font(bold=True, size=10, name='Calibri')
    info_font    = Font(size=9, name='Calibri')
    info_bold    = Font(bold=True, size=9, name='Calibri')
    header_font  = Font(bold=True, size=8.5, name='Calibri')
    body_font    = Font(size=9, name='Calibri')
    body_bold    = Font(bold=True, size=9, name='Calibri')
    total_font   = Font(bold=True, size=9.5, name='Calibri')

    thin   = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    total_fill  = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

    TOTAL_COLS = 12  # Sr + 11 statutory columns

    # ── Title block (centred, merged across all columns) ──
    ws.cell(row=1, column=1, value="FORM C").font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)

    ws.cell(row=2, column=1, value="[See Rule 4(c)]").font = subtitle_fnt
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)

    ws.cell(row=3, column=1, value="BONUS PAID REGISTER").font = subtitle_fnt
    ws.cell(row=3, column=1).alignment = center
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)

    ws.cell(row=4, column=1, value="(Under the Payment of Bonus Act, 1965)").font = info_font
    ws.cell(row=4, column=1).alignment = center
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TOTAL_COLS)

    # ── Establishment info block ──
    est = run.establishment
    info_rows = [
        ('Name of the Establishment:', (est.company_name or '').upper()),
        ('Address:',                   est.address or ''),
        ('Accounting Year:',           f"{run.fy_label} (01-Apr-{run.start_year} to 31-Mar-{run.end_year})"),
    ]
    r = 6
    for label, value in info_rows:
        ws.cell(row=r, column=1, value=label).font = info_bold
        ws.cell(row=r, column=1).alignment = left
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=value).font = info_font
        ws.cell(row=r, column=3).alignment = left
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=TOTAL_COLS)
        r += 1

    # Meta strip — bonus %, settings, payment date
    att_bits = []
    if run.include_holiday_attendance: att_bits.append('NPH')
    if run.att_include_ot_days:        att_bits.append('OT(÷8)')
    att_summary = ' + '.join(att_bits) if att_bits else 'Worked Only'

    if run.wage_use_full_gross:
        wage_summary = 'Full Gross'
    else:
        wb = ['Att×Rate']
        if run.wage_add_nph_wages:       wb.append('+NPH')
        if run.include_overtime_in_wage: wb.append('+OT')
        if run.wage_add_other_allowance: wb.append('+OtherAlw')
        wage_summary = ' '.join(wb)

    meta_parts = [
        f"Bonus Percentage: {run.bonus_percentage}%",
        f"Min Days (Sec. 8): {run.min_days_worked}",
        f"Attendance: {att_summary}",
        f"Wage: {wage_summary}",
    ]
    if run.wage_ceiling_per_month:
        meta_parts.append(f"Wage Ceiling: ₹{int(run.wage_ceiling_per_month):,}")
    if run.bonus_cap_per_employee:
        meta_parts.append(f"Bonus Cap: ₹{int(run.bonus_cap_per_employee):,}")
    if run.payment_date:
        meta_parts.append(f"Date of Payment: {run.payment_date.strftime('%d-%m-%Y')}")
    ws.cell(row=r, column=1, value='   |   '.join(meta_parts)).font = info_bold
    ws.cell(row=r, column=1).alignment = left
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    r += 2   # one blank row before the table

    # ── Table header (12 columns matching the HTML view) ──
    header_row = r
    headers = [
        'Sr\nNo.',
        'Name of the\nEmployee',
        "Father's /\nHusband's Name",
        'Designation',
        'No. of days\nworked in the\naccounting year',
        'Total salary\nor wage in respect\nof the accounting year\n(₹)',
        'Amount of\nbonus payable\nunder Sec. 10 or 11\n(₹)',
        'Deductions, if any,\non a/c of Puja bonus\nor other customary\nbonus',
        'Deductions on\naccount of financial\nloss, if any, caused\nby misconduct',
        'Amount\nactually paid\n(₹)',
        'Date on\nwhich paid',
        'Signature /\nThumb impression\nof the employee',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 56

    # ── Data rows ──
    data_start = header_row + 1
    payment_date_str = run.payment_date.strftime('%d-%m-%Y') if run.payment_date else ''
    t_wage  = 0
    t_bonus = 0
    t_paid  = 0
    for idx, e in enumerate(entries, 1):
        rr = data_start + idx - 1
        emp = e.employee
        bonus_val = float(e.override_amount if e.override_amount is not None else (e.bonus_at_actual or 0))
        wage_val  = float(e.total_basic_da or 0)
        cells = [
            (1, idx,                                                  center, body_font),
            (2, emp.name,                                             left,   body_bold),
            (3, emp.father_husband_name or '',                        left,   body_font),
            (4, emp.designation or '',                                left,   body_font),
            (5, int(e.total_days_worked or 0),                        center, body_font),
            (6, round(wage_val),                                      right,  body_font),
            (7, round(bonus_val),                                     right,  body_bold),
            (8, '',                                                   center, body_font),
            (9, '',                                                   center, body_font),
            (10, round(bonus_val),                                    right,  body_bold),
            (11, payment_date_str,                                    center, body_font),
            (12, '',                                                  center, body_font),
        ]
        for col, val, align, font in cells:
            c = ws.cell(row=rr, column=col, value=val)
            c.font = font
            c.alignment = align
            c.border = border
            if col in (6, 7, 10) and isinstance(val, (int, float)) and val:
                c.number_format = '#,##0'
        ws.row_dimensions[rr].height = 22
        t_wage  += wage_val
        t_bonus += bonus_val
        t_paid  += bonus_val

    # ── Total row ──
    total_row = data_start + len(entries)
    label_cell = ws.cell(row=total_row, column=1, value=f"TOTAL ({len(entries)} Employees)")
    label_cell.font = total_font
    label_cell.alignment = center
    label_cell.fill = total_fill
    label_cell.border = border
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    # Fill borders on the merged-into cells too (openpyxl needs each cell styled)
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = total_fill

    totals = [
        (6,  round(t_wage),  right),
        (7,  round(t_bonus), right),
        (8,  '',             center),
        (9,  '',             center),
        (10, round(t_paid),  right),
        (11, '',             center),
        (12, '',             center),
    ]
    for col, val, align in totals:
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = total_font
        c.alignment = align
        c.fill = total_fill
        c.border = border
        if isinstance(val, (int, float)) and val:
            c.number_format = '#,##0'
    ws.row_dimensions[total_row].height = 26

    # ── Signature block (3 rows below totals) ──
    sig_row = total_row + 3
    ws.cell(row=sig_row,     column=1, value='Signature of Employer:').font = info_bold
    ws.cell(row=sig_row + 1, column=1, value='_______________________').font = info_font
    ws.cell(row=sig_row,     column=10, value='Date:').font = info_bold
    ws.cell(row=sig_row,     column=11, value='_______________').font = info_font
    ws.cell(row=sig_row + 1, column=10, value='Place:').font = info_bold
    ws.cell(row=sig_row + 1, column=11, value='_______________').font = info_font

    # ── Generated footer ──
    footer_row = sig_row + 4
    ws.cell(row=footer_row, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   Vaishnavi Consultant"
            ).font = Font(size=8, italic=True, color='64748B', name='Calibri')
    ws.cell(row=footer_row, column=1).alignment = center
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=TOTAL_COLS)

    # ── Column widths (tuned for Legal landscape fit) ──
    widths = [5, 22, 22, 14, 11, 14, 14, 16, 18, 12, 11, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Print setup — Legal Landscape, fit-to-width ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = 5   # 5 = Legal
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5

    # Repeat header rows on every printed page
    ws.print_title_rows = f'{header_row}:{header_row}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = (run.establishment.company_name or 'Establishment').replace(' ', '_').replace('/', '_')[:60]
    filename = f"Form_C_{safe_name}_{run.fy_label.replace(' ', '_')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@bonus_bp.route('/bonus/<int:run_id>/form-c')
def bonus_form_c(run_id):
    """Form C — Bonus Paid Register (statutory format)."""
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = BonusEntry.query.filter_by(bonus_run_id=run_id).join(Employee).order_by(Employee.emp_code).all()
    # Only eligible employees on Form C
    entries = [e for e in entries if e.is_eligible]
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('bonus/form_c.html', run=run, entries=entries,
                           generated_on=generated_on)


# =====================================================
# EXCEL EXPORTS
# =====================================================

@bonus_bp.route('/bonus/<int:run_id>/statement/excel')
def bonus_statement_excel(run_id):
    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = BonusEntry.query.filter_by(bonus_run_id=run_id).join(Employee).order_by(Employee.emp_code).all()
    for e in entries:
        try:
            e.monthly = json.loads(e.monthly_data) if e.monthly_data else {}
        except Exception:
            e.monthly = {}
    months = _fy_month_keys(run)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Bonus Statement"

    bold = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin = Side(border_style='thin', color='94A3B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(months) + 4)
    ws.cell(row=1, column=1, value=f"BONUS STATEMENT — {run.establishment.company_name} — {run.fy_label}").font = Font(bold=True, size=13)
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=2, column=1, value=f"Percentage: {run.bonus_percentage}%  |  Wage Ceiling: ₹{run.wage_ceiling}  |  Min Wage Floor: ₹{run.min_wage_floor or '—'}  |  Eligibility Cap: ₹{run.eligibility_cap}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4 + len(months) + 4)

    # Header row
    headers = ['Sr', 'Emp Code', 'Name', 'Designation']
    for _, _, lbl, _ in months:
        headers.append(lbl)
    headers += ['Total Basic+DA', 'Total Capped', 'Bonus @ Ceiling', 'Bonus @ Actual']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Data
    r = 5
    tot_basic = tot_cap = tot_bc = tot_ba = 0.0
    for idx, e in enumerate(entries, 1):
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2, value=e.employee.emp_code).border = border
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=3, value=e.employee.name).border = border
        ws.cell(row=r, column=3).alignment = left
        ws.cell(row=r, column=4, value=e.employee.designation or '').border = border

        col = 5
        for _, _, _, key in months:
            m = e.monthly.get(key)
            val = m['basic_da'] if m and m.get('eligible') else (0 if not m else None)
            if val is None:
                ws.cell(row=r, column=col, value='')
            else:
                ws.cell(row=r, column=col, value=round(val, 0))
            ws.cell(row=r, column=col).border = border
            ws.cell(row=r, column=col).alignment = right
            col += 1

        ws.cell(row=r, column=col, value=round(e.total_basic_da, 0)).border = border
        ws.cell(row=r, column=col).alignment = right
        col += 1
        ws.cell(row=r, column=col, value=round(e.total_capped_wage, 0)).border = border
        ws.cell(row=r, column=col).alignment = right
        col += 1
        ws.cell(row=r, column=col, value=round(e.final_bonus_ceiling, 0)).border = border
        ws.cell(row=r, column=col).alignment = right
        ws.cell(row=r, column=col).font = bold
        col += 1
        ws.cell(row=r, column=col, value=round(e.final_bonus_actual, 0)).border = border
        ws.cell(row=r, column=col).alignment = right
        ws.cell(row=r, column=col).font = bold

        tot_basic += e.total_basic_da
        tot_cap += e.total_capped_wage
        tot_bc += e.final_bonus_ceiling
        tot_ba += e.final_bonus_actual
        r += 1

    # Totals
    ws.cell(row=r, column=1, value='TOTAL').font = bold
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4 + len(months))
    tot_col = 5 + len(months)
    for val in [tot_basic, tot_cap, tot_bc, tot_ba]:
        c = ws.cell(row=r, column=tot_col, value=round(val, 0))
        c.font = bold
        c.border = border
        c.alignment = right
        c.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
        tot_col += 1

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    for i in range(len(months)):
        ws.column_dimensions[chr(ord('E') + i)].width = 10

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5  # Legal
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Bonus_Statement_{run.establishment.company_name}_{run.fy_label}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# =============================================================================
# Vaishnavi-format Bonus Statement (Attendance × Daily Rate basis)
# -----------------------------------------------------------------------------
# Matches the SM_Bonus sample sheet's exact column layout the user uses for
# client delivery:
#
#   UAN/ESIC | Name | Father | Bonus %
#   |  Apr  : Attendance | Per day wage | Total Wage | Total Bonus  |
#   |  May  : Attendance | Per day wage | Total Wage | Total Bonus  |
#   |  ...  (12 months Apr..Mar)                                    |
#   | Grand Total: Total Attendance | Total Wage | Total Bonus
#
# Wage basis is "earned wage" = Attendance × Daily Rate (NOT Basic + DA).
# This is the right calculation for daily-wage establishments where the
# single BASIC head represents the entire wage. Computed on the fly from
# PayrollEntry rows for the FY months — no schema change, no impact on
# the existing Form C / statutory exporter.
# =============================================================================

def _compute_vaishnavi_bonus_data(run):
    """Read the persisted BonusEntry rows for this run and return them in the
    shape the Vaishnavi Excel builder expects.

    All math has already been done by _calculate_bonus_run (simple
    Attendance × Daily Rate × Bonus % basis). This helper just loads + sorts.

    If there are no BonusEntry rows yet — e.g. the run was just created and
    the calculation crashed midway — returns an empty list so the caller
    can flash a clear "click Recalculate first" message.
    """
    entries = (BonusEntry.query
               .filter_by(bonus_run_id=run.id)
               .join(Employee, Employee.id == BonusEntry.employee_id)
               .order_by(Employee.name)
               .all())

    rows = []
    for be in entries:
        try:
            monthly = json.loads(be.monthly_data) if be.monthly_data else {}
        except Exception:
            monthly = {}
        rows.append({
            'employee':         be.employee,
            'monthly':          monthly,
            'total_attendance': be.total_days_worked or 0,
            'total_wage':       be.total_basic_da or 0,
            'total_bonus':      (be.override_amount
                                 if be.override_amount is not None
                                 else (be.bonus_at_actual or 0)),
            'is_eligible':      bool(be.is_eligible),
        })
    return rows


@bonus_bp.route('/bonus/<int:run_id>/statement/vaishnavi-excel')
def bonus_vaishnavi_excel(run_id):
    """Vaishnavi-format Bonus Statement — Attendance × Daily Rate basis.
    Matches the SM_Bonus client-delivery sheet exactly.

    Computation is intentionally simple — NO CEILING, NO CAP:
      Per row :  Attendance | Daily Rate | Monthly Wage | Monthly Bonus
                 (× 12 months)
      Totals  :  Grand Total Attendance | Grand Total Wage | Grand Total Bonus

    Any error during build is caught and the user is redirected back to the
    bonus run page with a flash message instead of seeing a raw 500 page.
    """
    try:
        return _build_vaishnavi_excel(run_id)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        flash(f'Could not build the Vaishnavi Bonus Excel: {exc}. '
              f'Check that monthly payrolls for the FY have been finalized.', 'danger')
        return redirect(url_for('bonus.bonus_view', run_id=run_id))


def _build_vaishnavi_excel(run_id):
    """Actual builder — split out so the route handler can wrap it in a
    top-level try/except for graceful failure."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    run = BonusRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)

    rows = _compute_vaishnavi_bonus_data(run)
    months = _fy_month_keys(run)   # 12 tuples: (year, month, abbr, key)

    if not rows:
        flash('No finalized payroll data found in the FY '
              f'{run.fy_label} for this establishment. '
              'Make sure at least one monthly payroll between Apr '
              f'{run.start_year} and Mar {run.end_year} is finalized, '
              'then try again.', 'warning')
        return redirect(url_for('bonus.bonus_view', run_id=run_id))

    wb = Workbook()
    ws = wb.active
    ws.title = "Bonus Statement"

    # ── Styles ──────────────────────────────────────────────────────────
    bold        = Font(bold=True, size=10, name='Calibri')
    bold_white  = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
    body        = Font(size=9, name='Calibri')
    body_bold   = Font(bold=True, size=9, name='Calibri')
    title_font  = Font(bold=True, size=12, name='Calibri')
    thin        = Side(border_style='thin', color='94A3B8')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right       = Alignment(horizontal='right', vertical='center')
    left        = Alignment(horizontal='left', vertical='center')
    slate_fill  = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    light_fill  = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    gt_fill     = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

    # Layout: 4 KYC cols + 12 × 4 month cols + 3 grand-total cols = 55 cols
    KYC_COLS = 4
    MONTH_SUBCOLS = 4
    LAST_COL = KYC_COLS + len(months) * MONTH_SUBCOLS + 3   # 55

    # ── Row 1: Title ────────────────────────────────────────────────────
    title = (f"BONUS FOR CONTRACT WORKMEN FROM APRIL {run.start_year} "
             f"TO MARCH {run.end_year}")
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_COL)

    # ── Row 2: Establishment + meta line (single line summarising every
    #          active setting so an auditor can read it at a glance) ────
    att_parts = []
    if run.include_holiday_attendance: att_parts.append('NPH')
    if run.att_include_ot_days:        att_parts.append('OT(÷8)')
    att_label = ' + '.join(att_parts) if att_parts else 'Worked Only'
    if run.att_skip_zero: att_label += ', SkipZero'

    if run.wage_use_full_gross:
        wage_label = 'Full Gross'
    else:
        wparts = ['Att×Rate']
        if run.wage_add_nph_wages:         wparts.append('+NPH')
        if run.include_overtime_in_wage:   wparts.append('+OT')
        if run.wage_add_other_allowance:   wparts.append('+OtherAlw')
        wage_label = ' '.join(wparts)

    cap_parts = []
    if run.wage_ceiling_per_month:  cap_parts.append(f'WageCeil ₹{int(run.wage_ceiling_per_month)}')
    if run.bonus_cap_per_employee:  cap_parts.append(f'BonusCap ₹{int(run.bonus_cap_per_employee)}')
    cap_label = ' · '.join(cap_parts) if cap_parts else 'No cap, no ceiling'

    meta = (f"{run.establishment.company_name}  |  {run.fy_label}  |  "
            f"Bonus % : {run.bonus_percentage}%  |  "
            f"Attendance : {att_label}  |  Wage : {wage_label}  |  {cap_label}")
    ws.cell(row=2, column=1, value=meta).font = Font(size=9, italic=True, color='475569', name='Calibri')
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST_COL)

    # ── Row 3: Month group headers (merged across 4 sub-columns each) ──
    # KYC cells (cols 1..4) — span both header rows so they're tall
    kyc_headers = ['UAN / ESIC', 'Name', 'Father', 'Bonus %']
    for i, h in enumerate(kyc_headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = bold_white
        c.fill = slate_fill
        c.alignment = center
        c.border = border
        ws.merge_cells(start_row=3, start_column=i, end_row=4, end_column=i)

    # 12 month group headers
    col = KYC_COLS + 1
    for (yr, mth, abbr, key) in months:
        ws.cell(row=3, column=col, value=f"{abbr}-{str(yr)[-2:]}")
        ws.cell(row=3, column=col).font = bold_white
        ws.cell(row=3, column=col).fill = slate_fill
        ws.cell(row=3, column=col).alignment = center
        ws.cell(row=3, column=col).border = border
        ws.merge_cells(start_row=3, start_column=col,
                       end_row=3, end_column=col + MONTH_SUBCOLS - 1)
        col += MONTH_SUBCOLS

    # Grand total — 3 cols, single header on row 3
    ws.cell(row=3, column=col, value='GRAND TOTAL')
    ws.cell(row=3, column=col).font = bold_white
    ws.cell(row=3, column=col).fill = slate_fill
    ws.cell(row=3, column=col).alignment = center
    ws.cell(row=3, column=col).border = border
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)

    # ── Row 4: Sub-headers for each month (KYC already merged through) ─
    sub_labels = ['Attendance', 'Per day wage', 'Total Wage', 'Total Bonus']
    col = KYC_COLS + 1
    for _ in months:
        for j, lbl in enumerate(sub_labels):
            c = ws.cell(row=4, column=col + j, value=lbl)
            c.font = bold
            c.fill = light_fill
            c.alignment = center
            c.border = border
        col += MONTH_SUBCOLS
    # Grand total sub-headers
    for j, lbl in enumerate(['Total Attendance', 'Total Wage', 'Total Bonus']):
        c = ws.cell(row=4, column=col + j, value=lbl)
        c.font = bold
        c.fill = gt_fill
        c.alignment = center
        c.border = border

    # Row heights
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 32

    # ── Data rows ──────────────────────────────────────────────────────
    r = 5
    gt_attendance = gt_wage = gt_bonus = 0
    for row_data in rows:
        emp = row_data['employee']
        # UAN if available, else ESIC IP, else "—"
        primary_id = (emp.uan_number or emp.esic_ip_number or '—')

        # KYC columns
        for col_idx, val, align in [
            (1, primary_id,                left),
            (2, emp.name,                  left),
            (3, emp.father_husband_name,   left),
            (4, f"{run.bonus_percentage}%", center),
        ]:
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = body
            c.alignment = align
            c.border = border

        # Monthly data — 4 sub-columns per month
        col = KYC_COLS + 1
        for (yr, mth, abbr, key) in months:
            m = row_data['monthly'].get(key, {})
            cells = [
                m.get('attendance', '') or '',
                m.get('daily_rate', '') or '',
                m.get('monthly_wage', '') or '',
                m.get('monthly_bonus', '') or '',
            ]
            for j, val in enumerate(cells):
                c = ws.cell(row=r, column=col + j, value=val)
                c.font = body
                c.alignment = center if j == 0 else right
                c.border = border
                if isinstance(val, (int, float)) and val != 0 and j > 0:
                    c.number_format = '#,##0'
            col += MONTH_SUBCOLS

        # Grand totals — bold + green fill
        for j, val in enumerate([row_data['total_attendance'],
                                 row_data['total_wage'],
                                 row_data['total_bonus']]):
            c = ws.cell(row=r, column=col + j, value=val)
            c.font = body_bold
            c.alignment = right if j > 0 else center
            c.border = border
            c.fill = gt_fill
            if j > 0:
                c.number_format = '#,##0'

        # Flag ineligible rows by italicising the name
        if not row_data['is_eligible']:
            ws.cell(row=r, column=2).font = Font(size=9, italic=True, color='B45309', name='Calibri')
            ws.cell(row=r, column=2).value = f"{emp.name}  (< {run.min_days_worked} days)"

        gt_attendance += row_data['total_attendance']
        gt_wage       += row_data['total_wage']
        gt_bonus      += row_data['total_bonus']
        r += 1

    # ── Establishment totals row ───────────────────────────────────────
    totals_label_cell = ws.cell(row=r, column=1, value=f"TOTAL ({len(rows)} employees)")
    totals_label_cell.font = bold_white
    totals_label_cell.fill = slate_fill
    totals_label_cell.alignment = center
    totals_label_cell.border = border
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=KYC_COLS + len(months) * MONTH_SUBCOLS)

    col = KYC_COLS + len(months) * MONTH_SUBCOLS + 1
    for j, val in enumerate([gt_attendance, gt_wage, gt_bonus]):
        c = ws.cell(row=r, column=col + j, value=val)
        c.font = bold
        c.alignment = right if j > 0 else center
        c.border = border
        c.fill = gt_fill
        if j > 0:
            c.number_format = '#,##0'
    ws.row_dimensions[r].height = 22

    # ── Column widths ───────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 16  # UAN/ESIC
    ws.column_dimensions['B'].width = 22  # Name
    ws.column_dimensions['C'].width = 18  # Father
    ws.column_dimensions['D'].width = 9   # Bonus %
    col = KYC_COLS + 1
    for _ in months:
        ws.column_dimensions[get_column_letter(col)].width     = 8   # Attendance
        ws.column_dimensions[get_column_letter(col + 1)].width = 9   # Per day wage
        ws.column_dimensions[get_column_letter(col + 2)].width = 10  # Total Wage
        ws.column_dimensions[get_column_letter(col + 3)].width = 10  # Total Bonus
        col += MONTH_SUBCOLS
    ws.column_dimensions[get_column_letter(col)].width     = 10  # GT Attendance
    ws.column_dimensions[get_column_letter(col + 1)].width = 12  # GT Wage
    ws.column_dimensions[get_column_letter(col + 2)].width = 12  # GT Bonus

    # Freeze panes — keep KYC cols + 4 header rows in view while scrolling
    ws.freeze_panes = 'E5'

    # Print setup — Legal landscape, fit to one page wide
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5  # 5 = Legal in openpyxl
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_name = (run.establishment.company_name or 'Establishment').replace(' ', '_').replace('/', '_')[:60]
    filename = f"Bonus_{safe_name}_{run.fy_label.replace(' ', '_')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
