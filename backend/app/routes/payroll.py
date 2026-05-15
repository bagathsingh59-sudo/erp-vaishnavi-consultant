from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app import db
from app.models.payroll import (PayrollConfig, SalaryHead, EmployeeSalary,
                                 EmployeeSalaryHead, MonthlyPayroll, PayrollEntry,
                                 PayrollEntryHead, SalaryTemplate, SalaryTemplateHead)
from app.models.establishment import Establishment
from app.models.employee import Employee
from datetime import datetime, date
import calendar
import math
from app.user_context import (current_user_id, is_admin, user_establishments,
                               verify_est_ownership, get_user_est_ids, log_activity,
                               capture_est_from_url)
from app.utils.date_helpers import current_wage_month

payroll_bp = Blueprint('payroll', __name__)


# Role-agnostic hook: let ?establishment=X in URL restore session when lost.
# Works identically for admin and user — no role branches.
@payroll_bp.before_request
def _capture_url_establishment():
    if request.path and '/api/' in request.path:
        return None
    capture_est_from_url()
    return None


# =============================================
# PAYROLL CONFIGURATION (per Establishment)
# =============================================

@payroll_bp.route('/establishments/<int:est_id>/payroll-config', methods=['GET', 'POST'])
def payroll_config(est_id):
    """Configure payroll settings for an establishment"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est_id).first()

    if request.method == 'POST':
        if not config:
            config = PayrollConfig(establishment_id=est_id)
            db.session.add(config)

        # Salary Type & Structure
        config.salary_type = request.form.get('salary_type', 'monthly_fixed')
        config.salary_structure = request.form.get('salary_structure', 'with_heads')

        # Working Days
        config.working_days_basis = request.form.get('working_days_basis', 'calendar')
        if config.working_days_basis == 'custom':
            try:
                config.custom_working_days = int(request.form.get('custom_working_days', 26))
            except ValueError:
                config.custom_working_days = 26
        else:
            config.custom_working_days = None

        # Compliance Basis
        config.compliance_basis = request.form.get('compliance_basis', 'basic_da')
        config.include_ot_in_epf = 'include_ot_in_epf' in request.form
        config.include_ot_in_esic = 'include_ot_in_esic' in request.form
        # Keep legacy field in sync (OR of both)
        config.include_ot_in_compliance = config.include_ot_in_epf or config.include_ot_in_esic

        # Absence
        config.absence_deduction = 'absence_deduction' in request.form

        # OT
        config.ot_applicable = 'ot_applicable' in request.form
        if config.ot_applicable:
            config.ot_rate_type = request.form.get('ot_rate_type', 'double')
            config.ot_unit = request.form.get('ot_unit', 'hours')
            config.ot_base_wage = request.form.get('ot_base_wage', 'gross')
        else:
            config.ot_rate_type = None
            config.ot_unit = None
            config.ot_base_wage = 'gross'

        # Rest Day
        config.rest_day_type = request.form.get('rest_day_type', 'sunday')
        if config.rest_day_type == 'fixed_day':
            try:
                config.rest_day_weekday = int(request.form.get('rest_day_weekday', 6))
            except ValueError:
                config.rest_day_weekday = 6
        elif config.rest_day_type == 'sunday':
            config.rest_day_weekday = 6

        # Weekly Off (WO) Policy — all salary types
        config.wo_applicable = request.form.get('wo_applicable', '1') == '1'
        config.wo_type = request.form.get('wo_type', 'paid')
        config.wo_day = request.form.get('wo_day', 'sunday')
        config.wo_sandwich_rule = request.form.get('wo_sandwich_rule', '0') == '1'
        config.absence_divisor = request.form.get('absence_divisor', '30')
        try:
            config.wo_ot_rate = float(request.form.get('wo_ot_rate', 2.0))
        except ValueError:
            config.wo_ot_rate = 2.0
        # Keep deprecated field in sync
        config.weekly_off_policy = config.wo_type

        # Paid Holiday
        config.paid_holiday_type = request.form.get('paid_holiday_type', 'included')

        # Statutory
        config.epf_applicable = 'epf_applicable' in request.form
        config.esic_applicable = 'esic_applicable' in request.form
        config.pt_applicable = 'pt_applicable' in request.form
        if config.pt_applicable:
            config.pt_state = request.form.get('pt_state', 'karnataka')

        # EPF Rates
        if config.epf_applicable:
            config.epf_contribution_type = request.form.get('epf_contribution_type', 'ceiling')
            try:
                config.epf_employee_rate = float(request.form.get('epf_employee_rate', 12.0))
                config.epf_ac01_rate = float(request.form.get('epf_ac01_rate', 3.67))
                config.epf_eps_rate = float(request.form.get('epf_eps_rate', 8.33))
                config.epf_admin_rate = float(request.form.get('epf_admin_rate', 0.50))
                config.epf_edli_rate = float(request.form.get('epf_edli_rate', 0.50))
                config.epf_admin_min = float(request.form.get('epf_admin_min', 500))
                config.epf_wage_ceiling = float(request.form.get('epf_wage_ceiling', 15000))
            except ValueError:
                pass
            config.epf_employer_in_ctc = 'epf_employer_in_ctc' in request.form

        # ESIC Rates
        if config.esic_applicable:
            config.esic_contribution_type = request.form.get('esic_contribution_type', 'ceiling')
            try:
                config.esic_employer_rate = float(request.form.get('esic_employer_rate', 3.25))
                config.esic_employee_rate = float(request.form.get('esic_employee_rate', 0.75))
                config.esic_wage_ceiling = float(request.form.get('esic_wage_ceiling', 21000))
            except ValueError:
                pass

        db.session.commit()

        # Auto-create default salary heads if structure is with_heads and no heads exist
        if config.salary_structure == 'with_heads':
            existing = SalaryHead.query.filter_by(establishment_id=est_id).count()
            if existing == 0:
                _create_default_salary_heads(est_id)

        # If in setup wizard mode, redirect to establishment view with setup=complete
        setup_mode = request.args.get('setup') or request.form.get('setup_mode')
        if setup_mode:
            flash(f'Step 2 Complete! Payroll configured. Your establishment is ready to use.', 'success')
            return redirect(url_for('establishment.establishment_view', id=est_id, setup='complete'))
        else:
            flash(f'Payroll configuration saved for "{est.company_name}"!', 'success')
            return redirect(url_for('payroll.payroll_config', est_id=est_id))

    setup_mode = request.args.get('setup', '')
    return render_template('payroll/config.html', est=est, config=config, setup_mode=setup_mode)


def _create_default_salary_heads(est_id):
    """Create standard Indian salary heads for an establishment"""
    # (name, code, type, calc, pv, poh, compliance, esic_exclude, in_gross, order)
    defaults = [
        ('Basic', 'BASIC', 'earning', 'fixed', None, None, True, False, True, 1),
        ('Dearness Allowance', 'DA', 'earning', 'fixed', None, None, True, False, True, 2),
        ('House Rent Allowance', 'HRA', 'earning', 'fixed', None, None, False, False, True, 3),
        ('Conveyance Allowance', 'CONV', 'earning', 'fixed', None, None, False, False, True, 4),
        ('Other Allowance', 'OTH_ALW', 'earning', 'fixed', None, None, False, False, True, 5),
        ('Wash Allowance', 'WASH', 'earning', 'fixed', None, None, False, True, True, 6),
    ]

    for name, code, htype, ctype, pv, poh, compliance, esic_excl, in_gross, order in defaults:
        head = SalaryHead(
            establishment_id=est_id,
            name=name,
            short_code=code,
            head_type=htype,
            calc_type=ctype,
            percent_value=pv,
            percent_of_head_id=poh,
            is_for_compliance=compliance,
            exclude_from_esic=esic_excl,
            is_in_gross=in_gross,
            display_order=order
        )
        db.session.add(head)

    db.session.commit()


# =============================================
# SALARY HEADS MANAGEMENT
# =============================================

@payroll_bp.route('/establishments/<int:est_id>/salary-heads')
def salary_heads(est_id):
    """View and manage salary heads for an establishment"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est_id).first()
    heads = SalaryHead.query.filter_by(establishment_id=est_id).order_by(SalaryHead.display_order).all()
    return render_template('payroll/salary_heads.html', est=est, config=config, heads=heads)


@payroll_bp.route('/establishments/<int:est_id>/salary-heads/add', methods=['GET', 'POST'])
def salary_head_add(est_id):
    """Add a new salary head"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)

    if request.method == 'POST':
        max_order = db.session.query(db.func.max(SalaryHead.display_order)).filter_by(
            establishment_id=est_id).scalar() or 0

        head = SalaryHead(
            establishment_id=est_id,
            name=request.form['name'].strip(),
            short_code=request.form['short_code'].strip().upper(),
            head_type=request.form.get('head_type', 'earning'),
            calc_type=request.form.get('calc_type', 'fixed'),
            is_for_compliance='is_for_compliance' in request.form,
            is_in_gross='is_in_gross' in request.form,
            exclude_from_esic='exclude_from_esic' in request.form,
            display_order=max_order + 1
        )

        if head.calc_type == 'percent':
            try:
                head.percent_value = float(request.form.get('percent_value', 0))
                head.percent_of_head_id = int(request.form.get('percent_of_head_id')) if request.form.get('percent_of_head_id') else None
            except (ValueError, TypeError):
                pass

        db.session.add(head)
        db.session.commit()
        flash(f'Salary head "{head.name}" added!', 'success')
        return redirect(url_for('payroll.salary_heads', est_id=est_id))

    existing_heads = SalaryHead.query.filter_by(establishment_id=est_id, head_type='earning').all()
    return render_template('payroll/salary_head_form.html', est=est, head=None, mode='add',
                           existing_heads=existing_heads)


@payroll_bp.route('/establishments/<int:est_id>/salary-heads/<int:head_id>/edit', methods=['GET', 'POST'])
def salary_head_edit(est_id, head_id):
    """Edit a salary head"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    head = SalaryHead.query.get_or_404(head_id)

    if head.establishment_id != est_id:
        flash('Invalid salary head.', 'danger')
        return redirect(url_for('payroll.salary_heads', est_id=est_id))

    if request.method == 'POST':
        head.name = request.form['name'].strip()
        head.short_code = request.form['short_code'].strip().upper()
        head.head_type = request.form.get('head_type', 'earning')
        head.calc_type = request.form.get('calc_type', 'fixed')
        head.is_for_compliance = 'is_for_compliance' in request.form
        head.is_in_gross = 'is_in_gross' in request.form
        head.exclude_from_esic = 'exclude_from_esic' in request.form

        if head.calc_type == 'percent':
            try:
                head.percent_value = float(request.form.get('percent_value', 0))
                head.percent_of_head_id = int(request.form.get('percent_of_head_id')) if request.form.get('percent_of_head_id') else None
            except (ValueError, TypeError):
                pass
        else:
            head.percent_value = None
            head.percent_of_head_id = None

        db.session.commit()
        flash(f'Salary head "{head.name}" updated!', 'success')
        return redirect(url_for('payroll.salary_heads', est_id=est_id))

    existing_heads = SalaryHead.query.filter(
        SalaryHead.establishment_id == est_id,
        SalaryHead.head_type == 'earning',
        SalaryHead.id != head_id
    ).all()
    return render_template('payroll/salary_head_form.html', est=est, head=head, mode='edit',
                           existing_heads=existing_heads)


@payroll_bp.route('/establishments/<int:est_id>/salary-heads/<int:head_id>/delete', methods=['POST'])
def salary_head_delete(est_id, head_id):
    """Delete a salary head"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    head = SalaryHead.query.get_or_404(head_id)
    if head.establishment_id != est_id:
        flash('Invalid salary head.', 'danger')
        return redirect(url_for('payroll.salary_heads', est_id=est_id))

    name = head.name
    db.session.delete(head)
    db.session.commit()
    flash(f'Salary head "{name}" removed.', 'warning')
    return redirect(url_for('payroll.salary_heads', est_id=est_id))


@payroll_bp.route('/establishments/<int:est_id>/salary-heads/<int:head_id>/toggle', methods=['POST'])
def salary_head_toggle(est_id, head_id):
    """Toggle salary head active/inactive"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    head = SalaryHead.query.get_or_404(head_id)
    if head.establishment_id != est_id:
        flash('Invalid salary head.', 'danger')
        return redirect(url_for('payroll.salary_heads', est_id=est_id))

    head.is_active = not head.is_active
    db.session.commit()
    status = 'activated' if head.is_active else 'deactivated'
    flash(f'Salary head "{head.name}" {status}.', 'success')
    return redirect(url_for('payroll.salary_heads', est_id=est_id))


# =============================================
# EMPLOYEE SALARY ASSIGNMENT
# =============================================

@payroll_bp.route('/employees/<int:emp_id>/salary', methods=['GET', 'POST'])
def employee_salary(emp_id):
    """Assign or update salary for an employee"""
    employee = Employee.query.get_or_404(emp_id)
    est = employee.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    if not config:
        flash('Please configure payroll settings for this establishment first.', 'warning')
        return redirect(url_for('payroll.payroll_config', est_id=est.id))

    # Get current salary
    current_salary = EmployeeSalary.query.filter_by(
        employee_id=emp_id, is_current=True).first()

    # Get active salary heads for this establishment
    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    if request.method == 'POST':
        # Parse effective date
        try:
            effective_from = datetime.strptime(request.form['effective_from'], '%Y-%m-%d').date()
        except (ValueError, KeyError):
            flash('Please enter a valid effective date.', 'danger')
            return render_template('payroll/employee_salary.html',
                                   emp=employee, config=config, salary=current_salary, heads=heads)

        # Mark old salary as not current
        if current_salary:
            current_salary.is_current = False

        # Capture revision reason
        revision_reason = request.form.get('revision_reason', '').strip() or None

        # Per-employee salary type (falls back to config default if not set)
        raw_salary_type = request.form.get('salary_type', '') or config.salary_type
        # Both monthly_heads and monthly_package use head-wise breakup
        use_heads = (raw_salary_type in ('monthly_heads', 'monthly_package'))
        emp_salary_type = 'monthly_fixed' if raw_salary_type == 'monthly_heads' else raw_salary_type
        emp_wo_policy = request.form.get('weekly_off_policy', '') or (getattr(config, 'weekly_off_policy', 'paid') or 'paid')

        # Track which template was used (if any)
        applied_tmpl_id = request.form.get('salary_template_id', '') or None
        if applied_tmpl_id:
            try:
                applied_tmpl_id = int(applied_tmpl_id)
            except ValueError:
                applied_tmpl_id = None

        # Create new salary record
        new_salary = EmployeeSalary(
            employee_id=emp_id,
            effective_from=effective_from,
            is_current=True,
            revision_reason=revision_reason,
            salary_type=emp_salary_type,
            weekly_off_policy=emp_wo_policy if emp_salary_type == 'daily_wages' else None,
            salary_template_id=applied_tmpl_id,
        )

        if emp_salary_type == 'daily_wages':
            try:
                new_salary.daily_rate = max(0, float(request.form.get('daily_rate', 0)))
            except ValueError:
                new_salary.daily_rate = 0
        else:
            # Fixed salary option: no deduction for absence (monthly fixed employees)
            new_salary.no_absence_deduction = 'no_absence_deduction' in request.form

        if use_heads and heads:
            # Head-wise entry (user selected "Monthly with Heads" slab)
            total_gross = 0
            db.session.add(new_salary)
            db.session.flush()  # Get the ID

            for head in heads:
                try:
                    amount = max(0, float(request.form.get(f'head_{head.id}', 0)))
                except ValueError:
                    amount = 0

                head_value = EmployeeSalaryHead(
                    employee_salary_id=new_salary.id,
                    salary_head_id=head.id,
                    amount=amount
                )
                db.session.add(head_value)

                if head.is_in_gross and head.head_type == 'earning':
                    total_gross += amount

            new_salary.gross_salary = total_gross
        elif emp_salary_type != 'daily_wages':
            # Gross-only (monthly_fixed or monthly_package without heads)
            try:
                new_salary.gross_salary = max(0, float(request.form.get('gross_salary', 0)))
            except ValueError:
                new_salary.gross_salary = 0

        if not new_salary.id:
            db.session.add(new_salary)

        # Save per-employee statutory overrides (EPF/ESIC checkboxes)
        if config.esic_applicable:
            esic_checked = 'esic_for_employee' in request.form
            employee.esic_exempt = not esic_checked
            if employee.esic_exempt:
                employee.esic_exemption_reason = request.form.get('esic_exemption_reason', '').strip() or None
            else:
                employee.esic_exemption_reason = None

        log_activity('updated', 'salary', entity_id=emp_id,
                     entity_name=employee.name,
                     details=f'Gross: ₹{new_salary.gross_salary:,.0f}, Effective: {effective_from}, Reason: {revision_reason or "N/A"}',
                     establishment_id=est.id)
        db.session.commit()

        # If in setup wizard mode, redirect to employee view
        setup_mode = request.args.get('setup') or request.form.get('setup_mode')
        if setup_mode:
            flash(f'Setup Complete! Salary assigned for "{employee.name}". Employee is ready.', 'success')
            return redirect(url_for('employee.employee_view', id=emp_id, setup='complete'))
        else:
            flash(f'Salary updated for "{employee.name}"!', 'success')
            return redirect(url_for('payroll.employee_salary', emp_id=emp_id))

    # Get head values for current salary
    head_values = {}
    if current_salary and current_salary.head_values:
        for hv in current_salary.head_values:
            head_values[hv.salary_head_id] = hv.amount

    # Get available salary templates for this establishment
    salary_templates = SalaryTemplate.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryTemplate.name).all()

    setup_mode = request.args.get('setup', '')
    return render_template('payroll/employee_salary.html',
                           emp=employee, config=config, salary=current_salary,
                           heads=heads, head_values=head_values,
                           setup_mode=setup_mode, salary_templates=salary_templates,
                           today=date.today())


# =============================================
# SALARY TEMPLATES (Reusable Presets)
# =============================================

@payroll_bp.route('/establishment/<int:est_id>/salary-templates')
def salary_template_list(est_id):
    """List all salary templates for an establishment"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    templates = SalaryTemplate.query.filter_by(
        establishment_id=est.id
    ).order_by(SalaryTemplate.name).all()

    # Count employees using each template
    for tmpl in templates:
        tmpl._emp_count = EmployeeSalary.query.filter_by(
            salary_template_id=tmpl.id, is_current=True
        ).count()

    return render_template('payroll/salary_templates.html',
                           est=est, config=config, templates=templates)


@payroll_bp.route('/establishment/<int:est_id>/salary-template/add', methods=['GET', 'POST'])
def salary_template_add(est_id):
    """Create a new salary template"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    if not config:
        flash('Please configure payroll settings first.', 'warning')
        return redirect(url_for('payroll.payroll_config', est_id=est.id))

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    if request.method == 'POST':
        name = request.form.get('template_name', '').strip()
        if not name:
            flash('Template name is required.', 'danger')
            return render_template('payroll/salary_template_form.html',
                                   est=est, config=config, heads=heads, template=None)

        # Check duplicate name
        existing = SalaryTemplate.query.filter(
            SalaryTemplate.establishment_id == est.id,
            db.func.lower(SalaryTemplate.name) == name.lower()
        ).first()
        if existing:
            flash(f'A template named "{name}" already exists.', 'danger')
            return render_template('payroll/salary_template_form.html',
                                   est=est, config=config, heads=heads, template=None)

        salary_type = request.form.get('salary_type', config.salary_type)
        wo_policy = request.form.get('weekly_off_policy', '') or None

        tmpl = SalaryTemplate(
            establishment_id=est.id,
            name=name,
            salary_type=salary_type,
            weekly_off_policy=wo_policy if salary_type == 'daily_wages' else None,
        )

        # Daily rate
        if salary_type == 'daily_wages':
            try:
                tmpl.daily_rate = max(0, float(request.form.get('daily_rate', 0)))
            except ValueError:
                tmpl.daily_rate = 0

        # Head-wise amounts
        if config.salary_structure == 'gross_only':
            try:
                tmpl.gross_salary = max(0, float(request.form.get('gross_salary', 0)))
            except ValueError:
                tmpl.gross_salary = 0
        else:
            db.session.add(tmpl)
            db.session.flush()

            total_gross = 0
            for head in heads:
                try:
                    amount = max(0, float(request.form.get(f'head_{head.id}', 0)))
                except ValueError:
                    amount = 0

                th = SalaryTemplateHead(
                    salary_template_id=tmpl.id,
                    salary_head_id=head.id,
                    amount=amount
                )
                db.session.add(th)
                if head.is_in_gross and head.head_type == 'earning':
                    total_gross += amount

            tmpl.gross_salary = total_gross

        if not tmpl.id:
            db.session.add(tmpl)

        log_activity('created', 'salary_template', entity_id=tmpl.id,
                     entity_name=name,
                     details=f'Type: {salary_type}, Gross: ₹{tmpl.gross_salary:,.0f}',
                     establishment_id=est.id)
        db.session.commit()

        flash(f'Salary Template "{name}" created successfully!', 'success')
        return redirect(url_for('payroll.salary_template_list', est_id=est.id))

    return render_template('payroll/salary_template_form.html',
                           est=est, config=config, heads=heads, template=None, head_values={})


@payroll_bp.route('/salary-template/<int:tmpl_id>/edit', methods=['GET', 'POST'])
def salary_template_edit(tmpl_id):
    """Edit an existing salary template"""
    tmpl = SalaryTemplate.query.get_or_404(tmpl_id)
    est = tmpl.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    if request.method == 'POST':
        name = request.form.get('template_name', '').strip()
        if not name:
            flash('Template name is required.', 'danger')
            return render_template('payroll/salary_template_form.html',
                                   est=est, config=config, heads=heads, template=tmpl)

        # Check duplicate name (exclude self)
        existing = SalaryTemplate.query.filter(
            SalaryTemplate.establishment_id == est.id,
            db.func.lower(SalaryTemplate.name) == name.lower(),
            SalaryTemplate.id != tmpl.id
        ).first()
        if existing:
            flash(f'A template named "{name}" already exists.', 'danger')
            return render_template('payroll/salary_template_form.html',
                                   est=est, config=config, heads=heads, template=tmpl)

        tmpl.name = name
        tmpl.salary_type = request.form.get('salary_type', config.salary_type)
        tmpl.weekly_off_policy = request.form.get('weekly_off_policy', '') or None
        if tmpl.salary_type != 'daily_wages':
            tmpl.weekly_off_policy = None

        if tmpl.salary_type == 'daily_wages':
            try:
                tmpl.daily_rate = max(0, float(request.form.get('daily_rate', 0)))
            except ValueError:
                tmpl.daily_rate = 0

        if config.salary_structure == 'gross_only':
            try:
                tmpl.gross_salary = max(0, float(request.form.get('gross_salary', 0)))
            except ValueError:
                tmpl.gross_salary = 0
        else:
            # Delete old head values and re-create
            SalaryTemplateHead.query.filter_by(salary_template_id=tmpl.id).delete()
            total_gross = 0
            for head in heads:
                try:
                    amount = max(0, float(request.form.get(f'head_{head.id}', 0)))
                except ValueError:
                    amount = 0

                th = SalaryTemplateHead(
                    salary_template_id=tmpl.id,
                    salary_head_id=head.id,
                    amount=amount
                )
                db.session.add(th)
                if head.is_in_gross and head.head_type == 'earning':
                    total_gross += amount

            tmpl.gross_salary = total_gross

        log_activity('updated', 'salary_template', entity_id=tmpl.id,
                     entity_name=name,
                     details=f'Type: {tmpl.salary_type}, Gross: ₹{tmpl.gross_salary:,.0f}',
                     establishment_id=est.id)
        db.session.commit()

        flash(f'Salary Template "{name}" updated!', 'success')
        return redirect(url_for('payroll.salary_template_list', est_id=est.id))

    # Load head values for editing
    head_values = {}
    for th in tmpl.head_values:
        head_values[th.salary_head_id] = th.amount

    return render_template('payroll/salary_template_form.html',
                           est=est, config=config, heads=heads,
                           template=tmpl, head_values=head_values)


@payroll_bp.route('/salary-template/<int:tmpl_id>/delete', methods=['POST'])
def salary_template_delete(tmpl_id):
    """Delete a salary template"""
    tmpl = SalaryTemplate.query.get_or_404(tmpl_id)
    est = tmpl.establishment
    verify_est_ownership(est)

    # Check if any current employees use this template
    emp_count = EmployeeSalary.query.filter_by(
        salary_template_id=tmpl.id, is_current=True
    ).count()

    if emp_count > 0:
        flash(f'Cannot delete "{tmpl.name}" — {emp_count} employee(s) are using this template. '
              f'Please reassign their salary first.', 'danger')
        return redirect(url_for('payroll.salary_template_list', est_id=est.id))

    name = tmpl.name
    log_activity('deleted', 'salary_template', entity_id=tmpl.id,
                 entity_name=name, establishment_id=est.id)
    db.session.delete(tmpl)
    db.session.commit()

    flash(f'Template "{name}" deleted.', 'success')
    return redirect(url_for('payroll.salary_template_list', est_id=est.id))


@payroll_bp.route('/salary-template/<int:tmpl_id>/apply-bulk', methods=['POST'])
def salary_template_apply_bulk(tmpl_id):
    """Apply a salary template to multiple employees at once"""
    tmpl = SalaryTemplate.query.get_or_404(tmpl_id)
    est = tmpl.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    # Get template head values
    tmpl_head_values = {}
    for th in tmpl.head_values:
        tmpl_head_values[th.salary_head_id] = th.amount

    # Get selected employee IDs
    emp_ids = request.form.getlist('employee_ids')
    if not emp_ids:
        flash('No employees selected.', 'warning')
        return redirect(url_for('payroll.salary_template_list', est_id=est.id))

    effective_from_str = request.form.get('effective_from', '')
    try:
        effective_from = datetime.strptime(effective_from_str, '%Y-%m-%d').date()
    except ValueError:
        effective_from = date.today()

    applied_count = 0
    for emp_id in emp_ids:
        try:
            emp_id = int(emp_id)
        except ValueError:
            continue

        emp = Employee.query.get(emp_id)
        if not emp or emp.establishment_id != est.id:
            continue

        # Mark old salary as not current
        old_salary = EmployeeSalary.query.filter_by(
            employee_id=emp_id, is_current=True
        ).first()
        if old_salary:
            old_salary.is_current = False

        # Create new salary from template
        new_salary = EmployeeSalary(
            employee_id=emp_id,
            salary_template_id=tmpl.id,
            effective_from=effective_from,
            is_current=True,
            salary_type=tmpl.salary_type,
            weekly_off_policy=tmpl.weekly_off_policy,
            daily_rate=tmpl.daily_rate if tmpl.salary_type == 'daily_wages' else None,
            revision_reason=f'Applied template: {tmpl.name}',
        )

        if config.salary_structure == 'gross_only':
            new_salary.gross_salary = tmpl.gross_salary
            db.session.add(new_salary)
        else:
            db.session.add(new_salary)
            db.session.flush()

            total_gross = 0
            for head in heads:
                amount = tmpl_head_values.get(head.id, 0)
                esh = EmployeeSalaryHead(
                    employee_salary_id=new_salary.id,
                    salary_head_id=head.id,
                    amount=amount
                )
                db.session.add(esh)
                if head.is_in_gross and head.head_type == 'earning':
                    total_gross += amount

            new_salary.gross_salary = total_gross

        applied_count += 1

    db.session.commit()

    log_activity('applied', 'salary_template', entity_id=tmpl.id,
                 entity_name=tmpl.name,
                 details=f'Applied to {applied_count} employees, Effective: {effective_from}',
                 establishment_id=est.id)

    flash(f'Template "{tmpl.name}" applied to {applied_count} employee(s)!', 'success')
    return redirect(url_for('payroll.salary_template_list', est_id=est.id))


@payroll_bp.route('/api/salary-template/<int:tmpl_id>/data')
def salary_template_data(tmpl_id):
    """API: Return template data as JSON (for auto-fill on employee salary page)"""
    from flask import jsonify
    tmpl = SalaryTemplate.query.get_or_404(tmpl_id)
    est = tmpl.establishment
    verify_est_ownership(est)

    head_values = {}
    for th in tmpl.head_values:
        head_values[str(th.salary_head_id)] = th.amount

    return jsonify({
        'salary_type': tmpl.salary_type,
        'gross_salary': tmpl.gross_salary,
        'daily_rate': tmpl.daily_rate,
        'weekly_off_policy': tmpl.weekly_off_policy or '',
        'head_values': head_values,
    })


# =============================================
# BULK SALARY REVISION (Phase 2)
# =============================================

@payroll_bp.route('/establishment/<int:est_id>/salary-revision', methods=['GET'])
def salary_revision(est_id):
    """Bulk Salary Revision page — revise template or apply % increment"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    templates = SalaryTemplate.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryTemplate.name).all()

    # Count employees per template
    for tmpl in templates:
        tmpl._emp_count = EmployeeSalary.query.filter_by(
            salary_template_id=tmpl.id, is_current=True
        ).count()

    # Get employees without any template (manual salary)
    all_emp_ids_with_template = db.session.query(EmployeeSalary.employee_id).filter(
        EmployeeSalary.is_current == True,
        EmployeeSalary.salary_template_id.isnot(None)
    ).subquery()

    from app.models.employee import Employee as Emp
    manual_emps = Emp.query.filter(
        Emp.establishment_id == est.id,
        Emp.is_active == True,
        ~Emp.id.in_(db.session.query(all_emp_ids_with_template.c.employee_id))
    ).order_by(Emp.name).all()

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    return render_template('payroll/salary_revision.html',
                           est=est, config=config, templates=templates,
                           manual_emps=manual_emps, heads=heads)


@payroll_bp.route('/salary-template/<int:tmpl_id>/revise', methods=['GET', 'POST'])
def salary_template_revise(tmpl_id):
    """Revise a template and apply new rates to all linked employees"""
    tmpl = SalaryTemplate.query.get_or_404(tmpl_id)
    est = tmpl.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    # Current template head values
    old_head_values = {}
    for th in tmpl.head_values:
        old_head_values[th.salary_head_id] = th.amount

    # Employees currently using this template
    linked_salaries = EmployeeSalary.query.filter_by(
        salary_template_id=tmpl.id, is_current=True
    ).all()

    linked_employees = []
    for sal in linked_salaries:
        emp = Employee.query.get(sal.employee_id)
        if emp and emp.is_active:
            linked_employees.append({
                'id': emp.id,
                'name': emp.name,
                'emp_code': emp.primary_id,
                'uan_number': emp.uan_number,
                'esic_ip_number': emp.esic_ip_number,
                'current_gross': sal.gross_salary,
                'current_daily_rate': sal.daily_rate,
            })

    if request.method == 'POST':
        action = request.form.get('action', 'preview')

        # Parse new values
        new_daily_rate = 0
        if tmpl.salary_type == 'daily_wages':
            try:
                new_daily_rate = max(0, float(request.form.get('daily_rate', 0)))
            except ValueError:
                new_daily_rate = tmpl.daily_rate or 0

        new_head_values = {}
        new_gross = 0
        if config.salary_structure == 'gross_only':
            try:
                new_gross = max(0, float(request.form.get('gross_salary', 0)))
            except ValueError:
                new_gross = tmpl.gross_salary
        else:
            for head in heads:
                try:
                    amt = max(0, float(request.form.get(f'head_{head.id}', 0)))
                except ValueError:
                    amt = old_head_values.get(head.id, 0)
                new_head_values[head.id] = amt
                if head.is_in_gross and head.head_type == 'earning':
                    new_gross += amt

        # Parse effective date
        effective_str = request.form.get('effective_from', '')
        try:
            effective_from = datetime.strptime(effective_str, '%Y-%m-%d').date()
        except ValueError:
            effective_from = date.today()

        revision_reason = request.form.get('revision_reason', 'Minimum Wage Revision')

        if action == 'preview':
            # Build preview data
            preview = []
            for emp_info in linked_employees:
                change = {
                    'name': emp_info['name'],
                    'emp_code': emp_info['emp_code'],
                    'old_gross': emp_info['current_gross'],
                    'new_gross': new_gross,
                    'old_daily_rate': emp_info['current_daily_rate'],
                    'new_daily_rate': new_daily_rate,
                    'diff': new_gross - emp_info['current_gross'],
                }
                preview.append(change)

            # Build head comparison
            head_comparison = []
            for head in heads:
                old_amt = old_head_values.get(head.id, 0)
                new_amt = new_head_values.get(head.id, old_amt)
                head_comparison.append({
                    'name': head.name,
                    'code': head.short_code,
                    'old': old_amt,
                    'new': new_amt,
                    'diff': new_amt - old_amt,
                })

            return render_template('payroll/salary_revision_preview.html',
                                   est=est, config=config, tmpl=tmpl, heads=heads,
                                   preview=preview, head_comparison=head_comparison,
                                   new_gross=new_gross, new_daily_rate=new_daily_rate,
                                   new_head_values=new_head_values,
                                   effective_from=effective_from,
                                   revision_reason=revision_reason,
                                   emp_count=len(linked_employees))

        elif action == 'confirm':
            # 1. Update the template itself
            if tmpl.salary_type == 'daily_wages':
                tmpl.daily_rate = new_daily_rate

            if config.salary_structure == 'gross_only':
                tmpl.gross_salary = new_gross
            else:
                # Update template heads
                SalaryTemplateHead.query.filter_by(salary_template_id=tmpl.id).delete()
                total_gross = 0
                for head in heads:
                    amt = float(request.form.get(f'new_head_{head.id}', 0))
                    th = SalaryTemplateHead(
                        salary_template_id=tmpl.id,
                        salary_head_id=head.id,
                        amount=amt
                    )
                    db.session.add(th)
                    if head.is_in_gross and head.head_type == 'earning':
                        total_gross += amt
                tmpl.gross_salary = total_gross
                new_gross = total_gross
                # Rebuild new_head_values from hidden fields
                new_head_values = {}
                for head in heads:
                    new_head_values[head.id] = float(request.form.get(f'new_head_{head.id}', 0))

            # 2. Create new salary records for all linked employees
            revised_count = 0
            for sal in linked_salaries:
                emp = Employee.query.get(sal.employee_id)
                if not emp or not emp.is_active:
                    continue

                # Mark old as not current
                sal.is_current = False

                # Create new salary
                new_sal = EmployeeSalary(
                    employee_id=emp.id,
                    salary_template_id=tmpl.id,
                    effective_from=effective_from,
                    is_current=True,
                    salary_type=tmpl.salary_type,
                    weekly_off_policy=tmpl.weekly_off_policy,
                    daily_rate=new_daily_rate if tmpl.salary_type == 'daily_wages' else None,
                    revision_reason=revision_reason,
                )

                if config.salary_structure == 'gross_only':
                    new_sal.gross_salary = new_gross
                    db.session.add(new_sal)
                else:
                    db.session.add(new_sal)
                    db.session.flush()

                    calc_gross = 0
                    for head in heads:
                        amt = new_head_values.get(head.id, 0)
                        esh = EmployeeSalaryHead(
                            employee_salary_id=new_sal.id,
                            salary_head_id=head.id,
                            amount=amt
                        )
                        db.session.add(esh)
                        if head.is_in_gross and head.head_type == 'earning':
                            calc_gross += amt
                    new_sal.gross_salary = calc_gross

                revised_count += 1

            log_activity('revised', 'salary_template', entity_id=tmpl.id,
                         entity_name=tmpl.name,
                         details=f'Revised: {revised_count} employees, New Gross: ₹{new_gross:,.0f}, Effective: {effective_from}, Reason: {revision_reason}',
                         establishment_id=est.id)
            db.session.commit()

            flash(f'Template "{tmpl.name}" revised! {revised_count} employee(s) updated with new salary effective {effective_from.strftime("%d %b %Y")}.', 'success')
            return redirect(url_for('payroll.salary_template_list', est_id=est.id))

    return render_template('payroll/salary_revision_form.html',
                           est=est, config=config, tmpl=tmpl, heads=heads,
                           head_values=old_head_values,
                           linked_employees=linked_employees)


@payroll_bp.route('/establishment/<int:est_id>/salary-revision/percentage', methods=['POST'])
def salary_revision_percentage(est_id):
    """Apply percentage increment to selected employees"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True
    ).order_by(SalaryHead.display_order).all()

    try:
        pct = float(request.form.get('percentage', 0))
    except ValueError:
        pct = 0

    if pct <= 0 or pct > 100:
        flash('Please enter a valid percentage between 0 and 100.', 'danger')
        return redirect(url_for('payroll.salary_revision', est_id=est.id))

    emp_ids = request.form.getlist('employee_ids')
    if not emp_ids:
        flash('No employees selected.', 'warning')
        return redirect(url_for('payroll.salary_revision', est_id=est.id))

    effective_str = request.form.get('effective_from', '')
    try:
        effective_from = datetime.strptime(effective_str, '%Y-%m-%d').date()
    except ValueError:
        effective_from = date.today()

    revision_reason = request.form.get('revision_reason', f'{pct}% Increment')
    multiplier = 1 + (pct / 100)

    revised_count = 0
    for emp_id_str in emp_ids:
        try:
            emp_id = int(emp_id_str)
        except ValueError:
            continue

        emp = Employee.query.get(emp_id)
        if not emp or emp.establishment_id != est.id or not emp.is_active:
            continue

        old_sal = EmployeeSalary.query.filter_by(
            employee_id=emp_id, is_current=True
        ).first()
        if not old_sal:
            continue

        # Mark old as not current
        old_sal.is_current = False

        # Create new salary with percentage increase
        new_sal = EmployeeSalary(
            employee_id=emp_id,
            salary_template_id=old_sal.salary_template_id,
            effective_from=effective_from,
            is_current=True,
            salary_type=old_sal.salary_type,
            weekly_off_policy=old_sal.weekly_off_policy,
            revision_reason=revision_reason,
        )

        if old_sal.daily_rate:
            new_sal.daily_rate = round(old_sal.daily_rate * multiplier)

        if config.salary_structure == 'gross_only':
            new_sal.gross_salary = round(old_sal.gross_salary * multiplier)
            db.session.add(new_sal)
        else:
            db.session.add(new_sal)
            db.session.flush()

            total_gross = 0
            # Get old head values
            old_head_vals = {}
            for hv in old_sal.head_values:
                old_head_vals[hv.salary_head_id] = hv.amount

            for head in heads:
                old_amt = old_head_vals.get(head.id, 0)
                new_amt = round(old_amt * multiplier)
                esh = EmployeeSalaryHead(
                    employee_salary_id=new_sal.id,
                    salary_head_id=head.id,
                    amount=new_amt
                )
                db.session.add(esh)
                if head.is_in_gross and head.head_type == 'earning':
                    total_gross += new_amt

            new_sal.gross_salary = total_gross

        revised_count += 1

    log_activity('bulk_increment', 'salary', entity_id=est.id,
                 entity_name=est.company_name,
                 details=f'{pct}% increment applied to {revised_count} employees, Effective: {effective_from}',
                 establishment_id=est.id)
    db.session.commit()

    flash(f'{pct}% increment applied to {revised_count} employee(s)! Effective: {effective_from.strftime("%d %b %Y")}.', 'success')
    return redirect(url_for('payroll.salary_revision', est_id=est.id))


# =============================================
# MONTHLY PAYROLL PROCESSING
# =============================================

@payroll_bp.route('/payroll')
def payroll_list():
    """List all monthly payrolls — supports FY (Apr-Mar) filtering.
    Non-admin establishment enforcement is handled by the blueprint's
    before_request guard, so by this point session is guaranteed set
    (or admin, which sees everything)."""
    # Auto-scope to selected establishment
    scoped_est_id = session.get('selected_est_id')
    est_id = request.args.get('establishment', '')
    tab = request.args.get('tab', 'payroll')

    # Financial Year logic: FY 2025-26 = April 2025 to March 2026
    now = datetime.now()
    if now.month >= 4:
        default_fy = now.year
    else:
        default_fy = now.year - 1

    selected_fy = request.args.get('fy', default_fy, type=int)

    query = MonthlyPayroll.query.filter(
        MonthlyPayroll.establishment_id.in_(get_user_est_ids())
    )

    # Establishment filter
    if scoped_est_id:
        query = query.filter_by(establishment_id=scoped_est_id)
    elif est_id:
        query = query.filter_by(establishment_id=int(est_id))

    # FY filter: April YYYY to March YYYY+1
    query = query.filter(
        db.or_(
            db.and_(MonthlyPayroll.year == selected_fy, MonthlyPayroll.month >= 4),
            db.and_(MonthlyPayroll.year == selected_fy + 1, MonthlyPayroll.month <= 3)
        )
    )

    # Order: April first (month 4) to March last (month 3 of next year)
    payrolls = query.order_by(MonthlyPayroll.year, MonthlyPayroll.month).all()

    establishments = user_establishments().filter_by(is_active=True).order_by(Establishment.company_name).all()

    # FY display string
    fy_display = f"FY {selected_fy}-{str(selected_fy + 1)[-2:]}"

    return render_template('payroll/list.html',
                           payrolls=payrolls,
                           establishments=establishments,
                           selected_est_filter=str(scoped_est_id) if scoped_est_id else est_id,
                           selected_fy=selected_fy,
                           fy_display=fy_display,
                           active_tab=tab)


@payroll_bp.route('/payroll/create', methods=['GET', 'POST'])
def payroll_create():
    """Create a new monthly payroll.
    Establishment is taken from session; URL param and form field serve as
    fallbacks when session is lost (same logic for admin and user)."""
    selected_est_id = session.get('selected_est_id')

    # Explicit fallback — URL query param or form field (role-agnostic)
    if not selected_est_id:
        url_est = request.args.get('establishment') or request.form.get('establishment_id')
        if url_est and str(url_est).isdigit():
            selected_est_id = int(url_est)
            session['selected_est_id'] = selected_est_id

    if request.method == 'POST':
        est_id = int(request.form.get('establishment_id') or selected_est_id or 0)
        if not est_id:
            flash('Please select an establishment first.', 'warning')
            return redirect(url_for('establishment.establishment_list'))
        est = Establishment.query.get_or_404(est_id)
        verify_est_ownership(est)
        month = int(request.form['month'])
        year = int(request.form['year'])

        # Check if already exists
        existing = MonthlyPayroll.query.filter_by(
            establishment_id=est_id, month=month, year=year).first()
        if existing:
            flash('Payroll for this month already exists!', 'warning')
            return redirect(url_for('payroll.payroll_process', payroll_id=existing.id))

        # Check if config exists
        config = PayrollConfig.query.filter_by(establishment_id=est_id).first()
        if not config:
            flash('Please configure payroll settings for this establishment first.', 'warning')
            return redirect(url_for('payroll.payroll_config', est_id=est_id))

        # Calculate working days
        if config.working_days_basis == 'calendar':
            working_days = calendar.monthrange(year, month)[1]
        elif config.working_days_basis == 'fixed_26':
            working_days = 26
        elif config.working_days_basis == 'fixed_30':
            working_days = 30
        elif config.working_days_basis == 'custom':
            working_days = config.custom_working_days or 26
        else:
            working_days = calendar.monthrange(year, month)[1]

        # Create payroll
        payroll = MonthlyPayroll(
            establishment_id=est_id,
            month=month,
            year=year,
            working_days=working_days,
            status='draft'
        )
        db.session.add(payroll)
        db.session.flush()

        # Get all active employees for this establishment
        # FILTER: Only include employees whose date_of_joining is on or before the last day of this payroll month
        last_day_of_month = date(year, month, calendar.monthrange(year, month)[1])
        first_day_of_month = date(year, month, 1)

        employees = Employee.query.filter(
            Employee.establishment_id == est_id,
            Employee.is_active == True,
            Employee.date_of_joining <= last_day_of_month
        ).all()

        # Optional: copy data from an older payroll batch
        copy_from_id = request.form.get('copy_from_payroll_id', type=int)
        copy_source_map = {}  # employee_id -> source PayrollEntry
        if copy_from_id:
            src_payroll = MonthlyPayroll.query.filter_by(
                id=copy_from_id, establishment_id=est_id).first()
            if src_payroll:
                src_entries = PayrollEntry.query.filter_by(
                    monthly_payroll_id=src_payroll.id).all()
                copy_source_map = {e.employee_id: e for e in src_entries}

        for emp in employees:
            # Get current salary
            salary = EmployeeSalary.query.filter_by(
                employee_id=emp.id, is_current=True).first()

            # If employee joined mid-month, default present days = remaining days only
            if emp.date_of_joining > first_day_of_month:
                # Employee joined during this month — calculate remaining days
                remaining_days = (last_day_of_month - emp.date_of_joining).days + 1
                default_present = min(remaining_days, working_days)
            else:
                default_present = working_days

            # Copy-from: override defaults with source payroll values
            src_entry = copy_source_map.get(emp.id)
            if src_entry:
                days_present = src_entry.days_present
                paid_holidays = src_entry.paid_holidays or 0
                ot_hours = src_entry.ot_hours or 0
                rate_overrides = src_entry.rate_overrides if hasattr(src_entry, 'rate_overrides') else None
            else:
                days_present = default_present
                paid_holidays = 0
                ot_hours = 0
                rate_overrides = None

            entry = PayrollEntry(
                monthly_payroll_id=payroll.id,
                employee_id=emp.id,
                days_present=days_present,
                days_absent=0,
                paid_holidays=paid_holidays,
                ot_hours=ot_hours,
                total_payable_days=days_present,
                gross_salary=salary.gross_salary if salary else 0
            )
            if hasattr(entry, 'rate_overrides'):
                entry.rate_overrides = rate_overrides
            db.session.add(entry)

        payroll.total_employees = len(employees)
        db.session.commit()

        copy_msg = ''
        if copy_source_map:
            copied_count = sum(1 for emp in employees if emp.id in copy_source_map)
            copy_msg = f' (copied data for {copied_count} employees from previous payroll)'
        flash(f'Payroll created for {calendar.month_name[month]} {year} with {len(employees)} employees.{copy_msg}', 'success')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll.id))

    now = datetime.now()

    # If an establishment is selected in session, use it directly
    if selected_est_id:
        est = Establishment.query.get(selected_est_id)
        if est:
            verify_est_ownership(est)
            previous_payrolls = MonthlyPayroll.query.filter_by(
                establishment_id=est.id
            ).order_by(MonthlyPayroll.year.desc(), MonthlyPayroll.month.desc()).all()
            # Default to WAGE MONTH (previous calendar month) — you can't
            # process contributions for a running month in advance.
            wage_year, wage_month = current_wage_month()
            return render_template('payroll/create.html',
                                   establishment=est,
                                   current_month=wage_month,
                                   current_year=wage_year,
                                   previous_payrolls=previous_payrolls)

    # No establishment selected — redirect to establishment list
    flash('Please select an establishment first to create payroll.', 'info')
    return redirect(url_for('establishment.establishment_list'))


# ═════════════════════════════════════════════════════════════════
#  CREATE NIL PAYROLL — for months with no work / no employees
# ═════════════════════════════════════════════════════════════════
@payroll_bp.route('/payroll/create-nil', methods=['POST'])
def payroll_create_nil():
    """Create a NIL return payroll — no employees processed.
    Records only: EPF admin charge + consultant fee for that month."""
    try:
        est_id = int(request.form.get('establishment_id'))
        month = int(request.form.get('month'))
        year = int(request.form.get('year'))
        nil_fee = float(request.form.get('nil_fee_amount') or 0)
        nil_admin = float(request.form.get('nil_epf_admin') or 0)
    except (ValueError, TypeError):
        flash('Invalid input for NIL payroll. Please check the values.', 'danger')
        return redirect(url_for('payroll.payroll_create'))

    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)

    # Require fee to be entered (per user decision Q2)
    if nil_fee <= 0:
        flash('Please enter the NIL filing fee before creating the payroll.', 'warning')
        return redirect(url_for('payroll.payroll_create'))

    # Prevent duplicate payroll
    existing = MonthlyPayroll.query.filter_by(
        establishment_id=est_id, month=month, year=year).first()
    if existing:
        flash(f'A payroll already exists for {calendar.month_name[month]} {year}. '
              f'Delete it first if you want to create a NIL return for this month.', 'warning')
        return redirect(url_for('payroll.payroll_process', payroll_id=existing.id))

    # Create NIL payroll — skip attendance flow entirely
    payroll = MonthlyPayroll(
        establishment_id=est_id,
        month=month,
        year=year,
        working_days=calendar.monthrange(year, month)[1],
        status='draft',
        is_nil=True,
        nil_epf_admin=nil_admin,
        nil_fee_amount=nil_fee,
        # Zero out all totals (NIL means no employees, no wages)
        total_gross=0,
        total_employees=0,
        total_epf_employee=0,
        total_epf_employer=0,
        total_epf_ac01=0,
        total_epf_eps=0,
        total_epf_edli=0,
        total_epf_admin=nil_admin,  # Only admin charge applies
        total_esic_employee=0,
        total_esic_employer=0,
        total_pt=0,
        total_net_pay=0,
    )
    db.session.add(payroll)
    db.session.commit()

    log_activity('created', 'nil_payroll', entity_id=payroll.id,
                 entity_name=f'{calendar.month_name[month]} {year}',
                 details=f'NIL return — Admin ₹{nil_admin:,.0f} + Fee ₹{nil_fee:,.0f}',
                 establishment_id=est_id)

    flash(f'✓ NIL payroll created for {calendar.month_name[month]} {year}. '
          f'Admin charge ₹{nil_admin:,.0f} + Fee ₹{nil_fee:,.0f}. '
          f'Click Finalize to complete.', 'success')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll.id))


@payroll_bp.route('/payroll/<int:payroll_id>')
def payroll_process(payroll_id):
    """Process payroll — attendance entry and calculations"""
    import datetime as dt
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    # ══════════════════════════════════════════════════════════════
    # AUTO-SYNC: Silently add any active employees missing from this payroll.
    # Runs every time the page opens (unless finalized / NIL).
    # This handles: employees imported after payroll was created,
    # establishments with 300, 500, 1000+ employees — any count works.
    # ══════════════════════════════════════════════════════════════
    if payroll.status != 'finalized' and not payroll.is_nil:
        existing_emp_ids = {
            row[0] for row in
            db.session.query(PayrollEntry.employee_id).filter_by(
                monthly_payroll_id=payroll_id).all()
        }
        last_day_sync = date(payroll.year, payroll.month,
                             calendar.monthrange(payroll.year, payroll.month)[1])
        first_day_sync = date(payroll.year, payroll.month, 1)

        q_new = Employee.query.filter(
            Employee.establishment_id == est.id,
            Employee.is_active == True,
            Employee.date_of_joining <= last_day_sync
        )
        if existing_emp_ids:
            q_new = q_new.filter(~Employee.id.in_(list(existing_emp_ids)))
        new_emps = q_new.all()

        if new_emps:
            for emp_s in new_emps:
                sal_s = EmployeeSalary.query.filter_by(
                    employee_id=emp_s.id, is_current=True).first()
                if emp_s.date_of_joining > first_day_sync:
                    rem = (last_day_sync - emp_s.date_of_joining).days + 1
                    def_present = min(rem, payroll.working_days)
                else:
                    def_present = payroll.working_days
                ent_new = PayrollEntry(
                    monthly_payroll_id=payroll_id,
                    employee_id=emp_s.id,
                    days_present=def_present,
                    days_absent=0,
                    paid_holidays=0,
                    ot_hours=0,
                    total_payable_days=def_present,
                    gross_salary=sal_s.gross_salary if sal_s else 0
                )
                if hasattr(ent_new, 'rate_overrides'):
                    ent_new.rate_overrides = None
                db.session.add(ent_new)
            payroll.total_employees = len(existing_emp_ids) + len(new_emps)
            db.session.commit()

    # Load all entries — guaranteed to include every active employee now
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id).join(Employee).order_by(Employee.name).all()

    # Get salary heads for this establishment
    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True, head_type='earning'
    ).order_by(SalaryHead.display_order).all()

    # Calculate rest days and holidays for smart absent calculation (used in JS)
    _, num_days_in_month = calendar.monthrange(payroll.year, payroll.month)

    rest_day_count = 0
    if config and config.rest_day_type == 'rotation':
        rest_day_count = num_days_in_month // 7
    elif config and config.rest_day_type == 'fixed_day':
        weekday = config.rest_day_weekday if config.rest_day_weekday is not None else 6
        for d in range(1, num_days_in_month + 1):
            if dt.date(payroll.year, payroll.month, d).weekday() == weekday:
                rest_day_count += 1
    else:  # 'sunday' default
        for d in range(1, num_days_in_month + 1):
            if dt.date(payroll.year, payroll.month, d).weekday() == 6:
                rest_day_count += 1

    holiday_count = 0
    if payroll.holiday_dates:
        try:
            holiday_count = len([d.strip() for d in payroll.holiday_dates.split(',') if d.strip().isdigit()])
        except Exception:
            holiday_count = 0

    # Build red flag alerts for data quality issues
    red_flags = []
    if payroll.status != 'draft':
        for entry in entries:
            emp = entry.employee
            # Employee with 0 gross but days present
            if entry.days_present > 0 and entry.earned_gross == 0:
                red_flags.append(f"{emp.name} — worked {entry.days_present} days but Earned Gross is ₹0 (check salary config)")
            # Days present exceeds working days
            if entry.days_present > num_days_in_month:
                red_flags.append(f"{emp.name} — Days Present ({entry.days_present}) exceeds month days ({num_days_in_month})")
            # EPF applicable but no UAN
            if config and config.epf_applicable and entry.epf_employee > 0 and not emp.uan_number:
                red_flags.append(f"{emp.name} — EPF deducted (₹{entry.epf_employee:,.0f}) but UAN Number is missing")
            # ESIC applicable but no IP number
            if config and config.esic_applicable and entry.esic_employee > 0 and not emp.esic_ip_number:
                red_flags.append(f"{emp.name} — ESIC deducted (₹{entry.esic_employee:,.0f}) but ESIC IP Number is missing")
            # Very high OT (more than 50 hours)
            if entry.ot_hours and entry.ot_hours > 50:
                red_flags.append(f"{emp.name} — OT hours ({entry.ot_hours}) is unusually high (>50 hours)")
            # Net pay is negative
            if entry.net_pay < 0:
                red_flags.append(f"{emp.name} — Net Pay is negative (₹{entry.net_pay:,.0f}), deductions exceed earnings")

    return render_template('payroll/process.html',
                           payroll=payroll, est=est, config=config,
                           entries=entries, heads=heads,
                           num_days_in_month=num_days_in_month,
                           rest_day_count=rest_day_count,
                           holiday_count=holiday_count,
                           red_flags=red_flags)


@payroll_bp.route('/payroll/<int:payroll_id>/save-attendance', methods=['POST'])
def save_attendance(payroll_id):
    """Save attendance data and calculate payroll"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    working_days = payroll.working_days
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id).all()

    # --- Smart absent calculation helpers ---
    # Count rest days (Sundays / configured rest day) in the month
    import datetime as dt
    import json as _json
    _, num_days_in_month = calendar.monthrange(payroll.year, payroll.month)

    rest_day_count = 0
    if config.rest_day_type == 'rotation':
        rest_day_count = num_days_in_month // 7
    elif config.rest_day_type == 'fixed_day':
        weekday = config.rest_day_weekday if config.rest_day_weekday is not None else 6
        for d in range(1, num_days_in_month + 1):
            if dt.date(payroll.year, payroll.month, d).weekday() == weekday:
                rest_day_count += 1
    else:  # 'sunday' default
        for d in range(1, num_days_in_month + 1):
            if dt.date(payroll.year, payroll.month, d).weekday() == 6:
                rest_day_count += 1

    # Count national holidays saved for this month
    holiday_count = 0
    if payroll.holiday_dates:
        try:
            holiday_count = len([d.strip() for d in payroll.holiday_dates.split(',') if d.strip().isdigit()])
        except Exception:
            holiday_count = 0

    totals = {
        'gross': 0, 'epf_ee': 0,
        'epf_ac01': 0, 'epf_eps': 0, 'epf_admin': 0, 'epf_edli': 0, 'epf_er': 0,
        'esic_ee': 0, 'esic_er': 0, 'pt': 0, 'net': 0
    }

    for entry in entries:
        emp_id = entry.employee_id
        emp = entry.employee  # Employee object for exempt checks

        # Read attendance from form
        try:
            entry.days_present = float(request.form.get(f'present_{emp_id}', 0))
            entry.ot_hours = float(request.form.get(f'ot_{emp_id}', 0))
            entry.other_deduction = float(request.form.get(f'other_ded_{emp_id}', 0))
            entry.other_deduction_remark = request.form.get(f'other_ded_remark_{emp_id}', '').strip() or None
            entry.arrear_amount = float(request.form.get(f'arrear_{emp_id}', 0))
            entry.arrear_remark = request.form.get(f'arrear_remark_{emp_id}', '').strip() or None
        except ValueError:
            continue

        # --- SMART ABSENT CALCULATION ---
        # Formula: Month Days - Present = Remaining
        #          Remaining - Rest Days (Sundays) = After Rest
        #          After Rest - Holidays = Absent
        remaining = num_days_in_month - entry.days_present
        after_rest = remaining - rest_day_count
        after_holidays = after_rest - holiday_count
        entry.days_absent = max(0, after_holidays)

        # --- Fetch salary record & rate overrides ---
        salary = EmployeeSalary.query.filter_by(
            employee_id=emp_id, is_current=True).first()

        _rate_overrides = {}
        if hasattr(entry, 'rate_overrides') and entry.rate_overrides:
            try:
                _rate_overrides = _json.loads(entry.rate_overrides)
            except (ValueError, TypeError):
                _rate_overrides = {}

        # ═══════════════════════════════════════════════════════════
        # PATH A: UNIVERSAL TEMPLATE (rate_overrides exist)
        # Fully INDEPENDENT — no config dependency for core calculation.
        # Like EPF/ESIC portal: data comes from template, system just calculates.
        # ═══════════════════════════════════════════════════════════
        if _rate_overrides:
            _ovr_daily = _rate_overrides.get('daily_rate')
            _ovr_gross = _rate_overrides.get('gross', 0)
            _ovr_heads = _rate_overrides.get('heads')

            # Determine type from the override itself (NOT from config)
            is_daily_wages = bool(_ovr_daily)
            _eff_daily_rate = _ovr_daily

            # Total payable days: days_present + NPH (from form / upload)
            try:
                manual_ph = float(request.form.get(f'ph_{emp_id}', entry.paid_holidays or 0))
            except (ValueError, TypeError):
                manual_ph = entry.paid_holidays or 0
            entry.paid_holidays = manual_ph
            entry.total_payable_days = entry.days_present + entry.paid_holidays

            # Earned gross: rate × days (daily) or proportionate (gross)
            if _ovr_daily:
                entry.earned_gross = round(_ovr_daily * entry.total_payable_days)
                full_gross = round(_ovr_daily * working_days) if working_days > 0 else entry.earned_gross
            elif _ovr_gross:
                full_gross = _ovr_gross
                if working_days > 0:
                    entry.earned_gross = round((_ovr_gross / working_days) * entry.total_payable_days)
                else:
                    entry.earned_gross = round(_ovr_gross)
            else:
                full_gross = 0
                entry.earned_gross = 0
            entry.gross_salary = full_gross

            # Head-wise breakup
            PayrollEntryHead.query.filter_by(payroll_entry_id=entry.id).delete()
            _use_head_values = None

            if _ovr_heads:
                _use_head_values = []
                for head_id_str, amt in _ovr_heads.items():
                    try:
                        sh = db.session.get(SalaryHead, int(head_id_str))
                    except (ValueError, TypeError):
                        continue
                    if sh:
                        _use_head_values.append({
                            'salary_head': sh,
                            'salary_head_id': sh.id,
                            'amount': float(amt)
                        })
                for hv_dict in _use_head_values:
                    sh = hv_dict['salary_head']
                    amt = hv_dict['amount']
                    if sh and sh.head_type == 'earning' and sh.is_in_gross:
                        if _ovr_daily:
                            earned_amt = round(amt * entry.total_payable_days)
                        elif working_days > 0:
                            earned_amt = round((amt / working_days) * entry.total_payable_days)
                        else:
                            earned_amt = round(amt)
                        peh = PayrollEntryHead(
                            payroll_entry_id=entry.id,
                            salary_head_id=hv_dict['salary_head_id'],
                            full_amount=amt,
                            earned_amount=earned_amt
                        )
                        db.session.add(peh)
            elif _ovr_daily and entry.earned_gross > 0:
                # Daily wages — map entire earned to BASIC head
                basic_head = SalaryHead.query.filter_by(
                    establishment_id=est.id, short_code='BASIC', is_active=True
                ).first()
                if basic_head:
                    peh = PayrollEntryHead(
                        payroll_entry_id=entry.id,
                        salary_head_id=basic_head.id,
                        full_amount=_ovr_daily,
                        earned_amount=entry.earned_gross
                    )
                    db.session.add(peh)

            # OT: use config rates (statutory) but template rate for base
            entry.ot_amount = 0
            if entry.ot_hours and entry.ot_hours > 0 and working_days > 0:
                rate_multiplier = 2.0 if getattr(config, 'ot_rate_type', 'double') == 'double' else 1.0
                ot_unit = getattr(config, 'ot_unit', 'hours')
                if ot_unit == 'hours':
                    if _ovr_daily:
                        per_hour = _ovr_daily / 8
                    elif full_gross > 0:
                        per_hour = (full_gross / working_days) / 8
                    else:
                        per_hour = 0
                    entry.ot_amount = round(per_hour * entry.ot_hours * rate_multiplier)
                else:
                    if _ovr_daily:
                        per_day = _ovr_daily
                    elif full_gross > 0:
                        per_day = full_gross / working_days
                    else:
                        per_day = 0
                    entry.ot_amount = round(per_day * entry.ot_hours * rate_multiplier)

            entry.total_earnings = entry.earned_gross + entry.ot_amount + entry.arrear_amount

            # Compliance wages: use head breakup if available, else full earned gross
            if _use_head_values and config.compliance_basis != 'gross':
                compliance_total = sum(
                    hv_dict['amount'] for hv_dict in _use_head_values
                    if hv_dict['salary_head'] and hv_dict['salary_head'].is_for_compliance
                )
                if full_gross > 0:
                    compliance_wages = round(entry.earned_gross * (compliance_total / full_gross))
                else:
                    compliance_wages = entry.earned_gross
            else:
                compliance_wages = entry.earned_gross

            if getattr(config, 'include_ot_in_epf', False) and entry.ot_amount > 0:
                compliance_wages += entry.ot_amount

        # ═══════════════════════════════════════════════════════════
        # PATH B: NORMAL (monthly template / manual attendance entry)
        # Uses establishment config + employee salary settings as before
        # ═══════════════════════════════════════════════════════════
        else:
            emp_salary_type = (salary.salary_type if salary and salary.salary_type else config.salary_type) or 'monthly_fixed'
            is_daily_wages = (emp_salary_type == 'daily_wages')
            _eff_daily_rate = salary.daily_rate if salary else None
            _eff_gross = salary.gross_salary if salary else 0

            # Weekly off policy
            wo_policy = 'paid'
            if is_daily_wages:
                wo_policy = (salary.weekly_off_policy if salary and salary.weekly_off_policy else None) or getattr(config, 'weekly_off_policy', 'paid') or 'paid'
            wo_ot_days = 0

            if is_daily_wages and wo_policy == 'unpaid':
                if config.paid_holiday_type == 'separate':
                    try:
                        manual_ph = float(request.form.get(f'ph_{emp_id}', 0))
                        entry.paid_holidays = manual_ph
                    except ValueError:
                        entry.paid_holidays = holiday_count
                else:
                    entry.paid_holidays = holiday_count
                entry.total_payable_days = entry.days_present + entry.paid_holidays

            elif is_daily_wages and wo_policy == 'ot_rate':
                available_working_days = num_days_in_month - rest_day_count - holiday_count
                wo_ot_days = max(0, entry.days_present - available_working_days)
                regular_present = entry.days_present - wo_ot_days
                if config.paid_holiday_type == 'separate':
                    try:
                        manual_ph = float(request.form.get(f'ph_{emp_id}', 0))
                        entry.paid_holidays = manual_ph
                    except ValueError:
                        entry.paid_holidays = holiday_count
                else:
                    entry.paid_holidays = holiday_count
                entry.total_payable_days = regular_present + entry.paid_holidays

            else:
                if config.paid_holiday_type == 'separate':
                    try:
                        manual_ph = float(request.form.get(f'ph_{emp_id}', 0))
                        entry.paid_holidays = manual_ph
                    except ValueError:
                        entry.paid_holidays = rest_day_count + holiday_count
                else:
                    entry.paid_holidays = rest_day_count + holiday_count
                entry.total_payable_days = entry.days_present + entry.paid_holidays

            # Skip if no salary at all
            if not salary:
                entry.earned_gross = 0
                entry.net_pay = 0
                continue

            full_gross = _eff_gross or (salary.gross_salary if salary else 0)

            # Absence deduction
            emp_absence_deduction = config.absence_deduction
            if getattr(salary, 'no_absence_deduction', False):
                emp_absence_deduction = False

            # Earned gross
            if is_daily_wages and _eff_daily_rate:
                entry.earned_gross = round(_eff_daily_rate * entry.total_payable_days)
            elif emp_absence_deduction and working_days > 0:
                entry.earned_gross = round((full_gross / working_days) * entry.total_payable_days)
            else:
                entry.earned_gross = round(full_gross)
            entry.gross_salary = full_gross

            # Head-wise breakup
            PayrollEntryHead.query.filter_by(payroll_entry_id=entry.id).delete()
            _use_head_values = None
            if salary and salary.head_values:
                _use_head_values = []
                for hv in salary.head_values:
                    _use_head_values.append({
                        'salary_head': hv.salary_head,
                        'salary_head_id': hv.salary_head_id,
                        'amount': hv.amount
                    })

            if _use_head_values and working_days > 0:
                for hv_dict in _use_head_values:
                    sh = hv_dict['salary_head']
                    amt = hv_dict['amount']
                    if sh and sh.head_type == 'earning' and sh.is_in_gross:
                        if is_daily_wages and _eff_daily_rate:
                            earned_amt = round(amt * entry.days_present)
                        elif not emp_absence_deduction:
                            earned_amt = round(amt)
                        else:
                            earned_amt = round((amt / working_days) * entry.total_payable_days)
                        peh = PayrollEntryHead(
                            payroll_entry_id=entry.id,
                            salary_head_id=hv_dict['salary_head_id'],
                            full_amount=amt,
                            earned_amount=earned_amt
                        )
                        db.session.add(peh)
            elif is_daily_wages and _eff_daily_rate and entry.earned_gross > 0:
                basic_head = SalaryHead.query.filter_by(
                    establishment_id=est.id, short_code='BASIC', is_active=True
                ).first()
                if basic_head:
                    basic_earned = round(_eff_daily_rate * entry.days_present)
                    peh = PayrollEntryHead(
                        payroll_entry_id=entry.id,
                        salary_head_id=basic_head.id,
                        full_amount=_eff_daily_rate,
                        earned_amount=basic_earned
                    )
                    db.session.add(peh)

            # OT calculation
            entry.ot_amount = 0
            is_fixed_salary = getattr(salary, 'no_absence_deduction', False) if salary else False

            if not is_fixed_salary:
                if is_daily_wages and wo_policy == 'ot_rate' and wo_ot_days > 0 and _eff_daily_rate:
                    wo_rate_multiplier = 2.0 if config.ot_rate_type == 'double' else 1.0
                    entry.ot_amount = round(_eff_daily_rate * wo_ot_days * wo_rate_multiplier)

                if config.ot_applicable and entry.ot_hours > 0 and working_days > 0:
                    rate_multiplier = 2.0 if config.ot_rate_type == 'double' else 1.0

                    # Determine OT base: full gross (default) or Basic only
                    ot_base = full_gross
                    if getattr(config, 'ot_base_wage', 'gross') == 'basic_only' and _use_head_values:
                        basic_amt = next(
                            (hv['amount'] for hv in _use_head_values
                             if hv['salary_head'] and hv['salary_head'].short_code == 'BASIC'),
                            None
                        )
                        if basic_amt:
                            ot_base = basic_amt

                    if config.ot_unit == 'hours':
                        if is_daily_wages and _eff_daily_rate:
                            per_hour = _eff_daily_rate / 8
                        else:
                            per_hour = (ot_base / working_days) / 8
                        entry.ot_amount += round(per_hour * entry.ot_hours * rate_multiplier)
                    else:
                        if is_daily_wages and _eff_daily_rate:
                            per_day = _eff_daily_rate
                        else:
                            per_day = ot_base / working_days
                        entry.ot_amount += round(per_day * entry.ot_hours * rate_multiplier)
            else:
                entry.paid_holidays = 0

            entry.total_earnings = entry.earned_gross + entry.ot_amount + entry.arrear_amount

            # Compliance wages
            if config.compliance_basis == 'gross':
                compliance_wages = entry.earned_gross
            else:
                if _use_head_values:
                    compliance_total = sum(
                        hv_dict['amount'] for hv_dict in _use_head_values
                        if hv_dict['salary_head'] and hv_dict['salary_head'].is_for_compliance
                    )
                    if full_gross > 0:
                        compliance_ratio = compliance_total / full_gross
                        compliance_wages = round(entry.earned_gross * compliance_ratio)
                    else:
                        compliance_wages = entry.earned_gross
                else:
                    compliance_wages = entry.earned_gross

            if getattr(config, 'include_ot_in_epf', False) and entry.ot_amount > 0:
                compliance_wages += entry.ot_amount

        # ========================================
        # EPF Calculation (correct Indian EPF structure)
        # Employee: 12% of EPF wages (goes to A/c 01)
        # Employer:
        #   EPF A/c 01: 3.67%
        #   EPS A/c 10: 8.33%
        #   Admin Charge: 0.5% (minimum ₹500 per establishment)
        #   EDLI: 0.5%
        #   Total Employer: 13%
        # ========================================
        entry.epf_employee = 0
        entry.epf_ac01 = 0
        entry.epf_eps = 0
        entry.epf_admin = 0
        entry.epf_edli = 0
        entry.epf_employer = 0
        entry.epf_wages = 0

        if config.epf_applicable:
            # ── STATUTORY EPF WAGE RULES ────────────────────────────────
            # EPF (12%)      : base varies — full wages if 'higher' deduction,
            #                  else capped at establishment's epf_wage_ceiling
            # EPS (8.33%)    : ALWAYS capped at ₹15,000 (statutory EPS ceiling)
            # EDLI (0.5%)    : ALWAYS capped at ₹15,000 (statutory EDLI ceiling)
            # Admin (0.5%)   : Same base as EDLI (capped at ₹15,000)
            # ────────────────────────────────────────────────────────────
            EPF_STATUTORY_CEILING = 15000   # Hard limit for EPS + EDLI + Admin

            # EPF base
            if config.epf_contribution_type == 'higher':
                epf_wages = compliance_wages
            else:
                epf_wages = min(compliance_wages, config.epf_wage_ceiling)
            entry.epf_wages = epf_wages

            # EPS/EDLI base — always capped at ₹15,000 regardless of higher/regular
            eps_edli_wages = min(epf_wages, EPF_STATUTORY_CEILING)

            # Employee share: 12% of EPF wages (uses higher base if higher deduction)
            entry.epf_employee = round(epf_wages * config.epf_employee_rate / 100)

            # Employer breakdown:
            #   EPS (8.33% of capped wages) — MAX ₹1,250/employee
            #   A/c 01 = Total employer 12% × epf_wages − EPS
            #            (for higher deduction, A/c 01 absorbs the excess above ceiling)
            #   EDLI (0.5% of capped wages) — MAX ₹75/employee
            #   Admin (0.5% of capped wages)
            entry.epf_eps = round(eps_edli_wages * config.epf_eps_rate / 100)      # 8.33% × min(wages, 15000)
            entry.epf_ac01 = entry.epf_employee - entry.epf_eps                      # Balance: Employee 12% − EPS
            entry.epf_edli = round(eps_edli_wages * config.epf_edli_rate / 100)    # 0.5% × min(wages, 15000)

            # Admin charge: 0.5% of capped wages (same base as EDLI)
            entry.epf_admin = round(eps_edli_wages * config.epf_admin_rate / 100)

            # Total employer contribution
            entry.epf_employer = entry.epf_ac01 + entry.epf_eps + entry.epf_admin + entry.epf_edli

        # ========================================
        # ESIC Calculation
        # ESIC is on Gross MINUS excluded heads (e.g., Wash Allowance)
        # Employee: 0.75% rounded UP  =  round(esic_gross * 0.75% + 0.49, 0)
        # Employer: 3.25% rounded UP  =  round(esic_gross * 3.25% + 0.49, 0)
        # ========================================
        entry.esic_employee = 0
        entry.esic_employer = 0
        entry.esic_wages = 0

        if config.esic_applicable and not emp.esic_exempt:
            # Calculate ESIC gross: earned_gross minus excluded heads (Wash Allowance etc.)
            esic_excluded = 0
            if _use_head_values and working_days > 0:
                for hv_dict in _use_head_values:
                    sh = hv_dict['salary_head']
                    if sh and sh.exclude_from_esic:
                        if is_daily_wages and _eff_daily_rate:
                            esic_excluded += round(hv_dict['amount'] * entry.total_payable_days)
                        else:
                            esic_excluded += round((hv_dict['amount'] / working_days) * entry.total_payable_days)

            esic_gross = entry.earned_gross - esic_excluded

            # Include OT in ESIC wages if enabled in config
            if getattr(config, 'include_ot_in_esic', False) and entry.ot_amount > 0:
                esic_gross += entry.ot_amount

            # Higher deduction = ESIC on full wages (no ceiling check — deduct even above ₹21,000)
            esic_type = getattr(config, 'esic_contribution_type', 'ceiling')
            if esic_type == 'higher' or esic_gross <= config.esic_wage_ceiling:
                entry.esic_wages = esic_gross
                # Round UP formula: round(amount + 0.49, 0)
                entry.esic_employee = round(esic_gross * config.esic_employee_rate / 100 + 0.49)
                entry.esic_employer = round(esic_gross * config.esic_employer_rate / 100 + 0.49)

        # Professional Tax (Karnataka slab for now)
        entry.professional_tax = 0
        if config.pt_applicable:
            pt_state = config.pt_state or 'karnataka'
            entry.professional_tax = _calculate_pt(entry.earned_gross, pt_state, payroll.month)

        # Total deductions and net pay (only employee shares deducted from salary)
        entry.total_deductions = (entry.epf_employee + entry.esic_employee +
                                  entry.professional_tax + entry.other_deduction)
        entry.net_pay = round(entry.total_earnings - entry.total_deductions)

        # Accumulate totals
        totals['gross'] += entry.total_earnings
        totals['epf_ee'] += entry.epf_employee
        totals['epf_ac01'] += entry.epf_ac01
        totals['epf_eps'] += entry.epf_eps
        totals['epf_admin'] += entry.epf_admin
        totals['epf_edli'] += entry.epf_edli
        totals['epf_er'] += entry.epf_employer
        totals['esic_ee'] += entry.esic_employee
        totals['esic_er'] += entry.esic_employer
        totals['pt'] += entry.professional_tax
        totals['net'] += entry.net_pay

    # Admin charge minimum check: if total admin < ₹500, set to ₹500
    # Count employees who actually have EPF contribution (after calculation)
    epf_employee_count = sum(1 for e in entries if e.epf_wages > 0)
    if config.epf_applicable and epf_employee_count > 0 and totals['epf_admin'] < config.epf_admin_min:
        totals['epf_admin'] = config.epf_admin_min
        totals['epf_er'] = totals['epf_ac01'] + totals['epf_eps'] + totals['epf_admin'] + totals['epf_edli']

    # Update payroll totals (all rounded to whole numbers)
    payroll.total_gross = round(totals['gross'])
    payroll.total_epf_employee = round(totals['epf_ee'])
    payroll.total_epf_ac01 = round(totals['epf_ac01'])
    payroll.total_epf_eps = round(totals['epf_eps'])
    payroll.total_epf_admin = round(totals['epf_admin'])
    payroll.total_epf_edli = round(totals['epf_edli'])
    payroll.total_epf_employer = round(totals['epf_er'])
    payroll.total_esic_employee = round(totals['esic_ee'])
    payroll.total_esic_employer = round(totals['esic_er'])
    payroll.total_pt = round(totals['pt'])
    payroll.total_net_pay = round(totals['net'])
    payroll.status = 'processing'

    # Save Other Charges (additional billing to client — Other Income)
    payroll.other_charges_description = request.form.get('other_charges_description', '').strip() or None
    try:
        payroll.other_charges_amount = float(request.form.get('other_charges_amount', 0) or 0)
    except (ValueError, TypeError):
        payroll.other_charges_amount = 0

    log_activity('calculated', 'payroll', entity_id=payroll.id,
                 entity_name=f'{payroll.period_display}',
                 details=f'Employees: {payroll.total_employees}, Net Pay: ₹{payroll.total_net_pay:,.0f}',
                 establishment_id=payroll.establishment_id)
    db.session.commit()
    flash('Attendance saved and payroll calculated!', 'success')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


@payroll_bp.route('/payroll/<int:payroll_id>/save-holidays', methods=['POST'])
def save_holidays(payroll_id):
    """Save holiday dates for the payroll month"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    verify_est_ownership(payroll.establishment)
    payroll.holiday_dates = request.form.get('holiday_dates', '').strip() or None
    db.session.commit()
    flash(f'Holiday dates saved for {payroll.period_display}.', 'success')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


@payroll_bp.route('/payroll/<int:payroll_id>/save-epf-payment', methods=['POST'])
def save_epf_payment(payroll_id):
    """Save EPF payment date and calculate:
       - Interest u/s 7Q  @ 12% per annum (daily calculation)
       - Damages u/s 14B  @ 1% per month (partial month = full month)
    """
    import math
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    verify_est_ownership(payroll.establishment)

    payment_date_str = request.form.get('epf_payment_date', '').strip()
    if payment_date_str:
        from datetime import date as dt_date
        try:
            payment_date = dt_date.fromisoformat(payment_date_str)
            payroll.epf_payment_date = payment_date

            # Calculate delay days from due date (15th of next month)
            due_date = payroll.epf_due_date
            delay_days = max(0, (payment_date - due_date).days)
            payroll.epf_delay_days = delay_days

            if delay_days > 0:
                # Total EPF contribution (Employee + Employer shares)
                epf_total = payroll.total_epf_employee + payroll.total_epf_employer

                # Delay in months (partial month counts as full month)
                delay_months = math.ceil(delay_days / 30)

                # Interest u/s 7Q: 12% per annum, daily calculation
                # Formula: (EPF Amount × 12% × delay_days) / 365
                payroll.epf_interest_14b = round((epf_total * 12 * delay_days) / (100 * 365))

                # Damages u/s 14B: 1% per month of EPF contribution
                # Formula: EPF Amount × 1% × months of delay
                payroll.epf_damages_7q = round(epf_total * 1 * delay_months / 100)

                total_penalty = payroll.epf_interest_14b + payroll.epf_damages_7q
                flash(f'EPF paid {delay_days} days ({delay_months} month) late! '
                      f'Interest (7Q) ₹{payroll.epf_interest_14b:,.0f} + '
                      f'Damages (14B) ₹{payroll.epf_damages_7q:,.0f} = '
                      f'Total Penalty ₹{total_penalty:,.0f}', 'warning')
            else:
                payroll.epf_interest_14b = 0
                payroll.epf_damages_7q = 0
                flash(f'EPF paid on time ({payment_date.strftime("%d %b %Y")}). No penalty applicable.', 'success')
        except (ValueError, TypeError):
            flash('Invalid payment date format.', 'danger')
            return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))
    else:
        # Clear payment details if date removed
        payroll.epf_payment_date = None
        payroll.epf_delay_days = 0
        payroll.epf_interest_14b = 0
        payroll.epf_damages_7q = 0
        flash('EPF payment details cleared.', 'info')

    db.session.commit()
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


@payroll_bp.route('/payroll/<int:payroll_id>/finalize', methods=['POST'])
def payroll_finalize(payroll_id):
    """Finalize the payroll — lock it from further editing.

    Finalize Modal posts two optional/required fields:
      - holiday_dates (text): comma-separated day numbers (or empty = 0 holidays)
      - epf_payment_date (date): when EPF was/will be paid — required, future allowed

    If holiday_dates changed from what the payroll was calculated with,
    we transparently recalculate the payroll so totals reflect the new
    holiday count before locking.
    """
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    verify_est_ownership(payroll.establishment)

    # --- NIL payrolls: skip holiday/EPF date logic (no employees to recalc) ---
    if payroll.is_nil:
        payroll.status = 'finalized'
        log_activity('finalized', 'nil_payroll', entity_id=payroll.id,
                     entity_name=payroll.period_display,
                     details='NIL return locked',
                     establishment_id=payroll.establishment_id)
        db.session.commit()
        flash(f'NIL payroll for {payroll.period_display} has been finalized.', 'success')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    # --- Regular payrolls: capture holiday + EPF payment date from modal ---
    new_holiday_raw = request.form.get('holiday_dates', None)
    old_holiday = payroll.holiday_dates or ''

    # Normalise holiday inputs for comparison (strip whitespace, sort, dedupe)
    def _normalise_holidays(s):
        if not s:
            return ''
        parts = [p.strip() for p in str(s).split(',') if p.strip().isdigit()]
        return ','.join(sorted(set(parts), key=int))

    holiday_changed = False
    if new_holiday_raw is not None:
        normalised_new = _normalise_holidays(new_holiday_raw)
        normalised_old = _normalise_holidays(old_holiday)
        if normalised_new != normalised_old:
            payroll.holiday_dates = normalised_new or None
            holiday_changed = True

    # EPF Payment Date (required — future dates allowed)
    # Also computes 7Q Interest (12% p.a.) and 14B Damages (1% per month)
    # if payment is late — so penalty reflects in summary + reimbursement letter.
    epf_date_str = (request.form.get('epf_payment_date') or '').strip()
    if epf_date_str:
        try:
            import math
            payroll.epf_payment_date = datetime.strptime(epf_date_str, '%Y-%m-%d').date()
            # Due date = 15th of calendar month FOLLOWING the wage month
            if payroll.month == 12:
                due_date = date(payroll.year + 1, 1, 15)
            else:
                due_date = date(payroll.year, payroll.month + 1, 15)
            delay_days = max(0, (payroll.epf_payment_date - due_date).days)
            payroll.epf_delay_days = delay_days

            if delay_days > 0:
                # Total EPF remittance = Employee 12% + Employer 13% (all accounts)
                epf_total = (payroll.total_epf_employee or 0) + (payroll.total_epf_employer or 0)
                # Interest u/s 7Q: 12% per annum, daily basis
                #   interest = (epf_total × 12% × delay_days) / 365
                interest_7q = round((epf_total * 12 * delay_days) / (100 * 365))
                # Damages u/s 14B: 1% per month — partial month counts as full month
                delay_months = math.ceil(delay_days / 30)
                damages_14b = round(epf_total * 1 * delay_months / 100)
                # NOTE: DB column names kept for historical compatibility but
                # semantic mapping matches EPFO terminology:
                #   epf_interest_14b ← Interest u/s 7Q (was mis-named previously)
                #   epf_damages_7q   ← Damages u/s 14B
                payroll.epf_interest_14b = interest_7q
                payroll.epf_damages_7q = damages_14b
            else:
                payroll.epf_interest_14b = 0
                payroll.epf_damages_7q = 0
        except ValueError:
            pass

    # If holidays changed at finalize time, warn user to recalculate before locking
    # (Full recalc requires re-running the whole attendance pipeline — safer to
    # let user do it explicitly via Save & Calculate.)
    if holiday_changed:
        db.session.commit()   # still save the new holiday_dates + epf_date
        flash(f'Holidays updated to "{payroll.holiday_dates or "None"}". '
              f'Please click Save & Calculate to recompute totals, '
              f'then click Finalize again to lock the payroll.', 'warning')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    payroll.status = 'finalized'
    log_activity('finalized', 'payroll', entity_id=payroll.id,
                 entity_name=payroll.period_display,
                 details=f'Net Pay: ₹{payroll.total_net_pay:,.0f}, '
                         f'EPF Paid: {payroll.epf_payment_date or "—"}',
                 establishment_id=payroll.establishment_id)
    db.session.commit()

    flash(f'Payroll for {payroll.period_display} has been finalized.', 'success')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


@payroll_bp.route('/payroll/<int:payroll_id>/reopen', methods=['POST'])
def payroll_reopen(payroll_id):
    """Reopen a finalized payroll for corrections"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    verify_est_ownership(payroll.establishment)
    payroll.status = 'processing'
    db.session.commit()
    flash(f'Payroll for {payroll.period_display} has been reopened.', 'warning')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


# Secret password for deleting finalized payrolls
FINALIZED_DELETE_PASSWORD = 'Vaishnavi@2026'


@payroll_bp.route('/payroll/<int:payroll_id>/delete', methods=['POST'])
def payroll_delete(payroll_id):
    """Delete a monthly payroll and all its entries"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    verify_est_ownership(payroll.establishment)
    period = payroll.period_display
    est_name = payroll.establishment.company_name

    # For finalized payrolls, verify secret password
    if payroll.status == 'finalized':
        password = request.form.get('delete_password', '')
        if password != FINALIZED_DELETE_PASSWORD:
            flash('Incorrect password! Cannot delete finalized payroll.', 'danger')
            return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    # Delete payroll (cascade deletes entries)
    db.session.delete(payroll)
    db.session.commit()
    flash(f'Payroll for {period} ({est_name}) has been deleted.', 'warning')
    return redirect(url_for('payroll.payroll_list'))


@payroll_bp.route('/payroll/<int:payroll_id>/statement')
def payroll_statement(payroll_id):
    """Professional salary statement for client review.
    Excludes employees with zero attendance (no work = no salary to show).
    """
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    # Exclude zero-attendance employees from salary statement
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .filter(PayrollEntry.days_present > 0)\
        .join(Employee).order_by(Employee.name).all()

    # Get salary heads for this establishment (for column headers)
    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True, head_type='earning', is_in_gross=True
    ).order_by(SalaryHead.display_order).all()

    # Build head-wise data for each entry + attach salary type and daily rate for Rate column
    for entry in entries:
        entry.head_amounts = {}
        for peh in entry.head_breakup:
            entry.head_amounts[peh.salary_head_id] = peh
        cur_sal = EmployeeSalary.query.filter_by(
            employee_id=entry.employee_id, is_current=True).first()
        entry._salary_type = cur_sal.salary_type if cur_sal else (config.salary_type if config else 'monthly_fixed')
        entry._daily_rate = cur_sal.daily_rate if cur_sal and cur_sal.daily_rate else 0

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('payroll/statement.html',
                           payroll=payroll, est=est, config=config,
                           entries=entries, heads=heads,
                           generated_on=generated_on)


def _calculate_pt(gross, state='karnataka', month=None):
    """Professional Tax slab calculator — multi-state support.
    Returns monthly PT amount based on state-specific slabs.
    month (1-12) needed for states with Feb adjustment (e.g., Maharashtra).
    """
    state = (state or 'karnataka').lower().strip()

    # ── KARNATAKA ── Revised: Amendment Act 2025, effective 1-Apr-2025
    # Old threshold was ₹15,000. Raised to ₹25,000.
    # ≤₹24,999 → Nil | ≥₹25,000 → ₹200/month (₹300 in February)
    # Annual total: 11 × ₹200 + ₹300 = ₹2,500
    if state == 'karnataka':
        if gross < 25000:
            return 0
        else:
            if month == 2:
                return 300
            return 200

    # ── MAHARASHTRA ──
    # Men: ₹2,500/year. Women: Nil up to ₹25,000.
    # Slab: ≤7500=Nil, 7501-10000=175, 10001+=200 (except Feb=300)
    elif state == 'maharashtra':
        if gross <= 7500:
            return 0
        elif gross <= 10000:
            return 175
        else:
            # February gets ₹300 to make annual total ₹2,500
            if month == 2:
                return 300
            return 200

    # ── TAMIL NADU ──
    # ≤21000=Nil, 21001-30000=₹100, 30001-45000=₹235,
    # 45001-60000=₹510, 60001-75000=₹760, >75000=₹1095
    elif state == 'tamil_nadu':
        if gross <= 21000:
            return 0
        elif gross <= 30000:
            return 100
        elif gross <= 45000:
            return 235
        elif gross <= 60000:
            return 510
        elif gross <= 75000:
            return 760
        else:
            return 1095

    # ── ANDHRA PRADESH ──
    # ≤15000=Nil, 15001-20000=₹150, 20001-25000=₹200, >25000=₹200
    elif state == 'andhra_pradesh':
        if gross <= 15000:
            return 0
        elif gross <= 20000:
            return 150
        elif gross <= 25000:
            return 200
        else:
            return 200

    # ── TELANGANA ──
    # ≤15000=Nil, 15001-20000=₹150, 20001-25000=₹200, >25000=₹200
    elif state == 'telangana':
        if gross <= 15000:
            return 0
        elif gross <= 20000:
            return 150
        else:
            return 200

    # ── WEST BENGAL ──
    # ≤10000=Nil, 10001-15000=₹110, 15001-25000=₹130,
    # 25001-40000=₹150, >40000=₹200
    elif state == 'west_bengal':
        if gross <= 10000:
            return 0
        elif gross <= 15000:
            return 110
        elif gross <= 25000:
            return 130
        elif gross <= 40000:
            return 150
        else:
            return 200

    # ── GUJARAT ──
    # ≤12000=Nil, >12000=₹200
    elif state == 'gujarat':
        if gross <= 12000:
            return 0
        else:
            return 200

    # ── MADHYA PRADESH ──
    # ≤18750=Nil (annual ≤2.25L), 18751-25000=₹125, 25001-33333=₹167, >33333=₹208
    elif state == 'madhya_pradesh':
        if gross <= 18750:
            return 0
        elif gross <= 25000:
            return 125
        elif gross <= 33333:
            return 167
        else:
            return 208

    # ── RAJASTHAN ──
    # No PT in Rajasthan
    elif state == 'rajasthan':
        return 0

    # ── KERALA ──
    # ≤11999=Nil, 12000-17999=₹120, 18000-24999=₹180,
    # 25000-29999=₹250, ≥30000=₹270 (Feb=₹420 to make ₹3,240/yr)
    elif state == 'kerala':
        if gross < 12000:
            return 0
        elif gross < 18000:
            return 120
        elif gross < 25000:
            return 180
        elif gross < 30000:
            return 250
        else:
            if month == 2:
                return 420
            return 270

    # ── ASSAM ──
    # ≤10000=Nil, 10001-15000=₹150, 15001-25000=₹180, >25000=₹208
    elif state == 'assam':
        if gross <= 10000:
            return 0
        elif gross <= 15000:
            return 150
        elif gross <= 25000:
            return 180
        else:
            return 208

    # ── ODISHA ──
    # ≤13304=Nil, 13305-25000=₹125, >25000=₹200
    elif state == 'odisha':
        if gross <= 13304:
            return 0
        elif gross <= 25000:
            return 125
        else:
            return 200

    # ── JHARKHAND ──
    # ≤25000=Nil, 25001-41666=₹100, 41667-66666=₹150,
    # 66667-83333=₹175, >83333=₹208
    elif state == 'jharkhand':
        if gross <= 25000:
            return 0
        elif gross <= 41666:
            return 100
        elif gross <= 66666:
            return 150
        elif gross <= 83333:
            return 175
        else:
            return 208

    # ── CHHATTISGARH ──
    # Same as MP
    elif state == 'chhattisgarh':
        if gross <= 18750:
            return 0
        elif gross <= 25000:
            return 125
        elif gross <= 33333:
            return 167
        else:
            return 208

    # ── MEGHALAYA ──
    # ≤16666=Nil, 16667-41666=₹150, >41666=₹208
    elif state == 'meghalaya':
        if gross <= 16666:
            return 0
        elif gross <= 41666:
            return 150
        else:
            return 208

    # ── TRIPURA, MANIPUR, SIKKIM, MIZORAM ──
    elif state in ('tripura', 'manipur', 'sikkim', 'mizoram'):
        if gross <= 15000:
            return 0
        elif gross <= 25000:
            return 150
        else:
            return 208

    # ── PUDUCHERRY ──
    # No PT
    elif state in ('puducherry', 'delhi', 'haryana', 'punjab',
                    'himachal_pradesh', 'jammu_kashmir', 'uttarakhand',
                    'goa', 'arunachal_pradesh', 'nagaland'):
        return 0

    # ── BIHAR ──
    # ≤25000=Nil, 25001-50000=₹100, 50001-75000=₹150, >75000=₹200
    elif state == 'bihar':
        if gross <= 25000:
            return 0
        elif gross <= 50000:
            return 100
        elif gross <= 75000:
            return 150
        else:
            return 200

    # ── DEFAULT (Karnataka) ──
    else:
        if gross <= 15000:
            return 0
        return 200


# =============================================
# BULK ATTENDANCE UPLOAD (Excel Download/Upload)
# =============================================

@payroll_bp.route('/payroll/<int:payroll_id>/download-template')
def download_attendance_template(payroll_id):
    """Download Excel template pre-filled with employee names for bulk attendance entry"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from flask import send_file

    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    entries = PayrollEntry.query.filter_by(
        monthly_payroll_id=payroll_id
    ).join(Employee).order_by(Employee.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # ── Styling ──
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4338CA', end_color='4338CA', fill_type='solid')
    info_fill = PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid')
    info_font = Font(name='Calibri', bold=True, size=10, color='4338CA')
    data_font = Font(name='Calibri', size=10)
    input_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    lock_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # ── Info Header (Row 1-3) ──
    ws.merge_cells('A1:H1')
    ws['A1'] = f'ATTENDANCE TEMPLATE — {payroll.period_display}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='4338CA')

    ws.merge_cells('A2:H2')
    ws['A2'] = f'{est.company_name} | Working Days: {payroll.working_days}'
    ws['A2'].font = Font(name='Calibri', size=10, color='64748B')

    ws.merge_cells('A3:H3')
    ws['A3'] = 'Fill YELLOW columns only. Do NOT change Employee ID, Name or Gross Salary columns.'
    ws['A3'].font = Font(name='Calibri', bold=True, size=9, color='DC2626')

    # ── Build Column Headers (Row 5) ──
    headers = ['Sr.', 'UAN / ID', 'Employee Name', 'Gross Salary', 'Days Present']

    # Optional columns based on config
    has_ph = config and config.paid_holiday_type == 'separate'
    has_ot = config and config.ot_applicable

    if has_ph:
        headers.append('Paid Holidays')
    if has_ot:
        headers.append('OT (Hours/Days)')

    headers.append('Other Deduction')
    headers.append('Deduction Remark')

    # Column widths
    col_widths = {
        1: 5,    # Sr
        2: 12,   # Emp ID
        3: 30,   # Name
        4: 14,   # Gross
        5: 14,   # Days Present
    }
    next_col = 6
    if has_ph:
        col_widths[next_col] = 14
        next_col += 1
    if has_ot:
        col_widths[next_col] = 16
        next_col += 1
    col_widths[next_col] = 16       # Other Ded
    col_widths[next_col + 1] = 25   # Remark

    for col_idx, width in col_widths.items():
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width

    # Set column widths properly
    from openpyxl.utils import get_column_letter
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write headers (Row 5)
    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # ── Editable column indices (1-based, for yellow fill) ──
    editable_cols = set()
    present_col = 5
    editable_cols.add(present_col)
    next_data_col = 6
    ph_col = None
    ot_col = None
    if has_ph:
        ph_col = next_data_col
        editable_cols.add(ph_col)
        next_data_col += 1
    if has_ot:
        ot_col = next_data_col
        editable_cols.add(ot_col)
        next_data_col += 1
    other_ded_col = next_data_col
    remark_col = next_data_col + 1
    editable_cols.add(other_ded_col)
    editable_cols.add(remark_col)

    # ── Write Employee Data (Row 6 onwards) ──
    for idx, entry in enumerate(entries, 1):
        row = header_row + idx
        emp = entry.employee

        row_data = [
            idx,
            emp.uan_number or emp.esic_ip_number or emp.emp_code,
            emp.name,
            round(entry.gross_salary),
            entry.days_present if entry.days_present else payroll.working_days,
        ]
        if has_ph:
            row_data.append(entry.paid_holidays or 0)
        if has_ot:
            row_data.append(entry.ot_hours or 0)
        row_data.append(entry.other_deduction or 0)
        row_data.append(entry.other_deduction_remark or '')

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in editable_cols:
                cell.fill = input_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.fill = lock_fill
                if col_idx == 4:  # Gross Salary — right align
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'

    # ── Hidden reference row: payroll_id + employee_ids for upload mapping ──
    ref_sheet = wb.create_sheet('_ref')
    ref_sheet['A1'] = 'payroll_id'
    ref_sheet['B1'] = payroll_id
    ref_sheet['A2'] = 'working_days'
    ref_sheet['B2'] = payroll.working_days
    for idx, entry in enumerate(entries):
        ref_sheet.cell(row=idx + 3, column=1, value=entry.employee_id)
        ref_sheet.cell(row=idx + 3, column=2, value=entry.employee.name)
    ref_sheet.sheet_state = 'hidden'

    # ── Info Sheet (Instructions) ──
    inst = wb.create_sheet('Instructions')
    instructions = [
        ['BULK ATTENDANCE UPLOAD — INSTRUCTIONS'],
        [''],
        ['1.', 'Only fill the YELLOW (highlighted) columns in the Attendance sheet.'],
        ['2.', 'Do NOT change Sr No, Emp ID, Employee Name, or Gross Salary columns.'],
        ['3.', f'Days Present must be between 0 and {payroll.working_days} (working days).'],
        ['4.', 'Half days allowed — use 0.5 increments (e.g., 20.5, 21, 25.5).'],
        ['5.', 'Other Deduction is optional — enter amount like advance recovery, loan EMI etc.'],
        ['6.', 'Deduction Remark is optional — reason for deduction.'],
        ['7.', 'After filling, save the file and upload it back on the Payroll Processing page.'],
        ['8.', 'The system will auto-calculate: Absent days, Earned Gross, EPF, ESIC, PT, Net Pay.'],
        [''],
        ['IMPORTANT:'],
        ['', 'Do NOT add or delete rows. Do NOT rename the sheets.'],
        ['', 'Do NOT change the _ref (hidden) sheet — it maps employees to system records.'],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = inst.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=14, color='4338CA')
            elif value and str(value).startswith('IMPORTANT'):
                cell.font = Font(name='Calibri', bold=True, size=11, color='DC2626')
            else:
                cell.font = Font(name='Calibri', size=10)
    inst.column_dimensions['A'].width = 5
    inst.column_dimensions['B'].width = 80

    # Set Attendance as the active sheet
    wb.active = 0

    # Save to BytesIO and send
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Attendance_{est.company_name.replace(" ", "_")}_{payroll.period_display.replace(" ", "_")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@payroll_bp.route('/payroll/<int:payroll_id>/upload-attendance', methods=['POST'])
def upload_attendance(payroll_id):
    """Upload filled Excel template (monthly or universal hybrid) and populate attendance + rate overrides.
    For universal hybrid: reads rate/head columns and stores as rate_overrides JSON on PayrollEntry.
    Employees NOT in the uploaded file get 0 days attendance (absent full month)."""
    from openpyxl import load_workbook
    import io
    import json

    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    if payroll.status == 'finalized':
        flash('Cannot upload — payroll is finalized. Reopen first.', 'danger')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    file = request.files.get('attendance_file')
    if not file or not file.filename:
        flash('No file selected. Please choose an Excel file.', 'warning')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Invalid file format. Please upload an .xlsx Excel file.', 'danger')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception as e:
        flash(f'Could not read Excel file: {str(e)}', 'danger')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    # ── Detect template type ──
    is_universal = False
    col_map_from_file = None
    head_col_map_from_file = None
    salary_structure_from_file = None

    if '_colmap' in wb.sheetnames:
        # Universal hybrid template
        is_universal = True
        colmap_sheet = wb['_colmap']
        template_type = colmap_sheet['B1'].value
        salary_structure_from_file = colmap_sheet['B2'].value
        try:
            col_map_from_file = json.loads(colmap_sheet['B3'].value or '{}')
        except (json.JSONDecodeError, TypeError):
            col_map_from_file = {}
        try:
            head_col_map_raw = colmap_sheet['B4'].value
            if head_col_map_raw:
                head_col_map_from_file = json.loads(head_col_map_raw)
            else:
                head_col_map_from_file = {}
        except (json.JSONDecodeError, TypeError):
            head_col_map_from_file = {}

    elif '_ref' in wb.sheetnames:
        # Monthly template with _ref sheet
        ref_sheet = wb['_ref']
        file_payroll_id = ref_sheet['B1'].value
        if file_payroll_id and int(file_payroll_id) != payroll_id:
            flash(f'Template mismatch! This file was generated for a different payroll (ID: {file_payroll_id}).', 'danger')
            return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    # ── Build employee_id mapping from ref sheet (monthly template only) ──
    ref_emp_ids = None
    if not is_universal and '_ref' in wb.sheetnames:
        ref_sheet = wb['_ref']
        ref_emp_ids = []
        for row in ref_sheet.iter_rows(min_row=3, max_col=2):
            if row[0].value:
                ref_emp_ids.append(int(row[0].value))

    # ── Read Attendance Sheet ──
    if 'Attendance' not in wb.sheetnames:
        flash('Attendance sheet not found in the uploaded file.', 'danger')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    ws = wb['Attendance']

    # For universal template: column positions come from _colmap, NOT from config
    # For monthly template: use config to determine column layout
    if is_universal:
        has_ph = True   # Universal always has NPH & OT columns
        has_ot = True
    else:
        has_ph = config and config.paid_holiday_type == 'separate'
        has_ot = config and config.ot_applicable

    # ── Determine column positions ──
    if is_universal and col_map_from_file:
        # Use column map from the universal template's hidden sheet
        # col_map_from_file: {"sr": 1, "uan": 2, "esic_ip": 3, "emp_name": 4, ...}
        uan_col = col_map_from_file.get('uan', 2)
        esic_ip_col = col_map_from_file.get('esic_ip', 3)
        name_col = col_map_from_file.get('emp_name', 4)
        present_col = col_map_from_file.get('days_present')
        ph_col = col_map_from_file.get('nph')
        ot_col = col_map_from_file.get('ot_hours')
        other_ded_col = col_map_from_file.get('other_ded')
        remark_col = col_map_from_file.get('remark')
        rate_col = col_map_from_file.get('rate')  # For daily_wages / gross_only
        data_start_row = 6  # Row 4=group, Row 5=header, Row 6=data
    else:
        # Monthly template: Col 1=Sr, 2=EmpID, 3=Name, 4=Gross, 5=Present
        uan_col = 2  # UAN/ID column
        esic_ip_col = None
        name_col = 3
        present_col = 5
        next_col = 6
        ph_col = None
        ot_col = None
        rate_col = None
        if has_ph:
            ph_col = next_col
            next_col += 1
        if has_ot:
            ot_col = next_col
            next_col += 1
        other_ded_col = next_col
        remark_col = next_col + 1
        data_start_row = 6

    # ── Load all PayrollEntry records for this payroll ──
    entries_map = {}
    all_entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id).all()
    for entry in all_entries:
        entries_map[entry.employee_id] = entry

    # Track which employees were found in the upload (for zero-ing out missing ones)
    matched_emp_ids = set()

    # ── Parse rows and update entries ──
    updated_count = 0
    auto_added_count = 0   # employees auto-added because they joined after payroll creation
    rate_changes = []
    error_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=False), start=0):
        # Get Sr. value — skip if blank (end of data or empty row)
        sr_val = row[0].value
        if sr_val is None:
            # For universal template: continue scanning (might have blank rows in between)
            if is_universal:
                continue
            else:
                break  # Monthly template: end of data

        # ── Match employee ──
        emp_id = None

        if ref_emp_ids and row_idx < len(ref_emp_ids):
            # Monthly template: use _ref sheet mapping
            emp_id = ref_emp_ids[row_idx]
        else:
            # Universal or fallback: match by UAN → ESIC IP → Name
            uan_value = str(row[uan_col - 1].value).strip() if row[uan_col - 1].value else ''
            esic_value = ''
            if esic_ip_col:
                esic_value = str(row[esic_ip_col - 1].value).strip() if row[esic_ip_col - 1].value else ''
            emp_name_val = str(row[name_col - 1].value).strip() if row[name_col - 1].value else ''

            # Skip completely blank rows
            if not uan_value and not esic_value and not emp_name_val:
                continue

            for eid, entry_obj in entries_map.items():
                emp = entry_obj.employee
                # Match by UAN
                if uan_value and emp.uan_number and emp.uan_number.strip() == uan_value:
                    emp_id = eid
                    break
                # Match by ESIC IP
                if esic_value and emp.esic_ip_number and emp.esic_ip_number.strip() == esic_value:
                    emp_id = eid
                    break
                # Match by emp_code (for monthly template's ID column)
                if not is_universal and uan_value and emp.emp_code and emp.emp_code == uan_value:
                    emp_id = eid
                    break
            # Final fallback: match by name
            if not emp_id and emp_name_val:
                for eid, entry_obj in entries_map.items():
                    if entry_obj.employee.name.strip().upper() == emp_name_val.upper():
                        emp_id = eid
                        break

        if not emp_id or emp_id not in entries_map:
            # ── Auto-create PayrollEntry (universal template only) ──
            # Handles the common case where employees were imported AFTER the
            # payroll was created — payroll_create() only captures employees that
            # exist at creation time, so late-added employees have no entry.
            # We find the employee in the Employee table and create one on-the-fly.
            if is_universal and (uan_value or esic_value or emp_name_val):
                auto_emp = None
                if uan_value:
                    auto_emp = Employee.query.filter_by(
                        uan_number=uan_value,
                        establishment_id=est.id,
                        is_active=True
                    ).first()
                if not auto_emp and esic_value:
                    auto_emp = Employee.query.filter_by(
                        esic_ip_number=esic_value,
                        establishment_id=est.id,
                        is_active=True
                    ).first()
                if not auto_emp and emp_name_val:
                    auto_emp = Employee.query.filter(
                        db.func.upper(Employee.name) == emp_name_val.upper(),
                        Employee.establishment_id == est.id,
                        Employee.is_active == True
                    ).first()

                if auto_emp:
                    # Create missing PayrollEntry for this employee
                    salary_auto = EmployeeSalary.query.filter_by(
                        employee_id=auto_emp.id, is_current=True).first()
                    new_entry = PayrollEntry(
                        monthly_payroll_id=payroll_id,
                        employee_id=auto_emp.id,
                        days_present=payroll.working_days,
                        days_absent=0,
                        paid_holidays=0,
                        ot_hours=0,
                        total_payable_days=payroll.working_days,
                        gross_salary=salary_auto.gross_salary if salary_auto else 0
                    )
                    if hasattr(new_entry, 'rate_overrides'):
                        new_entry.rate_overrides = None
                    db.session.add(new_entry)
                    db.session.flush()
                    entries_map[auto_emp.id] = new_entry
                    emp_id = auto_emp.id
                    auto_added_count += 1
                    # Fall through — emp_id is now valid, attendance will be applied below
                else:
                    id_check = str(row[uan_col - 1].value).strip() if row[uan_col - 1].value else ''
                    name_check = str(row[name_col - 1].value).strip() if row[name_col - 1].value else ''
                    if id_check or name_check:
                        error_rows.append(f'Row {data_start_row + row_idx}: Not matched ({name_check or id_check})')
                    continue
            else:
                id_check = str(row[uan_col - 1].value).strip() if row[uan_col - 1].value else ''
                name_check = str(row[name_col - 1].value).strip() if row[name_col - 1].value else ''
                if id_check or name_check:
                    error_rows.append(f'Row {data_start_row + row_idx}: Not matched ({name_check or id_check})')
                continue

        entry = entries_map[emp_id]
        matched_emp_ids.add(emp_id)

        # ── Parse attendance values ──
        try:
            days_present = float(row[present_col - 1].value or 0) if present_col else 0
            days_present = max(0, min(days_present, payroll.working_days + 10))
        except (ValueError, TypeError):
            days_present = 0
            error_rows.append(f'Row {data_start_row + row_idx}: Invalid Days Present, set to 0')

        ph_value = 0
        if has_ph and ph_col:
            try:
                ph_value = float(row[ph_col - 1].value or 0)
            except (ValueError, TypeError):
                ph_value = 0

        ot_value = 0
        if has_ot and ot_col:
            try:
                ot_value = float(row[ot_col - 1].value or 0)
            except (ValueError, TypeError):
                ot_value = 0

        other_ded = 0
        if other_ded_col:
            try:
                other_ded = float(row[other_ded_col - 1].value or 0)
            except (ValueError, TypeError):
                other_ded = 0

        remark = ''
        if remark_col:
            try:
                remark = str(row[remark_col - 1].value or '').strip()
            except (ValueError, TypeError, IndexError):
                remark = ''

        # Update attendance on entry
        entry.days_present = days_present
        entry.ot_hours = ot_value
        entry.paid_holidays = ph_value
        entry.other_deduction = round(other_ded)
        entry.other_deduction_remark = remark or None

        # ── Parse rate overrides (universal template only) ──
        if is_universal:
            overrides = {}
            emp = entry.employee
            salary = EmployeeSalary.query.filter_by(employee_id=emp.id, is_current=True).first()

            # ── Read "Rate / Gross" column (always present) ──
            if rate_col:
                try:
                    rate_val = row[rate_col - 1].value
                    if rate_val is not None and str(rate_val).strip() != '':
                        rate_val = float(rate_val)
                        # Determine if this is daily_rate or gross based on employee type
                        emp_type = (salary.salary_type if salary and salary.salary_type else config.salary_type) or 'monthly_fixed'
                        if emp_type == 'daily_wages' or (salary and salary.daily_rate):
                            overrides['daily_rate'] = rate_val
                            if salary and salary.daily_rate and abs(rate_val - salary.daily_rate) > 0.5:
                                rate_changes.append(f'{emp.name}: Rate {round(salary.daily_rate)} → {round(rate_val)}')
                        else:
                            overrides['gross'] = rate_val
                            if salary and salary.gross_salary and abs(rate_val - salary.gross_salary) > 1:
                                rate_changes.append(f'{emp.name}: Gross {round(salary.gross_salary)} → {round(rate_val)}')
                except (ValueError, TypeError, IndexError):
                    pass

            # ── Read individual head columns (if present) ──
            if head_col_map_from_file:
                heads_override = {}
                for head_id_str, cidx in head_col_map_from_file.items():
                    try:
                        val = row[int(cidx) - 1].value
                        if val is not None and str(val).strip() != '':
                            heads_override[head_id_str] = float(val)
                    except (ValueError, TypeError, IndexError):
                        pass
                if heads_override:
                    overrides['heads'] = heads_override
                    # Calculate gross from heads if not already set from Rate column
                    if 'gross' not in overrides and 'daily_rate' not in overrides:
                        new_gross = sum(heads_override.values())
                        overrides['gross'] = new_gross
                        if salary and salary.gross_salary and abs(new_gross - salary.gross_salary) > 1:
                            rate_changes.append(f'{emp.name}: Gross {round(salary.gross_salary)} → {round(new_gross)}')

            # Store rate overrides as JSON on the payroll entry
            entry.rate_overrides = json.dumps(overrides) if overrides else None

            # ── Record salary history (EmployeeSalary) for this rate change ──
            # Effective from the first day of the payroll month so reports of that month
            # reflect the correct historical rate.
            if overrides and ('daily_rate' in overrides or 'gross' in overrides):
                from datetime import date as _date
                eff_date = _date(payroll.year, payroll.month, 1)
                new_daily = overrides.get('daily_rate')
                new_gross = overrides.get('gross')

                # Check if a salary row already exists with same effective_from
                existing = EmployeeSalary.query.filter_by(
                    employee_id=emp.id, effective_from=eff_date).first()

                # Determine if it's actually a change vs the current/base salary
                base_daily = salary.daily_rate if salary and salary.daily_rate else None
                base_gross = salary.gross_salary if salary and salary.gross_salary else None
                is_change = False
                if new_daily is not None and (base_daily is None or abs(float(new_daily) - float(base_daily)) > 0.009):
                    is_change = True
                if new_gross is not None and (base_gross is None or abs(float(new_gross) - float(base_gross)) > 0.5):
                    is_change = True

                if is_change or existing:
                    if existing:
                        # Update existing history row
                        if new_daily is not None:
                            existing.daily_rate = float(new_daily)
                        if new_gross is not None:
                            existing.gross_salary = float(new_gross)
                    else:
                        # Create new history row — copy structure from current salary
                        new_sal = EmployeeSalary(
                            employee_id=emp.id,
                            effective_from=eff_date,
                            gross_salary=float(new_gross) if new_gross is not None else (float(base_gross) if base_gross else 0),
                            daily_rate=float(new_daily) if new_daily is not None else base_daily,
                            salary_type=(salary.salary_type if salary else None),
                            is_current=False,
                        )
                        db.session.add(new_sal)
                        db.session.flush()

                    # Recompute is_current: latest effective_from <= today becomes current
                    from datetime import date as _today_date
                    today = _today_date.today()
                    all_sals = EmployeeSalary.query.filter_by(
                        employee_id=emp.id).order_by(
                        EmployeeSalary.effective_from.desc()).all()
                    current_found = False
                    for s in all_sals:
                        if not current_found and s.effective_from <= today:
                            s.is_current = True
                            current_found = True
                        else:
                            s.is_current = False

        updated_count += 1

    # ── Set 0 days for employees NOT in the upload (universal template only) ──
    zero_count = 0
    if is_universal:
        for eid, entry in entries_map.items():
            if eid not in matched_emp_ids:
                entry.days_present = 0
                entry.ot_hours = 0
                entry.paid_holidays = 0
                entry.rate_overrides = None
                zero_count += 1

    # Refresh total employee count (auto-adds may have increased it)
    payroll.total_employees = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id).count()
    db.session.commit()

    # ── Show result messages ──
    msg = f'Attendance uploaded for {updated_count} employees.'
    if auto_added_count > 0:
        msg += f' {auto_added_count} new employee(s) were automatically added to the payroll.'
    if zero_count > 0:
        msg += f' {zero_count} employees not in file — set to 0 days (absent).'
    flash(msg, 'success')

    if rate_changes:
        rate_msg = 'Rate changes detected: ' + '; '.join(rate_changes[:8])
        if len(rate_changes) > 8:
            rate_msg += f' (+{len(rate_changes) - 8} more)'
        flash(rate_msg, 'info')

    if error_rows:
        err_msg = 'Warnings: ' + '; '.join(error_rows[:5])
        if len(error_rows) > 5:
            err_msg += f' (+{len(error_rows) - 5} more)'
        flash(err_msg, 'warning')

    flash('Now click "Save & Calculate" to compute salary, EPF, ESIC, and Net Pay.', 'info')
    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


# ═════════════════════════════════════════════════════════════════
#  SYNC EMPLOYEES INTO PAYROLL
#  Adds PayrollEntry records for active employees registered after
#  the payroll was originally created.
# ═════════════════════════════════════════════════════════════════
@payroll_bp.route('/payroll/<int:payroll_id>/sync-employees', methods=['POST'])
def payroll_sync_employees(payroll_id):
    """Add PayrollEntry records for any active employees not yet in this payroll.
    Handles: employees uploaded/added after the payroll was first created."""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)

    if payroll.status == 'finalized':
        flash('Cannot sync — this payroll is already finalized. Reopen first.', 'danger')
        return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))

    # IDs already in this payroll
    existing_emp_ids = {
        row[0] for row in
        db.session.query(PayrollEntry.employee_id).filter_by(
            monthly_payroll_id=payroll_id).all()
    }

    last_day = date(payroll.year, payroll.month,
                    calendar.monthrange(payroll.year, payroll.month)[1])
    first_day = date(payroll.year, payroll.month, 1)

    # Find active employees for this establishment who have no entry yet
    query = Employee.query.filter(
        Employee.establishment_id == est.id,
        Employee.is_active == True,
        Employee.date_of_joining <= last_day
    )
    if existing_emp_ids:
        query = query.filter(~Employee.id.in_(list(existing_emp_ids)))

    new_employees = query.order_by(Employee.name).all()
    added = 0

    for emp in new_employees:
        salary = EmployeeSalary.query.filter_by(
            employee_id=emp.id, is_current=True).first()

        # Prorate for mid-month joiners
        if emp.date_of_joining > first_day:
            remaining = (last_day - emp.date_of_joining).days + 1
            default_present = min(remaining, payroll.working_days)
        else:
            default_present = payroll.working_days

        entry = PayrollEntry(
            monthly_payroll_id=payroll_id,
            employee_id=emp.id,
            days_present=default_present,
            days_absent=0,
            paid_holidays=0,
            ot_hours=0,
            total_payable_days=default_present,
            gross_salary=salary.gross_salary if salary else 0
        )
        if hasattr(entry, 'rate_overrides'):
            entry.rate_overrides = None
        db.session.add(entry)
        added += 1

    if added > 0:
        payroll.total_employees = len(existing_emp_ids) + added
        db.session.commit()
        flash(
            f'{added} employee(s) added to this payroll. '
            f'Total now: {payroll.total_employees} employees. '
            f'Download a fresh Universal Template and upload attendance.',
            'success'
        )
    else:
        flash('All active employees are already in this payroll — nothing to sync.', 'info')

    return redirect(url_for('payroll.payroll_process', payroll_id=payroll_id))


@payroll_bp.route('/payroll/<int:payroll_id>/download-universal-template')
def download_universal_template(payroll_id):
    """Download hybrid universal attendance template with rate/head columns.
    Pre-fills current employee data (KYC + Rates). Client modifies as needed.
    System uses template values when filled, falls back to stored values when blank.
    Employees NOT in the upload file get 0 days (absent full month)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from flask import send_file
    import json

    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    # Auto-sync: add any missing active employees before building the template
    if payroll.status != 'finalized' and not payroll.is_nil:
        _ex_ids = {r[0] for r in db.session.query(PayrollEntry.employee_id).filter_by(monthly_payroll_id=payroll_id).all()}
        _last = date(payroll.year, payroll.month, calendar.monthrange(payroll.year, payroll.month)[1])
        _first = date(payroll.year, payroll.month, 1)
        _q = Employee.query.filter(Employee.establishment_id == est.id, Employee.is_active == True, Employee.date_of_joining <= _last)
        if _ex_ids:
            _q = _q.filter(~Employee.id.in_(list(_ex_ids)))
        for _emp in _q.all():
            _sal = EmployeeSalary.query.filter_by(employee_id=_emp.id, is_current=True).first()
            _def = min(((_last - _emp.date_of_joining).days + 1), payroll.working_days) if _emp.date_of_joining > _first else payroll.working_days
            _ent = PayrollEntry(monthly_payroll_id=payroll_id, employee_id=_emp.id,
                                days_present=_def, days_absent=0, paid_holidays=0, ot_hours=0,
                                total_payable_days=_def, gross_salary=_sal.gross_salary if _sal else 0)
            if hasattr(_ent, 'rate_overrides'):
                _ent.rate_overrides = None
            db.session.add(_ent)
            _ex_ids.add(_emp.id)
        payroll.total_employees = len(_ex_ids)
        db.session.commit()

    # Get all active employees in this payroll with their salary records
    entries = PayrollEntry.query.filter_by(
        monthly_payroll_id=payroll_id
    ).join(Employee).order_by(Employee.name).all()

    # Get salary heads for this establishment (for with_heads salary structure)
    salary_heads = SalaryHead.query.filter_by(
        establishment_id=est.id, head_type='earning', is_active=True, is_in_gross=True
    ).order_by(SalaryHead.display_order).all()

    # UNIVERSAL = always show ALL columns regardless of config
    # Heads shown if establishment has any salary heads configured
    is_with_heads = len(salary_heads) > 0
    has_ph = True   # Always show NPH column in universal template
    has_ot = True    # Always show OT column in universal template

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # ── Styling ──
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
    kyc_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    rate_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    attend_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    ded_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    data_font = Font(name='Calibri', size=10)
    input_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
    rate_input_fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')
    lock_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # ═══════════════════════════════════════════
    # BUILD COLUMN STRUCTURE DYNAMICALLY
    # ═══════════════════════════════════════════
    # Columns are grouped:
    #   KYC: Sr, UAN, ESIC IP, Employee Name, Father/Husband Name
    #   RATE: Daily Rate / Gross -OR- individual heads (Basic, DA, HRA...)
    #   ATTENDANCE: Days Present, NPH, OT
    #   DEDUCTION: Other Ded, Remark

    col_map = {}  # column_key -> (col_index_1based, header_name, group, width)
    col_idx = 1

    # ── KYC Columns ──
    col_map['sr'] = (col_idx, 'Sr.', 'kyc', 5); col_idx += 1
    col_map['uan'] = (col_idx, 'UAN', 'kyc', 15); col_idx += 1
    col_map['esic_ip'] = (col_idx, 'ESIC IP', 'kyc', 14); col_idx += 1
    col_map['emp_name'] = (col_idx, 'Employee Name', 'kyc', 28); col_idx += 1
    col_map['father_name'] = (col_idx, 'Father/Husband', 'kyc', 22); col_idx += 1

    # ── Rate / Head Columns ──
    # ALWAYS include "Rate" column (daily_rate or gross for simple salary employees)
    # PLUS individual head columns if establishment has salary heads configured
    head_col_map = {}  # salary_head_id -> col_index
    col_map['rate'] = (col_idx, 'Rate / Gross', 'rate', 14); col_idx += 1
    if is_with_heads:
        for sh in salary_heads:
            key = f'head_{sh.id}'
            col_map[key] = (col_idx, sh.short_code or sh.name, 'rate', 12)
            head_col_map[sh.id] = col_idx
            col_idx += 1

    # ── Attendance Columns ──
    col_map['days_present'] = (col_idx, 'Days Present', 'attend', 13); col_idx += 1
    if has_ph:
        col_map['nph'] = (col_idx, 'NPH', 'attend', 8); col_idx += 1
    if has_ot:
        col_map['ot_hours'] = (col_idx, 'OT (Hrs/Days)', 'attend', 13); col_idx += 1

    # ── Deduction Columns ──
    col_map['other_ded'] = (col_idx, 'Other Ded.', 'deduct', 12); col_idx += 1
    col_map['remark'] = (col_idx, 'Remark', 'deduct', 22); col_idx += 1

    total_cols = col_idx - 1

    # ── Info Header (Rows 1-4) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f'UNIVERSAL ATTENDANCE TEMPLATE — {est.company_name}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='0D9488')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws['A2'] = 'Reusable for any month. Fill data and upload in Payroll Processing page. System matches by UAN / ESIC IP / Name.'
    ws['A2'].font = Font(name='Calibri', size=10, color='64748B')

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    ws['A3'] = 'YELLOW = Attendance (must fill)  |  PURPLE = Rates (fill if changed, blank = use stored rate)  |  Do NOT change Sr. column.'
    ws['A3'].font = Font(name='Calibri', bold=True, size=9, color='DC2626')

    # ── Group Header Row (Row 4) — color-coded section labels ──
    group_header_row = 4
    group_fills = {'kyc': kyc_fill, 'rate': rate_fill, 'attend': attend_fill, 'deduct': ded_fill}
    group_labels = {'kyc': 'KYC INFORMATION', 'rate': 'EARNING / RATE', 'attend': 'ATTENDANCE', 'deduct': 'DEDUCTIONS'}

    # Find column ranges for each group
    group_ranges = {}
    for key, (cidx, name, group, width) in col_map.items():
        if group not in group_ranges:
            group_ranges[group] = [cidx, cidx]
        else:
            group_ranges[group][1] = cidx

    for group, (start_c, end_c) in group_ranges.items():
        if start_c == end_c:
            cell = ws.cell(row=group_header_row, column=start_c, value=group_labels.get(group, ''))
        else:
            ws.merge_cells(start_row=group_header_row, start_column=start_c, end_row=group_header_row, end_column=end_c)
            cell = ws.cell(row=group_header_row, column=start_c, value=group_labels.get(group, ''))
        cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill = group_fills.get(group, header_fill)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        # Fill all cells in merged range with the same fill
        for c in range(start_c, end_c + 1):
            ws.cell(row=group_header_row, column=c).fill = group_fills.get(group, header_fill)
            ws.cell(row=group_header_row, column=c).border = thin_border

    # ── Column Headers (Row 5) ──
    header_row = 5
    for key, (cidx, name, group, width) in col_map.items():
        cell = ws.cell(row=header_row, column=cidx, value=name)
        cell.font = header_font
        cell.fill = group_fills.get(group, header_fill)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(cidx)].width = width

    # ═══════════════════════════════════════════
    # PRE-FILL EMPLOYEE DATA
    # ═══════════════════════════════════════════
    for idx, entry in enumerate(entries, 1):
        row = header_row + idx
        emp = entry.employee
        salary = EmployeeSalary.query.filter_by(employee_id=emp.id, is_current=True).first()

        # KYC columns (locked — grey background)
        ws.cell(row=row, column=col_map['sr'][0], value=idx).font = data_font
        ws.cell(row=row, column=col_map['sr'][0]).fill = lock_fill
        ws.cell(row=row, column=col_map['sr'][0]).border = thin_border
        ws.cell(row=row, column=col_map['sr'][0]).alignment = Alignment(horizontal='center')

        uan_cell = ws.cell(row=row, column=col_map['uan'][0], value=emp.uan_number or '')
        uan_cell.font = data_font; uan_cell.fill = lock_fill; uan_cell.border = thin_border

        esic_cell = ws.cell(row=row, column=col_map['esic_ip'][0], value=emp.esic_ip_number or '')
        esic_cell.font = data_font; esic_cell.fill = lock_fill; esic_cell.border = thin_border

        name_cell = ws.cell(row=row, column=col_map['emp_name'][0], value=emp.name)
        name_cell.font = data_font; name_cell.fill = lock_fill; name_cell.border = thin_border

        father_cell = ws.cell(row=row, column=col_map['father_name'][0], value=emp.father_husband_name or '')
        father_cell.font = data_font; father_cell.fill = lock_fill; father_cell.border = thin_border

        # ── Rate / Gross column (ALWAYS present — pre-fill with daily_rate or gross) ──
        rate_val = ''
        if salary:
            if salary.daily_rate:
                rate_val = round(salary.daily_rate)
            elif salary.gross_salary:
                rate_val = round(salary.gross_salary)
        c = ws.cell(row=row, column=col_map['rate'][0], value=rate_val)
        c.font = data_font; c.fill = rate_input_fill; c.border = thin_border
        c.alignment = Alignment(horizontal='right')
        if rate_val:
            c.number_format = '#,##0'

        # ── Head columns (if establishment has heads — pre-fill with stored amounts) ──
        if is_with_heads and head_col_map:
            if salary and salary.head_values:
                for hv in salary.head_values:
                    if hv.salary_head_id in head_col_map:
                        c = ws.cell(row=row, column=head_col_map[hv.salary_head_id], value=round(hv.amount) if hv.amount else '')
                        c.font = data_font; c.fill = rate_input_fill; c.border = thin_border
                        c.alignment = Alignment(horizontal='right')
                        c.number_format = '#,##0'
            # Fill any head columns that don't have values with empty purple cells
            for sh_id, cidx in head_col_map.items():
                if ws.cell(row=row, column=cidx).value is None:
                    c = ws.cell(row=row, column=cidx, value='')
                    c.font = data_font; c.fill = rate_input_fill; c.border = thin_border

        # Attendance columns (yellow — must fill)
        dp_cell = ws.cell(row=row, column=col_map['days_present'][0], value='')
        dp_cell.font = data_font; dp_cell.fill = input_fill; dp_cell.border = thin_border
        dp_cell.alignment = Alignment(horizontal='center')

        if has_ph and 'nph' in col_map:
            ph_cell = ws.cell(row=row, column=col_map['nph'][0], value='')
            ph_cell.font = data_font; ph_cell.fill = input_fill; ph_cell.border = thin_border
            ph_cell.alignment = Alignment(horizontal='center')

        if has_ot and 'ot_hours' in col_map:
            ot_cell = ws.cell(row=row, column=col_map['ot_hours'][0], value='')
            ot_cell.font = data_font; ot_cell.fill = input_fill; ot_cell.border = thin_border
            ot_cell.alignment = Alignment(horizontal='center')

        # Deduction columns (yellow — optional)
        ded_cell = ws.cell(row=row, column=col_map['other_ded'][0], value='')
        ded_cell.font = data_font; ded_cell.fill = input_fill; ded_cell.border = thin_border
        ded_cell.alignment = Alignment(horizontal='right')

        rem_cell = ws.cell(row=row, column=col_map['remark'][0], value='')
        rem_cell.font = data_font; rem_cell.fill = input_fill; rem_cell.border = thin_border

    # ── Extra blank rows (for new employees client may add) ──
    for idx in range(len(entries) + 1, len(entries) + 11):
        row = header_row + idx
        for key, (cidx, name, group, width) in col_map.items():
            c = ws.cell(row=row, column=cidx, value='')
            c.font = data_font
            c.border = thin_border
            c.fill = input_fill if group != 'kyc' else lock_fill
            c.alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=col_map['sr'][0], value=idx)

    # ── Hidden _colmap sheet — stores column structure for upload parsing ──
    ref = wb.create_sheet('_colmap')
    ref['A1'] = 'template_type'
    ref['B1'] = 'universal_hybrid'
    ref['A2'] = 'salary_structure'
    ref['B2'] = 'with_heads' if is_with_heads else 'gross_only'
    # Store column mapping as JSON
    col_export = {}
    for key, (cidx, name, group, width) in col_map.items():
        col_export[key] = cidx
    ref['A3'] = 'col_map_json'
    ref['B3'] = json.dumps(col_export)
    # Store head column mapping
    if head_col_map:
        ref['A4'] = 'head_col_map_json'
        ref['B4'] = json.dumps({str(k): v for k, v in head_col_map.items()})
    ref.sheet_state = 'hidden'

    # ── Instructions Sheet ──
    inst = wb.create_sheet('Instructions')
    instructions = [
        ['UNIVERSAL ATTENDANCE TEMPLATE — INSTRUCTIONS'],
        [''],
        ['This template can be reused every month for this establishment.'],
        ['Just update attendance and rates each month, save, and upload.'],
        [''],
        ['COLOR GUIDE:'],
        ['', 'GREY columns (KYC) — UAN, ESIC IP, Name, Father Name. Pre-filled from system. Do not change unless wrong.'],
        ['', ''],
        ['', 'PURPLE columns (Rate/Heads) — Pre-filled with current rates from system.'],
        ['', '  → "Rate / Gross" = Daily Rate for daily-wages employees, Monthly Gross for salaried.'],
        ['', '  → Head columns (BASIC, DA, HRA, etc.) = individual salary head amounts (if applicable).'],
        ['', '  → If rate is SAME as last month: leave as-is. System uses the pre-filled rate.'],
        ['', '  → If rate CHANGED this month: update the value. System uses the NEW rate.'],
        ['', '  → If blank: system falls back to stored rate from employee records.'],
        ['', '  → NOTE: For daily-wages employees, just fill "Rate / Gross". Head columns are optional.'],
        ['', ''],
        ['', 'YELLOW columns (Attendance) — Fill for each employee.'],
        ['', '  → Days Present = actual days worked (half days allowed: 20.5, 21, 25.5).'],
        ['', '  → Days Present BLANK or 0 = employee absent full month (0 salary).'],
        ['', '  → NPH = National/Paid Holidays (if applicable).'],
        ['', '  → OT = Overtime hours or days.'],
        ['', ''],
        ['', 'YELLOW columns (Deductions) — Optional.'],
        ['', '  → Other Ded. = Advance recovery, Loan EMI, Canteen, etc.'],
        ['', '  → Remark = reason for deduction.'],
        [''],
        ['HOW TO USE:'],
        ['1.', 'Fill Rate (purple) and Attendance (yellow) columns for each employee.'],
        ['2.', 'Remove employees who did not work at all (or leave Days Present blank/0).'],
        ['3.', 'Save the Excel file.'],
        ['4.', 'Go to Payroll Processing page for the month → click "Upload Attendance".'],
        ['5.', 'Upload this file. System matches employees by UAN → ESIC IP → Name.'],
        ['6.', 'Employees NOT in the file automatically get 0 days (absent full month).'],
        ['7.', 'Click "Save & Calculate" to compute Gross, EPF, ESIC, PT, Net Pay.'],
        [''],
        ['IMPORTANT:'],
        ['', 'Do NOT delete the hidden sheets — they help the system read column positions.'],
        ['', 'You can reorder employee rows but do NOT change column headers or sheet names.'],
        ['', 'This same file works for ANY month — just update the numbers and re-upload.'],
        ['', 'Rate changes are shown after upload so you can verify before calculating.'],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = inst.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=14, color='0D9488')
            elif value and ('IMPORTANT' in str(value) or 'COLOR GUIDE' in str(value) or 'HOW TO USE' in str(value)):
                cell.font = Font(name='Calibri', bold=True, size=11, color='DC2626')
            elif value and str(value).startswith('  →'):
                cell.font = Font(name='Calibri', size=10, color='7C3AED')
            else:
                cell.font = Font(name='Calibri', size=10)
    inst.column_dimensions['A'].width = 5
    inst.column_dimensions['B'].width = 90

    wb.active = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Universal_Template_{est.company_name.replace(" ", "_")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
