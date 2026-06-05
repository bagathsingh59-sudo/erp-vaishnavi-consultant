"""
Paid Leave (Earned Leave with Wages) routes — modelled on the Bonus module.

Annual statement, calendar year (Jan → Dec), per establishment.
  PL Amount = (Total Attendance ÷ divisor) × December Daily Rate
  Eligible  = Total Attendance ≥ threshold (default 240 per Sec. 79)

Three layout modes for the Excel statement:
  • mixed           — single sheet, employees mixed
  • separate_sheets — two sheets: "Eligible" and "Not Eligible"
  • top_bottom      — single sheet, eligible block above non-eligible block

Plus a manual "add days" feature so the consultant can push specific
employees over the threshold when the client requests it.
"""
import io
import json
import calendar
from datetime import datetime

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file)

from app import db
from app.models.paid_leave import PaidLeaveRun, PaidLeaveEntry
from app.models.establishment import Establishment
from app.models.employee import Employee
from app.models.payroll import MonthlyPayroll, PayrollEntry
from app.user_context import (user_establishments, verify_est_ownership,
                              capture_est_from_url)

paid_leave_bp = Blueprint('paid_leave', __name__)


@paid_leave_bp.before_request
def _capture_url_establishment():
    if request.path and '/api/' in request.path:
        return None
    capture_est_from_url()
    return None


# ────────────────────────────────────────────────────────────────────────
# Engine — calculates all entries for a PaidLeaveRun
# ────────────────────────────────────────────────────────────────────────
def _calculate_paid_leave_run(run):
    """(Re)build PaidLeaveEntry rows by scanning Jan-Dec payroll data."""
    inc_nph    = bool(run.include_holiday_attendance)
    skip_zero  = bool(run.skip_zero_attendance)
    threshold  = run.eligibility_threshold or 240
    divisor    = run.eligibility_divisor or 20

    # All payrolls in the calendar year for this establishment
    payrolls = MonthlyPayroll.query.filter(
        MonthlyPayroll.establishment_id == run.establishment_id,
        MonthlyPayroll.year == run.year,
        MonthlyPayroll.month >= 1,
        MonthlyPayroll.month <= 12,
    ).all()
    if not payrolls:
        # Wipe entries — there's nothing to compute
        PaidLeaveEntry.query.filter_by(paid_leave_run_id=run.id).delete()
        run.total_employees = 0
        run.eligible_employees = 0
        run.total_pl_amount = 0
        db.session.commit()
        return

    payroll_by_id = {p.id: p for p in payrolls}
    december_payroll = next((p for p in payrolls if p.month == 12), None)

    # Existing manual additions per employee (preserved across recalcs)
    existing_manuals = {
        e.employee_id: e.manual_addition or 0
        for e in PaidLeaveEntry.query.filter_by(paid_leave_run_id=run.id).all()
    }

    # All entries this year
    entries = PayrollEntry.query.filter(
        PayrollEntry.monthly_payroll_id.in_([p.id for p in payrolls])
    ).all()

    # Group per employee
    emp_data = {}   # emp_id -> { monthly_dict, base_attendance }
    for entry in entries:
        payroll = payroll_by_id.get(entry.monthly_payroll_id)
        if not payroll:
            continue
        emp_id = entry.employee_id
        if not emp_id:
            continue
        present = float(entry.days_present or 0)
        ph      = float(entry.paid_holidays or 0)
        att     = present + ph if inc_nph else present

        bucket = emp_data.setdefault(emp_id, {
            'monthly': {},
            'base_attendance': 0.0,
        })
        bucket['monthly'][f'{payroll.year}-{payroll.month:02d}'] = round(att, 1)
        bucket['base_attendance'] += att

    # December rate per employee — pulled from the December payroll
    december_rate_by_emp = {}
    if december_payroll:
        dec_entries = [e for e in entries
                       if e.monthly_payroll_id == december_payroll.id]
        wd = december_payroll.working_days or 26
        for e in dec_entries:
            gross = float(e.gross_salary or 0)
            if gross <= 0:
                december_rate_by_emp[e.employee_id] = 0.0
            elif gross <= 2000:
                december_rate_by_emp[e.employee_id] = gross
            else:
                december_rate_by_emp[e.employee_id] = round(gross / wd, 2)

    # Wipe and rebuild
    PaidLeaveEntry.query.filter_by(paid_leave_run_id=run.id).delete()
    db.session.flush()

    total_emp = 0
    eligible_emp = 0
    total_pl = 0.0

    for emp_id, d in emp_data.items():
        base_att = d['base_attendance']
        if skip_zero and base_att <= 0:
            continue

        manual_add = float(existing_manuals.get(emp_id, 0))
        total_att = base_att + manual_add
        eligible_att = round(total_att / divisor, 2) if divisor else 0
        dec_rate = float(december_rate_by_emp.get(emp_id, 0))
        pl_amount = round(eligible_att * dec_rate)

        is_eligible = total_att >= threshold
        reason = None
        if not is_eligible:
            reason = (f'Total attendance {int(total_att)} days '
                      f'(threshold {threshold})')

        ple = PaidLeaveEntry(
            paid_leave_run_id=run.id,
            employee_id=emp_id,
            monthly_data=json.dumps(d['monthly']),
            base_attendance=round(base_att, 2),
            manual_addition=manual_add,
            total_attendance=round(total_att, 2),
            eligible_attendance=eligible_att,
            december_rate=dec_rate,
            pl_amount=pl_amount,
            is_eligible=is_eligible,
            ineligibility_reason=reason,
        )
        db.session.add(ple)
        total_emp += 1
        if is_eligible:
            eligible_emp += 1
            total_pl += pl_amount

    run.total_employees    = total_emp
    run.eligible_employees = eligible_emp
    run.total_pl_amount    = round(total_pl, 2)
    db.session.commit()


# ────────────────────────────────────────────────────────────────────────
# Routes: list / create / view / recalc / finalize / delete
# ────────────────────────────────────────────────────────────────────────
@paid_leave_bp.route('/paid-leave')
def paid_leave_list():
    runs = PaidLeaveRun.query.join(Establishment).filter(
        Establishment.id.in_([e.id for e in user_establishments().all()])
    ).order_by(PaidLeaveRun.year.desc(), PaidLeaveRun.id.desc()).all()
    return render_template('paid_leave/list.html', runs=runs)


@paid_leave_bp.route('/paid-leave/new', methods=['GET', 'POST'])
def paid_leave_new():
    ests = user_establishments().order_by(Establishment.company_name).all()

    if request.method == 'POST':
        est_id = request.form.get('establishment_id', type=int)
        year = request.form.get('year', type=int)
        inc_nph = 'include_holiday_attendance' in request.form
        skip_zero = 'skip_zero_attendance' in request.form
        threshold = request.form.get('eligibility_threshold', type=int) or 240
        divisor = request.form.get('eligibility_divisor', type=int) or 20
        layout_mode = request.form.get('layout_mode', 'top_bottom').strip()

        est = Establishment.query.get_or_404(est_id)
        verify_est_ownership(est)

        existing = PaidLeaveRun.query.filter_by(
            establishment_id=est_id, year=year
        ).first()
        if existing:
            flash(f'A paid-leave run for {est.company_name} {year} '
                  'already exists. Open it or delete it first.', 'warning')
            return redirect(url_for('paid_leave.paid_leave_view',
                                    run_id=existing.id))

        run = PaidLeaveRun(
            establishment_id=est_id,
            year=year,
            include_holiday_attendance=inc_nph,
            skip_zero_attendance=skip_zero,
            eligibility_threshold=threshold,
            eligibility_divisor=divisor,
            layout_mode=layout_mode,
            status='draft',
        )
        db.session.add(run)
        db.session.commit()

        _calculate_paid_leave_run(run)
        flash(f'Paid Leave run created for {est.company_name} — {year}',
              'success')
        return redirect(url_for('paid_leave.paid_leave_view', run_id=run.id))

    current_year = datetime.now().year
    years = list(range(current_year - 5, current_year + 1))
    return render_template('paid_leave/create.html', establishments=ests,
                           years=years, current_year=current_year)


@paid_leave_bp.route('/paid-leave/<int:run_id>')
def paid_leave_view(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = (PaidLeaveEntry.query
               .filter_by(paid_leave_run_id=run_id)
               .join(Employee)
               .order_by(Employee.name)
               .all())
    for e in entries:
        try:
            e.monthly = json.loads(e.monthly_data) if e.monthly_data else {}
        except Exception:
            e.monthly = {}
    return render_template('paid_leave/view.html', run=run, entries=entries)


@paid_leave_bp.route('/paid-leave/<int:run_id>/recalculate', methods=['POST'])
def paid_leave_recalculate(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    if run.status == 'finalized':
        flash('Cannot recalculate a finalized run. Unfinalize first.',
              'warning')
        return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))

    run.include_holiday_attendance = 'include_holiday_attendance' in request.form
    run.skip_zero_attendance       = 'skip_zero_attendance' in request.form
    run.eligibility_threshold = (request.form.get('eligibility_threshold',
                                                   type=int)
                                  or run.eligibility_threshold)
    run.eligibility_divisor = (request.form.get('eligibility_divisor',
                                                 type=int)
                                or run.eligibility_divisor)
    run.layout_mode = (request.form.get('layout_mode',
                                          run.layout_mode).strip()
                        or run.layout_mode)
    db.session.commit()

    _calculate_paid_leave_run(run)
    flash('Paid leave recalculated with updated settings.', 'success')
    return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))


@paid_leave_bp.route('/paid-leave/<int:run_id>/entry/<int:entry_id>/manual',
                     methods=['POST'])
def paid_leave_set_manual(run_id, entry_id):
    """Manually add (or remove) days for one employee — used when the
    client asks to bump an employee over the 240-day threshold."""
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    if run.status == 'finalized':
        flash('Cannot edit a finalized run.', 'warning')
        return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))

    entry = PaidLeaveEntry.query.filter_by(id=entry_id,
                                             paid_leave_run_id=run_id).first_or_404()
    try:
        manual = float(request.form.get('manual_addition') or 0)
    except ValueError:
        manual = 0.0
    if manual < 0:
        manual = 0.0
    entry.manual_addition = manual
    entry.remarks = (request.form.get('remarks') or '').strip()[:200] or None

    # Recompute this row only — total attendance, eligibility, pl_amount
    entry.total_attendance = round((entry.base_attendance or 0) + manual, 2)
    divisor = run.eligibility_divisor or 20
    entry.eligible_attendance = round(entry.total_attendance / divisor, 2) if divisor else 0
    entry.pl_amount = round(entry.eligible_attendance * (entry.december_rate or 0))
    threshold = run.eligibility_threshold or 240
    entry.is_eligible = entry.total_attendance >= threshold
    entry.ineligibility_reason = (
        None if entry.is_eligible
        else f'Total attendance {int(entry.total_attendance)} days (threshold {threshold})'
    )
    db.session.commit()

    # Refresh run-level totals
    rebuilt = PaidLeaveEntry.query.filter_by(paid_leave_run_id=run_id).all()
    run.eligible_employees = sum(1 for x in rebuilt if x.is_eligible)
    run.total_pl_amount = round(sum(x.pl_amount for x in rebuilt if x.is_eligible), 2)
    db.session.commit()

    flash(f'Updated manual addition for {entry.employee.name}.', 'success')
    return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))


@paid_leave_bp.route('/paid-leave/<int:run_id>/finalize', methods=['POST'])
def paid_leave_finalize(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    payment_date_str = request.form.get('payment_date')
    if payment_date_str:
        try:
            run.payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    run.status = 'finalized'
    run.finalized_at = datetime.utcnow()
    db.session.commit()
    flash('Paid leave run finalized.', 'success')
    return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))


@paid_leave_bp.route('/paid-leave/<int:run_id>/unfinalize', methods=['POST'])
def paid_leave_unfinalize(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    run.status = 'draft'
    run.finalized_at = None
    db.session.commit()
    flash('Unfinalized — you can now edit.', 'info')
    return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))


@paid_leave_bp.route('/paid-leave/<int:run_id>/delete', methods=['POST'])
def paid_leave_delete(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    db.session.delete(run)
    db.session.commit()
    flash('Paid leave run deleted.', 'info')
    return redirect(url_for('paid_leave.paid_leave_list'))


# ────────────────────────────────────────────────────────────────────────
# Excel exports
# ────────────────────────────────────────────────────────────────────────
def _sort_entries(entries, layout_mode):
    """Order entries per the chosen layout."""
    if layout_mode == 'mixed':
        return sorted(entries, key=lambda e: (e.employee.name or '').upper())
    # 'top_bottom' and 'separate_sheets' both want eligible-first ordering
    eligibles = sorted([e for e in entries if e.is_eligible],
                       key=lambda e: (e.employee.name or '').upper())
    non_eligibles = sorted([e for e in entries if not e.is_eligible],
                            key=lambda e: (e.employee.name or '').upper())
    return eligibles, non_eligibles


def _build_pl_workbook(run, entries):
    """Build the Vaishnavi-format Paid Leave Statement workbook.
    Returns a BytesIO + filename."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    months = [(m, calendar.month_abbr[m]) for m in range(1, 13)]
    KYC_COLS = 3            # UAN/ESIC | Name | (we'll add Father in remarks later if needed)
    MONTH_COLS = 12
    TAIL_COLS = 5           # Total Att | Eligible Att | Rate | PL Amount | Remarks
    LAST = KYC_COLS + MONTH_COLS + TAIL_COLS

    # ── Styles ─────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=14, name='Calibri')
    sub_font   = Font(size=9, italic=True, color='475569', name='Calibri')
    head_font  = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
    body_font  = Font(size=9, name='Calibri')
    bold_body  = Font(bold=True, size=9, name='Calibri')

    thin = Side(border_style='thin', color='94A3B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    slate_fill   = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    eligible_fill = PatternFill('solid', start_color='DCFCE7', end_color='DCFCE7')
    not_elig_fill = PatternFill('solid', start_color='FEE2E2', end_color='FEE2E2')
    section_fill = PatternFill('solid', start_color='F1F5F9', end_color='F1F5F9')

    def write_sheet_header(ws, title_text):
        ws.cell(row=1, column=1, value=title_text).font = title_font
        ws.cell(row=1, column=1).alignment = center
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)

        meta = (f"{run.establishment.company_name}  |  Year {run.year}  |  "
                f"NPH in attendance: {'YES' if run.include_holiday_attendance else 'NO'}  |  "
                f"Eligibility: ≥ {run.eligibility_threshold} days  |  "
                f"Divisor: {run.eligibility_divisor}")
        ws.cell(row=2, column=1, value=meta).font = sub_font
        ws.cell(row=2, column=1).alignment = center
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)

    def write_table_header(ws, header_row):
        headers = (['UAN / ESIC', 'Employee Name']
                   + [f"{abbr}-{str(run.year)[-2:]}" for _, abbr in months]
                   + ['Total\nAttendance', 'Eligible\nAtt (÷' + str(run.eligibility_divisor) + ')',
                      'Rate\n(Dec)', 'PL\nAmount', 'Remarks'])
        # We have 2 KYC labels + 12 months + 5 tail = 19, but KYC_COLS=3. Adjust.
        # Actually KYC = 2 columns here (UAN/ESIC, Name).  Re-fix LAST.
        for col, text in enumerate(headers, 1):
            c = ws.cell(row=header_row, column=col, value=text)
            c.font = head_font
            c.fill = slate_fill
            c.alignment = center
            c.border = border
        ws.row_dimensions[header_row].height = 32

    def write_row(ws, row_idx, entry, is_section=False):
        emp = entry.employee
        uan_or_esic = emp.uan_number or emp.esic_ip_number or '—'
        cells = [
            (1, uan_or_esic, left, body_font),
            (2, emp.name,    left, bold_body),
        ]
        for col, val, align, font in cells:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = font
            c.alignment = align
            c.border = border

        try:
            mdata = json.loads(entry.monthly_data) if entry.monthly_data else {}
        except Exception:
            mdata = {}
        for j, (mth, _) in enumerate(months):
            key = f"{run.year}-{mth:02d}"
            val = mdata.get(key)
            c = ws.cell(row=row_idx, column=3 + j,
                        value=int(round(val)) if val else 0)
            c.font = body_font
            c.alignment = center
            c.border = border

        tail_col = 3 + MONTH_COLS
        total_att = entry.total_attendance or 0
        if (entry.manual_addition or 0) > 0:
            total_disp = f"{int(round(entry.base_attendance or 0))}+{int(entry.manual_addition)}={int(total_att)}"
        else:
            total_disp = int(round(total_att))
        ws.cell(row=row_idx, column=tail_col,     value=total_disp).font = bold_body
        ws.cell(row=row_idx, column=tail_col + 1, value=entry.eligible_attendance or 0).number_format = '0.00'
        ws.cell(row=row_idx, column=tail_col + 2, value=round(entry.december_rate or 0)).number_format = '#,##0'
        ws.cell(row=row_idx, column=tail_col + 3, value=int(entry.pl_amount or 0)).number_format = '#,##0'
        ws.cell(row=row_idx, column=tail_col + 4, value=entry.remarks or '')

        for col in range(tail_col, tail_col + 5):
            c = ws.cell(row=row_idx, column=col)
            if not c.font or c.font.name != 'Calibri':
                c.font = body_font
            c.alignment = center if col != tail_col + 4 else left
            c.border = border
            if entry.is_eligible:
                c.fill = eligible_fill
            elif not is_section:
                c.fill = not_elig_fill
        if entry.is_eligible:
            for col in range(1, tail_col):
                ws.cell(row=row_idx, column=col).fill = eligible_fill
        elif not is_section:
            for col in range(1, tail_col):
                ws.cell(row=row_idx, column=col).fill = not_elig_fill

    def write_section_band(ws, row_idx, label):
        c = ws.cell(row=row_idx, column=1, value=label)
        c.font = Font(bold=True, size=11, color='1E293B', name='Calibri')
        c.alignment = center
        c.fill = section_fill
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=LAST)
        ws.row_dimensions[row_idx].height = 22

    # Recompute LAST since KYC labels are 2 (UAN, Name) not 3
    LAST = 2 + MONTH_COLS + TAIL_COLS   # 19

    def setup_print(ws):
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 5
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        # Column widths
        widths = [16, 22] + [6] * 12 + [11, 10, 9, 11, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    if run.layout_mode == 'separate_sheets':
        ws_e = wb.active
        ws_e.title = "Eligible"
        ws_n = wb.create_sheet("Not Eligible")

        eligibles, non_eligibles = _sort_entries(entries, 'top_bottom')

        for ws, group, hdr in [
            (ws_e, eligibles,      "PAID LEAVE STATEMENT — ELIGIBLE EMPLOYEES"),
            (ws_n, non_eligibles,  "PAID LEAVE STATEMENT — NOT ELIGIBLE EMPLOYEES"),
        ]:
            write_sheet_header(ws, hdr)
            write_table_header(ws, 4)
            for i, e in enumerate(group):
                write_row(ws, 5 + i, e)
            setup_print(ws)
            ws.freeze_panes = 'C5'

    elif run.layout_mode == 'top_bottom':
        ws = wb.active
        ws.title = "Paid Leave Statement"
        write_sheet_header(ws,
            f"PAID LEAVE STATEMENT — {run.establishment.company_name.upper()} — {run.year}")
        write_table_header(ws, 4)

        eligibles, non_eligibles = _sort_entries(entries, 'top_bottom')
        row = 5
        if eligibles:
            write_section_band(ws, row, f"ELIGIBLE ({len(eligibles)})")
            row += 1
            for e in eligibles:
                write_row(ws, row, e)
                row += 1
        if non_eligibles:
            write_section_band(ws, row, f"NOT ELIGIBLE ({len(non_eligibles)})")
            row += 1
            for e in non_eligibles:
                write_row(ws, row, e)
                row += 1
        setup_print(ws)
        ws.freeze_panes = 'C5'

    else:  # mixed
        ws = wb.active
        ws.title = "Paid Leave Statement"
        write_sheet_header(ws,
            f"PAID LEAVE STATEMENT — {run.establishment.company_name.upper()} — {run.year}")
        write_table_header(ws, 4)
        sorted_entries = _sort_entries(entries, 'mixed')
        for i, e in enumerate(sorted_entries):
            write_row(ws, 5 + i, e)
        setup_print(ws)
        ws.freeze_panes = 'C5'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    safe = (run.establishment.company_name or 'Establishment').replace(' ', '_').replace('/', '_')[:60]
    filename = f"Paid_Leave_Statement_{safe}_{run.year}.xlsx"
    return out, filename


@paid_leave_bp.route('/paid-leave/<int:run_id>/statement/excel')
def paid_leave_statement_excel(run_id):
    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = (PaidLeaveEntry.query
               .filter_by(paid_leave_run_id=run_id)
               .join(Employee)
               .order_by(Employee.name)
               .all())
    if not entries:
        flash('No data to export. Make sure Jan-Dec payrolls are finalized.',
              'warning')
        return redirect(url_for('paid_leave.paid_leave_view', run_id=run_id))
    out, name = _build_pl_workbook(run, entries)
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=name)


@paid_leave_bp.route('/paid-leave/<int:run_id>/form-15')
def paid_leave_form15(run_id):
    """Karnataka Factories Act — Form 15 (Register of Leave with Wages).
    One row per eligible employee with statutory columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    run = PaidLeaveRun.query.get_or_404(run_id)
    verify_est_ownership(run.establishment)
    entries = (PaidLeaveEntry.query
               .filter_by(paid_leave_run_id=run_id, is_eligible=True)
               .join(Employee)
               .order_by(Employee.emp_code)
               .all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Form 15"
    LAST = 11

    title_font   = Font(bold=True, size=14, name='Calibri')
    subtitle_fnt = Font(bold=True, size=10, name='Calibri')
    info_font    = Font(size=9, name='Calibri')
    info_bold    = Font(bold=True, size=9, name='Calibri')
    header_font  = Font(bold=True, size=8.5, name='Calibri')
    body_font    = Font(size=9, name='Calibri')
    body_bold    = Font(bold=True, size=9, name='Calibri')

    thin = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    ws.cell(row=1, column=1, value="FORM 15").font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)

    ws.cell(row=2, column=1, value="[See Rule 100]").font = subtitle_fnt
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)

    ws.cell(row=3, column=1, value="REGISTER OF LEAVE WITH WAGES").font = subtitle_fnt
    ws.cell(row=3, column=1).alignment = center
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=LAST)

    ws.cell(row=4, column=1, value="(Under the Karnataka Factories Rules — Sec. 79 of the Factories Act, 1948)").font = info_font
    ws.cell(row=4, column=1).alignment = center
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=LAST)

    est = run.establishment
    info_rows = [
        ('Name of the Factory / Establishment:', (est.company_name or '').upper()),
        ('Address:', est.address or ''),
        ('Calendar Year:', f"{run.year} (01-Jan-{run.year} to 31-Dec-{run.year})"),
    ]
    r = 6
    for label, value in info_rows:
        ws.cell(row=r, column=1, value=label).font = info_bold
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=value).font = info_font
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST)
        r += 1

    meta_parts = [
        f"Eligibility threshold: {run.eligibility_threshold} days",
        f"Leave divisor: 1 day for every {run.eligibility_divisor} days worked",
    ]
    if run.payment_date:
        meta_parts.append(f"Date of payment: {run.payment_date.strftime('%d-%m-%Y')}")
    ws.cell(row=r, column=1, value='   |   '.join(meta_parts)).font = info_bold
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST)
    r += 2

    headers = [
        'Sr\nNo.',
        'Name of the\nWorker',
        "Father's /\nHusband's Name",
        'Designation',
        'UAN / ESIC\nIP No.',
        'Days worked\nin the calendar\nyear',
        'Leave\nearned\n(Days)',
        'Daily rate\n(₹) (Dec.)',
        'Wages payable\nfor the leave\nperiod (₹)',
        'Date of\npayment',
        'Signature / Thumb\nimpression of the\nworker',
    ]
    header_row = r
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 56

    data_start = header_row + 1
    payment_str = run.payment_date.strftime('%d-%m-%Y') if run.payment_date else ''
    t_days = t_leave = t_wages = 0
    for idx, e in enumerate(entries, 1):
        rr = data_start + idx - 1
        emp = e.employee
        primary_id = emp.uan_number or emp.esic_ip_number or '—'
        cells = [
            (1, idx, center, body_font),
            (2, emp.name, left, body_bold),
            (3, emp.father_husband_name or '', left, body_font),
            (4, emp.designation or '', left, body_font),
            (5, primary_id, center, body_font),
            (6, int(round(e.total_attendance or 0)), center, body_font),
            (7, float(e.eligible_attendance or 0), center, body_font),
            (8, round(e.december_rate or 0), right, body_font),
            (9, int(e.pl_amount or 0), right, body_bold),
            (10, payment_str, center, body_font),
            (11, '', center, body_font),
        ]
        for col, val, align, font in cells:
            c = ws.cell(row=rr, column=col, value=val)
            c.font = font
            c.alignment = align
            c.border = border
            if col in (8, 9):
                c.number_format = '#,##0'
            if col == 7:
                c.number_format = '0.00'
        ws.row_dimensions[rr].height = 22
        t_days += float(e.total_attendance or 0)
        t_leave += float(e.eligible_attendance or 0)
        t_wages += float(e.pl_amount or 0)

    total_row = data_start + len(entries)
    label = ws.cell(row=total_row, column=1,
                    value=f"TOTAL ({len(entries)} workers)")
    label.font = body_bold
    label.fill = header_fill
    label.alignment = center
    label.border = border
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = header_fill

    totals = [
        (6, int(round(t_days)), center),
        (7, round(t_leave, 2),  center),
        (8, '', center),
        (9, int(t_wages),       right),
        (10, '', center),
        (11, '', center),
    ]
    for col, val, align in totals:
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = body_bold
        c.alignment = align
        c.border = border
        c.fill = header_fill
        if col in (9,):
            c.number_format = '#,##0'

    # Signature block
    sig = total_row + 3
    ws.cell(row=sig, column=1, value='Signature of Occupier / Manager:').font = info_bold
    ws.cell(row=sig + 1, column=1, value='_______________________').font = info_font
    ws.cell(row=sig, column=8, value='Date:').font = info_bold
    ws.cell(row=sig, column=9, value='_______________').font = info_font
    ws.cell(row=sig + 1, column=8, value='Place:').font = info_bold
    ws.cell(row=sig + 1, column=9, value='_______________').font = info_font

    foot = sig + 4
    ws.cell(row=foot, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   Vaishnavi Consultant"
            ).font = Font(size=8, italic=True, color='64748B', name='Calibri')
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=LAST)

    widths = [5, 20, 18, 14, 14, 11, 9, 11, 13, 11, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{header_row}:{header_row}'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    safe = (est.company_name or 'Establishment').replace(' ', '_').replace('/', '_')[:60]
    filename = f"Form_15_Leave_Register_{safe}_{run.year}.xlsx"
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
