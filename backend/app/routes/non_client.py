"""
Non-Client Quick Returns — Routes
===================================
Process EPF ECR + ESIC monthly return for one-off establishments
that are NOT regular clients in the system.

Key rules (updated):
  • NO ceiling enforcement — uses actual wages from the uploaded sheet.
    EPF: 12% on actual wages (EPS pension cap still ₹15,000 by statute).
    ESIC: any employee with an IP number is included regardless of wages.
  • ESIC EE rounded UP: math.ceil(wages × 0.0075)  [= Round(Sal×0.75%+0.49,0)]
  • Comprehensive template: Monthly Fixed / Daily Wages / Monthly Heads / OT / Holidays
"""

import io
import json
import math
import calendar
import traceback

from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.worksheet.datavalidation import DataValidation
    _HAS_DV = True
except ImportError:
    _HAS_DV = False

try:
    import xlwt
    _HAS_XLWT = True
except ImportError:
    _HAS_XLWT = False

from flask import (Blueprint, render_template, request, redirect,
                   url_for, flash, send_file)

from app import db
from app.models.non_client import NonClientReturn
from app.auth import login_required
from app.user_context import current_user_id, is_admin

non_client_bp = Blueprint('non_client', __name__)

# ── Statutory rates ────────────────────────────────────────────────────────────
EPF_CEILING  = 15000    # Used ONLY for EPS pension base (statutory pension ceiling)
EPF_EE_RATE  = 0.12     # Employee EPF contribution
EPS_RATE     = 0.0833   # Employer EPS portion
ADMIN_RATE   = 0.005    # EPF Admin charges (0.5%, min ₹500)
EDLI_RATE    = 0.005    # EDLI (0.5%)
ESIC_EE_RATE = 0.0075   # Employee ESIC (0.75%)
ESIC_ER_RATE = 0.0325   # Employer ESIC (3.25%)

# ── Default config (from Configuration sheet if present, else these) ───────────
DEFAULT_CONFIG = {
    'salary_type':  'MF',   # MF=Monthly Fixed / DW=Daily Wages / MH=Monthly Heads
    'working_days': 26,
    'ot_rate':      2.0,    # OT wage multiplier (2 = double time)
    'ot_in_epf':    False,
    'ot_in_esic':   False,
}


# ══════════════════════════════════════════════════════════════════════════════
#  CALCULATION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _user_query():
    q = NonClientReturn.query
    if not is_admin():
        uid = current_user_id()
        if uid:
            q = q.filter(NonClientReturn.user_id == uid)
    return q


def _calc_epf(epf_wages: float, days_present: int) -> dict:
    """
    EPF — NO ceiling on EPF wages. Uses actual wages passed in.
    EPS pension base is still capped at ₹15,000 per statute.
    """
    w        = epf_wages
    epf_ee   = round(w * EPF_EE_RATE)
    eps_base = min(w, EPF_CEILING)          # Pension always capped at 15000
    eps      = round(eps_base * EPS_RATE)
    er_diff  = epf_ee - eps                 # Employer AC-I (3.67%)
    ncp_days = max(0, 26 - days_present)
    return {
        'epf_wages':  int(round(w)),
        'eps_wages':  int(round(eps_base)),  # EPS wage base (capped)
        'edli_wages': int(round(w)),
        'epf_ee':     epf_ee,
        'eps':        eps,
        'er_diff':    er_diff,
        'ncp_days':   ncp_days,
    }


def _calc_esic(esic_wages: float) -> dict:
    """
    ESIC — NO ceiling exclusion. Calculate for all employees with an IP number.
    EE rounded UP (math.ceil) = Round(Salary×0.75%+0.49,0) per ESIC portal rule.
    ER also rounded UP for consistency.
    """
    esic_ee = math.ceil(esic_wages * ESIC_EE_RATE)   # 0.75% — always round up
    esic_er = math.ceil(esic_wages * ESIC_ER_RATE)   # 3.25% — always round up
    return {'esic_ee': esic_ee, 'esic_er': esic_er, 'covered': True}


def _clean_name(name: str) -> str:
    """Keep only alphabets + space (ESIC/EPF portal requirement)."""
    return ' '.join(
        ''.join(c for c in part if c.isalpha()).strip()
        for part in (name or '').split()
    ).upper()


def _process_rows(raw_rows: list, config: dict = None) -> tuple:
    """
    Calculate EPF + ESIC for all employees.
    Supports: Monthly Fixed (MF), Daily Wages (DW), Monthly Heads (MH), OT, Holiday wages.
    No ceiling enforcement — dynamic based on actual wages in the sheet.
    """
    if config is None:
        config = DEFAULT_CONFIG.copy()

    sal_default  = config.get('salary_type', 'MF')
    wd_default   = max(1, int(config.get('working_days', 26) or 26))
    ot_rate      = float(config.get('ot_rate', 2.0) or 2.0)
    ot_in_epf    = bool(config.get('ot_in_epf', False))
    ot_in_esic   = bool(config.get('ot_in_esic', False))

    employees = []
    totals = {
        'epf_ee': 0, 'eps': 0, 'er_diff': 0, 'edli_wages': 0,
        'esic_ee': 0, 'esic_er': 0,
        'gross_total': 0.0,
        'count_epf': 0, 'count_esic': 0, 'count_total': 0,
        'admin_charges': 0, 'edli_admin': 0,
    }

    for row in raw_rows:
        name    = str(row.get('name', '')).strip()
        uan     = str(row.get('uan', '')).strip()
        ip_no   = str(row.get('ip_no', '')).strip()
        days    = int(float(row.get('days', 0) or 0))
        other_d = float(row.get('other_ded', 0) or 0)
        remarks = str(row.get('remarks', '')).strip()

        # Salary type (per-row or config default)
        raw_type = (row.get('salary_type') or '').strip().upper()[:2]
        sal_type = raw_type if raw_type in ('MF', 'DW', 'MH') else sal_default

        # Working days (per-row override or config default)
        wd_row = row.get('working_days_row')
        working_days = max(1, int(float(wd_row))) if wd_row else wd_default

        # Salary heads
        basic   = float(row.get('basic', 0) or 0)
        da      = float(row.get('da', 0) or 0)
        hra     = float(row.get('hra', 0) or 0)
        conv    = float(row.get('conveyance', 0) or 0)
        oa      = float(row.get('other_allow', 0) or 0)
        heads_total = basic + da + hra + conv + oa

        daily_rate   = float(row.get('daily_rate', 0) or 0)
        gross_direct = row.get('gross_direct')          # None = not entered
        ot_hours     = float(row.get('ot_hours', 0) or 0)
        ot_days_cnt  = float(row.get('ot_days', 0) or 0)
        ot_amt_dir   = row.get('ot_amount_direct')      # None = not entered
        holiday_w    = float(row.get('holiday_wages', 0) or 0)

        # ── BASE GROSS ─────────────────────────────────────────
        if gross_direct is not None and float(gross_direct) > 0:
            base_gross = float(gross_direct)            # Already earned — no proration
        elif sal_type == 'DW' and daily_rate > 0:
            base_gross = daily_rate * days              # Daily rate × days present
        elif heads_total > 0:
            base_gross = heads_total * (days / working_days)   # Prorate for absence
        else:
            base_gross = 0.0

        # ── OT AMOUNT ──────────────────────────────────────────
        if ot_amt_dir is not None:
            ot_amount = float(ot_amt_dir)
        elif ot_hours > 0 and daily_rate > 0:
            ot_amount = round(ot_hours * (daily_rate / 8.0) * ot_rate, 2)
        elif ot_days_cnt > 0 and daily_rate > 0:
            ot_amount = round(ot_days_cnt * daily_rate * ot_rate, 2)
        else:
            ot_amount = 0.0

        total_gross = round(base_gross + ot_amount + holiday_w, 2)

        # ── EPF WAGES (no ceiling — use actual) ───────────────
        epf_ov = row.get('epf_wages_override')
        if epf_ov is not None:
            epf_wages = float(epf_ov)
        elif ot_in_epf:
            epf_wages = base_gross + ot_amount
        else:
            epf_wages = base_gross

        # ── ESIC WAGES (no ceiling — use actual) ──────────────
        esic_ov = row.get('esic_wages_override')
        if esic_ov is not None:
            esic_wages = float(esic_ov)
        elif ot_in_esic:
            esic_wages = base_gross + ot_amount
        else:
            esic_wages = base_gross

        has_epf  = bool(uan)
        has_esic = bool(ip_no)              # No ceiling exclusion

        epf_data  = _calc_epf(epf_wages, days) if has_epf  else {}
        esic_data = _calc_esic(esic_wages)     if has_esic else {}

        emp = {
            'name':          name,
            'uan':           uan,
            'ip_no':         ip_no,
            'days':          days,
            'salary_type':   sal_type,
            'working_days':  working_days,
            'gross':         total_gross,
            'base_gross':    round(base_gross, 2),
            'ot_amount':     round(ot_amount, 2),
            'holiday_wages': round(holiday_w, 2),
            'epf_wages':     epf_data.get('epf_wages', 0),
            'eps_wages':     epf_data.get('eps_wages', 0),
            'edli_wages':    epf_data.get('edli_wages', 0),
            'epf_ee':        epf_data.get('epf_ee', 0),
            'eps':           epf_data.get('eps', 0),
            'er_diff':       epf_data.get('er_diff', 0),
            'ncp_days':      epf_data.get('ncp_days', 0),
            'esic_wages':    round(esic_wages, 2) if has_esic else 0,
            'esic_ee':       esic_data.get('esic_ee', 0),
            'esic_er':       esic_data.get('esic_er', 0),
            'esic_covered':  has_esic,
            'other_ded':     round(other_d, 2),
            'remarks':       remarks,
            'has_epf':       has_epf,
            'has_esic':      has_esic,
        }
        employees.append(emp)

        totals['gross_total'] += total_gross
        totals['count_total'] += 1
        if has_epf:
            totals['epf_ee']     += epf_data.get('epf_ee', 0)
            totals['eps']        += epf_data.get('eps', 0)
            totals['er_diff']    += epf_data.get('er_diff', 0)
            totals['edli_wages'] += epf_data.get('edli_wages', 0)
            totals['count_epf']  += 1
        if has_esic:
            totals['esic_ee']    += esic_data.get('esic_ee', 0)
            totals['esic_er']    += esic_data.get('esic_er', 0)
            totals['count_esic'] += 1

    total_epf_wages         = sum(e['epf_wages'] for e in employees)
    totals['admin_charges'] = max(500.0, round(total_epf_wages * ADMIN_RATE, 2))
    totals['edli_admin']    = round(total_epf_wages * EDLI_RATE, 2)
    totals['gross_total']   = round(totals['gross_total'], 2)

    return employees, totals


def _build_ecr_lines(employees: list, pf_code: str) -> str:
    lines = []
    for e in employees:
        if not e.get('has_epf') or not e.get('uan'):
            continue
        line = '#~#'.join([
            str(e['uan']),
            _clean_name(e['name']),
            str(int(round(e['gross']))),
            str(e['epf_wages']),
            str(e['eps_wages']),
            str(e['edli_wages']),
            str(e['epf_ee']),
            str(e['eps']),
            str(e['er_diff']),
            str(e['ncp_days']),
            '0',
        ])
        lines.append(line)
    return '\n'.join(lines)


def _build_esic_rows(employees: list) -> list:
    """Build ESIC MC Template rows. No ceiling blocking."""
    rows = []
    for e in employees:
        if not e.get('ip_no'):
            continue
        days  = int(math.ceil(e['days'] or 0))
        wages = int(round(e['esic_wages'] or 0))
        reason_code = '11' if (days == 0 and wages == 0) else ''
        rows.append({
            'ip_number':        e['ip_no'].strip(),
            'ip_name':          _clean_name(e['name']),
            'no_of_days':       str(days),
            'total_wages':      str(wages),
            'reason_code':      reason_code,
            'last_working_day': '',
        })
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  ESIC XLS GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

def _generate_esic_xls_nc(rows: list, month: int, year: int, est_name: str) -> io.BytesIO:
    if not _HAS_XLWT:
        raise RuntimeError('xlwt not installed — cannot generate .xls ESIC template')

    wb  = xlwt.Workbook(encoding='utf-8')
    ws  = wb.add_sheet('Sheet1')
    ws.col(0).width = 5000
    ws.col(1).width = 10000
    ws.col(2).width = 5000
    ws.col(3).width = 6000
    ws.col(4).width = 8000
    ws.col(5).width = 7000

    hdr = xlwt.easyxf('font: bold on; align: wrap on, vert centre, horiz centre;')
    for ci, h in enumerate([
        'IP Number \n(10 Digits)',
        'IP Name\n( Only alphabets and space )',
        'No of Days for which wages paid/payable during the month',
        'Total Monthly Wages',
        'Reason Code for Zero workings days(numeric only; '
        'provide 0 for all other reasons- Click on the link for reference)',
        'Last Working Day\n( Format DD/MM/YYYY  or DD-MM-YYYY)',
    ]):
        ws.write(0, ci, h, hdr)

    for ri, r in enumerate(rows, 1):
        ws.write(ri, 0, r['ip_number'])
        ws.write(ri, 1, r['ip_name'])
        ws.write(ri, 2, r['no_of_days'])
        ws.write(ri, 3, r['total_wages'])
        ws.write(ri, 4, r['reason_code'])
        ws.write(ri, 5, r['last_working_day'])

    ws2   = wb.add_sheet('Instructions & Reason Codes')
    bold  = xlwt.easyxf('font: bold on;')
    ws2.write(0, 0, f'Establishment: {est_name}', bold)
    ws2.write(0, 2, f'Period: {calendar.month_name[month]} {year}', bold)
    ws2.write(2, 0, 'Reason', bold)
    ws2.write(2, 1, 'Code', bold)
    ws2.write(2, 2, 'Note', bold)
    for ri, (reason, code, note) in enumerate([
        ('Without Reason',              '0',  'Leave last working day as blank'),
        ('On Leave',                    '1',  'Leave last working day as blank'),
        ('Left Service',                '2',  'Provide last working day (dd/mm/yyyy).'),
        ('Retired',                     '3',  'Provide last working day (dd/mm/yyyy).'),
        ('Out of Coverage',             '4',  'Valid only April/October wage period.'),
        ('Expired',                     '5',  'Provide last working day (dd/mm/yyyy).'),
        ('Non Implemented area',        '6',  'Provide last working day (dd/mm/yyyy).'),
        ('Compliance by Immed. Employer','7', 'Leave last working day as blank'),
        ('Suspension of work',          '8',  'Leave last working day as blank'),
        ('Strike/Lockout',              '9',  'Leave last working day as blank'),
        ('Retrenchment',                '10', 'Provide last working day (dd/mm/yyyy).'),
        ('No Work',                     '11', 'Leave last working day as blank'),
        ('Doesnt Belong To This Employer','12','Leave last working day as blank'),
        ('Duplicate IP',                '13', 'Leave last working day as blank'),
    ], 3):
        ws2.write(ri, 0, reason)
        ws2.write(ri, 1, code)
        ws2.write(ri, 2, note)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  COMPREHENSIVE INPUT TEMPLATE  (3 sheets)
# ══════════════════════════════════════════════════════════════════════════════

def _generate_input_template(month: int, year: int, est_name: str) -> io.BytesIO:
    """
    3-sheet Excel template:
      Sheet 1 — Configuration  (salary type defaults, OT settings)
      Sheet 2 — Employee Data  (22 columns: MF / DW / MH / OT / Holiday)
      Sheet 3 — Notes & Help
    """
    wb = openpyxl.Workbook()

    thin   = Side(style='thin',   color='CBD5E1')
    thick  = Side(style='medium', color='94A3B8')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_t  = Border(left=thick, right=thick, top=thick, bottom=thick)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_w = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ─────────────────────────────────────────────────────────
    # SHEET 1 — CONFIGURATION
    # ─────────────────────────────────────────────────────────
    cfg = wb.active
    cfg.title = 'Configuration'
    cfg.column_dimensions['A'].width = 32
    cfg.column_dimensions['B'].width = 26
    cfg.column_dimensions['C'].width = 58

    # Title
    cfg.merge_cells('A1:C1')
    t = cfg['A1']
    t.value     = f'Non-Client Return — Configuration  |  {est_name}  |  {calendar.month_name[month]} {year}'
    t.font      = Font(bold=True, size=11, color='FFFFFF')
    t.fill      = PatternFill('solid', fgColor='1e40af')
    t.alignment = center
    cfg.row_dimensions[1].height = 26

    # Column header row
    for ci, h in enumerate(['Setting', 'Value (edit this column)', 'Notes / Options'], 1):
        c = cfg.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, size=9, color='1e3a5f')
        c.fill = PatternFill('solid', fgColor='DBEAFE')
        c.border = bdr
        c.alignment = center
    cfg.row_dimensions[2].height = 18

    SETTINGS = [
        ('Default Salary Type',    'Monthly Fixed',
         'Monthly Fixed  |  Daily Wages  |  Monthly Heads'),
        ('Working Days in Month',  '26',
         '26 (standard)  |  30  |  31  |  actual calendar days'),
        ('OT Rate Multiplier',     '2',
         '1 = single time  |  1.5 = time-and-half  |  2 = double time'),
        ('Include OT in EPF',      'No',
         'Yes / No — should OT wages be part of EPF base?'),
        ('Include OT in ESIC',     'No',
         'Yes / No — should OT wages be part of ESIC base?'),
    ]
    for ri, (s, v, note) in enumerate(SETTINGS, 3):
        c_s = cfg.cell(row=ri, column=1, value=s)
        c_v = cfg.cell(row=ri, column=2, value=v)
        c_n = cfg.cell(row=ri, column=3, value=note)
        c_s.font = Font(size=9, bold=True, color='374151')
        c_v.font = Font(size=10, bold=True, color='1e3a5f')
        c_n.font = Font(size=8,  italic=True, color='6b7280')
        for c in (c_s, c_v, c_n):
            c.border = bdr
            c.alignment = left_w
        c_v.alignment = Alignment(horizontal='center', vertical='center')
        cfg.row_dimensions[ri].height = 20

    # Dropdowns
    if _HAS_DV:
        try:
            dv1 = DataValidation(type='list',
                                  formula1='"Monthly Fixed,Daily Wages,Monthly Heads"',
                                  allow_blank=True)
            cfg.add_data_validation(dv1)
            dv1.add(cfg['B3'])
            dv2 = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
            cfg.add_data_validation(dv2)
            dv2.add(cfg['B6'])
            dv2.add(cfg['B7'])
        except Exception:
            pass

    cfg.merge_cells('A9:C9')
    n = cfg['A9']
    n.value     = ('NOTE: These are defaults for ALL employees. '
                   'You may override Salary Type and Working Days per employee '
                   'in the "Employee Data" sheet columns E & F.')
    n.font      = Font(size=9, bold=True, color='b45309')
    n.fill      = PatternFill('solid', fgColor='FEF3C7')
    n.alignment = left_w
    cfg.row_dimensions[9].height = 28

    # ─────────────────────────────────────────────────────────
    # SHEET 2 — EMPLOYEE DATA
    # ─────────────────────────────────────────────────────────
    ws = wb.create_sheet(title='Employee Data')

    # Column definitions: (header_text, width, group_id)
    # Groups: 0=sno  1=info  2=setup  3=daily  4=heads  5=gross  6=ot  7=hol  8=ovr  9=ded
    COLS = [
        ('S.No',                   5,  0),
        ('Employee Name *',       22,  1),
        ('UAN',                   14,  1),
        ('ESIC IP Number',        14,  1),
        ('Salary Type\n(MF/DW/MH)',12, 2),
        ('Working Days\nin Month', 11, 2),
        ('Days Present *',        11,  2),
        ('Daily Rate\n(₹ / day)', 12,  3),
        ('Basic Salary',          13,  4),
        ('DA',                    10,  4),
        ('HRA',                   10,  4),
        ('Conveyance',            12,  4),
        ('Other Allow',           12,  4),
        ('Gross Wages\n(Direct)', 14,  5),
        ('OT Hours',              10,  6),
        ('OT Days',               10,  6),
        ('OT Amount (₹)',         12,  6),
        ('Holiday Wages',         12,  7),
        ('EPF Wages\n(Override)', 14,  8),
        ('ESIC Wages\n(Override)',14,  8),
        ('Other\nDeduction',      12,  9),
        ('Remarks',               18,  9),
    ]
    TOTAL_COLS = len(COLS)

    # Group style map: group_id → (bg, fg, label)
    GSTYLE = {
        0: ('475569', 'FFFFFF', ''),
        1: ('1e40af', 'FFFFFF', 'EMPLOYEE INFO'),
        2: ('166534', 'FFFFFF', 'ATTENDANCE & SETUP'),
        3: ('b45309', 'FFFFFF', 'DAILY WAGES'),
        4: ('312e81', 'FFFFFF', 'SALARY HEADS  (for MF / MH)'),
        5: ('134e4a', 'FFFFFF', 'DIRECT GROSS'),
        6: ('c2410c', 'FFFFFF', 'OVERTIME'),
        7: ('9d174d', 'FFFFFF', 'HOLIDAY WAGES'),
        8: ('6b21a8', 'FFFFFF', 'COMPLIANCE OVERRIDE'),
        9: ('374151', 'FFFFFF', 'DEDUCTIONS'),
    }

    # Row 1 — Title
    ws.merge_cells(f'A1:{get_column_letter(TOTAL_COLS)}1')
    tr = ws['A1']
    tr.value     = (f'Non-Client EPF/ESIC Input  |  {est_name}  |  '
                    f'{calendar.month_name[month]} {year}')
    tr.font      = Font(bold=True, size=11, color='1e3a5f')
    tr.fill      = PatternFill('solid', fgColor='DBEAFE')
    tr.alignment = center
    ws.row_dimensions[1].height = 26

    # Row 2 — Instructions
    ws.merge_cells(f'A2:{get_column_letter(TOTAL_COLS)}2')
    ir = ws['A2']
    ir.value = (
        'Configuration sheet sets defaults.  * = Required.  '
        'MF=Monthly Fixed (prorate by days), DW=Daily Wages (Rate×Days), '
        'MH=Monthly Heads (heads prorated).  '
        'Fill Gross Direct OR heads OR daily rate — NOT both.  '
        'Leave Override columns blank for auto-calculation.'
    )
    ir.font      = Font(size=8, italic=True, color='374151')
    ir.fill      = PatternFill('solid', fgColor='EFF6FF')
    ir.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Row 3 — Group sub-headers
    ws.row_dimensions[3].height = 16
    grp_cols = {}   # group_id → [col indices]
    for ci, (_, _, gid) in enumerate(COLS, 1):
        grp_cols.setdefault(gid, []).append(ci)

    for gid, ci_list in grp_cols.items():
        bg, fg, lbl = GSTYLE[gid]
        start = get_column_letter(ci_list[0])
        end   = get_column_letter(ci_list[-1])
        if len(ci_list) > 1:
            ws.merge_cells(f'{start}3:{end}3')
        cell = ws[f'{start}3']
        cell.value     = lbl
        cell.font      = Font(bold=True, size=7, color=fg)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = bdr

    # Row 4 — Column headers
    ws.row_dimensions[4].height = 40
    for ci, (hdr, width, gid) in enumerate(COLS, 1):
        bg, fg, _ = GSTYLE[gid]
        cell = ws.cell(row=4, column=ci, value=hdr)
        cell.font      = Font(bold=True, size=8, color=fg)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = center
        cell.border    = bdr
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Row 5 — Sample row
    SAMPLE = [
        '1', 'Sample Employee', '100123456789', '1234567890',
        'MF', '26', '26',
        '',                         # Daily Rate
        '12000', '3000', '2000', '1000', '2000',  # Heads
        '',                         # Gross Direct (blank = use heads)
        '2', '', '',                # OT Hours, Days, Amount
        '',                         # Holiday
        '', '',                     # EPF/ESIC override (blank = auto)
        '0', 'Sample — delete this row',
    ]
    ws.row_dimensions[5].height = 18
    sf = PatternFill('solid', fgColor='F1F5F9')
    for ci, v in enumerate(SAMPLE, 1):
        c = ws.cell(row=5, column=ci, value=v)
        c.font      = Font(color='94A3B8', italic=True, size=8)
        c.fill      = sf
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = bdr

    # Rows 6-55 — Data rows
    fills = [PatternFill('solid', fgColor='FFFFFF'),
             PatternFill('solid', fgColor='F8FAFC')]
    for ri in range(6, 56):
        ws.row_dimensions[ri].height = 18
        fill = fills[ri % 2]
        for ci in range(1, TOTAL_COLS + 1):
            c = ws.cell(row=ri, column=ci, value='')
            c.fill      = fill
            c.border    = bdr
            c.alignment = Alignment(horizontal='center', vertical='center')
        # Auto S.No
        sno = ws.cell(row=ri, column=1, value=ri - 5)
        sno.font = Font(color='CBD5E1', size=8)

    # Salary type dropdown for column E (index 5), rows 6-55
    if _HAS_DV:
        try:
            dv_t = DataValidation(type='list', formula1='"MF,DW,MH"', allow_blank=True)
            dv_t.error      = 'Use MF, DW, or MH'
            dv_t.errorTitle = 'Invalid type'
            ws.add_data_validation(dv_t)
            dv_t.sqref = 'E6:E55'
        except Exception:
            pass

    ws.freeze_panes = 'B6'

    # ─────────────────────────────────────────────────────────
    # SHEET 3 — NOTES & HELP
    # ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title='Notes & Help')
    ws3.column_dimensions['A'].width = 26
    ws3.column_dimensions['B'].width = 60
    ws3.column_dimensions['C'].width = 38

    hf = Font(bold=True, color='FFFFFF', size=9)
    hfill = PatternFill('solid', fgColor='1e40af')
    for ci, h in enumerate(['Topic', 'Details', 'Example / Notes'], 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.border = bdr
        c.alignment = center
    ws3.row_dimensions[1].height = 18

    NOTES = [
        ('— SALARY TYPES —',   '', ''),
        ('MF  Monthly Fixed',
         'Fixed monthly salary. System prorates for absent days: Earned = Gross × Days÷WorkingDays.',
         'Gross=20000, Days=24÷26 → Earned=18,461'),
        ('DW  Daily Wages',
         'Earned = Daily Rate × Days Present. No proration needed.',
         'Rate=800, Days=24 → Earned=19,200'),
        ('MH  Monthly Heads',
         'Sum all heads (Basic+DA+HRA+Conv+Other) then prorate: Earned = Sum × Days÷WD.',
         'Sum=20000, Days=24÷26 → 18,461'),
        ('', '', ''),
        ('— GROSS DIRECT —',   '', ''),
        ('Gross Wages (Direct)',
         'If filled, overrides ALL head/daily calculations. Enter the EARNED amount (already prorated).',
         'Enter 18461 directly if already calculated'),
        ('', '', ''),
        ('— OVERTIME —',       '', ''),
        ('OT Amount (Direct)',  'Overrides OT Hours/Days calculation. Enter final OT wages.',   ''),
        ('OT Hours',
         'OT = Hours × (Daily Rate ÷ 8) × OT Multiplier  (needs Daily Rate filled)',
         '4 hrs, Rate=800, Mult=2 → OT=800'),
        ('OT Days',
         'OT = Days × Daily Rate × OT Multiplier  (needs Daily Rate filled)',
         '1 day, Rate=800, Mult=2 → OT=1600'),
        ('OT in EPF / ESIC',   'Configure in Configuration sheet.', ''),
        ('', '', ''),
        ('— EPF  (NO CEILING) —', '', ''),
        ('EE Contribution',    '12% of EPF wages — ACTUAL wages used, no ₹15,000 cap.', '20000 × 12% = 2400'),
        ('EPS  (Employer)',    '8.33% of wages — EPS pension base STILL capped at ₹15,000.', 'Max EPS = 15000×8.33%=1250'),
        ('EPF AC-I  (Diff)',   '3.67% of EPF wages (= EPF EE − EPS)', ''),
        ('Admin Charges',      '0.5% of total EPF wages — minimum ₹500', ''),
        ('EDLI',               '0.5% of EPF wages', ''),
        ('NCP Days',           'max(0,  26 − Days Present)', ''),
        ('', '', ''),
        ('— ESIC  (NO CEILING) —', '', ''),
        ('EE Contribution',
         '0.75% — ROUNDED UP: math.ceil(Wages × 0.0075)  i.e. Round(Wages×0.75%+0.49,0)',
         '85.01 → 86,   85.90 → 86'),
        ('ER Contribution',    '3.25% — rounded UP (math.ceil)', ''),
        ('No Ceiling',
         'ALL employees with ESIC IP are included regardless of wages. '
         'Even if gross > ₹21,000 — ESIC is still calculated.',
         '25000 × 0.75% = ceil(187.5) = 188'),
        ('', '', ''),
        ('— OVERRIDES —',      '', ''),
        ('EPF Wages Override',
         'If filled, EPF is calculated on this value instead of gross.',
         'Enter 15000 to cap manually'),
        ('ESIC Wages Override',
         'If filled, ESIC is calculated on this value instead of gross.', ''),
    ]

    for ri, (a, b, c_val) in enumerate(NOTES, 2):
        ca = ws3.cell(row=ri, column=1, value=a)
        cb = ws3.cell(row=ri, column=2, value=b)
        cc = ws3.cell(row=ri, column=3, value=c_val)
        if a.startswith('—'):
            for cx in (ca, cb, cc):
                cx.font = Font(bold=True, size=9, color='1e40af')
                cx.fill = PatternFill('solid', fgColor='EFF6FF')
        elif a == '':
            pass
        else:
            ca.font = Font(bold=True, size=9)
            cb.font = Font(size=9)
            cc.font = Font(size=9, color='059669', italic=True)
        for cx in (ca, cb, cc):
            cx.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cx.border = bdr
        ws3.row_dimensions[ri].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _parse_uploaded_excel(file_obj) -> tuple:
    """
    Parse the comprehensive non-client Excel template.
    Returns (config_dict, raw_rows).
    Backward-compatible: old simple template (without Config sheet / extra columns)
    is also handled — missing columns default to 0 / None.
    """
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        raise ValueError(f'Could not open Excel file: {e}')

    # ── 1. Parse Configuration sheet ─────────────────────────
    config = DEFAULT_CONFIG.copy()
    for sh in ('Configuration', 'Config', 'configuration', 'config'):
        if sh in wb.sheetnames:
            cws = wb[sh]
            for ri in range(1, 20):
                k = cws.cell(row=ri, column=1).value
                v = cws.cell(row=ri, column=2).value
                if not k:
                    continue
                ks = str(k).lower().strip()
                vs = str(v or '').strip()
                vl = vs.lower()
                if 'salary type' in ks:
                    if 'daily' in vl:         config['salary_type'] = 'DW'
                    elif 'head' in vl:        config['salary_type'] = 'MH'
                    else:                     config['salary_type'] = 'MF'
                elif 'working days' in ks:
                    try:    config['working_days'] = int(float(vs))
                    except: pass
                elif 'ot rate' in ks or 'overtime rate' in ks:
                    try:    config['ot_rate'] = float(vs)
                    except: pass
                elif 'epf' in ks and 'ot' in ks:
                    config['ot_in_epf']  = vl in ('yes', 'y', 'true', '1')
                elif 'esic' in ks and 'ot' in ks:
                    config['ot_in_esic'] = vl in ('yes', 'y', 'true', '1')
            break

    # ── 2. Find Employee Data sheet ───────────────────────────
    ws = None
    for sh in ('Employee Data', 'employee data', 'Data', 'data'):
        if sh in wb.sheetnames:
            ws = wb[sh]
            break
    if ws is None:
        ws = wb.active

    # ── 3. Find header row ────────────────────────────────────
    # The official template has TWO header-like rows:
    #   Row 3 = group sub-headers ("EMPLOYEE INFO", "ATTENDANCE & SETUP"…)
    #   Row 4 = real column headers ("Employee Name *", "Days Present *"…)
    # The old scan matched any cell containing "employee" and so latched
    # onto "EMPLOYEE INFO" in the group row, building a column map from the
    # group labels → "Employee Name column not found". We now require the
    # real header row to carry BOTH an employee/name column AND a days
    # column, which the group row never does.
    def _row_texts(ri):
        return [str(ws.cell(row=ri, column=ci).value or '').lower().replace('\n', ' ').strip()
                for ci in range(1, 30)]

    header_row = None
    for ri in range(1, 16):
        texts = _row_texts(ri)
        has_name = any(('employee name' in t) or ('emp name' in t) or (t == 'name')
                       for t in texts)
        has_days = any('days present' in t for t in texts)
        if has_name and has_days:
            header_row = ri
            break
    # Fallback — any row that has an explicit "employee name" cell
    if header_row is None:
        for ri in range(1, 16):
            if any('employee name' in t for t in _row_texts(ri)):
                header_row = ri
                break

    if header_row is None:
        raise ValueError(
            "Header row with 'Employee Name' not found. "
            "Please use the official Non-Client template."
        )

    # ── 4. Build column map ───────────────────────────────────
    # Normalise newlines → spaces so multi-line headers like
    # "Gross Wages\n(Direct)" match cleanly.
    col_map = {}
    for ci in range(1, 30):
        v = ws.cell(row=header_row, column=ci).value
        if v:
            col_map[str(v).lower().replace('\n', ' ').strip()] = ci

    def _fc(*kws):
        """Find a column by keyword. Tries an EXACT (asterisk-stripped)
        match first so short keys like 'da' don't latch onto 'days
        present' / 'working days' via substring, then falls back to a
        substring match."""
        norm = {k.replace('*', '').strip(): v for k, v in col_map.items()}
        for kw in kws:                       # exact, normalised
            if kw in norm:
                return norm[kw]
        for kw in kws:                       # substring fallback
            for k, v in col_map.items():
                if kw in k:
                    return v
        return None

    c_name  = _fc('employee name', 'emp name', 'name')
    c_uan   = _fc('uan')
    c_ip    = _fc('esic ip', 'ip number', 'ip no')
    c_type  = _fc('salary type')
    c_wdays = _fc('working days in month', 'working days')
    c_days  = _fc('days present')
    c_drate = _fc('daily rate')
    c_basic = _fc('basic salary', 'basic')
    c_da    = _fc('da', 'dearness')
    c_hra   = _fc('hra')
    c_conv  = _fc('conveyance')
    c_oalw  = _fc('other allow', 'other allowance')
    c_gross = _fc('gross wages (direct)', 'gross wages', 'gross')
    c_oth   = _fc('ot hours')
    c_otd   = _fc('ot days')
    c_ota   = _fc('ot amount')
    c_hol   = _fc('holiday wages', 'holiday')
    c_epfw  = _fc('epf wages (override)', 'epf wages')
    c_esicw = _fc('esic wages (override)', 'esic wages')
    c_othd  = _fc('other deduction', 'other ded')
    c_rem   = _fc('remarks')

    if not c_name:
        raise ValueError("'Employee Name' column not found. Use the official template.")
    if not c_days:
        raise ValueError("'Days Present' column not found. Use the official template.")

    # ── 5. Parse rows ─────────────────────────────────────────
    def gv(ri, ci):
        return ws.cell(row=ri, column=ci).value if ci else None

    def flt(v):
        if v is None or str(v).strip() == '':
            return None
        try:
            return float(str(v).replace(',', ''))
        except (ValueError, TypeError):
            return None

    def flt0(v):
        r = flt(v)
        return r if r is not None else 0.0

    rows = []
    for ri in range(header_row + 1, ws.max_row + 1):
        name_v = gv(ri, c_name)
        if not name_v or not str(name_v).strip():
            continue
        name_s = str(name_v).strip()
        if 'sample' in name_s.lower():
            continue
        if name_s.lower() in ('employee name', 'emp name', 'name'):
            continue

        rows.append({
            'name':               name_s,
            'uan':                str(gv(ri, c_uan) or '').strip(),
            'ip_no':              str(gv(ri, c_ip) or '').strip(),
            'salary_type':        str(gv(ri, c_type) or '').strip().upper()[:2] or None,
            'working_days_row':   flt(gv(ri, c_wdays)),
            'days':               int(flt0(gv(ri, c_days))),
            'daily_rate':         flt0(gv(ri, c_drate)),
            'basic':              flt0(gv(ri, c_basic)),
            'da':                 flt0(gv(ri, c_da)),
            'hra':                flt0(gv(ri, c_hra)),
            'conveyance':         flt0(gv(ri, c_conv)),
            'other_allow':        flt0(gv(ri, c_oalw)),
            'gross_direct':       flt(gv(ri, c_gross)),   # None = not entered
            'ot_hours':           flt0(gv(ri, c_oth)),
            'ot_days':            flt0(gv(ri, c_otd)),
            'ot_amount_direct':   flt(gv(ri, c_ota)),     # None = not entered
            'holiday_wages':      flt0(gv(ri, c_hol)),
            'epf_wages_override': flt(gv(ri, c_epfw)),
            'esic_wages_override':flt(gv(ri, c_esicw)),
            'other_ded':          flt0(gv(ri, c_othd)),
            'remarks':            str(gv(ri, c_rem) or '').strip(),
        })

    if not rows:
        raise ValueError(
            'No employee data found. '
            'Fill in employee rows below the header in the "Employee Data" sheet. '
            'Rows that say "Sample" are skipped automatically.'
        )

    return config, rows


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@non_client_bp.route('/non-client-returns')
@login_required
def nc_list():
    uid    = current_user_id()
    admin  = is_admin()
    search = request.args.get('search', '').strip()

    q = _user_query().order_by(
        NonClientReturn.year.desc(),
        NonClientReturn.month.desc(),
        NonClientReturn.created_at.desc()
    )
    if search:
        q = q.filter(NonClientReturn.est_name.ilike(f'%{search}%'))

    records = q.all()
    now = datetime.utcnow()
    return render_template('non_client_returns.html',
                           records=records, search=search,
                           is_admin=admin, now=now,
                           MONTHS=list(calendar.month_name)[1:])


@non_client_bp.route('/non-client-returns/create', methods=['POST'])
@login_required
def nc_create():
    uid       = current_user_id()
    est_name  = request.form.get('est_name', '').strip()
    pf_code   = request.form.get('pf_code', '').strip()
    esic_code = request.form.get('esic_code', '').strip()
    month_str = request.form.get('month', '')
    year_str  = request.form.get('year', '')
    fee_str   = request.form.get('fee_charged', '').strip()
    notes     = request.form.get('notes', '').strip()

    if not est_name:
        flash('Establishment name is required.', 'danger')
        return redirect(url_for('non_client.nc_list'))

    try:
        month = int(month_str)
        year  = int(year_str)
        if not (1 <= month <= 12) or not (2000 <= year <= 2099):
            raise ValueError
    except (ValueError, TypeError):
        flash('Please select a valid month and year.', 'danger')
        return redirect(url_for('non_client.nc_list'))

    fee = 0.0
    if fee_str:
        try:    fee = float(fee_str)
        except: fee = 0.0

    record = NonClientReturn(
        user_id=uid, est_name=est_name,
        pf_code=pf_code or None, esic_code=esic_code or None,
        month=month, year=year,
        fee_charged=fee, notes=notes or None, status='pending',
    )
    db.session.add(record)
    db.session.commit()

    flash(f'Record created for {est_name} — {record.period_label}.', 'success')
    return redirect(url_for('non_client.nc_detail', record_id=record.id))


@non_client_bp.route('/non-client-returns/<int:record_id>')
@login_required
def nc_detail(record_id):
    rec       = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    employees = rec.get_employees()
    totals    = rec.get_totals()
    esic_rows = rec.get_esic_rows()
    return render_template('non_client_detail.html',
                           rec=rec, employees=employees,
                           totals=totals, esic_rows=esic_rows)


@non_client_bp.route('/non-client-returns/<int:record_id>/download-template')
@login_required
def nc_download_template(record_id):
    rec = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    try:
        buf = _generate_input_template(rec.month, rec.year, rec.est_name)
    except Exception as e:
        flash(f'Could not generate template: {e}', 'danger')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    safe  = ''.join(c if c.isalnum() or c in ('_', '-') else '_' for c in rec.est_name)
    fname = f'NCReturn_Template_{safe}_{rec.month:02d}{rec.year}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@non_client_bp.route('/non-client-returns/<int:record_id>/upload', methods=['POST'])
@login_required
def nc_upload_process(record_id):
    rec      = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    uploaded = request.files.get('data_file')

    if not uploaded or not uploaded.filename:
        flash('Please choose an Excel file to upload.', 'warning')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    lower = uploaded.filename.lower()
    if not (lower.endswith('.xlsx') or lower.endswith('.xls')):
        flash('Only .xlsx or .xls files are supported.', 'danger')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    try:
        config, raw_rows = _parse_uploaded_excel(uploaded)          # ← updated
        employees, totals = _process_rows(raw_rows, config)         # ← updated

        rec.employees_json  = json.dumps(employees)
        rec.totals_json     = json.dumps(totals)
        rec.ecr_text        = _build_ecr_lines(employees, rec.pf_code or '')
        rec.esic_json       = json.dumps(_build_esic_rows(employees))
        rec.status          = 'processed'
        rec.source_filename = uploaded.filename
        rec.updated_at      = datetime.utcnow()
        db.session.commit()

        flash(
            f'Processed {totals["count_total"]} employees — '
            f'{totals["count_epf"]} EPF, {totals["count_esic"]} ESIC. '
            'Download your ECR and ESIC files below.',
            'success'
        )
    except ValueError as e:
        flash(str(e), 'danger')
        rec.status = 'error'
        db.session.commit()
    except Exception as e:
        traceback.print_exc()
        flash(f'Processing error: {e}', 'danger')
        rec.status = 'error'
        db.session.commit()

    return redirect(url_for('non_client.nc_detail', record_id=record_id))


@non_client_bp.route('/non-client-returns/<int:record_id>/download-ecr')
@login_required
def nc_download_ecr(record_id):
    rec = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    if not rec.ecr_text:
        flash('ECR data not available. Upload and process an Excel file first.', 'warning')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    buf = io.BytesIO(rec.ecr_text.encode('utf-8'))
    buf.seek(0)
    pf_s  = (rec.pf_code or 'NOPF').replace('/', '').replace(' ', '')
    fname = f'ECR_{pf_s}_{rec.month:02d}{str(rec.year)[-2:]}.txt'
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='text/plain')


@non_client_bp.route('/non-client-returns/<int:record_id>/download-esic')
@login_required
def nc_download_esic(record_id):
    rec       = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    esic_rows = rec.get_esic_rows()
    if not esic_rows:
        flash('ESIC data not available. Upload and process an Excel file first.', 'warning')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    try:
        buf = _generate_esic_xls_nc(esic_rows, rec.month, rec.year, rec.est_name)
    except RuntimeError as e:
        flash(str(e), 'danger')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    safe  = ''.join(c if c.isalnum() or c in ('_', '-') else '_' for c in rec.est_name)
    fname = f'MC_Template_{safe}_{rec.month:02d}{rec.year}.xls'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.ms-excel')


def _statement_summary(rec):
    """Build the EPF + ESIC payable summary dict for the client statement
    from the stored totals. Splits employee vs employer share clearly."""
    t = rec.get_totals()
    epf_ee    = round(t.get('epf_ee', 0))
    epf_ac01  = round(t.get('er_diff', 0))       # Employer EPF A/c 01 (3.67%)
    epf_eps   = round(t.get('eps', 0))           # Employer EPS A/c 10 (8.33%)
    epf_admin = round(t.get('admin_charges', 0)) # A/c 02 admin (0.5%, min ₹500)
    epf_edli  = round(t.get('edli_admin', 0))    # A/c 21 EDLI (0.5%)
    epf_er    = epf_ac01 + epf_eps + epf_admin + epf_edli
    esic_ee   = round(t.get('esic_ee', 0))
    esic_er   = round(t.get('esic_er', 0))
    return {
        'count_total': t.get('count_total', 0),
        'count_epf':   t.get('count_epf', 0),
        'count_esic':  t.get('count_esic', 0),
        'gross_total': round(t.get('gross_total', 0)),
        # EPF
        'epf_ee':    epf_ee,
        'epf_ac01':  epf_ac01,
        'epf_eps':   epf_eps,
        'epf_admin': epf_admin,
        'epf_edli':  epf_edli,
        'epf_er':    epf_er,
        'epf_total': epf_ee + epf_er,
        # ESIC
        'esic_ee':    esic_ee,
        'esic_er':    esic_er,
        'esic_total': esic_ee + esic_er,
        # Grand total
        'grand_total': epf_ee + epf_er + esic_ee + esic_er,
    }


def _statement_employee_rows(rec):
    """Per-employee contribution rows for the detailed statement.
    EPF employer (per employee) = A/c 01 (er_diff) + A/c 10 (eps);
    admin (A/c 02) + EDLI (A/c 21) are establishment-level, shown in the
    summary footer only."""
    rows = []
    for e in rec.get_employees():
        epf_ee = round(e.get('epf_ee', 0))
        epf_er = round(e.get('er_diff', 0) + e.get('eps', 0))
        esic_ee = round(e.get('esic_ee', 0))
        esic_er = round(e.get('esic_er', 0))
        rows.append({
            'name':       e.get('name', ''),
            'uan':        e.get('uan', ''),
            'ip_no':      e.get('ip_no', ''),
            'days':       e.get('days', 0),
            'gross':      round(e.get('gross', 0)),
            'epf_wages':  round(e.get('epf_wages', 0)),
            'epf_ee':     epf_ee,
            'epf_er':     epf_er,
            'esic_wages': round(e.get('esic_wages', 0)),
            'esic_ee':    esic_ee,
            'esic_er':    esic_er,
            'total':      epf_ee + epf_er + esic_ee + esic_er,
            'has_epf':    e.get('has_epf', bool(e.get('uan'))),
            'has_esic':   e.get('has_esic', bool(e.get('ip_no'))),
        })
    return rows


@non_client_bp.route('/non-client-returns/<int:record_id>/statement')
@login_required
def nc_statement(record_id):
    """Client-facing EPF + ESIC statement — employee-wise detail with
    employee/employer contribution split, plus a summary footer.
    HTML preview by default; ?format=excel for the workbook."""
    rec = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    if rec.status != 'processed':
        flash('Process an Excel file first to generate the statement.', 'warning')
        return redirect(url_for('non_client.nc_detail', record_id=record_id))

    s = _statement_summary(rec)
    emp_rows = _statement_employee_rows(rec)
    if request.args.get('format') == 'excel':
        return _nc_statement_excel(rec, s, emp_rows)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('non_client_statement.html', rec=rec, s=s,
                           emp_rows=emp_rows, generated_on=generated_on)


def _nc_statement_excel(rec, s, emp_rows):
    """Detailed employee-wise EPF + ESIC statement (Legal landscape):
    one row per employee with EPF EE/ER and ESIC EE/ER, a TOTAL row, then
    a compact payable summary band."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "EPF-ESIC Statement"
    title_f = Font(bold=True, size=13, name='Calibri')
    sub_f   = Font(size=9, italic=True, color='475569', name='Calibri')
    grp_f   = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
    hdr_f   = Font(bold=True, size=8.5, color='FFFFFF', name='Calibri')
    body    = Font(size=8.5, name='Calibri')
    name_f  = Font(bold=True, size=8.5, name='Calibri')
    bold    = Font(bold=True, size=9, name='Calibri')
    info_b  = Font(bold=True, size=9, name='Calibri')
    info_v  = Font(size=9, name='Calibri')
    thin = Side(border_style='thin', color='94A3B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    rightA = Alignment(horizontal='right', vertical='center')
    leftA  = Alignment(horizontal='left', vertical='center')
    slate  = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    epf_h  = PatternFill('solid', start_color='1E40AF', end_color='1E40AF')
    esic_h = PatternFill('solid', start_color='9333EA', end_color='9333EA')
    band   = PatternFill('solid', start_color='F1F5F9', end_color='F1F5F9')
    green  = PatternFill('solid', start_color='D9EAD3', end_color='D9EAD3')

    # 12 columns: # | Name | UAN | ESIC IP | Days | Gross | EPF Wages |
    #             EPF-EE | EPF-ER | ESIC-EE | ESIC-ER | Total
    LAST = 12
    ws.cell(row=1, column=1, value="EPF & ESIC MONTHLY STATEMENT").font = title_f
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)
    ws.cell(row=2, column=1,
            value=f"{(rec.est_name or '').upper()}  |  {rec.period_label}  |  "
                  f"PF: {rec.pf_code or '—'}   ESIC: {rec.esic_code or '—'}").font = sub_f
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)

    # Group header row (row 4): blank over 1-7, EPF over 8-9, ESIC over 10-11, blank 12
    gr = 4
    ws.merge_cells(start_row=gr, start_column=8, end_row=gr, end_column=9)
    gc = ws.cell(row=gr, column=8, value="EPF (₹)"); gc.font = grp_f; gc.fill = epf_h
    gc.alignment = center; gc.border = border
    ws.cell(row=gr, column=9).fill = epf_h; ws.cell(row=gr, column=9).border = border
    ws.merge_cells(start_row=gr, start_column=10, end_row=gr, end_column=11)
    gc2 = ws.cell(row=gr, column=10, value="ESIC (₹)"); gc2.font = grp_f; gc2.fill = esic_h
    gc2.alignment = center; gc2.border = border
    ws.cell(row=gr, column=11).fill = esic_h; ws.cell(row=gr, column=11).border = border

    # Column header (row 5)
    hr = 5
    headers = ['#', 'Employee Name', 'UAN', 'ESIC IP', 'Days', 'Gross\nWages',
               'EPF\nWages', 'EE\n(12%)', 'ER\n(13%)', 'EE\n(0.75%)', 'ER\n(3.25%)', 'Total\n(₹)']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_f; cell.alignment = center; cell.border = border
        cell.fill = epf_h if c in (8, 9) else (esic_h if c in (10, 11) else slate)
    ws.row_dimensions[hr].height = 30

    t = {'gross': 0, 'epfw': 0, 'epf_ee': 0, 'epf_er': 0, 'esic_ee': 0, 'esic_er': 0, 'tot': 0}
    rr = hr + 1
    for i, e in enumerate(emp_rows, 1):
        vals = [
            (1, i, center), (2, e['name'], leftA),
            (3, e['uan'] or '—', center), (4, e['ip_no'] or '—', center),
            (5, e['days'], center), (6, e['gross'], rightA),
            (7, e['epf_wages'] if e['has_epf'] else 0, rightA),
            (8, e['epf_ee'] if e['has_epf'] else 0, rightA),
            (9, e['epf_er'] if e['has_epf'] else 0, rightA),
            (10, e['esic_ee'] if e['has_esic'] else 0, rightA),
            (11, e['esic_er'] if e['has_esic'] else 0, rightA),
            (12, e['total'], rightA),
        ]
        for c, v, al in vals:
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = name_f if c == 2 else body
            cell.alignment = al; cell.border = border
            if c in (6, 7, 8, 9, 10, 11, 12):
                cell.number_format = '#,##0'
        t['gross'] += e['gross']; t['epfw'] += (e['epf_wages'] if e['has_epf'] else 0)
        t['epf_ee'] += (e['epf_ee'] if e['has_epf'] else 0); t['epf_er'] += (e['epf_er'] if e['has_epf'] else 0)
        t['esic_ee'] += (e['esic_ee'] if e['has_esic'] else 0); t['esic_er'] += (e['esic_er'] if e['has_esic'] else 0)
        t['tot'] += e['total']
        rr += 1

    # TOTAL row
    tc = ws.cell(row=rr, column=1, value=f"TOTAL ({len(emp_rows)})")
    tc.font = bold; tc.fill = green; tc.alignment = center; tc.border = border
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    for c in range(2, 6):
        ws.cell(row=rr, column=c).fill = green; ws.cell(row=rr, column=c).border = border
    for c, v in [(6, t['gross']), (7, t['epfw']), (8, t['epf_ee']), (9, t['epf_er']),
                 (10, t['esic_ee']), (11, t['esic_er']), (12, t['tot'])]:
        cell = ws.cell(row=rr, column=c, value=v)
        cell.font = bold; cell.fill = green; cell.alignment = rightA; cell.border = border
        cell.number_format = '#,##0'
    rr += 2

    # ── Payable summary band ───────────────────────────────────────────
    def sum_line(label, val, fill=None, bold_row=False):
        nonlocal rr
        lc = ws.cell(row=rr, column=2, value=label)
        lc.font = bold if bold_row else info_v
        lc.alignment = leftA
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=9)
        vc = ws.cell(row=rr, column=10, value=val)
        vc.font = bold if bold_row else info_v; vc.alignment = rightA; vc.number_format = '#,##0'
        ws.merge_cells(start_row=rr, start_column=10, end_row=rr, end_column=12)
        if fill:
            for cc in range(2, 13):
                ws.cell(row=rr, column=cc).fill = fill
        rr += 1

    hb = ws.cell(row=rr, column=2, value="PAYABLE SUMMARY")
    hb.font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
    hb.fill = slate; hb.alignment = leftA
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=12)
    for cc in range(2, 13):
        ws.cell(row=rr, column=cc).fill = slate
    rr += 1
    sum_line("EPF — Employee Share (12%)", s['epf_ee'])
    sum_line("EPF — Employer Share (A/c 01 3.67% + EPS 8.33%)", s['epf_ac01'] + s['epf_eps'])
    sum_line("EPF — Admin Charges (A/c 02, 0.5% min ₹500)", s['epf_admin'])
    sum_line("EPF — EDLI (A/c 21, 0.5%)", s['epf_edli'])
    sum_line("TOTAL EPF PAYABLE", s['epf_total'], green, True)
    sum_line("ESIC — Employee Share (0.75%)", s['esic_ee'])
    sum_line("ESIC — Employer Share (3.25%)", s['esic_er'])
    sum_line("TOTAL ESIC PAYABLE", s['esic_total'], green, True)
    sum_line("GRAND TOTAL PAYABLE (EPF + ESIC)", s['grand_total'],
             PatternFill('solid', start_color='0F172A', end_color='0F172A'), True)
    # white text on the grand-total row
    ws.cell(row=rr - 1, column=2).font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
    ws.cell(row=rr - 1, column=10).font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')

    rr += 1
    ws.cell(row=rr, column=2, value='Prepared by Vaishnavi Consultant').font = info_b
    ws.cell(row=rr, column=10, value='Date: ' + datetime.now().strftime('%d-%m-%Y')).font = info_v
    ws.cell(row=rr, column=10).alignment = rightA

    widths = [4, 24, 15, 14, 6, 11, 11, 10, 10, 10, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.paperSize = 5  # Legal
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{gr}:{hr}'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = ''.join(c if c.isalnum() or c in ('_', '-') else '_' for c in rec.est_name)[:50]
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f"EPF_ESIC_Statement_{safe}_{rec.month:02d}{rec.year}.xlsx")


@non_client_bp.route('/non-client-returns/<int:record_id>/delete', methods=['POST'])
@login_required
def nc_delete(record_id):
    rec = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    est = rec.est_name
    db.session.delete(rec)
    db.session.commit()
    flash(f'Record deleted: {est}', 'success')
    return redirect(url_for('non_client.nc_list'))


@non_client_bp.route('/non-client-returns/<int:record_id>/update-fee', methods=['POST'])
@login_required
def nc_update_fee(record_id):
    rec = _user_query().filter(NonClientReturn.id == record_id).first_or_404()
    try:
        rec.fee_charged = float(request.form.get('fee_charged', 0) or 0)
        db.session.commit()
        flash('Fee updated.', 'success')
    except ValueError:
        flash('Invalid fee value.', 'danger')
    return redirect(url_for('non_client.nc_detail', record_id=record_id))
