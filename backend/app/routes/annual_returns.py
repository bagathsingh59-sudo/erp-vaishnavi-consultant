"""
Annual / Statutory Returns
  • EPF Form 3A  — Member's annual contribution card (per employee)
  • EPF Form 6A  — Consolidated annual statement of contributions
  • Gratuity     — Liability report (config-driven divisor) + Form F nomination
  • LWF          — Karnataka Labour Welfare Fund annual return

All are READ-ONLY aggregations of already-finalized payrolls (EPF/LWF) or
of the employee master (Gratuity) — no new tables, nothing persisted.

EPF contribution year runs MARCH → FEBRUARY (not the April–March FY).
"""
import io
import calendar
from datetime import datetime, date

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file)

from app import db
from app.models.establishment import Establishment
from app.models.employee import Employee
from app.models.payroll import (MonthlyPayroll, PayrollEntry, PayrollEntryHead,
                                 PayrollConfig, SalaryHead)
from app.user_context import (user_establishments, verify_est_ownership,
                              capture_est_from_url)
from app.utils.naming import short_est_code

annual_bp = Blueprint('annual_returns', __name__)


@annual_bp.before_request
def _capture_url_establishment():
    if request.path and '/api/' in request.path:
        return None
    capture_est_from_url()
    return None


# ──────────────────────────────────────────────────────────────────────────
# Shared helpers
# ──────────────────────────────────────────────────────────────────────────
def _epf_year_months(start_year):
    """EPF contribution year: March(start_year) → February(start_year+1).
    Returns ordered list of (year, month, 'Mon')."""
    out = []
    for m in range(3, 13):                       # Mar .. Dec
        out.append((start_year, m, calendar.month_abbr[m]))
    for m in range(1, 3):                         # Jan, Feb (next year)
        out.append((start_year + 1, m, calendar.month_abbr[m]))
    return out


def _finalized_payrolls(est_id, month_pairs):
    """Return {(year, month): MonthlyPayroll} for finalized payrolls among
    the requested (year, month) pairs."""
    if not month_pairs:
        return {}
    payrolls = MonthlyPayroll.query.filter(
        MonthlyPayroll.establishment_id == est_id,
        MonthlyPayroll.status == 'finalized',
    ).all()
    wanted = {(y, m) for (y, m, _a) in month_pairs}
    return {(p.year, p.month): p for p in payrolls if (p.year, p.month) in wanted}


def _aggregate_epf(est_id, start_year):
    """Build per-employee EPF month-wise data for the EPF year.
    Returns (months, emp_rows) where each emp_row is:
      { employee, monthly: {(y,m): {...}}, totals: {...} }
    """
    months = _epf_year_months(start_year)
    pmap = _finalized_payrolls(est_id, months)
    payroll_ids = [p.id for p in pmap.values()]

    entries = PayrollEntry.query.filter(
        PayrollEntry.monthly_payroll_id.in_(payroll_ids)
    ).all() if payroll_ids else []

    # Map payroll_id → (year, month)
    pid_to_ym = {p.id: (p.year, p.month) for p in pmap.values()}

    emp = {}
    for e in entries:
        ym = pid_to_ym.get(e.monthly_payroll_id)
        if not ym or not e.employee_id:
            continue
        if not e.epf_employee and not e.epf_wages:
            # employee not in EPF this month — still recorded as NCP-ish blank
            pass
        bucket = emp.setdefault(e.employee_id, {
            'employee': None,
            'monthly': {},
            't_wages': 0.0, 't_ee': 0.0, 't_ac01': 0.0, 't_eps': 0.0,
        })
        eps_wages = min(float(e.epf_wages or 0), 15000.0)
        bucket['monthly'][ym] = {
            'epf_wages': float(e.epf_wages or 0),
            'eps_wages': eps_wages,
            'ee': float(e.epf_employee or 0),
            'ac01': float(e.epf_ac01 or 0),
            'eps': float(e.epf_eps or 0),
        }
        bucket['t_wages'] += float(e.epf_wages or 0)
        bucket['t_ee']    += float(e.epf_employee or 0)
        bucket['t_ac01']  += float(e.epf_ac01 or 0)
        bucket['t_eps']   += float(e.epf_eps or 0)

    # Attach employee objects, drop those with zero EPF all year
    emp_ids = list(emp.keys())
    emps = {x.id: x for x in Employee.query.filter(Employee.id.in_(emp_ids)).all()} if emp_ids else {}
    rows = []
    for eid, b in emp.items():
        if b['t_ee'] <= 0 and b['t_wages'] <= 0:
            continue
        b['employee'] = emps.get(eid)
        if not b['employee']:
            continue
        rows.append(b)
    rows.sort(key=lambda r: (r['employee'].name or '').upper())
    return months, rows


def _gratuity_divisor(config):
    """Gratuity 15-days-wages divisor, derived from establishment config.
    Statutory norm is 26 (monthly-rated). We honour the establishment's
    absence_divisor where it's 26 or 30, else default to 26."""
    ad = (getattr(config, 'absence_divisor', '26') or '26')
    if ad == '30':
        return 30
    if ad == '26':
        return 26
    return 26   # 'calendar' or anything else → gratuity standard


def _latest_basic_da(est_id):
    """Return {employee_id: (basic_da_amount, payroll_period)} from the most
    recent finalized payroll per employee — used as 'last drawn Basic+DA'
    for gratuity. Falls back to earned_gross if no Basic/DA heads exist."""
    # compliance heads (Basic, DA) for this establishment
    comp_heads = SalaryHead.query.filter_by(
        establishment_id=est_id, is_for_compliance=True, is_active=True
    ).all()
    comp_ids = {h.id for h in comp_heads}

    payrolls = (MonthlyPayroll.query
                .filter_by(establishment_id=est_id, status='finalized')
                .order_by(MonthlyPayroll.year.desc(), MonthlyPayroll.month.desc())
                .all())
    latest = {}   # emp_id → (basic_da, "Mon YYYY")
    seen = set()
    for p in payrolls:
        entries = PayrollEntry.query.filter_by(monthly_payroll_id=p.id).all()
        # head amounts for this payroll
        peh = PayrollEntryHead.query.filter(
            PayrollEntryHead.payroll_entry_id.in_([e.id for e in entries])
        ).all() if entries else []
        head_by_entry = {}
        for h in peh:
            head_by_entry.setdefault(h.payroll_entry_id, []).append(h)
        for e in entries:
            if e.employee_id in seen:
                continue
            # Basic+DA from compliance heads; else earned_gross
            bd = 0.0
            rows = head_by_entry.get(e.id, [])
            comp_rows = [h for h in rows if h.salary_head_id in comp_ids]
            if comp_rows:
                bd = sum(float(h.earned_amount or 0) for h in comp_rows)
            else:
                bd = float(e.earned_gross or 0)
            latest[e.employee_id] = (round(bd), p.period_display)
            seen.add(e.employee_id)
    return latest


# ──────────────────────────────────────────────────────────────────────────
# Landing page
# ──────────────────────────────────────────────────────────────────────────
@annual_bp.route('/annual-returns')
def annual_returns_home():
    ests = user_establishments().order_by(Establishment.company_name).all()
    current_year = datetime.now().year
    sel_est = request.args.get('establishment_id', type=int)
    sel_year = request.args.get('year', type=int) or (current_year - 1)
    est = Establishment.query.get(sel_est) if sel_est else None
    if est:
        verify_est_ownership(est)

    # EPF contribution year choices (each a March-Feb year). The list must
    # reach back to the establishment's commencement (date of registration),
    # since some clients have records from FY 1980-81. When a single
    # establishment is selected we use its own commencement; otherwise we use
    # the earliest commencement among all visible establishments. Falls back
    # to the last 6 years when no registration date is recorded.
    earliest_year = None
    if est and est.date_of_registration:
        earliest_year = est.date_of_registration.year
    elif not est:
        reg = [e.date_of_registration for e in ests if e.date_of_registration]
        if reg:
            earliest_year = min(reg).year
    if earliest_year is None:
        earliest_year = current_year - 5
    earliest_year = min(earliest_year, current_year, sel_year)
    years = list(range(earliest_year, current_year + 1))

    return render_template('annual_returns/home.html',
                           establishments=ests, years=years,
                           sel_est=est, sel_year=sel_year,
                           current_year=current_year)


# ──────────────────────────────────────────────────────────────────────────
# EPF Form 6A — Consolidated annual statement
# ──────────────────────────────────────────────────────────────────────────
@annual_bp.route('/establishment/<int:est_id>/epf-6a')
def epf_form_6a(est_id):
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    start_year = request.args.get('year', type=int) or (datetime.now().year - 1)
    months, rows = _aggregate_epf(est_id, start_year)

    if request.args.get('format') == 'excel':
        return _epf_6a_excel(est, start_year, months, rows)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('annual_returns/epf_6a.html',
                           est=est, start_year=start_year, end_year=start_year + 1,
                           rows=rows, generated_on=generated_on)


def _epf_6a_excel(est, start_year, months, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Form 6A"
    title_f = Font(bold=True, size=14, name='Calibri')
    sub_f   = Font(bold=True, size=10, name='Calibri')
    info_b  = Font(bold=True, size=9, name='Calibri')
    info_v  = Font(size=9, name='Calibri')
    hdr_f   = Font(bold=True, size=8.5, color='FFFFFF', name='Calibri')
    body    = Font(size=9, name='Calibri')
    bold    = Font(bold=True, size=9, name='Calibri')
    thin = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    slate  = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    band   = PatternFill('solid', start_color='F1F5F9', end_color='F1F5F9')
    green  = PatternFill('solid', start_color='D9EAD3', end_color='D9EAD3')

    LAST = 9
    ws.cell(row=1, column=1, value="FORM 6A").font = title_f
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)
    ws.cell(row=2, column=1, value="ANNUAL CONSOLIDATED STATEMENT OF CONTRIBUTIONS").font = sub_f
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)
    ws.cell(row=3, column=1, value="(Para 43 of the EPF Scheme, 1952 & Para 20 of the EPS, 1995)").font = info_v
    ws.cell(row=3, column=1).alignment = center
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=LAST)

    info = [
        ('Name & Address of Establishment:', (est.company_name or '').upper() + (('  —  ' + est.address) if est.address else '')),
        ('Code No. (PF):', est.pf_code or '—'),
        ('Currency Period (EPF Year):', f"March {start_year} to February {start_year + 1}"),
    ]
    r = 5
    for lab, val in info:
        ws.cell(row=r, column=1, value=lab).font = info_b
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=val).font = info_v
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST)
        r += 1
    r += 1

    headers = ['Sl.\nNo.', 'Account No.\n(UAN)', 'Name of the Member',
               'Wages on which\ncontributions\npayable (₹)',
               "Worker's Share\nEPF (₹)",
               "Employer's Share\nEPF A/c 1 (₹)",
               "Pension Fund\nA/c 10 (₹)",
               'Total\nContribution (₹)', 'Remarks']
    hr = r
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_f; cell.fill = slate; cell.alignment = center; cell.border = border
    ws.row_dimensions[hr].height = 40

    t_w = t_ee = t_ac01 = t_eps = t_tot = 0.0
    rr = hr + 1
    for i, b in enumerate(rows, 1):
        emp = b['employee']
        total = b['t_ee'] + b['t_ac01'] + b['t_eps']
        vals = [
            (1, i, center),
            (2, emp.uan_number or emp.esic_ip_number or '—', center),
            (3, emp.name, left),
            (4, round(b['t_wages']), right),
            (5, round(b['t_ee']), right),
            (6, round(b['t_ac01']), right),
            (7, round(b['t_eps']), right),
            (8, round(total), right),
            (9, '', left),
        ]
        for c, v, al in vals:
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = body; cell.alignment = al; cell.border = border
            if c in (4, 5, 6, 7, 8):
                cell.number_format = '#,##0'
        t_w += b['t_wages']; t_ee += b['t_ee']; t_ac01 += b['t_ac01']
        t_eps += b['t_eps']; t_tot += total
        rr += 1

    # total row
    tcell = ws.cell(row=rr, column=1, value=f"TOTAL ({len(rows)} members)")
    tcell.font = bold; tcell.alignment = center; tcell.fill = green; tcell.border = border
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    for c in range(2, 4):
        ws.cell(row=rr, column=c).fill = green; ws.cell(row=rr, column=c).border = border
    for c, v in [(4, t_w), (5, t_ee), (6, t_ac01), (7, t_eps), (8, t_tot)]:
        cell = ws.cell(row=rr, column=c, value=round(v))
        cell.font = bold; cell.alignment = right; cell.fill = green; cell.border = border
        cell.number_format = '#,##0'
    ws.cell(row=rr, column=9, value='').fill = green
    ws.cell(row=rr, column=9).border = border

    # signature
    sig = rr + 3
    ws.cell(row=sig, column=1, value='Signature of the Employer:').font = info_b
    ws.cell(row=sig + 1, column=1, value='_______________________').font = info_v
    ws.cell(row=sig, column=7, value='Date:').font = info_b
    ws.cell(row=sig, column=8, value='_______________').font = info_v
    foot = sig + 3
    ws.cell(row=foot, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   Vaishnavi Consultant").font = Font(size=8, italic=True, color='64748B', name='Calibri')
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=LAST)

    widths = [5, 18, 26, 14, 13, 14, 13, 13, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.paperSize = 5
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{hr}:{hr}'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = short_est_code(est.company_name)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Form_6A_{safe}_{start_year}-{start_year+1}.xlsx")


# ──────────────────────────────────────────────────────────────────────────
# EPF Form 3A — Per-member annual contribution card
# ──────────────────────────────────────────────────────────────────────────
@annual_bp.route('/establishment/<int:est_id>/epf-3a')
def epf_form_3a(est_id):
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    start_year = request.args.get('year', type=int) or (datetime.now().year - 1)
    months, rows = _aggregate_epf(est_id, start_year)
    return _epf_3a_excel(est, start_year, months, rows)


def _epf_3a_excel(est, start_year, months, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.pagebreak import Break

    wb = Workbook(); ws = wb.active; ws.title = "Form 3A"
    title_f = Font(bold=True, size=13, name='Calibri')
    sub_f   = Font(bold=True, size=10, name='Calibri')
    info_b  = Font(bold=True, size=9, name='Calibri')
    info_v  = Font(size=9, name='Calibri')
    hdr_f   = Font(bold=True, size=8.5, color='FFFFFF', name='Calibri')
    body    = Font(size=9, name='Calibri')
    bold    = Font(bold=True, size=9, name='Calibri')
    thin = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center')
    slate  = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    green  = PatternFill('solid', start_color='D9EAD3', end_color='D9EAD3')

    LAST = 6
    n_members = len(rows)
    row = 1
    for idx, b in enumerate(rows):
        emp = b['employee']
        # Header per member
        ws.cell(row=row, column=1, value="FORM 3A").font = title_f
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST)
        row += 1
        ws.cell(row=row, column=1, value="MEMBER'S ANNUAL CONTRIBUTION CARD").font = sub_f
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST)
        row += 1
        # member info
        for lab, val in [
            ('Establishment:', (est.company_name or '').upper()),
            ('PF Code:', est.pf_code or '—'),
            ('EPF Year:', f"March {start_year} – February {start_year+1}"),
            ('Member Name:', emp.name),
            ('UAN / Account No.:', emp.uan_number or '—'),
            ("Father's / Husband's Name:", emp.father_husband_name or '—'),
        ]:
            ws.cell(row=row, column=1, value=lab).font = info_b
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=3, value=val).font = info_v
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=LAST)
            row += 1

        headers = ['Month', 'Wages (₹)', "Worker's Share\nEPF (₹)",
                   "Employer's Share\nEPF A/c 1 (₹)", "Pension Fund\nA/c 10 (₹)", 'NCP\nDays']
        hr = row
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=hr, column=c, value=h)
            cell.font = hdr_f; cell.fill = slate; cell.alignment = center; cell.border = border
        ws.row_dimensions[hr].height = 30
        row += 1

        t_w = t_ee = t_ac01 = t_eps = 0.0
        for (y, m, abbr) in months:
            md = b['monthly'].get((y, m))
            label = f"{abbr}-{str(y)[-2:]}"
            if md:
                vals = [label, round(md['epf_wages']), round(md['ee']),
                        round(md['ac01']), round(md['eps']), '']
                t_w += md['epf_wages']; t_ee += md['ee']
                t_ac01 += md['ac01']; t_eps += md['eps']
            else:
                vals = [label, '', '', '', '', '']
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = body
                cell.alignment = left if c == 1 else (center if c == 6 else right)
                cell.border = border
                if c in (2, 3, 4, 5) and isinstance(v, (int, float)):
                    cell.number_format = '#,##0'
            row += 1

        # total
        tc = ws.cell(row=row, column=1, value='TOTAL'); tc.font = bold; tc.fill = green; tc.alignment = center; tc.border = border
        for c, v in [(2, t_w), (3, t_ee), (4, t_ac01), (5, t_eps)]:
            cell = ws.cell(row=row, column=c, value=round(v))
            cell.font = bold; cell.fill = green; cell.alignment = right; cell.border = border
            cell.number_format = '#,##0'
        ws.cell(row=row, column=6, value='').fill = green; ws.cell(row=row, column=6).border = border
        # `row` is now the TOTAL row of this member's card. Two member cards
        # print per page: the 1st of each pair gets a small in-page gap, the
        # 2nd ends the page with a manual page break so the next member always
        # starts at the top of a fresh page (cards never split across pages).
        is_last = (idx == n_members - 1)
        second_of_pair = (idx % 2 == 1)
        if not is_last:
            if second_of_pair:
                ws.row_breaks.append(Break(id=row))   # page break after this card
                row += 1                              # next card → top of new page
            else:
                row += 3                              # gap, partner card same page

    widths = [14, 13, 15, 16, 14, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'portrait'; ws.page_setup.paperSize = 9  # A4
    # Fixed scale (no fit-to-page) so the two-cards-per-page manual breaks are
    # always honoured; 92% leaves a safe margin for long names that wrap.
    ws.page_setup.scale = 92
    ws.print_options.horizontalCentered = True
    ws.print_area = f'A1:{get_column_letter(LAST)}{row}'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = short_est_code(est.company_name)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Form_3A_{safe}_{start_year}-{start_year+1}.xlsx")


# ──────────────────────────────────────────────────────────────────────────
# Gratuity — Liability report + Form F
# ──────────────────────────────────────────────────────────────────────────
def _gratuity_rows(est_id, as_on):
    config = PayrollConfig.query.filter_by(establishment_id=est_id).first()
    divisor = _gratuity_divisor(config)
    basic_da = _latest_basic_da(est_id)
    employees = (Employee.query
                 .filter_by(establishment_id=est_id, is_active=True)
                 .order_by(Employee.name).all())
    rows = []
    for emp in employees:
        doj = emp.date_of_joining
        end = emp.date_of_exit or as_on
        if not doj:
            continue
        # completed years of service
        days = (end - doj).days
        years_exact = days / 365.25
        completed_years = int(years_exact)
        # round up if >= 6 months over a completed year (Gratuity Act rounding)
        frac = years_exact - completed_years
        rounded_years = completed_years + (1 if frac >= 0.5 else 0)
        bd = basic_da.get(emp.id, (0, '—'))
        last_bd = bd[0]
        eligible = rounded_years >= 5
        gratuity = round(last_bd * 15 / divisor * rounded_years) if eligible else 0
        # statutory cap ₹20 lakh
        gratuity = min(gratuity, 2000000)
        rows.append({
            'employee': emp,
            'doj': doj, 'as_on': end,
            'years_exact': round(years_exact, 2),
            'rounded_years': rounded_years,
            'last_basic_da': last_bd,
            'bd_period': bd[1],
            'eligible': eligible,
            'gratuity': gratuity,
        })
    return rows, divisor


@annual_bp.route('/establishment/<int:est_id>/gratuity')
def gratuity_report(est_id):
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    as_on_str = request.args.get('as_on')
    try:
        as_on = datetime.strptime(as_on_str, '%Y-%m-%d').date() if as_on_str else date.today()
    except ValueError:
        as_on = date.today()
    rows, divisor = _gratuity_rows(est_id, as_on)

    if request.args.get('format') == 'excel':
        return _gratuity_excel(est, rows, divisor, as_on)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('annual_returns/gratuity.html',
                           est=est, rows=rows, divisor=divisor, as_on=as_on,
                           generated_on=generated_on)


def _gratuity_excel(est, rows, divisor, as_on):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Gratuity Liability"
    title_f = Font(bold=True, size=14, name='Calibri')
    sub_f   = Font(size=9, italic=True, color='475569', name='Calibri')
    info_b  = Font(bold=True, size=9, name='Calibri')
    info_v  = Font(size=9, name='Calibri')
    hdr_f   = Font(bold=True, size=8.5, color='FFFFFF', name='Calibri')
    body    = Font(size=9, name='Calibri')
    bold    = Font(bold=True, size=9, name='Calibri')
    thin = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    slate  = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    green  = PatternFill('solid', start_color='D9EAD3', end_color='D9EAD3')
    amber  = PatternFill('solid', start_color='FEF3C7', end_color='FEF3C7')

    LAST = 9
    ws.cell(row=1, column=1, value="GRATUITY LIABILITY STATEMENT").font = title_f
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)
    ws.cell(row=2, column=1,
            value=f"{(est.company_name or '').upper()}  |  As on {as_on.strftime('%d-%m-%Y')}  |  "
                  f"Formula: (Last Basic+DA) × 15 / {divisor} × completed years  |  Eligibility: 5 years  |  Cap: ₹20,00,000").font = sub_f
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)

    headers = ['Sl.', 'Emp Code', 'Name', 'Date of\nJoining',
               'Service\n(Years)', 'Last drawn\nBasic + DA (₹)',
               'Gratuity\nPayable (₹)', 'Eligibility', 'Remarks']
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_f; cell.fill = slate; cell.alignment = center; cell.border = border
    ws.row_dimensions[hr].height = 30

    total_liab = 0.0
    rr = hr + 1
    for i, b in enumerate(rows, 1):
        emp = b['employee']
        vals = [
            (1, i, center),
            (2, emp.emp_code or '', center),
            (3, emp.name, left),
            (4, b['doj'].strftime('%d-%m-%Y') if b['doj'] else '', center),
            (5, b['rounded_years'], center),
            (6, b['last_basic_da'], right),
            (7, b['gratuity'], right),
            (8, 'Eligible' if b['eligible'] else 'Not yet (<5 yrs)', center),
            (9, '', left),
        ]
        for c, v, al in vals:
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = body; cell.alignment = al; cell.border = border
            if c in (6, 7):
                cell.number_format = '#,##0'
            if c == 8 and not b['eligible']:
                cell.fill = amber
        total_liab += b['gratuity']
        rr += 1

    tc = ws.cell(row=rr, column=1, value=f"TOTAL LIABILITY ({len(rows)} employees)")
    tc.font = bold; tc.alignment = center; tc.fill = green; tc.border = border
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    for c in range(2, 7):
        ws.cell(row=rr, column=c).fill = green; ws.cell(row=rr, column=c).border = border
    cell = ws.cell(row=rr, column=7, value=round(total_liab))
    cell.font = bold; cell.alignment = right; cell.fill = green; cell.border = border
    cell.number_format = '#,##0'
    for c in (8, 9):
        ws.cell(row=rr, column=c, value='').fill = green
        ws.cell(row=rr, column=c).border = border

    foot = rr + 3
    ws.cell(row=foot, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   Vaishnavi Consultant").font = Font(size=8, italic=True, color='64748B', name='Calibri')
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=LAST)

    widths = [5, 11, 24, 13, 9, 15, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'landscape'; ws.page_setup.paperSize = 5
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{hr}:{hr}'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = short_est_code(est.company_name)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Gratuity_{safe}_{as_on.strftime('%Y%m%d')}.xlsx")


@annual_bp.route('/establishment/<int:est_id>/gratuity-form-f')
def gratuity_form_f(est_id):
    """Form F — Nomination (per employee). One sheet, one nomination form
    per active employee, ready for the worker to fill in nominee details."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    employees = (Employee.query
                 .filter_by(establishment_id=est_id, is_active=True)
                 .order_by(Employee.name).all())

    wb = Workbook(); ws = wb.active; ws.title = "Form F"
    title_f = Font(bold=True, size=13, name='Calibri')
    sub_f   = Font(bold=True, size=10, name='Calibri')
    lab_f   = Font(bold=True, size=9, name='Calibri')
    val_f   = Font(size=9, name='Calibri')
    sm_f    = Font(size=8, italic=True, color='64748B', name='Calibri')
    thin = Side(border_style='thin', color='94A3B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    LAST = 4
    row = 1
    for emp in employees:
        ws.cell(row=row, column=1, value="FORM F").font = title_f
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 1
        ws.cell(row=row, column=1, value="NOMINATION — [See Sub-Rule (1) of Rule 6]").font = sub_f
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 1
        ws.cell(row=row, column=1, value="(Payment of Gratuity Act, 1972)").font = sm_f
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 2

        for lab, val in [
            ('To, The Employer:', (est.company_name or '').upper()),
            ('Name of Employee:', emp.name),
            ("Father's / Husband's Name:", emp.father_husband_name or '—'),
            ('Designation:', emp.designation or '—'),
            ('Date of Joining:', emp.date_of_joining.strftime('%d-%m-%Y') if emp.date_of_joining else '—'),
            ('Department / Branch:', emp.department or '—'),
            ('Marital Status:', emp.marital_status or '—'),
        ]:
            ws.cell(row=row, column=1, value=lab).font = lab_f
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=3, value=val).font = val_f
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=LAST)
            row += 1

        row += 1
        ws.cell(row=row, column=1,
                value="I hereby nominate the person(s) mentioned below to receive the gratuity payable after my death:").font = val_f
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 1

        nom_hdr = ['Name of Nominee', 'Relationship', 'Age', 'Share of Gratuity (%)']
        for c, h in enumerate(nom_hdr, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = lab_f; cell.alignment = center; cell.border = border
        row += 1
        for _ in range(3):   # 3 blank nominee rows
            for c in range(1, LAST + 1):
                ws.cell(row=row, column=c, value='').border = border
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='Signature / Thumb impression of Employee: ______________________').font = val_f
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 1
        ws.cell(row=row, column=1, value='Date: ______________      Place: ______________').font = val_f
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST); row += 3

    for i, w in enumerate([26, 16, 10, 18], 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'portrait'; ws.page_setup.paperSize = 9

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = short_est_code(est.company_name)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"Form_F_{safe}.xlsx")


# ──────────────────────────────────────────────────────────────────────────
# LWF — Karnataka Labour Welfare Fund annual return
# ──────────────────────────────────────────────────────────────────────────
LWF_EMPLOYEE = 40    # Karnataka: employee contribution ₹40 / year
LWF_EMPLOYER = 60    # Karnataka: employer contribution ₹60 / year


@annual_bp.route('/establishment/<int:est_id>/lwf')
def lwf_return(est_id):
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)
    year = request.args.get('year', type=int) or (datetime.now().year - 1)
    employees = (Employee.query
                 .filter_by(establishment_id=est_id, is_active=True)
                 .order_by(Employee.name).all())

    if request.args.get('format') == 'excel':
        return _lwf_excel(est, employees, year)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('annual_returns/lwf.html',
                           est=est, employees=employees, year=year,
                           emp_rate=LWF_EMPLOYEE, empr_rate=LWF_EMPLOYER,
                           generated_on=generated_on)


def _lwf_excel(est, employees, year):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "LWF Return"
    title_f = Font(bold=True, size=14, name='Calibri')
    sub_f   = Font(bold=True, size=10, name='Calibri')
    note_f  = Font(size=9, italic=True, color='475569', name='Calibri')
    info_b  = Font(bold=True, size=9, name='Calibri')
    info_v  = Font(size=9, name='Calibri')
    hdr_f   = Font(bold=True, size=8.5, color='FFFFFF', name='Calibri')
    body    = Font(size=9, name='Calibri')
    bold    = Font(bold=True, size=9, name='Calibri')
    thin = Side(border_style='thin', color='475569')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
    slate  = PatternFill('solid', start_color='1E293B', end_color='1E293B')
    green  = PatternFill('solid', start_color='D9EAD3', end_color='D9EAD3')

    LAST = 6
    ws.cell(row=1, column=1, value="FORM D").font = title_f
    ws.cell(row=1, column=1).alignment = center
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST)
    ws.cell(row=2, column=1, value="ANNUAL STATEMENT OF CONTRIBUTION — LABOUR WELFARE FUND").font = sub_f
    ws.cell(row=2, column=1).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=LAST)
    ws.cell(row=3, column=1, value="(The Karnataka Labour Welfare Fund Act, 1965)").font = note_f
    ws.cell(row=3, column=1).alignment = center
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=LAST)

    info = [
        ('Name of Establishment:', (est.company_name or '').upper()),
        ('Address:', est.address or '—'),
        ('Year:', str(year)),
        ('Contribution Rate:', f"Employee ₹{LWF_EMPLOYEE} + Employer ₹{LWF_EMPLOYER} = ₹{LWF_EMPLOYEE + LWF_EMPLOYER} per employee per year"),
    ]
    r = 5
    for lab, val in info:
        ws.cell(row=r, column=1, value=lab).font = info_b
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=val).font = info_v
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST)
        r += 1
    r += 1

    headers = ['Sl.', 'Emp Code', 'Name of Employee',
               "Employee\nContribution (₹)", "Employer\nContribution (₹)", 'Total (₹)']
    hr = r
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_f; cell.fill = slate; cell.alignment = center; cell.border = border
    ws.row_dimensions[hr].height = 28

    rr = hr + 1
    for i, emp in enumerate(employees, 1):
        vals = [(1, i, center), (2, emp.emp_code or '', center), (3, emp.name, left),
                (4, LWF_EMPLOYEE, right), (5, LWF_EMPLOYER, right), (6, LWF_EMPLOYEE + LWF_EMPLOYER, right)]
        for c, v, al in vals:
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = body; cell.alignment = al; cell.border = border
            if c in (4, 5, 6):
                cell.number_format = '#,##0'
        rr += 1

    n = len(employees)
    tc = ws.cell(row=rr, column=1, value=f"TOTAL ({n} employees)")
    tc.font = bold; tc.alignment = center; tc.fill = green; tc.border = border
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=3)
    for c in range(2, 4):
        ws.cell(row=rr, column=c).fill = green; ws.cell(row=rr, column=c).border = border
    for c, v in [(4, n * LWF_EMPLOYEE), (5, n * LWF_EMPLOYER), (6, n * (LWF_EMPLOYEE + LWF_EMPLOYER))]:
        cell = ws.cell(row=rr, column=c, value=v)
        cell.font = bold; cell.alignment = right; cell.fill = green; cell.border = border
        cell.number_format = '#,##0'

    sig = rr + 3
    ws.cell(row=sig, column=1, value='Signature of Employer:').font = info_b
    ws.cell(row=sig + 1, column=1, value='_______________________').font = info_v
    ws.cell(row=sig, column=5, value='Date:').font = info_b
    ws.cell(row=sig, column=6, value='_______________').font = info_v
    foot = sig + 3
    ws.cell(row=foot, column=1,
            value=f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}   |   Vaishnavi Consultant").font = Font(size=8, italic=True, color='64748B', name='Calibri')
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=LAST)

    for i, w in enumerate([5, 12, 28, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.page_setup.orientation = 'portrait'; ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{hr}:{hr}'

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe = short_est_code(est.company_name)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"LWF_{safe}_{year}.xlsx")
