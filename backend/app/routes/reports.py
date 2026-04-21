from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session
from app import db
from app.models.payroll import (PayrollConfig, SalaryHead, EmployeeSalary,
                                 EmployeeSalaryHead, MonthlyPayroll, PayrollEntry,
                                 PayrollEntryHead)
from app.models.establishment import Establishment
from app.models.employee import Employee
from app.user_context import verify_est_ownership
from datetime import datetime
import calendar
import io
import math

reports_bp = Blueprint('reports', __name__)

import re

def _clean_name_for_statutory(name):
    """Remove salutations (Mr., Mrs., Ms., Smt., Shri, etc.) and special characters
    from employee name for EPF/ESIC statutory templates.
    EPF & ESIC portals accept only alphabets and spaces."""
    if not name:
        return ''
    cleaned = name.strip().upper()
    # Remove common salutations (longer patterns first to avoid partial match)
    cleaned = re.sub(r'^(MRS\.?|MR\.?|MS\.?|SMT\.?|SHRI\.?|DR\.?|KUMARI\.?|KUM\.?)\s*', '', cleaned, flags=re.IGNORECASE)
    # Remove anything that is not a letter or space (ESIC/EPF accept only alphabets + space)
    cleaned = re.sub(r'[^A-Z\s]', '', cleaned)
    # Collapse multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


@reports_bp.app_context_processor
def inject_helpers():
    """Make helper functions available in all templates"""
    import datetime as dt
    def _is_sunday(year, month, day):
        return dt.date(year, month, day).weekday() == 6
    return dict(_is_sunday=_is_sunday)


# =============================================
# QUICK REPORTS — pick Month/Year + Report Type and jump instantly
# =============================================

# Report type catalog: key -> (label, endpoint, category)
QUICK_REPORT_TYPES = [
    # Statutory (Government)
    ('form_b_view',      'Form B — Wage Register (View)',      'reports.form_b_view',              'Statutory'),
    ('form_b_excel',     'Form B — Wage Register (Excel)',     'reports.form_b_excel',             'Statutory'),
    ('form_d_view',      'Form D — Attendance Register (View)','reports.form_d_view',              'Statutory'),
    ('form_d_excel',     'Form D — Attendance Register (Excel)','reports.form_d_excel',            'Statutory'),
    ('form_d_2625_view', 'Form D (26-25) — Attendance Register (View)','reports.form_d_2625_view', 'Statutory'),
    ('form_d_2625_excel','Form D (26-25) — Attendance Register (Excel)','reports.form_d_2625_excel','Statutory'),
    ('attendance_view',  'Attendance Sheet (View)',            'reports.attendance_view',          'Statutory'),
    ('attendance_excel', 'Attendance Sheet (Excel)',           'reports.attendance_excel',         'Statutory'),
    # Salary Statements
    ('statement_f2',     'Salary Statement — Format 2',        'reports.statement_format2',        'Salary Statement'),
    ('statement_f3',     'Salary Statement — Format 3',        'reports.statement_format3',        'Salary Statement'),
    # EPF / ESIC
    ('epf_ecr_view',     'EPF ECR (View)',                     'reports.epf_ecr_view',             'EPF / ESIC'),
    ('epf_ecr_text',     'EPF ECR (Text File)',                'reports.epf_ecr_text',             'EPF / ESIC'),
    ('epf_ecr_csv',      'EPF ECR (CSV)',                      'reports.epf_ecr_csv',              'EPF / ESIC'),
    ('esic_view',        'ESIC Contribution (View)',           'reports.esic_view',                'EPF / ESIC'),
    ('esic_excel',       'ESIC Contribution (Excel)',          'reports.esic_excel',               'EPF / ESIC'),
    # Payslips
    ('payslip_xix',      'Payslip — Form XIX',                 'reports.payslip_form_xix',         'Payslips'),
    ('payslip_pro',      'Payslip — Professional',             'reports.payslip_professional',     'Payslips'),
    # Compliance
    ('compliance',       'Compliance Statement (Monthly)',     'reports.compliance_monthly',       'Compliance'),
]


@reports_bp.route('/reports/quick', methods=['GET', 'POST'])
def quick_reports():
    """Dynamic report launcher — pick month + year + report type and jump."""
    selected_est_id = session.get('selected_est_id')
    if not selected_est_id:
        flash('Please select an establishment first to use Quick Reports.', 'info')
        return redirect(url_for('establishment.establishment_list'))

    est = Establishment.query.get_or_404(selected_est_id)
    verify_est_ownership(est)

    # All available payrolls for this establishment (for "existing months" highlight)
    payrolls = MonthlyPayroll.query.filter_by(
        establishment_id=est.id
    ).order_by(MonthlyPayroll.year.desc(), MonthlyPayroll.month.desc()).all()

    # Build a set of existing (year, month) tuples for quick lookup
    existing_set = {(p.year, p.month): p.id for p in payrolls}

    now = datetime.now()
    current_year = now.year

    # Available years: from earliest payroll (or 2019) up to current year
    if payrolls:
        earliest_year = min(p.year for p in payrolls)
    else:
        earliest_year = 2019
    year_range = list(range(current_year, min(earliest_year, 2019) - 1, -1))

    if request.method == 'POST':
        try:
            month = int(request.form.get('month'))
            year = int(request.form.get('year'))
        except (ValueError, TypeError):
            flash('Please select a valid month and year.', 'warning')
            return redirect(url_for('reports.quick_reports'))

        report_key = request.form.get('report_type', '')
        # Find the report entry
        report_entry = next((r for r in QUICK_REPORT_TYPES if r[0] == report_key), None)
        if not report_entry:
            flash('Please select a valid report type.', 'warning')
            return redirect(url_for('reports.quick_reports'))

        # Look up the payroll batch
        payroll = MonthlyPayroll.query.filter_by(
            establishment_id=est.id, month=month, year=year).first()

        if not payroll:
            flash(f'No payroll exists for {calendar.month_name[month]} {year}. '
                  f'Please create the payroll first.', 'warning')
            return redirect(url_for('reports.quick_reports'))

        # Redirect to the actual report route
        _, _, endpoint, _ = report_entry
        return redirect(url_for(endpoint, payroll_id=payroll.id))

    # Group report types by category for the UI
    grouped = {}
    for key, label, endpoint, category in QUICK_REPORT_TYPES:
        grouped.setdefault(category, []).append((key, label))

    return render_template('reports/quick_reports.html',
                           est=est,
                           year_range=year_range,
                           current_month=now.month,
                           current_year=current_year,
                           existing_set=existing_set,
                           payrolls=payrolls,
                           grouped_reports=grouped)


# =============================================
# REPORT: Form B — Wage Register (Government)
# =============================================

def _get_payroll_data(payroll_id, include_zero=False):
    """Common helper to fetch all data needed for reports.
    By default, excludes employees with zero attendance (days_present == 0).
    Set include_zero=True to include them (e.g., for ESIC template).
    """
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    query = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id).join(Employee).order_by(Employee.name)

    # Filter out zero-attendance employees for salary statements, wage registers, etc.
    if not include_zero:
        query = query.filter(PayrollEntry.days_present > 0)

    entries = query.all()

    # Get earning heads for this establishment
    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True, head_type='earning', is_in_gross=True
    ).order_by(SalaryHead.display_order).all()

    # Build head-wise data for each entry + daily rate lookup
    import json as _json
    for entry in entries:
        entry.head_amounts = {}
        for peh in entry.head_breakup:
            entry.head_amounts[peh.salary_head_id] = peh

        # Attach daily rate for daily wages display in reports
        # Priority: rate_overrides JSON on this entry > historical EmployeeSalary > current salary
        cur_sal = EmployeeSalary.query.filter_by(
            employee_id=entry.employee_id, is_current=True).first()
        entry._salary_type = cur_sal.salary_type if cur_sal else (config.salary_type if config else 'monthly_fixed')

        # Start with stored salary defaults
        _daily_rate = cur_sal.daily_rate if cur_sal and cur_sal.daily_rate else 0
        _eff_gross = cur_sal.gross_salary if cur_sal else (entry.gross_salary or 0)

        # Apply per-entry rate overrides (historical — reflects what was uploaded for THIS month)
        if hasattr(entry, 'rate_overrides') and entry.rate_overrides:
            try:
                _ro = _json.loads(entry.rate_overrides)
                if 'daily_rate' in _ro and _ro['daily_rate']:
                    _daily_rate = float(_ro['daily_rate'])
                if 'gross' in _ro and _ro['gross']:
                    _eff_gross = float(_ro['gross'])
            except (ValueError, TypeError):
                pass

        entry._daily_rate = _daily_rate
        entry._effective_gross = _eff_gross

        # Calculate NPH amount for report display
        ph_count = entry.paid_holidays or 0
        working_days = payroll.working_days or 26
        if ph_count > 0:
            if entry._salary_type == 'daily_wages' and entry._daily_rate:
                entry._nph_amount = round(ph_count * entry._daily_rate)
            elif entry.gross_salary and working_days > 0:
                entry._nph_amount = round(ph_count * (entry.gross_salary / working_days))
            else:
                entry._nph_amount = 0
        else:
            entry._nph_amount = 0

    return payroll, est, config, entries, heads


@reports_bp.route('/payroll/<int:payroll_id>/report/form-b')
def form_b_view(payroll_id):
    """Form B Wage Register — HTML view for printing as PDF"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)

    # Map heads to Form B columns
    head_map = _map_heads_to_form_b(heads)

    # Calculate Form B rows
    rows = _build_form_b_rows(entries, heads, head_map, config, payroll)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('reports/form_b.html',
                           payroll=payroll, est=est, config=config,
                           entries=entries, rows=rows,
                           head_map=head_map,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/form-b/excel')
def form_b_excel(payroll_id):
    """Form B Wage Register — Excel download"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)

    # Map heads to Form B columns
    head_map = _map_heads_to_form_b(heads)

    # Calculate Form B rows
    rows = _build_form_b_rows(entries, heads, head_map, config, payroll)

    # Generate Excel
    output = _generate_form_b_excel(payroll, est, config, entries, rows, head_map)

    filename = f"Form_B_{est.company_name}_{payroll.month_name}_{payroll.year}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =============================================
# ATTENDANCE: Random Marks Generation Logic
# =============================================

def _get_rest_days(year, month, num_days, config):
    """Get list of rest day numbers based on config"""
    import datetime as dt
    rest_days = []

    if config.rest_day_type == 'rotation':
        # Every 7th day starting from day 7
        for d in range(7, num_days + 1, 7):
            rest_days.append(d)
    elif config.rest_day_type == 'fixed_day':
        weekday = config.rest_day_weekday if config.rest_day_weekday is not None else 6
        for d in range(1, num_days + 1):
            if dt.date(year, month, d).weekday() == weekday:
                rest_days.append(d)
    else:  # 'sunday' default
        for d in range(1, num_days + 1):
            if dt.date(year, month, d).weekday() == 6:
                rest_days.append(d)

    return rest_days


def _get_holiday_days(payroll):
    """Parse holiday_dates string into list of day numbers"""
    if not payroll.holiday_dates:
        return []
    try:
        return [int(d.strip()) for d in payroll.holiday_dates.split(',') if d.strip().isdigit()]
    except Exception:
        return []


def _generate_random_marks(days_present, days_absent, num_days, rest_days, holiday_days):
    """Generate natural-looking random P/A/R/H marks for attendance"""
    import random

    marks = {}  # day -> mark

    # 1. Place R on rest days
    for d in rest_days:
        if 1 <= d <= num_days:
            marks[d] = 'R'

    # 2. Place H on holiday days (skip if already rest day)
    for d in holiday_days:
        if 1 <= d <= num_days and d not in marks:
            marks[d] = 'H'

    # 3. Get available working days
    available = [d for d in range(1, num_days + 1) if d not in marks]

    # 4. Calculate how many P and A to place
    present = int(round(days_present))
    absent = int(round(days_absent))

    # Adjust if totals don't match available days
    total_needed = present + absent
    if total_needed > len(available):
        present = len(available) - absent
        if present < 0:
            present = 0
            absent = len(available)
    elif total_needed < len(available):
        # Extra days — assign as present
        present = len(available) - absent

    # 5. Create randomized assignment
    # Strategy: distribute A's randomly among available days for natural look
    assignments = ['P'] * present + ['A'] * absent

    # Pad if needed
    while len(assignments) < len(available):
        assignments.append('P')
    assignments = assignments[:len(available)]

    # Shuffle for random distribution
    random.shuffle(assignments)

    # 6. Assign to available days
    for i, d in enumerate(available):
        marks[d] = assignments[i]

    return marks


def _build_attendance_data(payroll, config, entries, include_zero=False):
    """Build attendance marks for all employees.
    By default, excludes employees with zero attendance.
    """
    _, num_days = calendar.monthrange(payroll.year, payroll.month)
    rest_days = _get_rest_days(payroll.year, payroll.month, num_days, config)
    holiday_days = _get_holiday_days(payroll)

    attendance = []
    for entry in entries:
        # Skip zero-attendance employees unless explicitly included
        if not include_zero and entry.days_present <= 0:
            continue

        marks = _generate_random_marks(
            entry.days_present, entry.days_absent,
            num_days, rest_days, holiday_days
        )
        # Calculate hours (8 hours per present day)
        hours = int(round(entry.days_present)) * 8

        attendance.append({
            'entry': entry,
            'emp': entry.employee,
            'marks': marks,
            'days_present': entry.days_present,
            'days_absent': entry.days_absent,
            'paid_holidays': entry.paid_holidays,
            'total_days': entry.total_payable_days,
            'hours': hours,
        })

    return attendance, num_days, rest_days, holiday_days


# =============================================
# REPORT: Form D — Attendance Register (Govt)
# =============================================

@reports_bp.route('/payroll/<int:payroll_id>/report/form-d')
def form_d_view(payroll_id):
    """Form D Attendance Register — Government format HTML"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, num_days, rest_days, holiday_days = _build_attendance_data(payroll, config, entries)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('reports/form_d.html',
                           payroll=payroll, est=est, config=config,
                           attendance=attendance, num_days=num_days,
                           rest_days=rest_days, holiday_days=holiday_days,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/form-d/excel')
def form_d_excel(payroll_id):
    """Form D Attendance Register — Excel download"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, num_days, rest_days, holiday_days = _build_attendance_data(payroll, config, entries)

    output = _generate_form_d_excel(payroll, est, config, attendance, num_days, rest_days, holiday_days)

    filename = f"Form_D_{est.company_name}_{payroll.month_name}_{payroll.year}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =============================================
# REPORT: Form D (26-25) — Attendance Register (26th to 25th cycle)
# For establishments that follow ESIC/factory attendance cycle
# Jan-2026 payroll → 26-Dec-2025 to 25-Jan-2026
# =============================================

def _build_attendance_data_2625(payroll, config, entries, include_zero=False):
    """Build attendance marks for the 26-25 cycle.
    For payroll month M/Y: period = 26th of (M-1) to 25th of M.
    Returns (attendance, date_list, rest_indices, holiday_indices)
    where date_list = [(year, month, day), ...] for each column.
    """
    import datetime as dt

    # Calculate previous month (for the 26th start)
    if payroll.month == 1:
        prev_month, prev_year = 12, payroll.year - 1
    else:
        prev_month, prev_year = payroll.month - 1, payroll.year

    # Days in previous month (to know how many days from 26 to end)
    _, prev_month_days = calendar.monthrange(prev_year, prev_month)

    # Build full date list: 26th-prev to end-prev, then 1st-current to 25th-current
    date_list = []
    # Part 1: 26th of previous month to last day of previous month
    for d in range(26, prev_month_days + 1):
        date_list.append((prev_year, prev_month, d))
    # Part 2: 1st of current month to 25th of current month
    for d in range(1, 26):
        date_list.append((payroll.year, payroll.month, d))

    num_cols = len(date_list)  # Typically 30 or 31 days

    # Determine rest days (Sundays or as per config)
    rest_indices = set()   # 0-based index into date_list
    for idx, (y, m, d) in enumerate(date_list):
        date_obj = dt.date(y, m, d)
        if config and config.rest_day_type == 'fixed_day':
            weekday = config.rest_day_weekday if config.rest_day_weekday is not None else 6
            if date_obj.weekday() == weekday:
                rest_indices.add(idx)
        elif config and config.rest_day_type == 'rotation':
            # Every 7th day from the start of the period
            if (idx + 1) % 7 == 0:
                rest_indices.add(idx)
        else:
            # Default: Sundays
            if date_obj.weekday() == 6:
                rest_indices.add(idx)

    # Holiday indices — holidays from the payroll's holiday_dates are day numbers
    # of the CURRENT month. Map them to indices in our date_list.
    holiday_indices = set()
    if payroll.holiday_dates:
        try:
            h_days = [int(x.strip()) for x in payroll.holiday_dates.split(',') if x.strip().isdigit()]
            for hd in h_days:
                if 1 <= hd <= 25:
                    # This holiday falls in Part 2 (current month 1-25)
                    part1_len = prev_month_days - 26 + 1
                    idx = part1_len + hd - 1
                    if 0 <= idx < num_cols:
                        holiday_indices.add(idx)
        except Exception:
            pass

    attendance = []
    for entry in entries:
        if not include_zero and entry.days_present <= 0:
            continue

        # Generate marks using index-based system for 26-25 period
        marks = _generate_random_marks_indexed(
            entry.days_present, entry.days_absent,
            num_cols, rest_indices, holiday_indices
        )
        hours = int(round(entry.days_present)) * 8

        attendance.append({
            'entry': entry,
            'emp': entry.employee,
            'marks': marks,          # {0-based index: 'P'/'A'/'R'/'H'}
            'days_present': entry.days_present,
            'days_absent': entry.days_absent,
            'paid_holidays': entry.paid_holidays,
            'total_days': entry.total_payable_days,
            'hours': hours,
        })

    return attendance, date_list, num_cols, rest_indices, holiday_indices


def _generate_random_marks_indexed(days_present, days_absent, num_cols, rest_indices, holiday_indices):
    """Generate P/A/R/H marks for index-based (0..num_cols-1) date list.
    Similar to _generate_random_marks but works with 0-based indices."""
    import random

    marks = {}

    # 1. Place R on rest days
    for idx in rest_indices:
        if 0 <= idx < num_cols:
            marks[idx] = 'R'

    # 2. Place H on holidays (skip if already rest day)
    for idx in holiday_indices:
        if 0 <= idx < num_cols and idx not in marks:
            marks[idx] = 'H'

    # 3. Get available working day indices
    available = [i for i in range(num_cols) if i not in marks]

    # 4. Calculate P and A
    present = int(round(days_present))
    absent = int(round(days_absent))

    total_needed = present + absent
    if total_needed > len(available):
        present = len(available) - absent
        if present < 0:
            present = 0
            absent = len(available)
    elif total_needed < len(available):
        present = len(available) - absent

    # 5. Randomize
    assignments = ['P'] * present + ['A'] * absent
    while len(assignments) < len(available):
        assignments.append('P')
    assignments = assignments[:len(available)]
    random.shuffle(assignments)

    # 6. Assign
    for i, idx in enumerate(available):
        marks[idx] = assignments[i]

    return marks


@reports_bp.route('/payroll/<int:payroll_id>/report/form-d-2625')
def form_d_2625_view(payroll_id):
    """Form D (26-25) Attendance Register — HTML view"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, date_list, num_cols, rest_indices, holiday_indices = \
        _build_attendance_data_2625(payroll, config, entries)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('reports/form_d_2625.html',
                           payroll=payroll, est=est, config=config,
                           attendance=attendance, date_list=date_list,
                           num_cols=num_cols, rest_indices=rest_indices,
                           holiday_indices=holiday_indices,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/form-d-2625/excel')
def form_d_2625_excel(payroll_id):
    """Form D (26-25) Attendance Register — Excel download"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, date_list, num_cols, rest_indices, holiday_indices = \
        _build_attendance_data_2625(payroll, config, entries)

    output = _generate_form_d_2625_excel(payroll, est, config, attendance,
                                          date_list, num_cols, rest_indices, holiday_indices)

    # Previous month for filename
    if payroll.month == 1:
        prev_m_name = 'Dec'
    else:
        prev_m_name = calendar.month_abbr[payroll.month - 1]

    filename = f"Form_D_26-25_{est.company_name}_{prev_m_name}-{payroll.month_name}_{payroll.year}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =============================================
# BULK B+D: Combined HTML view for multiple payrolls
# User sets Landscape + Legal in browser print dialog once → single PDF
# =============================================

@reports_bp.route('/reports/bulk-bd', methods=['GET', 'POST'])
def bulk_bd():
    """Pick establishment + months → combined Form B/D HTML page, user prints once."""
    from app.user_context import user_establishments
    user_ests = user_establishments().order_by(Establishment.company_name).all()

    if request.method == 'POST':
        payroll_ids = request.form.getlist('payroll_ids', type=int)
        include_b = request.form.get('include_b') == 'on'
        include_d = request.form.get('include_d') == 'on'

        if not payroll_ids:
            flash('Please select at least one payroll month', 'warning')
            return redirect(url_for('reports.bulk_bd'))
        if not (include_b or include_d):
            flash('Please select at least one of Form B or Form D', 'warning')
            return redirect(url_for('reports.bulk_bd'))

        # Build a list of rendered section snippets (ownership enforced in _get_payroll_data)
        sections = []
        est_name = ''
        for pid in payroll_ids:
            payroll, est, config, entries, heads = _get_payroll_data(pid)
            est_name = est.company_name
            if include_b:
                head_map = _map_heads_to_form_b(heads)
                rows = _build_form_b_rows(entries, heads, head_map, config, payroll)
                sections.append({
                    'kind': 'B',
                    'payroll': payroll, 'est': est, 'config': config,
                    'entries': entries, 'rows': rows, 'head_map': head_map,
                })
            if include_d:
                attendance, num_days, rest_days, holiday_days = _build_attendance_data(payroll, config, entries)
                sections.append({
                    'kind': 'D',
                    'payroll': payroll, 'est': est, 'config': config,
                    'attendance': attendance, 'num_days': num_days,
                    'rest_days': rest_days, 'holiday_days': holiday_days,
                })

        generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
        return render_template('reports/bulk_bd_combined.html',
                               sections=sections, est_name=est_name,
                               generated_on=generated_on)

    # GET: show selection form
    est_payrolls = {}
    for e in user_ests:
        prs = MonthlyPayroll.query.filter_by(establishment_id=e.id).order_by(
            MonthlyPayroll.year.desc(), MonthlyPayroll.month.desc()
        ).all()
        est_payrolls[e.id] = prs

    return render_template('reports/bulk_bd.html',
                           establishments=user_ests,
                           est_payrolls=est_payrolls)



# =============================================
# REPORT: Attendance Register (Professional)
# =============================================

@reports_bp.route('/payroll/<int:payroll_id>/report/attendance')
def attendance_view(payroll_id):
    """Attendance Register — Professional format HTML"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, num_days, rest_days, holiday_days = _build_attendance_data(payroll, config, entries)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('reports/attendance_register.html',
                           payroll=payroll, est=est, config=config,
                           attendance=attendance, num_days=num_days,
                           rest_days=rest_days, holiday_days=holiday_days,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/attendance/excel')
def attendance_excel(payroll_id):
    """Attendance Register — Professional Excel download"""
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    attendance, num_days, rest_days, holiday_days = _build_attendance_data(payroll, config, entries)

    output = _generate_attendance_excel(payroll, est, config, attendance, num_days, rest_days, holiday_days)

    filename = f"Attendance_Register_{est.company_name}_{payroll.month_name}_{payroll.year}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _generate_form_d_excel(payroll, est, config, attendance, num_days, rest_days, holiday_days):
    """Generate Form D Excel — exact government format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = f"Form D - {payroll.month_name} {payroll.year}"

    # Styles — readable font sizes for Legal paper print
    title_font = Font(name='Arial', size=14, bold=True)
    subtitle_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=9, bold=True)
    data_font = Font(name='Arial', size=9)
    bold_font = Font(name='Arial', size=9, bold=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow like Form D
    rest_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    holiday_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    # Title area
    # Columns: A=Sl, B=Name, C=Relay, D=Place, E..E+num_days-1=Days, then Summary, Remarks, Signature
    day_start_col = 5  # Column E
    last_day_col = day_start_col + num_days - 1
    summary_col = last_day_col + 1
    remarks_col = summary_col + 1
    sig_col = remarks_col + 1
    last_col_letter = get_column_letter(sig_col)

    # Row 1-2: Title
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'FORM D'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'FORMAT OF ATTENDANCE REGISTER'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center

    # Row 4: Establishment details
    ws.merge_cells('A4:B4')
    ws['A4'] = f'Name of Establishment: {est.company_name}'
    ws['A4'].font = Font(name='Arial', size=9, bold=True)

    mid_col = get_column_letter(day_start_col + num_days // 3)
    ws[f'{mid_col}4'] = f'Name of Owner: {est.address or ""}'
    ws[f'{mid_col}4'].font = Font(name='Arial', size=9, bold=True)

    far_col = get_column_letter(day_start_col + (2 * num_days) // 3)
    ws[f'{far_col}4'] = f'LIN ______________'
    ws[f'{far_col}4'].font = Font(name='Arial', size=9)

    # Row 5: Period
    _, last_day_num = calendar.monthrange(payroll.year, payroll.month)
    ws.merge_cells('A5:D5')
    ws['A5'] = f'For the Period From 01.{payroll.month:02d}.{payroll.year} To {last_day_num}.{payroll.month:02d}.{payroll.year}'
    ws['A5'].font = Font(name='Arial', size=9, bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 10  # Emp Code
    ws.column_dimensions['B'].width = 22  # Name
    ws.column_dimensions['C'].width = 11  # Designation
    ws.column_dimensions['D'].width = 10  # Place of Work

    for d in range(num_days):
        col = get_column_letter(day_start_col + d)
        ws.column_dimensions[col].width = 4

    ws.column_dimensions[get_column_letter(summary_col)].width = 9
    ws.column_dimensions[get_column_letter(remarks_col)].width = 9
    ws.column_dimensions[get_column_letter(sig_col)].width = 14

    # Row 7: "Date" header spanning day columns
    header_row1 = 7
    ws.merge_cells(f'{get_column_letter(day_start_col)}{header_row1}:{get_column_letter(last_day_col)}{header_row1}')
    ws[f'{get_column_letter(day_start_col)}{header_row1}'] = 'Date'
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].font = header_font
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].alignment = center
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].border = thin_border

    # Row 8: Main headers
    header_row = 8
    main_headers = [
        ('A', 'Emp.\nCode'),
        ('B', 'Name'),
        ('C', 'Designation'),
        ('D', 'Place of\nWork'),
    ]

    for col, text in main_headers:
        ws.merge_cells(f'{col}{header_row}:{col}{header_row + 1}')
        cell = ws[f'{col}{header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    # Day number headers
    for d in range(1, num_days + 1):
        col = get_column_letter(day_start_col + d - 1)
        cell = ws[f'{col}{header_row}']
        cell.value = d
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        if d in rest_days:
            cell.fill = rest_fill
        elif d in holiday_days:
            cell.fill = holiday_fill
        else:
            cell.fill = header_fill

    # Summary, Remarks, Signature headers
    ws.merge_cells(f'{get_column_letter(summary_col)}{header_row}:{get_column_letter(summary_col)}{header_row + 1}')
    ws[f'{get_column_letter(summary_col)}{header_row}'] = 'Summary No.\nOf Days'
    ws[f'{get_column_letter(summary_col)}{header_row}'].font = header_font
    ws[f'{get_column_letter(summary_col)}{header_row}'].alignment = center
    ws[f'{get_column_letter(summary_col)}{header_row}'].border = thin_border
    ws[f'{get_column_letter(summary_col)}{header_row}'].fill = header_fill

    ws.merge_cells(f'{get_column_letter(remarks_col)}{header_row}:{get_column_letter(remarks_col)}{header_row + 1}')
    ws[f'{get_column_letter(remarks_col)}{header_row}'] = 'Remarks\nNo of hours'
    ws[f'{get_column_letter(remarks_col)}{header_row}'].font = header_font
    ws[f'{get_column_letter(remarks_col)}{header_row}'].alignment = center
    ws[f'{get_column_letter(remarks_col)}{header_row}'].border = thin_border
    ws[f'{get_column_letter(remarks_col)}{header_row}'].fill = header_fill

    ws.merge_cells(f'{get_column_letter(sig_col)}{header_row}:{get_column_letter(sig_col)}{header_row + 1}')
    ws[f'{get_column_letter(sig_col)}{header_row}'] = 'Signature of\nRegister Keeper'
    ws[f'{get_column_letter(sig_col)}{header_row}'].font = header_font
    ws[f'{get_column_letter(sig_col)}{header_row}'].alignment = center
    ws[f'{get_column_letter(sig_col)}{header_row}'].border = thin_border
    ws[f'{get_column_letter(sig_col)}{header_row}'].fill = header_fill

    # In/Out sub-rows (row 9)
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'] = 'In:'
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'].font = Font(name='Arial', size=7)
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'].border = thin_border

    ws.row_dimensions[header_row].height = 40
    ws.row_dimensions[header_row + 1].height = 14

    # Column number row (row 10)
    num_row = header_row + 2
    col_nums = {'A': '1', 'B': '2', 'C': '3', 'D': '4'}
    for col, num in col_nums.items():
        ws[f'{col}{num_row}'] = num
        ws[f'{col}{num_row}'].font = Font(name='Arial', size=7, bold=True)
        ws[f'{col}{num_row}'].alignment = center
        ws[f'{col}{num_row}'].border = thin_border

    # Day column numbers (5, 6, 7...)
    for d in range(1, num_days + 1):
        col = get_column_letter(day_start_col + d - 1)
        ws[f'{col}{num_row}'] = ''
        ws[f'{col}{num_row}'].border = thin_border

    ws[f'{get_column_letter(summary_col)}{num_row}'] = '8'
    ws[f'{get_column_letter(summary_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(summary_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(summary_col)}{num_row}'].border = thin_border

    ws[f'{get_column_letter(remarks_col)}{num_row}'] = '9'
    ws[f'{get_column_letter(remarks_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(remarks_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(remarks_col)}{num_row}'].border = thin_border

    ws[f'{get_column_letter(sig_col)}{num_row}'] = '10'
    ws[f'{get_column_letter(sig_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(sig_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(sig_col)}{num_row}'].border = thin_border

    # ---- Data Rows ----
    data_start = num_row + 1
    t_days = 0
    t_hours = 0

    for idx, att in enumerate(attendance):
        r = data_start + idx
        emp = att['emp']
        ws.row_dimensions[r].height = 22

        # Emp Code (internal code preferred, fallback to system code)
        emp_code = emp.internal_emp_code if emp.use_internal_code and emp.internal_emp_code else emp.emp_code
        ws[f'A{r}'] = emp_code
        ws[f'A{r}'].font = data_font
        ws[f'A{r}'].alignment = center
        ws[f'A{r}'].border = thin_border

        ws[f'B{r}'] = emp.name.upper()
        ws[f'B{r}'].font = bold_font
        ws[f'B{r}'].alignment = left_align
        ws[f'B{r}'].border = thin_border

        ws[f'C{r}'] = emp.designation or ''
        ws[f'C{r}'].font = data_font
        ws[f'C{r}'].alignment = center
        ws[f'C{r}'].border = thin_border

        ws[f'D{r}'] = ''
        ws[f'D{r}'].border = thin_border

        # Day marks
        for d in range(1, num_days + 1):
            col = get_column_letter(day_start_col + d - 1)
            mark = att['marks'].get(d, '')
            cell = ws[f'{col}{r}']
            cell.value = mark
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border

            # Color coding
            if mark == 'R':
                cell.fill = rest_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='FF0000')
            elif mark == 'H':
                cell.fill = holiday_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='1B5E20')
            elif mark == 'A':
                cell.font = Font(name='Arial', size=8, color='FF0000')

        # Summary
        ws[f'{get_column_letter(summary_col)}{r}'] = int(round(att['days_present']))
        ws[f'{get_column_letter(summary_col)}{r}'].font = bold_font
        ws[f'{get_column_letter(summary_col)}{r}'].alignment = center
        ws[f'{get_column_letter(summary_col)}{r}'].border = thin_border

        ws[f'{get_column_letter(remarks_col)}{r}'] = att['hours']
        ws[f'{get_column_letter(remarks_col)}{r}'].font = bold_font
        ws[f'{get_column_letter(remarks_col)}{r}'].alignment = center
        ws[f'{get_column_letter(remarks_col)}{r}'].border = thin_border

        ws[f'{get_column_letter(sig_col)}{r}'] = ''
        ws[f'{get_column_letter(sig_col)}{r}'].border = thin_border

        t_days += int(round(att['days_present']))
        t_hours += att['hours']

    # ---- Totals Row ----
    total_r = data_start + len(attendance)
    ws.row_dimensions[total_r].height = 22

    ws.merge_cells(f'A{total_r}:D{total_r}')
    ws[f'A{total_r}'] = ''
    ws[f'A{total_r}'].border = thin_border
    ws[f'A{total_r}'].fill = total_fill
    for c in ['B', 'C', 'D']:
        ws[f'{c}{total_r}'].border = thin_border
        ws[f'{c}{total_r}'].fill = total_fill

    for d in range(1, num_days + 1):
        col = get_column_letter(day_start_col + d - 1)
        ws[f'{col}{total_r}'].border = thin_border
        ws[f'{col}{total_r}'].fill = total_fill

    ws[f'{get_column_letter(summary_col)}{total_r}'] = t_days
    ws[f'{get_column_letter(summary_col)}{total_r}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'{get_column_letter(summary_col)}{total_r}'].alignment = center
    ws[f'{get_column_letter(summary_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(summary_col)}{total_r}'].fill = total_fill

    ws[f'{get_column_letter(remarks_col)}{total_r}'] = t_hours
    ws[f'{get_column_letter(remarks_col)}{total_r}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'{get_column_letter(remarks_col)}{total_r}'].alignment = center
    ws[f'{get_column_letter(remarks_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(remarks_col)}{total_r}'].fill = total_fill

    ws[f'{get_column_letter(sig_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(sig_col)}{total_r}'].fill = total_fill

    # Print setup — Legal Size Landscape
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5  # Legal (8.5 x 14 inches)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f'{header_row}:{header_row + 1}'
    ws.page_margins = PageMargins(left=0.5, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_form_d_2625_excel(payroll, est, config, attendance, date_list, num_cols, rest_indices, holiday_indices):
    """Generate Form D (26-25) Excel — government format with 26th-to-25th date columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active

    # Previous month name for title
    if payroll.month == 1:
        prev_month, prev_year = 12, payroll.year - 1
    else:
        prev_month, prev_year = payroll.month - 1, payroll.year
    prev_month_name = calendar.month_name[prev_month]

    ws.title = f"Form D 26-25 {payroll.month_name} {payroll.year}"

    # Styles
    title_font = Font(name='Arial', size=14, bold=True)
    subtitle_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=9, bold=True)
    data_font = Font(name='Arial', size=9)
    bold_font = Font(name='Arial', size=9, bold=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    total_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    rest_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    holiday_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    # Columns: A=Sl, B=Name, C=Designation, D=Place, E..E+num_cols-1=Days, then Summary, Remarks, Signature
    day_start_col = 5  # Column E
    last_day_col = day_start_col + num_cols - 1
    summary_col = last_day_col + 1
    remarks_col = summary_col + 1
    sig_col = remarks_col + 1
    last_col_letter = get_column_letter(sig_col)

    # Row 1-2: Title
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = 'FORM D'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'FORMAT OF ATTENDANCE REGISTER (26-25 Cycle)'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center

    # Row 4: Establishment details
    ws.merge_cells('A4:B4')
    ws['A4'] = f'Name of Establishment: {est.company_name}'
    ws['A4'].font = Font(name='Arial', size=9, bold=True)

    mid_col = get_column_letter(day_start_col + num_cols // 3)
    ws[f'{mid_col}4'] = f'Name of Owner: {est.address or ""}'
    ws[f'{mid_col}4'].font = Font(name='Arial', size=9, bold=True)

    far_col = get_column_letter(day_start_col + (2 * num_cols) // 3)
    ws[f'{far_col}4'] = f'LIN ______________'
    ws[f'{far_col}4'].font = Font(name='Arial', size=9)

    # Row 5: Period — 26th prev to 25th current
    _, prev_month_days = calendar.monthrange(prev_year, prev_month)
    ws.merge_cells('A5:D5')
    ws['A5'] = f'For the Period From 26.{prev_month:02d}.{prev_year} To 25.{payroll.month:02d}.{payroll.year}'
    ws['A5'].font = Font(name='Arial', size=9, bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 10
    for i in range(num_cols):
        col = get_column_letter(day_start_col + i)
        ws.column_dimensions[col].width = 4
    ws.column_dimensions[get_column_letter(summary_col)].width = 9
    ws.column_dimensions[get_column_letter(remarks_col)].width = 9
    ws.column_dimensions[get_column_letter(sig_col)].width = 14

    # Row 7: "Date" header spanning day columns
    header_row1 = 7
    ws.merge_cells(f'{get_column_letter(day_start_col)}{header_row1}:{get_column_letter(last_day_col)}{header_row1}')
    ws[f'{get_column_letter(day_start_col)}{header_row1}'] = 'Date'
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].font = header_font
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].alignment = center
    ws[f'{get_column_letter(day_start_col)}{header_row1}'].border = thin_border

    # Row 8: Main headers
    header_row = 8
    main_headers = [
        ('A', 'Emp.\nCode'),
        ('B', 'Name'),
        ('C', 'Designation'),
        ('D', 'Place of\nWork'),
    ]
    for col, text in main_headers:
        ws.merge_cells(f'{col}{header_row}:{col}{header_row + 1}')
        cell = ws[f'{col}{header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    # Day number headers — show actual date number (26, 27, ... 31, 1, 2, ... 25)
    for i in range(num_cols):
        col = get_column_letter(day_start_col + i)
        y, m, d = date_list[i]
        cell = ws[f'{col}{header_row}']
        cell.value = d
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        if i in rest_indices:
            cell.fill = rest_fill
        elif i in holiday_indices:
            cell.fill = holiday_fill
        else:
            cell.fill = header_fill

    # Summary, Remarks, Signature headers
    for col_idx, label in [(summary_col, 'Summary No.\nOf Days'), (remarks_col, 'Remarks\nNo of hours'), (sig_col, 'Signature of\nRegister Keeper')]:
        c = get_column_letter(col_idx)
        ws.merge_cells(f'{c}{header_row}:{c}{header_row + 1}')
        ws[f'{c}{header_row}'] = label
        ws[f'{c}{header_row}'].font = header_font
        ws[f'{c}{header_row}'].alignment = center
        ws[f'{c}{header_row}'].border = thin_border
        ws[f'{c}{header_row}'].fill = header_fill

    # In/Out sub-row (row 9)
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'] = 'In:'
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'].font = Font(name='Arial', size=7)
    ws[f'{get_column_letter(day_start_col)}{header_row + 1}'].border = thin_border

    ws.row_dimensions[header_row].height = 40
    ws.row_dimensions[header_row + 1].height = 14

    # Column number row (row 10)
    num_row = header_row + 2
    col_nums = {'A': '1', 'B': '2', 'C': '3', 'D': '4'}
    for col, num in col_nums.items():
        ws[f'{col}{num_row}'] = num
        ws[f'{col}{num_row}'].font = Font(name='Arial', size=7, bold=True)
        ws[f'{col}{num_row}'].alignment = center
        ws[f'{col}{num_row}'].border = thin_border
    for i in range(num_cols):
        col = get_column_letter(day_start_col + i)
        ws[f'{col}{num_row}'] = ''
        ws[f'{col}{num_row}'].border = thin_border
    ws[f'{get_column_letter(summary_col)}{num_row}'] = '8'
    ws[f'{get_column_letter(summary_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(summary_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(summary_col)}{num_row}'].border = thin_border
    ws[f'{get_column_letter(remarks_col)}{num_row}'] = '9'
    ws[f'{get_column_letter(remarks_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(remarks_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(remarks_col)}{num_row}'].border = thin_border
    ws[f'{get_column_letter(sig_col)}{num_row}'] = '10'
    ws[f'{get_column_letter(sig_col)}{num_row}'].font = Font(name='Arial', size=7, bold=True)
    ws[f'{get_column_letter(sig_col)}{num_row}'].alignment = center
    ws[f'{get_column_letter(sig_col)}{num_row}'].border = thin_border

    # ---- Data Rows ----
    data_start = num_row + 1
    t_days = 0
    t_hours = 0

    for idx, att in enumerate(attendance):
        r = data_start + idx
        emp = att['emp']
        ws.row_dimensions[r].height = 22

        emp_code = emp.internal_emp_code if emp.use_internal_code and emp.internal_emp_code else emp.emp_code
        ws[f'A{r}'] = emp_code
        ws[f'A{r}'].font = data_font
        ws[f'A{r}'].alignment = center
        ws[f'A{r}'].border = thin_border

        ws[f'B{r}'] = emp.name.upper()
        ws[f'B{r}'].font = bold_font
        ws[f'B{r}'].alignment = left_align
        ws[f'B{r}'].border = thin_border

        ws[f'C{r}'] = emp.designation or ''
        ws[f'C{r}'].font = data_font
        ws[f'C{r}'].alignment = center
        ws[f'C{r}'].border = thin_border

        ws[f'D{r}'] = ''
        ws[f'D{r}'].border = thin_border

        # Day marks (0-based index)
        for i in range(num_cols):
            col = get_column_letter(day_start_col + i)
            mark = att['marks'].get(i, '')
            cell = ws[f'{col}{r}']
            cell.value = mark
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            if mark == 'R':
                cell.fill = rest_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='FF0000')
            elif mark == 'H':
                cell.fill = holiday_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='1B5E20')
            elif mark == 'A':
                cell.font = Font(name='Arial', size=8, color='FF0000')

        # Summary
        ws[f'{get_column_letter(summary_col)}{r}'] = int(round(att['days_present']))
        ws[f'{get_column_letter(summary_col)}{r}'].font = bold_font
        ws[f'{get_column_letter(summary_col)}{r}'].alignment = center
        ws[f'{get_column_letter(summary_col)}{r}'].border = thin_border

        ws[f'{get_column_letter(remarks_col)}{r}'] = att['hours']
        ws[f'{get_column_letter(remarks_col)}{r}'].font = bold_font
        ws[f'{get_column_letter(remarks_col)}{r}'].alignment = center
        ws[f'{get_column_letter(remarks_col)}{r}'].border = thin_border

        ws[f'{get_column_letter(sig_col)}{r}'] = ''
        ws[f'{get_column_letter(sig_col)}{r}'].border = thin_border

        t_days += int(round(att['days_present']))
        t_hours += att['hours']

    # ---- Totals Row ----
    total_r = data_start + len(attendance)
    ws.row_dimensions[total_r].height = 22

    ws.merge_cells(f'A{total_r}:D{total_r}')
    ws[f'A{total_r}'] = ''
    ws[f'A{total_r}'].border = thin_border
    ws[f'A{total_r}'].fill = total_fill
    for c in ['B', 'C', 'D']:
        ws[f'{c}{total_r}'].border = thin_border
        ws[f'{c}{total_r}'].fill = total_fill
    for i in range(num_cols):
        col = get_column_letter(day_start_col + i)
        ws[f'{col}{total_r}'].border = thin_border
        ws[f'{col}{total_r}'].fill = total_fill
    ws[f'{get_column_letter(summary_col)}{total_r}'] = t_days
    ws[f'{get_column_letter(summary_col)}{total_r}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'{get_column_letter(summary_col)}{total_r}'].alignment = center
    ws[f'{get_column_letter(summary_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(summary_col)}{total_r}'].fill = total_fill
    ws[f'{get_column_letter(remarks_col)}{total_r}'] = t_hours
    ws[f'{get_column_letter(remarks_col)}{total_r}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'{get_column_letter(remarks_col)}{total_r}'].alignment = center
    ws[f'{get_column_letter(remarks_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(remarks_col)}{total_r}'].fill = total_fill
    ws[f'{get_column_letter(sig_col)}{total_r}'].border = thin_border
    ws[f'{get_column_letter(sig_col)}{total_r}'].fill = total_fill

    # Print setup — Legal Size Landscape
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5  # Legal
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f'{header_row}:{header_row + 1}'
    ws.page_margins = PageMargins(left=0.5, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_attendance_excel(payroll, est, config, attendance, num_days, rest_days, holiday_days):
    """Generate Professional Attendance Register Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {payroll.month_name} {payroll.year}"

    # Styles
    title_font = Font(name='Arial', size=13, bold=True)
    subtitle_font = Font(name='Arial', size=10, bold=True)
    header_font = Font(name='Arial', size=8, bold=True)
    data_font = Font(name='Arial', size=8)
    bold_font = Font(name='Arial', size=8, bold=True)

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    total_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
    rest_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    holiday_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    summary_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')

    # Layout: A=Sr, B=Name, C=UAN, D=ESIC, E..=Days, then Pres, Abs, PH, Total
    day_start = 5
    sum_start = day_start + num_days
    last_col = get_column_letter(sum_start + 3)

    # Title
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'ATTENDANCE REGISTER'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = est.company_name.upper()
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = est.address or ''
    ws['A3'].font = Font(name='Arial', size=9)
    ws['A3'].alignment = center

    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = f'Month: {payroll.month_name} {payroll.year}  |  Working Days: {payroll.working_days}'
    ws['A4'].font = Font(name='Arial', size=10, bold=True, color='1E3A8A')
    ws['A4'].alignment = center

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    for d in range(num_days):
        ws.column_dimensions[get_column_letter(day_start + d)].width = 3.5
    for i in range(4):
        ws.column_dimensions[get_column_letter(sum_start + i)].width = 6

    # Header
    header_row = 6
    for col, text in [('A', 'Sr'), ('B', 'Employee Name'), ('C', 'UAN'), ('D', 'ESIC IP')]:
        cell = ws[f'{col}{header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    import datetime as dt
    for d in range(1, num_days + 1):
        col = get_column_letter(day_start + d - 1)
        cell = ws[f'{col}{header_row}']
        cell.value = d
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        if d in rest_days:
            cell.fill = rest_fill
        elif d in holiday_days:
            cell.fill = holiday_fill
        else:
            cell.fill = header_fill

    for i, text in enumerate(['Pres', 'Abs', 'PH', 'Total']):
        col = get_column_letter(sum_start + i)
        cell = ws[f'{col}{header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = summary_fill

    ws.row_dimensions[header_row].height = 22

    # Data rows
    data_start_row = header_row + 1
    t_pres = t_abs = t_ph = t_total = 0

    for idx, att in enumerate(attendance):
        r = data_start_row + idx
        emp = att['emp']
        ws.row_dimensions[r].height = 20

        ws[f'A{r}'] = idx + 1
        ws[f'A{r}'].font = data_font
        ws[f'A{r}'].alignment = center
        ws[f'A{r}'].border = thin_border

        ws[f'B{r}'] = emp.name
        ws[f'B{r}'].font = bold_font
        ws[f'B{r}'].alignment = left_align
        ws[f'B{r}'].border = thin_border

        ws[f'C{r}'] = emp.uan_number or ''
        ws[f'C{r}'].font = data_font
        ws[f'C{r}'].alignment = center
        ws[f'C{r}'].border = thin_border

        ws[f'D{r}'] = emp.esic_ip_number or 'NA'
        ws[f'D{r}'].font = data_font
        ws[f'D{r}'].alignment = center
        ws[f'D{r}'].border = thin_border

        for d in range(1, num_days + 1):
            col = get_column_letter(day_start + d - 1)
            mark = att['marks'].get(d, '')
            cell = ws[f'{col}{r}']
            cell.value = mark
            cell.font = data_font
            cell.alignment = center
            cell.border = thin_border
            if mark == 'R':
                cell.fill = rest_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='FF0000')
            elif mark == 'H':
                cell.fill = holiday_fill
                cell.font = Font(name='Arial', size=8, bold=True, color='1B5E20')
            elif mark == 'A':
                cell.font = Font(name='Arial', size=8, color='FF0000')

        pres = int(round(att['days_present']))
        abs_d = int(round(att['days_absent']))
        ph = int(round(att['paid_holidays']))
        total = int(round(att['total_days']))

        for i, val in enumerate([pres, abs_d, ph, total]):
            col = get_column_letter(sum_start + i)
            cell = ws[f'{col}{r}']
            cell.value = val
            cell.font = bold_font
            cell.alignment = center
            cell.border = thin_border

        t_pres += pres
        t_abs += abs_d
        t_ph += ph
        t_total += total

    # Totals row
    total_r = data_start_row + len(attendance)
    ws.merge_cells(f'A{total_r}:D{total_r}')
    ws[f'A{total_r}'] = f'TOTAL ({len(attendance)} Employees)'
    ws[f'A{total_r}'].font = bold_font
    ws[f'A{total_r}'].alignment = center
    ws[f'A{total_r}'].border = thin_border
    ws[f'A{total_r}'].fill = total_fill
    for c in ['B', 'C', 'D']:
        ws[f'{c}{total_r}'].border = thin_border
        ws[f'{c}{total_r}'].fill = total_fill

    for d in range(1, num_days + 1):
        col = get_column_letter(day_start + d - 1)
        ws[f'{col}{total_r}'].border = thin_border
        ws[f'{col}{total_r}'].fill = total_fill

    for i, val in enumerate([t_pres, t_abs, t_ph, t_total]):
        col = get_column_letter(sum_start + i)
        cell = ws[f'{col}{total_r}']
        cell.value = val
        cell.font = bold_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = total_fill

    # Legend
    legend_r = total_r + 2
    ws[f'A{legend_r}'] = 'Legend:'
    ws[f'A{legend_r}'].font = bold_font
    ws.merge_cells(f'B{legend_r}:J{legend_r}')
    ws[f'B{legend_r}'] = 'P = Present  |  A = Absent  |  R = Rest/Weekly Off  |  H = Holiday'
    ws[f'B{legend_r}'].font = Font(name='Arial', size=8)

    # Print setup
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f'{header_row}:{header_row}'
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _map_heads_to_form_b(heads):
    """Map dynamic salary heads to fixed Form B columns"""
    mapping = {
        'basic': None,          # Basic
        'special_basic': None,  # Special Basic
        'da': None,             # DA
        'hra': None,            # HRA
        'others': [],           # All other earning heads (combined)
    }

    for head in heads:
        code = head.short_code.upper()
        if code == 'BASIC':
            mapping['basic'] = head
        elif code in ('DA', 'DEARNESS ALLOWANCE'):
            mapping['da'] = head
        elif code == 'HRA':
            mapping['hra'] = head
        elif code in ('SPECIAL BASIC', 'SPL BASIC', 'SPLBASIC'):
            mapping['special_basic'] = head
        else:
            mapping['others'].append(head)

    return mapping


def _build_form_b_rows(entries, heads, head_map, config, payroll):
    """Build row data for Form B"""
    rows = []
    working_days = payroll.working_days or 0

    for entry in entries:
        emp = entry.employee

        # Get head-wise earned amounts
        basic_amt = 0
        spl_basic_amt = 0
        da_amt = 0
        hra_amt = 0
        others_amt = 0
        nph_amt = 0  # National Paid Holiday amount

        if head_map['basic'] and head_map['basic'].id in entry.head_amounts:
            basic_amt = entry.head_amounts[head_map['basic'].id].earned_amount

        if head_map['special_basic'] and head_map['special_basic'].id in entry.head_amounts:
            spl_basic_amt = entry.head_amounts[head_map['special_basic'].id].earned_amount

        if head_map['da'] and head_map['da'].id in entry.head_amounts:
            da_amt = entry.head_amounts[head_map['da'].id].earned_amount

        if head_map['hra'] and head_map['hra'].id in entry.head_amounts:
            hra_amt = entry.head_amounts[head_map['hra'].id].earned_amount

        for oh in head_map['others']:
            if oh.id in entry.head_amounts:
                others_amt += entry.head_amounts[oh.id].earned_amount

        # NPH amount: calculate from paid holidays and rate
        ph_count = entry.paid_holidays or 0
        if ph_count > 0:
            working_days = payroll.working_days or 26
            if hasattr(entry, '_salary_type') and entry._salary_type == 'daily_wages' and hasattr(entry, '_daily_rate') and entry._daily_rate:
                nph_amt = round(ph_count * entry._daily_rate)
            elif entry.gross_salary and working_days > 0:
                nph_amt = round(ph_count * (entry.gross_salary / working_days))
            else:
                nph_amt = 0
        else:
            nph_amt = 0

        row = {
            'sl': 0,  # Will be set in loop
            'emp_code': emp.internal_emp_code if emp.use_internal_code and emp.internal_emp_code else emp.emp_code,
            'uan': emp.uan_number or '',
            'esic_ip': emp.esic_ip_number or 'NA',
            'name': emp.name,
            'rate': (entry._daily_rate
                     if hasattr(entry, '_salary_type') and entry._salary_type == 'daily_wages'
                        and hasattr(entry, '_daily_rate') and entry._daily_rate
                     else (entry._effective_gross if hasattr(entry, '_effective_gross') and entry._effective_gross else entry.gross_salary)),
            'days_worked': entry.days_present,
            'ph_days': ph_count,
            'ot_days': entry.ot_hours,  # OT in days (as per Form B)
            'basic': basic_amt,
            'spl_basic': spl_basic_amt,
            'da': da_amt,
            'ot_amount': entry.ot_amount,
            'hra': hra_amt,
            'others': others_amt,
            'nph': nph_amt,
            'gross': entry.total_earnings,
            'pf': entry.epf_employee,
            'esic': entry.esic_employee,
            'society': 0,
            'income_tax': 0,
            'insurance': 0,
            'lwf': 0,
            'recoveries': entry.other_deduction,
            'total_ded': entry.total_deductions,
            'net_pay': entry.net_pay,
        }
        rows.append(row)

    return rows


def _generate_form_b_excel(payroll, est, config, entries, rows, head_map):
    """Generate Form B Excel file — professional Legal-size print format.
    Uses Emp. Code instead of UAN/ESIC IP. Reduced to 27 columns (A-AA) for better fit."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Form B - {payroll.month_name} {payroll.year}"

    # ---- Styles (11pt for readable Legal print) ----
    title_font = Font(name='Arial', size=14, bold=True)
    subtitle_font = Font(name='Arial', size=11, bold=True)
    header_font = Font(name='Arial', size=10, bold=True)
    data_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)
    small_font = Font(name='Arial', size=9)
    name_font = Font(name='Arial', size=11, bold=True)
    gross_font = Font(name='Arial', size=11, bold=True, color='006100')
    net_font = Font(name='Arial', size=11, bold=True, color='1A237E')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    total_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    gross_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    net_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ded_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # ---- Row 1-3: Title ----
    # Reduced to A-AA (27 columns) — removed separate ESIC IP column
    last_col = 'AA'
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'FORM B'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = 'FORMAT OF WAGE REGISTER'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = '[Part-A: For all Establishments]'
    ws['A3'].font = Font(name='Arial', size=10, bold=True)
    ws['A3'].alignment = center

    # ---- Row 5-7: Establishment Details ----
    ws.merge_cells('A5:D5')
    ws['A5'] = 'Name of the Establishment:'
    ws['A5'].font = bold_font
    ws.merge_cells('E5:N5')
    ws['E5'] = est.company_name.upper()
    ws['E5'].font = Font(name='Arial', size=11, bold=True)

    ws.merge_cells('A6:D6')
    ws['A6'] = 'Name of Owner:'
    ws['A6'].font = bold_font
    ws.merge_cells('E6:N6')
    ws['E6'] = est.address or ''
    ws['E6'].font = data_font

    ws.merge_cells('A7:D7')
    ws['A7'] = 'LIN/PAN:'
    ws['A7'].font = bold_font
    ws.merge_cells('E7:N7')
    ws['E7'] = est.pan_number or ''
    ws['E7'].font = data_font

    # ---- Row 5-7 Right Side ----
    ws.merge_cells(f'R5:{last_col}5')
    ws['R5'] = 'Rate of Minimum Wages and since the date…………'
    ws['R5'].font = Font(name='Arial', size=8, italic=True)

    # ---- Row 9: Wage Period ----
    ws.merge_cells('A9:D9')
    ws['A9'] = 'Wage period From'
    ws['A9'].font = bold_font

    _, last_day = calendar.monthrange(payroll.year, payroll.month)
    period_from = f'01-{payroll.month_name[:3]}-{payroll.year}'
    period_to = f'{last_day}-{payroll.month_name[:3]}-{payroll.year}'

    ws['E9'] = period_from
    ws['E9'].font = Font(name='Arial', size=10, bold=True)
    ws['F9'] = 'To'
    ws['F9'].font = bold_font
    ws['F9'].alignment = center
    ws['G9'] = period_to
    ws['G9'].font = Font(name='Arial', size=10, bold=True)

    ws.merge_cells('I9:N9')
    ws['I9'] = '(Monthly/Fortnight/Weekly/Daily/Piece Rated)'
    ws['I9'].font = small_font

    ws.merge_cells(f'R9:{last_col}9')
    ws['R9'] = f'Salary for the Month of: - {payroll.month_name}-{payroll.year}'
    ws['R9'].font = Font(name='Arial', size=10, bold=True, color='FF0000')

    # ---- Data Table ----
    # COLUMNS (A-AA = 27 columns):
    # A=Sr, B=Emp.Code, C=Name, D=Rate, E=Days Worked, F=PH Days, G=OT Days,
    # H=Basic, I=Spl Basic, J=DA, K=OT Amt, L=HRA, M=Others, N=NPH,
    # O=Gross, P=PF, Q=ESIC, R=Society, S=Income Tax, T=Insurance, U=LWF, V=Recoveries,
    # W=Total Ded, X=Net Pay, Y=Receipt/Bank, Z=Date of Payment, AA=Remarks

    header_row = 11
    sub_header_row = 12

    # Optimized column widths for Legal landscape with 11pt font
    col_widths = {
        'A': 5, 'B': 11, 'C': 22, 'D': 10, 'E': 8, 'F': 6, 'G': 7,
        'H': 10, 'I': 9, 'J': 9, 'K': 10, 'L': 8, 'M': 8, 'N': 9,
        'O': 13, 'P': 9, 'Q': 8, 'R': 7, 'S': 8, 'T': 8, 'U': 6, 'V': 10,
        'W': 11, 'X': 13, 'Y': 13, 'Z': 10, 'AA': 7
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Header Row 1 (merged headers) ---
    headers_r1 = [
        ('A', 'A', 'Sr.\nNo.'),
        ('B', 'B', 'Emp.\nCode'),
        ('C', 'C', 'Name'),
        ('D', 'D', 'Rate of\nWages'),
        ('E', 'E', 'Days\nWorked'),
        ('F', 'F', 'Paid\nHol.'),
        ('G', 'G', 'OT\nDays'),
        ('H', 'H', 'Basic'),
        ('I', 'I', 'Spl.\nBasic'),
        ('J', 'J', 'DA'),
        ('K', 'K', 'OT\nAmount'),
        ('L', 'L', 'HRA'),
        ('M', 'M', 'Others'),
        ('N', 'N', 'NPH\nAmt.'),
        ('O', 'O', 'Gross\nWages'),
    ]

    for start_col, end_col, text in headers_r1:
        ws.merge_cells(f'{start_col}{header_row}:{end_col}{sub_header_row}')
        cell = ws[f'{start_col}{header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    # Gross Wages header — green tint
    ws[f'O{header_row}'].fill = gross_fill

    # Deductions header (merged P-V)
    ws.merge_cells(f'P{header_row}:V{header_row}')
    ws[f'P{header_row}'] = 'Deductions'
    ws[f'P{header_row}'].font = header_font
    ws[f'P{header_row}'].alignment = center
    ws[f'P{header_row}'].border = thin_border
    ws[f'P{header_row}'].fill = ded_fill

    # Deduction sub-headers (row 12)
    ded_headers = [
        ('P', 'PF'), ('Q', 'ESIC'), ('R', 'Society'),
        ('S', 'Income\nTax'), ('T', 'Insur.'), ('U', 'LWF'), ('V', 'Recov.')
    ]
    for col, text in ded_headers:
        cell = ws[f'{col}{sub_header_row}']
        cell.value = text
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = ded_fill

    # Total Deductions
    ws.merge_cells(f'W{header_row}:W{sub_header_row}')
    ws[f'W{header_row}'] = 'Total\nDed.'
    ws[f'W{header_row}'].font = header_font
    ws[f'W{header_row}'].alignment = center
    ws[f'W{header_row}'].border = thin_border
    ws[f'W{header_row}'].fill = ded_fill

    # Net Payment
    ws.merge_cells(f'X{header_row}:X{sub_header_row}')
    ws[f'X{header_row}'] = 'Net\nPayment'
    ws[f'X{header_row}'].font = header_font
    ws[f'X{header_row}'].alignment = center
    ws[f'X{header_row}'].border = thin_border
    ws[f'X{header_row}'].fill = net_fill

    # Receipt / Bank
    ws.merge_cells(f'Y{header_row}:Y{sub_header_row}')
    ws[f'Y{header_row}'] = 'Receipt /\nBank\nTrans. ID'
    ws[f'Y{header_row}'].font = Font(name='Arial', size=8, bold=True)
    ws[f'Y{header_row}'].alignment = center
    ws[f'Y{header_row}'].border = thin_border
    ws[f'Y{header_row}'].fill = header_fill

    # Date of Payment
    ws.merge_cells(f'Z{header_row}:Z{sub_header_row}')
    ws[f'Z{header_row}'] = 'Date of\nPayment'
    ws[f'Z{header_row}'].font = header_font
    ws[f'Z{header_row}'].alignment = center
    ws[f'Z{header_row}'].border = thin_border
    ws[f'Z{header_row}'].fill = header_fill

    # Remarks
    ws.merge_cells(f'AA{header_row}:AA{sub_header_row}')
    ws[f'AA{header_row}'] = 'Rmk.'
    ws[f'AA{header_row}'].font = header_font
    ws[f'AA{header_row}'].alignment = center
    ws[f'AA{header_row}'].border = thin_border
    ws[f'AA{header_row}'].fill = header_fill

    # Header row heights
    ws.row_dimensions[header_row].height = 36
    ws.row_dimensions[sub_header_row].height = 28

    # ---- Data Rows ----
    data_start_row = 13
    totals = {
        'days_worked': 0, 'ph_days': 0, 'ot_days': 0,
        'basic': 0, 'spl_basic': 0, 'da': 0, 'ot_amount': 0,
        'hra': 0, 'others': 0, 'nph': 0, 'gross': 0,
        'pf': 0, 'esic': 0, 'society': 0, 'income_tax': 0,
        'insurance': 0, 'lwf': 0, 'recoveries': 0,
        'total_ded': 0, 'net_pay': 0
    }

    for idx, row_data in enumerate(rows):
        r = data_start_row + idx
        sl = idx + 1

        ws.row_dimensions[r].height = 26

        # Build cells: A-AA (27 columns)
        data_cells = [
            ('A', sl, center, data_font, None),
            ('B', row_data.get('emp_code', ''), center, data_font, None),
            ('C', row_data['name'], left_align, name_font, None),
            ('D', float(row_data['rate']) if row_data['rate'] else 0, right_align, data_font, None),
            ('E', row_data['days_worked'], center, data_font, None),
            ('F', round(row_data.get('ph_days', 0)), center, data_font, None),
            ('G', round(row_data['ot_days']) if row_data['ot_days'] else 0, center, data_font, None),
            ('H', round(row_data['basic']), right_align, data_font, None),
            ('I', round(row_data['spl_basic']), right_align, data_font, None),
            ('J', round(row_data['da']), right_align, data_font, None),
            ('K', round(row_data['ot_amount']), right_align, data_font, None),
            ('L', round(row_data['hra']), right_align, data_font, None),
            ('M', round(row_data['others']), right_align, data_font, None),
            ('N', round(row_data['nph']), right_align, data_font, None),
            ('O', round(row_data['gross']), right_align, gross_font, gross_fill),
            ('P', round(row_data['pf']), right_align, data_font, None),
            ('Q', round(row_data['esic']), right_align, data_font, None),
            ('R', round(row_data['society']), right_align, data_font, None),
            ('S', round(row_data['income_tax']), right_align, data_font, None),
            ('T', round(row_data['insurance']), right_align, data_font, None),
            ('U', round(row_data['lwf']), right_align, data_font, None),
            ('V', round(row_data['recoveries']), right_align, data_font, None),
            ('W', round(row_data['total_ded']), right_align, bold_font, None),
            ('X', round(row_data['net_pay']), right_align, net_font, net_fill),
            ('Y', '', center, data_font, None),
            ('Z', '', center, data_font, None),
            ('AA', '', center, data_font, None),
        ]

        for col, val, align, font, fill in data_cells:
            cell = ws[f'{col}{r}']
            cell.value = val
            cell.font = font
            cell.alignment = align
            cell.border = thin_border
            if fill:
                cell.fill = fill
            # Number format — rate column (D) keeps decimals, other amounts rounded
            if col == 'D':
                cell.number_format = '#,##0.00'
            elif col in ('H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X'):
                cell.number_format = '#,##0'

        # Accumulate totals
        for key in totals:
            totals[key] += row_data.get(key, 0)

    # ---- Totals Row ----
    total_row = data_start_row + len(rows)
    ws.row_dimensions[total_row].height = 28

    ws.merge_cells(f'A{total_row}:C{total_row}')

    total_cells = [
        ('A', '', center),
        ('D', '', center),
        ('E', round(totals['days_worked']), center),
        ('F', round(totals['ph_days']), center),
        ('G', round(totals['ot_days']), center),
        ('H', round(totals['basic']), right_align),
        ('I', round(totals['spl_basic']), right_align),
        ('J', round(totals['da']), right_align),
        ('K', round(totals['ot_amount']), right_align),
        ('L', round(totals['hra']), right_align),
        ('M', round(totals['others']), right_align),
        ('N', round(totals['nph']), right_align),
        ('O', round(totals['gross']), right_align),
        ('P', round(totals['pf']), right_align),
        ('Q', round(totals['esic']), right_align),
        ('R', round(totals['society']), right_align),
        ('S', round(totals['income_tax']), right_align),
        ('T', round(totals['insurance']), right_align),
        ('U', round(totals['lwf']), right_align),
        ('V', round(totals['recoveries']), right_align),
        ('W', round(totals['total_ded']), right_align),
        ('X', round(totals['net_pay']), right_align),
        ('Y', '', center),
        ('Z', '', center),
        ('AA', '', center),
    ]

    ws[f'A{total_row}'] = f'TOTAL ({len(rows)} Employees)'
    ws[f'A{total_row}'].font = bold_font
    ws[f'A{total_row}'].alignment = center
    ws[f'A{total_row}'].border = thin_border
    ws[f'A{total_row}'].fill = total_fill

    for col, val, align in total_cells:
        if col == 'A':
            continue
        cell = ws[f'{col}{total_row}']
        cell.value = val
        cell.font = bold_font
        cell.alignment = align
        cell.border = thin_border
        cell.fill = total_fill

    # Apply border and fill to merged cells B-C in total row
    for col in ['B', 'C']:
        cell = ws[f'{col}{total_row}']
        cell.border = thin_border
        cell.fill = total_fill
    # Number format for total amount cells
    for col in ('H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X'):
        ws[f'{col}{total_row}'].number_format = '#,##0'

    # ---- Print Setup — Legal Size Landscape ----
    from openpyxl.worksheet.page import PageMargins, PrintPageSetup
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 5  # 5 = Legal (8.5 x 14 inches)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f'{header_row}:{sub_header_row}'

    # Margins: enough left margin to avoid printer cut-off
    ws.page_margins = PageMargins(left=0.5, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================
# Salary Statement Format 2 — Modern Professional
# =============================================

@reports_bp.route('/payroll/<int:payroll_id>/report/statement-format2')
def statement_format2(payroll_id):
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/statement_format2.html',
                           payroll=payroll, est=est, config=config,
                           entries=entries, heads=heads,
                           generated_on=generated_on)


# =============================================
# Salary Statement Format 3 — Compact Executive
# =============================================

@reports_bp.route('/payroll/<int:payroll_id>/report/statement-format3')
def statement_format3(payroll_id):
    payroll, est, config, entries, heads = _get_payroll_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/statement_format3.html',
                           payroll=payroll, est=est, config=config,
                           entries=entries, heads=heads,
                           generated_on=generated_on)


# =============================================
# REPORT 3: EPF ECR — Text & CSV for EPFO Portal
# =============================================
# ECR (Electronic Challan cum Return) format:
# UAN | Member Name | Gross Wages | EPF Wages | EPS Wages | EDLI Wages |
# EPF Contribution(EE) | EPS Contribution | EPF Diff (ER A/c01) | NCP Days | Refund

def _validate_uan(uan):
    """Validate UAN format: must be exactly 12 digits."""
    if not uan:
        return ['UAN is missing']
    uan = str(uan).strip()
    errors = []
    # Remove scientific notation (e.g., 1.02217E+11)
    if 'E' in uan.upper() or 'e' in uan:
        errors.append('UAN in scientific notation — re-import from EPFO')
    elif not uan.isdigit():
        errors.append(f'UAN "{uan}" is not numeric')
    elif len(uan) != 12:
        errors.append(f'UAN "{uan}" is {len(uan)} digits (must be 12)')
    return errors


def _validate_esic_ip(ip):
    """Validate ESIC IP Number format: must be exactly 10 digits."""
    if not ip:
        return ['ESIC IP is missing']
    ip = str(ip).strip()
    errors = []
    if not ip.isdigit():
        errors.append(f'ESIC IP "{ip}" is not numeric')
    elif len(ip) != 10:
        errors.append(f'ESIC IP "{ip}" is {len(ip)} digits (must be 10)')
    return errors


def _build_ecr_data(payroll_id):
    """Build ECR row data for all employees in a payroll.
    Excludes employees with zero attendance (no work = no EPF contribution).
    Adds validation flags for UAN format, name mismatches, duplicate UAN,
    and Age 58+ EPS check.
    """
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .join(Employee).order_by(Employee.name).all()

    # ---- Pass 1: Detect duplicate UANs across ALL employees in this payroll ----
    uan_name_map = {}   # { uan_number: [emp_name, emp_name, ...] }
    for entry in entries:
        emp = entry.employee
        if emp.uan_number and entry.days_present > 0:
            uan_clean = emp.uan_number.strip()
            if uan_clean not in uan_name_map:
                uan_name_map[uan_clean] = []
            uan_name_map[uan_clean].append(emp.name)
    # Only keep UANs that appear more than once (duplicates)
    duplicate_uans = {uan: names for uan, names in uan_name_map.items() if len(names) > 1}

    # ---- Payroll period end date for age calculation ----
    # Age as on last day of the payroll month
    last_day = calendar.monthrange(payroll.year, payroll.month)[1]
    period_end = datetime(payroll.year, payroll.month, last_day).date()

    rows = []
    skipped = []
    for entry in entries:
        emp = entry.employee
        # Skip employees without UAN (EPF requires UAN)
        if not emp.uan_number:
            skipped.append({'name': emp.name, 'reason': 'No UAN assigned'})
            continue
        # Skip employees with zero attendance — no EPF contribution for zero work
        if entry.days_present <= 0:
            continue

        gross_wages = int(round(entry.earned_gross))
        epf_wages = int(round(entry.epf_wages)) if entry.epf_wages else gross_wages
        # EPS wages = same as EPF wages (capped at 15000 by system)
        eps_wages = epf_wages
        edli_wages = epf_wages
        epf_ee = int(round(entry.epf_employee))          # Employee 12%
        eps_contribution = int(round(entry.epf_eps))      # EPS 8.33%
        # A/c 01 = 12% − 8.33% (ensures perfect balance, no rounding mismatch)
        epf_er_diff = epf_ee - eps_contribution            # 3.67% derived
        ncp_days = int(round(entry.days_absent))          # Non-Contributing Period days
        refund = 0

        # Validation flags
        warnings = []
        warnings.extend(_validate_uan(emp.uan_number))
        # Name mismatch with EPFO records
        if emp.epfo_name and emp.has_name_mismatch and not emp.name_mismatch_accepted:
            warnings.append(f'Name mismatch: ERP "{emp.name}" vs EPFO "{emp.epfo_name}"')
        # Zero EPF contribution check
        if epf_ee == 0 and gross_wages > 0:
            warnings.append('Zero EPF contribution on non-zero wages')

        # --- NEW: Duplicate UAN check ---
        uan_clean = emp.uan_number.strip()
        if uan_clean in duplicate_uans:
            other_names = [n for n in duplicate_uans[uan_clean] if n != emp.name]
            if other_names:
                warnings.append(f'Duplicate UAN — shared with {", ".join(other_names)}')

        # --- NEW: Age 58+ EPS check ---
        if emp.date_of_birth:
            age_years = (period_end - emp.date_of_birth).days / 365.25
            if age_years >= 58:
                age_display = int(age_years)
                warnings.append(f'Age {age_display}+ — EPS contribution should be ₹0 (pension stops at 58)')
                if eps_contribution > 0:
                    warnings.append(f'EPS ₹{eps_contribution:,} charged but employee is above 58 — verify with EPFO')

        rows.append({
            'uan': uan_clean,
            'name': _clean_name_for_statutory(emp.name),
            'gross_wages': gross_wages,
            'epf_wages': epf_wages,
            'eps_wages': eps_wages,
            'edli_wages': edli_wages,
            'epf_ee': epf_ee,
            'eps_contribution': eps_contribution,
            'epf_er_diff': epf_er_diff,
            'ncp_days': ncp_days,
            'refund': refund,
            'warnings': warnings,
        })

    return payroll, est, config, entries, rows, skipped


@reports_bp.route('/payroll/<int:payroll_id>/report/epf-ecr-view')
def epf_ecr_view(payroll_id):
    """HTML preview of EPF ECR data with validation flags"""
    payroll, est, config, entries, rows, skipped = _build_ecr_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/epf_ecr.html',
                           payroll=payroll, est=est, config=config,
                           rows=rows, entries=entries, skipped=skipped,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/epf-ecr-text')
def epf_ecr_text(payroll_id):
    """Download EPF ECR as pipe-delimited text file (for EPFO portal upload)"""
    payroll, est, config, entries, rows, skipped = _build_ecr_data(payroll_id)

    lines = []
    for r in rows:
        line = '#~#'.join([
            str(r['uan']),
            r['name'],
            str(r['gross_wages']),
            str(r['epf_wages']),
            str(r['eps_wages']),
            str(r['edli_wages']),
            str(r['epf_ee']),
            str(r['eps_contribution']),
            str(r['epf_er_diff']),
            str(r['ncp_days']),
            str(r['refund']),
        ])
        lines.append(line)

    content = '\n'.join(lines)
    output = io.BytesIO(content.encode('utf-8'))
    output.seek(0)

    # Short filename for EPF portal (under 20 chars): ECR_PFCODE_MMYY.txt
    pf_code = (est.pf_code or 'NOPF').replace('/', '').replace(' ', '')
    filename = f"ECR_{pf_code}_{payroll.month:02d}{str(payroll.year)[-2:]}.txt"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='text/plain')


@reports_bp.route('/payroll/<int:payroll_id>/report/epf-ecr-csv')
def epf_ecr_csv(payroll_id):
    """Download EPF ECR as CSV file"""
    payroll, est, config, entries, rows, skipped = _build_ecr_data(payroll_id)

    import csv
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        'UAN', 'Member Name', 'Gross Wages', 'EPF Wages', 'EPS Wages',
        'EDLI Wages', 'EPF Contribution (EE)', 'EPS Contribution',
        'EPF Contribution Diff (ER)', 'NCP Days', 'Refund of Advances'
    ])

    for r in rows:
        writer.writerow([
            r['uan'], r['name'], r['gross_wages'], r['epf_wages'],
            r['eps_wages'], r['edli_wages'], r['epf_ee'],
            r['eps_contribution'], r['epf_er_diff'], r['ncp_days'],
            r['refund']
        ])

    byte_output = io.BytesIO(output.getvalue().encode('utf-8'))
    byte_output.seek(0)

    # Short filename for EPF portal (under 20 chars): ECR_PFCODE_MMYY.csv
    pf_code = (est.pf_code or 'NOPF').replace('/', '').replace(' ', '')
    filename = f"ECR_{pf_code}_{payroll.month:02d}{str(payroll.year)[-2:]}.csv"
    return send_file(byte_output, as_attachment=True, download_name=filename,
                     mimetype='text/csv')


# =============================================
# REPORT 4: ESIC Monthly Contribution — Excel for ESIC Portal
# =============================================
# Standard ESIC portal upload format:
# IP Number | IP Name | No of Days | Total Monthly Wages |
# Reason Code (0=worked) | Last Working Day (blank if working)

@reports_bp.route('/payroll/<int:payroll_id>/report/esic-view')
def esic_view(payroll_id):
    """HTML preview of ESIC contribution data.
    NOTE: ESIC includes ALL employees — even zero attendance (Reason Code 11 = No Work).
    """
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
    # Include ALL employees (including zero attendance) — ESIC requires code 11 for zero work
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .join(Employee).order_by(Employee.name).all()

    rows, skipped = _build_esic_rows(entries, payroll)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/esic_template.html',
                           payroll=payroll, est=est, config=config,
                           rows=rows, entries=entries, skipped=skipped,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/esic-excel')
def esic_excel(payroll_id):
    """Download ESIC MC Template as .xls (Excel 97-2003) for ESIC portal upload.
    NOTE: ESIC includes ALL employees — even zero attendance (Reason Code 11 = No Work).
    """
    try:
        payroll = MonthlyPayroll.query.get_or_404(payroll_id)
        est = payroll.establishment
        verify_est_ownership(est)
        config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
        # Include ALL employees (including zero attendance)
        entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
            .join(Employee).order_by(Employee.name).all()

        rows, skipped = _build_esic_rows(entries, payroll, round_up=True)

        if not rows:
            flash('No employees with ESIC IP Number found. Cannot generate template.', 'warning')
            return redirect(request.referrer or url_for('reports.esic_view', payroll_id=payroll_id))

        output = _generate_esic_xls(payroll, est, config, rows)

        # Clean filename — remove special characters that may cause issues
        safe_name = ''.join(c if c.isalnum() or c in ('_', '-') else '_' for c in est.company_name)
        filename = f"MC_Template_{safe_name}_{payroll.month_name}_{payroll.year}.xls"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.ms-excel')
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error generating ESIC template: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('payroll.payroll_list'))


def _build_esic_rows(entries, payroll, round_up=False):
    """Build ESIC row data matching exact ESIC portal format.
    Returns (rows, skipped) with validation warnings per row.
    Includes: IP format check, duplicate IP detection, wage ceiling check.
    round_up: If True, use math.ceil for attendance days (ESIC portal requirement).
    Reason Codes (as per ESIC portal):
        0  = Without Reason (default for working employees)
        1  = On Leave
        2  = Left Service
        3  = Retired
        4  = Out of Coverage
        5  = Expired (Died in Service)
        6  = Non Implemented area
        7  = Compliance by Immediate Employer
        8  = Suspension of work
        9  = Strike/Lockout
        10 = Retrenchment
        11 = No Work
        12 = Doesnt Belong To This Employer
        13 = Duplicate IP
    """
    # ---- Pass 1: Detect duplicate ESIC IP Numbers ----
    ip_name_map = {}   # { ip_number: [emp_name, emp_name, ...] }
    for entry in entries:
        emp = entry.employee
        if emp.esic_ip_number:
            ip_clean = emp.esic_ip_number.strip()
            if ip_clean not in ip_name_map:
                ip_name_map[ip_clean] = []
            ip_name_map[ip_clean].append(emp.name)
    # Only keep IPs that appear more than once (duplicates)
    duplicate_ips = {ip: names for ip, names in ip_name_map.items() if len(names) > 1}

    # ESIC wage ceiling — employees earning above this are not covered
    ESIC_WAGE_CEILING = 21000

    rows = []
    skipped = []
    for entry in entries:
        emp = entry.employee
        if not emp.esic_ip_number:
            skipped.append({'name': emp.name, 'reason': 'No ESIC IP assigned'})
            continue

        total_days = math.ceil(entry.total_payable_days or 0) if round_up else int(round(entry.total_payable_days or 0))
        total_wages = int(round(entry.earned_gross or 0))

        reason_code = ''
        last_working_day = ''

        # If employee has exited during this month with zero days
        if emp.date_of_exit and emp.date_of_exit.year == payroll.year and emp.date_of_exit.month == payroll.month:
            # Only provide last working day if 0 days wages paid
            if total_days == 0:
                last_working_day = emp.date_of_exit.strftime('%d/%m/%Y')
            if emp.exit_reason == 'Resigned':
                reason_code = '2'   # Left Service
            elif emp.exit_reason == 'Terminated':
                reason_code = '2'   # Left Service
            elif emp.exit_reason == 'Retired':
                reason_code = '3'   # Retired
            elif emp.exit_reason == 'Deceased':
                reason_code = '5'   # Expired
            elif emp.exit_reason == 'Absconded':
                reason_code = '2'   # Left Service
            else:
                reason_code = '2'
        elif total_days == 0 and total_wages == 0:
            # Zero wages but not exited — could be No Work or On Leave
            reason_code = '11'  # No Work

        # Wages above ESIC ceiling — ESIC contribution = 0
        # When payroll already calculated ESIC as 0 (ceiling type), report wages as 0
        # and reason code as 11 (No Work / Not contributing)
        esic_ee = int(round(entry.esic_employee or 0))
        esic_er = int(round(entry.esic_employer or 0))
        if total_wages > ESIC_WAGE_CEILING and esic_ee == 0 and esic_er == 0:
            total_wages = 0
            total_days = 0
            reason_code = '11'  # No Work — wages crossed ceiling, not covered

        # Validation flags
        warnings = []
        warnings.extend(_validate_esic_ip(emp.esic_ip_number))
        # Wages sanity checks
        if total_wages > 0 and total_days == 0:
            warnings.append('Wages reported but zero days worked')
        if total_days > 0 and total_wages == 0:
            warnings.append('Days worked but zero wages')

        # --- Duplicate ESIC IP check ---
        ip_clean = emp.esic_ip_number.strip()
        if ip_clean in duplicate_ips:
            other_names = [n for n in duplicate_ips[ip_clean] if n != emp.name]
            if other_names:
                warnings.append(f'Duplicate ESIC IP — shared with {", ".join(other_names)}')

        rows.append({
            'ip_number': ip_clean,
            'ip_name': _clean_name_for_statutory(emp.name),
            'no_of_days': str(total_days),
            'total_wages': str(total_wages),
            'reason_code': str(reason_code) if reason_code else '',
            'last_working_day': last_working_day,
            'warnings': warnings,
        })

    return rows, skipped


def _generate_esic_xls(payroll, est, config, rows):
    """Generate ESIC MC Template in .xls (Excel 97-2003) format.
    Exact match to ESIC portal template:
    - Sheet 1: Header row + data (all TEXT columns)
    - Sheet 2: Instructions & Reason Codes
    """
    import xlwt

    wb = xlwt.Workbook(encoding='utf-8')

    # ---- Sheet 1: Data ----
    ws = wb.add_sheet('Sheet1')

    # Column widths (in 256ths of character width)
    ws.col(0).width = 5000   # IP Number
    ws.col(1).width = 10000  # IP Name
    ws.col(2).width = 5000   # No of Days
    ws.col(3).width = 6000   # Total Monthly Wages
    ws.col(4).width = 8000   # Reason Code
    ws.col(5).width = 7000   # Last Working Day

    # Header style
    hdr_style = xlwt.easyxf(
        'font: bold on; align: wrap on, vert centre, horiz centre;'
    )

    # Exact ESIC portal header texts
    headers = [
        'IP Number \n(10 Digits)',
        'IP Name\n( Only alphabets and space )',
        'No of Days for which wages paid/payable during the month',
        'Total Monthly Wages',
        'Reason Code for Zero workings days(numeric only; provide 0 for all other reasons- Click on the link for reference)',
        'Last Working Day\n( Format DD/MM/YYYY  or DD-MM-YYYY)',
    ]

    for ci, hdr in enumerate(headers):
        ws.write(0, ci, hdr, hdr_style)

    # Data rows — all as TEXT (strings)
    for ri, r in enumerate(rows, 1):
        ws.write(ri, 0, r['ip_number'])
        ws.write(ri, 1, r['ip_name'])
        ws.write(ri, 2, r['no_of_days'])
        ws.write(ri, 3, r['total_wages'])
        ws.write(ri, 4, r['reason_code'])
        ws.write(ri, 5, r['last_working_day'])

    # ---- Sheet 2: Instructions & Reason Codes ----
    ws2 = wb.add_sheet('Instructions & Reason Codes')

    ws2.col(0).width = 15000
    ws2.col(1).width = 3000
    ws2.col(2).width = 25000

    bold = xlwt.easyxf('font: bold on;')

    # Reason codes table
    ws2.write(1, 0, 'Reason', bold)
    ws2.write(1, 1, 'Code', bold)
    ws2.write(1, 2, 'Note', bold)

    reason_codes = [
        ('Without Reason', '0', 'Leave last working day as blank'),
        ('On Leave', '1', 'Leave last working day as blank'),
        ('Left Service', '2', 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'),
        ('Retired', '3', 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'),
        ('Out of Coverage', '4', 'Please provide last working day (dd/mm/yyyy). IP will not appear from next contribution period. This option is valid only if Wage Period is April/October.'),
        ('Expired', '5', 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'),
        ('Non Implemented area', '6', 'Please provide last working day (dd/mm/yyyy).'),
        ('Compliance by Immediate Employer', '7', 'Leave last working day as blank'),
        ('Suspension of work', '8', 'Leave last working day as blank'),
        ('Strike/Lockout', '9', 'Leave last working day as blank'),
        ('Retrenchment', '10', 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'),
        ('No Work', '11', 'Leave last working day as blank'),
        ('Doesnt Belong To This Employer', '12', 'Leave last working day as blank'),
        ('Duplicate IP', '13', 'Leave last working day as blank'),
    ]

    for ri, (reason, code, note) in enumerate(reason_codes, 2):
        ws2.write(ri, 0, reason)
        ws2.write(ri, 1, code)
        ws2.write(ri, 2, note)

    # Instructions section
    instr_row = len(reason_codes) + 5
    ws2.write(instr_row, 0, 'Instructions to fill in the excel file:', bold)
    instructions = [
        '1. Enter IP number, IP name, No. of Days, Total Monthly Wages, Reason for 0 wages (If Wages=0) & Last Working Day.',
        '2. Number of days must be a whole number. Fractions should be rounded up to next higher whole number.',
        '3. All Employees currently mapped in the system must be entered in the excel sheet.',
        '4. Reasons are to be assigned numeric code and date has to be provided as mentioned in the table above.',
        '5. Once 0 wages given and last working day is mentioned (reason codes 2,3,4,5,10) IP will be removed from employer record.',
        '6. If IP has worked for part of the month (at least 1 day), last working day should NOT be mentioned.',
        '7. IP Contribution and Employer contribution calculation will be automatically done by the system.',
        '8. Date column format is dd/mm/yyyy or dd-mm-yyyy. Pad single digit dates with 0. Eg: 02/05/2010',
        '9. Excel file should be saved in .xls format (Excel 97-2003).',
        '10. All columns including date column should be in Text format.',
    ]
    for i, instr in enumerate(instructions):
        ws2.write(instr_row + 1 + i, 0, instr)

    # Save to BytesIO — ESIC XLS
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================
# REPORT 5: PAY SLIPS (Form XIX Government + Professional)
# =============================================

def _build_payslip_data(payroll_id):
    """Build payslip data for all employees"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .join(Employee).order_by(Employee.name).all()

    # Get earning heads
    heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True, head_type='earning', is_in_gross=True
    ).order_by(SalaryHead.display_order).all()

    # Get deduction heads
    ded_heads = SalaryHead.query.filter_by(
        establishment_id=est.id, is_active=True, head_type='deduction'
    ).order_by(SalaryHead.display_order).all()

    slips = []
    for entry in entries:
        emp = entry.employee

        # Build head-wise earnings
        earnings = []
        for head in heads:
            peh = PayrollEntryHead.query.filter_by(
                payroll_entry_id=entry.id, salary_head_id=head.id
            ).first()
            if peh and peh.earned_amount > 0:
                earnings.append({
                    'name': head.name,
                    'code': head.short_code,
                    'amount': peh.earned_amount
                })

        # Build deductions list
        deductions = []
        if entry.epf_employee > 0:
            deductions.append({'name': 'EPF Employee', 'amount': entry.epf_employee})
        if entry.esic_employee > 0:
            deductions.append({'name': 'ESIC Employee', 'amount': entry.esic_employee})
        if entry.professional_tax > 0:
            deductions.append({'name': 'Professional Tax', 'amount': entry.professional_tax})

        # Other deduction heads from salary heads
        for dh in ded_heads:
            peh = PayrollEntryHead.query.filter_by(
                payroll_entry_id=entry.id, salary_head_id=dh.id
            ).first()
            if peh and peh.earned_amount > 0:
                deductions.append({'name': dh.name, 'amount': peh.earned_amount})

        if entry.other_deduction > 0:
            deductions.append({
                'name': entry.other_deduction_remark or 'Other Deduction',
                'amount': entry.other_deduction
            })

        # Daily rate calculation
        working_days = payroll.working_days or 26
        daily_rate = round(entry.gross_salary / working_days) if working_days > 0 else 0

        slips.append({
            'emp': emp,
            'entry': entry,
            'earnings': earnings,
            'deductions': deductions,
            'daily_rate': daily_rate,
        })

    return payroll, est, config, slips, heads


@reports_bp.route('/payroll/<int:payroll_id>/report/payslip-form-xix')
def payslip_form_xix(payroll_id):
    """Form XIX Pay Slip — Government format, 2 slips per page"""
    payroll, est, config, slips, heads = _build_payslip_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/payslip_form_xix.html',
                           payroll=payroll, est=est, config=config,
                           slips=slips, generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/payslip-professional')
def payslip_professional(payroll_id):
    """Professional Pay Slip — Modern format, 2 slips per page"""
    payroll, est, config, slips, heads = _build_payslip_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/payslip_professional.html',
                           payroll=payroll, est=est, config=config,
                           slips=slips, heads=heads,
                           generated_on=generated_on)


@reports_bp.route('/payroll/<int:payroll_id>/report/payslip-elegant')
def payslip_elegant(payroll_id):
    """Elegant Pay Slip — Premium format with amount in words, 1 slip per page"""
    payroll, est, config, slips, heads = _build_payslip_data(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/payslip_elegant.html',
                           payroll=payroll, est=est, config=config,
                           slips=slips, heads=heads,
                           generated_on=generated_on)


# =============================================
# REPORT 6: EPF/ESIC Employer Share Reimbursement Letter
# =============================================

def _build_reimbursement_data(payroll_id):
    """Build reimbursement letter data from payroll"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .join(Employee).order_by(Employee.name).all()

    # Count EPF and ESIC employees separately (they may differ)
    epf_count = 0
    esic_count = 0
    total_epf_wages = 0
    total_esic_wages = 0
    total_epf_ac01 = 0      # 3.67%
    total_epf_edli = 0      # 0.5%
    total_epf_eps = 0       # 8.33%
    total_epf_admin = 0     # Admin A/c 21
    total_esic_employer = 0 # 3.25%

    for entry in entries:
        emp = entry.employee
        # EPF applicable employees
        if entry.epf_employee > 0 or entry.epf_employer > 0:
            epf_count += 1
            total_epf_wages += entry.epf_wages
            total_epf_ac01 += entry.epf_ac01
            total_epf_edli += entry.epf_edli
            total_epf_eps += entry.epf_eps
            total_epf_admin += entry.epf_admin

        # ESIC applicable employees
        if entry.esic_employee > 0 or entry.esic_employer > 0:
            esic_count += 1
            total_esic_wages += entry.esic_wages
            total_esic_employer += entry.esic_employer

    # EPF (3.67% + 0.5%) = A/c 01 + EDLI
    epf_367_05 = total_epf_ac01 + total_epf_edli

    # EPF Employer Refund = A/c 01 + EDLI + EPS + Admin
    epf_employer_refund = epf_367_05 + total_epf_eps + total_epf_admin

    # ESIC Employer Refund = ESIC 3.25%
    esic_employer_refund = total_esic_employer

    # Total Refund
    total_refund = epf_employer_refund + esic_employer_refund

    # For the gross wages for ESIC — use earned_gross of ESIC-applicable employees
    total_esic_gross = 0
    total_epf_gross = 0
    for entry in entries:
        if entry.epf_employee > 0 or entry.epf_employer > 0:
            total_epf_gross += entry.earned_gross
        if entry.esic_employee > 0 or entry.esic_employer > 0:
            total_esic_gross += entry.earned_gross

    data = {
        'epf_count': epf_count,
        'esic_count': esic_count,
        'total_epf_wages': total_epf_wages,
        'total_esic_wages': total_esic_wages,
        'total_epf_gross': total_epf_gross,
        'total_esic_gross': total_esic_gross,
        'total_epf_ac01': total_epf_ac01,
        'total_epf_edli': total_epf_edli,
        'epf_367_05': epf_367_05,
        'total_epf_eps': total_epf_eps,
        'total_epf_admin': total_epf_admin,
        'epf_employer_refund': epf_employer_refund,
        'total_esic_employer': total_esic_employer,
        'esic_employer_refund': esic_employer_refund,
        'total_refund': total_refund,
    }

    return payroll, est, config, data


@reports_bp.route('/payroll/<int:payroll_id>/report/reimbursement')
def reimbursement_view(payroll_id):
    """EPF/ESIC Employer Share Reimbursement Letter — HTML view.
    Available only when payroll is FINALIZED."""
    payroll, est, config, data = _build_reimbursement_data(payroll_id)

    # Only allow viewing/downloading when payroll is finalized
    if payroll.status != 'finalized':
        flash('Reimbursement letter can be generated only after payroll is finalized. '
              f'Current status: {payroll.status.title()}. Please finalize the payroll first.', 'warning')
        return redirect(url_for('payroll.payroll_list'))

    letter_date = datetime.now().strftime('%d-%b-%Y')
    return render_template('reports/reimbursement.html',
                           payroll=payroll, est=est, config=config,
                           data=data, letter_date=letter_date)


@reports_bp.route('/reports/reimbursement-multi')
def reimbursement_multi():
    """Multi-month combined reimbursement letter"""
    ids_str = request.args.get('ids', '')
    if not ids_str:
        flash('No months selected.', 'warning')
        return redirect(url_for('payroll.payroll_list', tab='reimbursement'))

    payroll_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    if not payroll_ids:
        flash('Invalid selection.', 'warning')
        return redirect(url_for('payroll.payroll_list', tab='reimbursement'))

    # Build combined data from multiple months
    payrolls = MonthlyPayroll.query.filter(MonthlyPayroll.id.in_(payroll_ids))\
        .order_by(MonthlyPayroll.year, MonthlyPayroll.month).all()

    if not payrolls:
        flash('No payrolls found.', 'warning')
        return redirect(url_for('payroll.payroll_list', tab='reimbursement'))

    est = payrolls[0].establishment
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    # Build row-wise data for each month and grand totals
    rows = []
    grand = {
        'epf_ac01': 0, 'epf_eps': 0, 'epf_edli': 0, 'epf_admin': 0,
        'epf_employer_refund': 0, 'esic_employer': 0, 'esic_employer_refund': 0,
        'total_refund': 0
    }

    for p in payrolls:
        _, _, _, d = _build_reimbursement_data(p.id)
        row = {
            'period': f"{p.month_name[:3]}-{p.year}",
            'emp_count': d['epf_count'],
            'epf_ac01': d['total_epf_ac01'],
            'epf_eps': d['total_epf_eps'],
            'epf_edli': d['total_epf_edli'],
            'epf_admin': d['total_epf_admin'],
            'epf_employer_refund': d['epf_employer_refund'],
            'esic_employer': d['total_esic_employer'],
            'esic_employer_refund': d['esic_employer_refund'],
            'total_refund': d['total_refund'],
        }
        rows.append(row)
        for k in grand:
            grand[k] += row.get(k, 0)

    # Period display
    if len(payrolls) == 1:
        period_display = f"{payrolls[0].month_name} {payrolls[0].year}"
    else:
        period_display = f"{payrolls[0].month_name[:3]} {payrolls[0].year} to {payrolls[-1].month_name[:3]} {payrolls[-1].year}"

    letter_date = datetime.now().strftime('%d-%b-%Y')

    return render_template('reports/reimbursement_multi.html',
                           est=est, config=config,
                           rows=rows, grand=grand,
                           period_display=period_display,
                           letter_date=letter_date)


# =============================================
# REPORT 7: MONTHLY & ANNUAL COMPLIANCE STATEMENT
# =============================================

def _build_monthly_compliance(payroll_id):
    """Build detailed monthly compliance data — employee-wise EPF + ESIC"""
    payroll = MonthlyPayroll.query.get_or_404(payroll_id)
    est = payroll.establishment
    verify_est_ownership(est)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
    entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll_id)\
        .join(Employee).order_by(Employee.name).all()

    rows = []
    totals = {
        'epf_wages': 0, 'esic_wages': 0, 'earned_gross': 0,
        'epf_ee': 0, 'epf_ac01': 0, 'epf_eps': 0, 'epf_edli': 0,
        'epf_admin': 0, 'epf_er_total': 0,
        'esic_ee': 0, 'esic_er': 0,
        'total_ee': 0, 'total_er': 0, 'total_deposit': 0,
    }

    for entry in entries:
        emp = entry.employee
        epf_er = entry.epf_ac01 + entry.epf_eps + entry.epf_edli + entry.epf_admin
        total_ee = entry.epf_employee + entry.esic_employee
        total_er = epf_er + entry.esic_employer
        total_deposit = total_ee + total_er

        row = {
            'emp': emp,
            'entry': entry,
            'epf_er': epf_er,
            'total_ee': total_ee,
            'total_er': total_er,
            'total_deposit': total_deposit,
        }
        rows.append(row)

        totals['epf_wages'] += entry.epf_wages
        totals['esic_wages'] += entry.esic_wages
        totals['earned_gross'] += entry.earned_gross
        totals['epf_ee'] += entry.epf_employee
        totals['epf_ac01'] += entry.epf_ac01
        totals['epf_eps'] += entry.epf_eps
        totals['epf_edli'] += entry.epf_edli
        totals['epf_admin'] += entry.epf_admin
        totals['epf_er_total'] += epf_er
        totals['esic_ee'] += entry.esic_employee
        totals['esic_er'] += entry.esic_employer
        totals['total_ee'] += total_ee
        totals['total_er'] += total_er
        totals['total_deposit'] += total_deposit

    return payroll, est, config, rows, totals


@reports_bp.route('/payroll/<int:payroll_id>/report/compliance')
def compliance_monthly(payroll_id):
    """Monthly Compliance Statement — HTML view"""
    payroll, est, config, rows, totals = _build_monthly_compliance(payroll_id)
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')
    return render_template('reports/compliance_monthly.html',
                           payroll=payroll, est=est, config=config,
                           rows=rows, totals=totals,
                           generated_on=generated_on)


@reports_bp.route('/establishment/<int:est_id>/report/compliance-annual')
def compliance_annual(est_id):
    """Annual Compliance Statement — April to March FY summary"""
    est = Establishment.query.get_or_404(est_id)
    config = PayrollConfig.query.filter_by(establishment_id=est.id).first()

    # Determine financial year from query param or current date
    fy = request.args.get('fy', None)
    if fy:
        fy_start_year = int(fy)
    else:
        now = datetime.now()
        fy_start_year = now.year if now.month >= 4 else now.year - 1

    fy_end_year = fy_start_year + 1
    fy_label = f'{fy_start_year}-{str(fy_end_year)[-2:]}'

    # Build 12 months: Apr(start_year) to Mar(end_year)
    months_order = []
    for m in range(4, 13):  # Apr to Dec
        months_order.append((m, fy_start_year))
    for m in range(1, 4):   # Jan to Mar
        months_order.append((m, fy_end_year))

    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    monthly_data = []
    grand = {
        'employees': 0, 'gross': 0,
        'epf_wages': 0, 'esic_wages': 0,
        'epf_ee': 0, 'epf_ac01': 0, 'epf_eps': 0,
        'epf_edli': 0, 'epf_admin': 0, 'epf_er': 0,
        'esic_ee': 0, 'esic_er': 0,
        'total_ee': 0, 'total_er': 0, 'total_deposit': 0,
    }

    for month_num, year in months_order:
        payroll = MonthlyPayroll.query.filter_by(
            establishment_id=est.id, month=month_num, year=year
        ).first()

        label = f'{month_names[month_num]}-{year}'

        if payroll and payroll.status in ['processing', 'finalized']:
            # Fetch entries to compute detailed breakup
            entries = PayrollEntry.query.filter_by(monthly_payroll_id=payroll.id).all()

            m_epf_wages = sum(e.epf_wages for e in entries)
            m_esic_wages = sum(e.esic_wages for e in entries)
            m_epf_ee = sum(e.epf_employee for e in entries)
            m_epf_ac01 = sum(e.epf_ac01 for e in entries)
            m_epf_eps = sum(e.epf_eps for e in entries)
            m_epf_edli = sum(e.epf_edli for e in entries)
            m_epf_admin = sum(e.epf_admin for e in entries)
            m_epf_er = m_epf_ac01 + m_epf_eps + m_epf_edli + m_epf_admin
            m_esic_ee = sum(e.esic_employee for e in entries)
            m_esic_er = sum(e.esic_employer for e in entries)
            m_total_ee = m_epf_ee + m_esic_ee
            m_total_er = m_epf_er + m_esic_er
            m_total_deposit = m_total_ee + m_total_er

            row = {
                'month': label,
                'employees': payroll.total_employees,
                'gross': payroll.total_gross,
                'epf_wages': m_epf_wages,
                'esic_wages': m_esic_wages,
                'epf_ee': m_epf_ee,
                'epf_ac01': m_epf_ac01,
                'epf_eps': m_epf_eps,
                'epf_edli': m_epf_edli,
                'epf_admin': m_epf_admin,
                'epf_er': m_epf_er,
                'esic_ee': m_esic_ee,
                'esic_er': m_esic_er,
                'total_ee': m_total_ee,
                'total_er': m_total_er,
                'total_deposit': m_total_deposit,
                'has_data': True,
                'payroll_id': payroll.id,
            }

            # Accumulate grand totals
            for key in grand:
                if key in row:
                    grand[key] += row[key]
        else:
            row = {
                'month': label,
                'has_data': False,
            }

        monthly_data.append(row)

    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('reports/compliance_annual.html',
                           est=est, config=config,
                           fy_label=fy_label,
                           fy_start_year=fy_start_year,
                           monthly_data=monthly_data,
                           grand=grand,
                           generated_on=generated_on)
