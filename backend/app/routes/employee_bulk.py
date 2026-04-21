from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from app import db
from app.models.employee import Employee
from app.models.establishment import Establishment
from app.models.payroll import (SalaryTemplate, SalaryTemplateHead, SalaryHead,
                                 PayrollConfig, EmployeeSalary, EmployeeSalaryHead)
from app.user_context import get_user_est_ids, user_establishments, verify_est_ownership, current_user_id, log_activity
from datetime import datetime, date
import io
import csv
import re

employee_bulk_bp = Blueprint('employee_bulk', __name__)

TEMPLATE_COLUMNS = [
    'Establishment Name or PF Code*',
    'Employee Name (as per Aadhaar)*',
    "Father's / Husband Name*",
    'Gender (Male/Female/Other)*',
    'Date of Birth (DD-MM-YYYY)*',
    'Date of Joining (DD-MM-YYYY)*',
    'UAN Number',
    'ESIC IP Number',
    # --- Salary Fields ---
    'Salary Type (Daily/Monthly/MonthlyHeads/CTC)',
    'Daily Rate',
    'Gross Salary (Monthly)',
    'Weekly Off (Paid/Unpaid/OT Rate)',
    # --- Head-wise columns (for MonthlyHeads) ---
    'Basic',
    'DA',
    'HRA',
    'Conveyance',
    'Other Allowance',
    'Washing Allowance',
    # --- CTC ---
    'CTC Amount (Monthly)',
    # --- WO Policy ---
    'WO Applicable (Yes/No)',
    'WO Type (Paid/Unpaid)',
    'WO Day (Sunday/Monday/etc/Rotational)',
    # --- Personal / KYC ---
    'Aadhaar Number',
    'PAN Number',
    'Mobile Number',
    'Email',
    'Marital Status',
    'Designation',
    'Department',
    'Address',
    'Bank Name',
    'Account Number',
    'IFSC Code',
    'Internal Emp Code'
]

SAMPLE_ROW = [
    'ABC Enterprises Pvt Ltd',
    'RAMESH KUMAR',
    'SURESH KUMAR',
    'Male',
    '15-06-1990',
    '01-04-2020',
    '100012345678',
    '71000012340000606',
    'Daily',
    '500',
    '',
    'Paid',
    # Head-wise (blank for Daily)
    '', '', '', '', '', '',
    # CTC (blank for Daily)
    '',
    # WO
    'Yes', 'Paid', 'Sunday',
    '987654321012',
    'ABCDE1234F',
    '9876543210',
    'ramesh@email.com',
    'Married',
    'Helper',
    'Production',
    '123, Main Road, City',
    'State Bank of India',
    '12345678901234',
    'SBIN0001234',
    'EMP-101'
]

COL_WIDTHS = [35, 30, 30, 22, 25, 25, 18, 22, 28, 14, 18, 22, 14, 14, 14, 16, 18, 18, 18, 18, 18, 28, 16, 14, 16, 25, 14, 18, 18, 30, 25, 20, 14, 16]


@employee_bulk_bp.route('/employees/download-template')
def download_template():
    """Download employee import template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sample_font = Font(name='Calibri', size=10, color='888888', italic=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TEMPLATE_COLUMNS))
    title_cell = ws['A1']
    title_cell.value = 'Vaishnavi Consultant ERP - Employee Import Template'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='10B981')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Instructions
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(TEMPLATE_COLUMNS))
    ws['A2'].value = 'Instructions: Fill data from Row 4. Row 3 is sample (delete before upload). Fields with * are required. At least UAN or ESIC IP must be provided.'
    ws['A2'].font = Font(name='Calibri', size=10, color='E53E3E', italic=True)
    ws.row_dimensions[2].height = 22

    # Headers
    ws.row_dimensions[3].height = 30
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_letters = [chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26) for i in range(1, len(TEMPLATE_COLUMNS) + 1)]
    for i in range(len(TEMPLATE_COLUMNS)):
        letter = chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
        ws.column_dimensions[letter].width = COL_WIDTHS[i]

    # Sample rows — one per salary type
    SAMPLE_ROWS = [
        SAMPLE_ROW,  # Daily wages
        [  # Monthly Fixed
            'ABC Enterprises Pvt Ltd', 'SURESH PATIL', 'MOHAN PATIL', 'Male',
            '20-03-1988', '15-06-2021', '100098765432', '71000098760000707',
            'Monthly', '', '12000', '',
            '', '', '', '', '', '',
            '',
            'Yes', 'Paid', 'Sunday',
            '876543210123', 'XYZAB5678C', '8765432109', '', 'Single',
            'Operator', 'Assembly', '456, MG Road, Town',
            'Bank of India', '98765432101234', 'BKID0001234', 'EMP-102'
        ],
        [  # Monthly Heads
            'ABC Enterprises Pvt Ltd', 'MEENA SHARMA', 'RAJESH SHARMA', 'Female',
            '10-11-1992', '01-01-2023', '100055566677', '',
            'MonthlyHeads', '', '', '',
            '6000', '0', '3000', '1600', '4400', '1000',
            '',
            'Yes', 'Unpaid', 'Rotational',
            '765432109876', 'PQRST1234D', '7654321098', 'meena@email.com', 'Married',
            'Supervisor', 'QC', '789, Nehru Nagar, City',
            'HDFC Bank', '55667788990011', 'HDFC0002345', 'EMP-103'
        ],
        [  # CTC
            'ABC Enterprises Pvt Ltd', 'AMIT JOSHI', 'PRAKASH JOSHI', 'Male',
            '05-08-1995', '10-04-2024', '100011122233', '71000011220000808',
            'CTC', '', '', '',
            '', '', '', '', '', '',
            '20000',
            'Yes', 'Paid', 'Sunday',
            '654321098765', 'LMNOP6789E', '6543210987', 'amit@email.com', 'Single',
            'Accountant', 'Finance', '321, Station Road, City',
            'ICICI Bank', '11223344556677', 'ICIC0003456', 'EMP-104'
        ],
    ]
    sample_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    for s_idx, s_row in enumerate(SAMPLE_ROWS):
        for col_idx, value in enumerate(s_row, 1):
            cell = ws.cell(row=4 + s_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.border = thin_border

    ws.cell(row=8, column=1, value='<-- Delete sample rows above and start here').font = Font(size=9, color='999999', italic=True)
    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Vaishnavi_ERP_Employee_Import_Template.xlsx')


@employee_bulk_bp.route('/employees/export')
def export_employees():
    """Export employees to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    est_id = request.args.get('establishment', '')
    filter_status = request.args.get('status', 'all')

    # Only export employees belonging to user's establishments
    allowed_est_ids = get_user_est_ids()
    query = Employee.query.filter(Employee.establishment_id.in_(allowed_est_ids))
    if est_id:
        query = query.filter_by(establishment_id=int(est_id))
    if filter_status == 'active':
        query = query.filter_by(is_active=True)
    elif filter_status == 'inactive':
        query = query.filter_by(is_active=False)

    employees = query.order_by(Employee.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )

    export_cols = ['Sr No', 'UAN Number', 'Employee Name', "Father/Husband", 'Gender', 'DOB', 'DOJ',
                   'Establishment', 'ESIC IP', 'Emp Code',
                   'Salary Type', 'Daily Rate', 'Gross Salary', 'Weekly Off',
                   'Basic', 'DA', 'HRA', 'Conveyance', 'Other Allowance', 'Washing Allowance',
                   'CTC Amount',
                   'WO Applicable', 'WO Type', 'WO Day',
                   'Aadhaar', 'PAN', 'Mobile', 'Email', 'Marital Status',
                   'Designation', 'Department', 'Address', 'Bank', 'Account No', 'IFSC',
                   'Internal Code', 'Status']

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_cols))
    ws['A1'].value = f'Vaishnavi Consultant ERP - Employees Export ({datetime.now().strftime("%d-%m-%Y")})'
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color='10B981')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(export_cols))
    ws['A2'].value = f'Total Records: {len(employees)}'
    ws['A2'].font = Font(size=10, color='666666')

    # Headers
    for col_idx, col_name in enumerate(export_cols, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    for row_idx, emp in enumerate(employees, 4):
        # Get current salary
        cur_sal = None
        for s in emp.salaries:
            if s.is_current:
                cur_sal = s
                break
        sal_type_label = ''
        daily_rate_val = ''
        gross_val = ''
        wo_val = ''
        head_vals = {'BASIC': '', 'DA': '', 'HRA': '', 'CONV': '', 'OTH_ALW': '', 'WASH': ''}
        ctc_val = ''
        wo_appl_val = ''
        wo_type_val = ''
        wo_day_val = ''
        if cur_sal:
            # Determine display type: check if heads exist to differentiate Monthly vs MonthlyHeads
            sal_heads = EmployeeSalaryHead.query.filter_by(employee_salary_id=cur_sal.id).all()
            has_sal_heads = len(sal_heads) > 0
            if cur_sal.salary_type == 'daily_wages':
                sal_type_label = 'Daily'
            elif cur_sal.salary_type == 'monthly_package':
                sal_type_label = 'CTC'
            elif cur_sal.salary_type == 'monthly_fixed' and has_sal_heads:
                sal_type_label = 'MonthlyHeads'
            else:
                sal_type_label = 'Monthly'
            daily_rate_val = cur_sal.daily_rate or ''
            gross_val = cur_sal.gross_salary or ''
            wo_val = (cur_sal.weekly_off_policy or '').title() if cur_sal.salary_type == 'daily_wages' else ''
            # Fill head-wise amounts
            if has_sal_heads:
                for sh in sal_heads:
                    head_obj = SalaryHead.query.get(sh.salary_head_id)
                    if head_obj and head_obj.short_code in head_vals:
                        head_vals[head_obj.short_code] = sh.amount or ''
            # CTC = gross + employer contributions (approximate), or store gross as CTC for now
            if cur_sal.salary_type == 'monthly_package':
                ctc_val = cur_sal.gross_salary or ''
            # WO fields
            if cur_sal.wo_applicable is not None:
                wo_appl_val = 'Yes' if cur_sal.wo_applicable else 'No'
            if cur_sal.wo_type:
                wo_type_val = cur_sal.wo_type.title()
            if cur_sal.wo_day:
                wo_day_val = cur_sal.wo_day.title()

        row_data = [
            row_idx - 3, emp.uan_number or '', emp.name, emp.father_husband_name, emp.gender,
            emp.date_of_birth.strftime('%d-%m-%Y'), emp.date_of_joining.strftime('%d-%m-%Y'),
            emp.establishment.company_name, emp.esic_ip_number or '', emp.emp_code,
            sal_type_label, daily_rate_val, gross_val, wo_val,
            head_vals['BASIC'], head_vals['DA'], head_vals['HRA'], head_vals['CONV'],
            head_vals['OTH_ALW'], head_vals['WASH'],
            ctc_val,
            wo_appl_val, wo_type_val, wo_day_val,
            emp.aadhaar_number or '', emp.pan_number or '', emp.mobile_number or '', emp.email or '',
            emp.marital_status or '',
            emp.designation or '', emp.department or '', emp.address or '', emp.bank_name or '',
            emp.bank_account_number or '', emp.bank_ifsc_code or '',
            emp.internal_emp_code or '',
            'Active' if emp.is_active else 'Exited'
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Employees_Export_{datetime.now().strftime("%d%m%Y_%H%M")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@employee_bulk_bp.route('/employees/import', methods=['GET', 'POST'])
def import_employees():
    """Import employees from Excel"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Please select a file.', 'danger')
            return redirect(url_for('employee_bulk.import_employees'))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                flash('File is empty.', 'danger')
                return redirect(url_for('employee_bulk.import_employees'))

            # Find header row
            header_idx = None
            headers = []
            for idx, row in enumerate(rows[:5]):
                row_str = [str(c).strip().lower() if c else '' for c in row]
                if any('employee name' in cell or 'employeename' in cell or 'name' in cell for cell in row_str):
                    header_idx = idx
                    headers = [str(c).strip() if c else '' for c in row]
                    break

            if header_idx is None:
                flash('Could not find header row with "Employee Name".', 'danger')
                return redirect(url_for('employee_bulk.import_employees'))

            # Build active establishments lookup (only user's own)
            est_lookup = {}
            for est in user_establishments().filter_by(is_active=True).all():
                est_lookup[est.company_name.lower()] = est.id
                if est.pf_code:
                    est_lookup[est.pf_code.lower()] = est.id

            # Map columns
            col_map = {}
            for idx, h in enumerate(headers):
                hl = h.lower().replace('_', '').replace(' ', '').replace('*', '')
                if 'establishment' in hl or 'pfcode' in hl or 'company' in hl:
                    col_map['establishment'] = idx
                elif 'employeename' in hl or ('name' in hl and 'father' not in hl and 'husband' not in hl and 'bank' not in hl and 'nominee' not in hl):
                    col_map['name'] = idx
                elif 'father' in hl or 'husband' in hl:
                    col_map['father'] = idx
                elif 'gender' in hl:
                    col_map['gender'] = idx
                elif 'birth' in hl or 'dob' in hl:
                    col_map['dob'] = idx
                elif 'joining' in hl or 'doj' in hl:
                    col_map['doj'] = idx
                elif 'uan' in hl:
                    col_map['uan'] = idx
                elif 'esic' in hl or 'ipnumber' in hl:
                    col_map['esic'] = idx
                elif 'aadhaar' in hl or 'aadhar' in hl:
                    col_map['aadhaar'] = idx
                elif 'pan' in hl:
                    col_map['pan'] = idx
                elif 'mobile' in hl or 'phone' in hl:
                    col_map['mobile'] = idx
                elif 'email' in hl:
                    col_map['email'] = idx
                elif 'designation' in hl:
                    col_map['designation'] = idx
                elif 'department' in hl:
                    col_map['department'] = idx
                elif 'bankname' in hl or (hl == 'bank'):
                    col_map['bank_name'] = idx
                elif 'account' in hl:
                    col_map['account'] = idx
                elif 'ifsc' in hl:
                    col_map['ifsc'] = idx
                elif 'salarytype' in hl or 'saltype' in hl:
                    col_map['salary_type'] = idx
                elif 'dailyrate' in hl or 'rateper' in hl:
                    col_map['daily_rate'] = idx
                elif 'grosssalary' in hl or 'gross' in hl or 'monthlysalary' in hl:
                    col_map['gross_salary'] = idx
                elif 'weeklyoff' in hl or 'wopolicy' in hl:
                    col_map['weekly_off'] = idx
                elif 'marital' in hl:
                    col_map['marital'] = idx
                elif 'address' in hl:
                    col_map['address'] = idx
                elif 'internalcode' in hl or 'internalempcode' in hl or 'empcode' in hl:
                    col_map['internal_code'] = idx
                elif hl == 'basic':
                    col_map['head_basic'] = idx
                elif hl == 'da':
                    col_map['head_da'] = idx
                elif hl == 'hra':
                    col_map['head_hra'] = idx
                elif 'conveyance' in hl:
                    col_map['head_conv'] = idx
                elif 'otherallowance' in hl or 'othalw' in hl:
                    col_map['head_oth_alw'] = idx
                elif 'washingallowance' in hl or 'wash' in hl:
                    col_map['head_wash'] = idx
                elif 'ctcamount' in hl or (hl == 'ctc'):
                    col_map['ctc_amount'] = idx
                elif 'woapplicable' in hl:
                    col_map['wo_applicable'] = idx
                elif 'wotype' in hl or 'wopaytype' in hl:
                    col_map['wo_type'] = idx
                elif 'woday' in hl:
                    col_map['wo_day'] = idx

            imported = 0
            skipped = 0
            errors = []

            for row_num, row in enumerate(rows[header_idx + 1:], 1):
                try:
                    def get(key):
                        idx = col_map.get(key)
                        if idx is not None and idx < len(row) and row[idx]:
                            val = str(row[idx]).strip()
                            if val.endswith('.0'):
                                val = val[:-2]
                            return val if val and val != '0' else None
                        return None

                    name = get('name')
                    if not name:
                        continue

                    # Find establishment
                    est_val = get('establishment')
                    est_id = None
                    if est_val:
                        est_id = est_lookup.get(est_val.lower())
                    if not est_id:
                        errors.append(f'Row {row_num}: Establishment "{est_val}" not found')
                        continue

                    father = get('father')
                    gender = get('gender')
                    if not father or not gender:
                        errors.append(f'Row {row_num}: Missing father/husband name or gender')
                        continue

                    # Parse dates
                    from app.routes.bulk import _parse_date
                    dob_val = row[col_map['dob']] if col_map.get('dob') is not None and col_map['dob'] < len(row) else None
                    doj_val = row[col_map['doj']] if col_map.get('doj') is not None and col_map['doj'] < len(row) else None
                    dob = _parse_date(dob_val)
                    doj = _parse_date(doj_val)

                    if not dob or not doj:
                        errors.append(f'Row {row_num}: Invalid date of birth or joining')
                        continue

                    uan = get('uan')
                    esic = get('esic')
                    if not uan and not esic:
                        errors.append(f'Row {row_num}: At least UAN or ESIC IP required')
                        continue

                    # Check duplicate by UAN or name+establishment
                    if uan:
                        existing = Employee.query.filter_by(uan_number=uan).first()
                        if existing:
                            skipped += 1
                            continue

                    emp_code = Employee.generate_emp_code()

                    emp = Employee(
                        emp_code=emp_code,
                        establishment_id=est_id,
                        name=name.upper(),
                        father_husband_name=father.upper(),
                        gender=gender.capitalize(),
                        date_of_birth=dob,
                        date_of_joining=doj,
                        uan_number=uan,
                        esic_ip_number=esic,
                        aadhaar_number=get('aadhaar'),
                        pan_number=get('pan').upper() if get('pan') else None,
                        mobile_number=get('mobile'),
                        email=get('email'),
                        marital_status=get('marital'),
                        designation=get('designation'),
                        department=get('department'),
                        address=get('address'),
                        bank_name=get('bank_name'),
                        bank_account_number=get('account'),
                        bank_ifsc_code=get('ifsc').upper() if get('ifsc') else None,
                        internal_emp_code=get('internal_code'),
                        is_active=True
                    )
                    db.session.add(emp)
                    db.session.flush()  # Get emp.id for salary

                    # Create salary record if salary data provided
                    sal_type_raw = (get('salary_type') or '').lower().strip()
                    daily_rate_val = get('daily_rate')
                    gross_val = get('gross_salary')
                    ctc_val = get('ctc_amount')
                    has_heads = any(get(k) for k in ('head_basic', 'head_da', 'head_hra', 'head_conv', 'head_oth_alw', 'head_wash'))

                    if sal_type_raw or daily_rate_val or gross_val or ctc_val or has_heads:
                        # Determine salary type
                        if sal_type_raw in ('daily', 'daily_wages', 'dailywages'):
                            emp_salary_type = 'daily_wages'
                        elif sal_type_raw in ('ctc', 'package', 'monthly_package', 'monthlypackage'):
                            emp_salary_type = 'monthly_package'
                        elif sal_type_raw in ('monthlyheads', 'monthly_heads', 'heads'):
                            emp_salary_type = 'monthly_fixed'  # monthly_heads maps to monthly_fixed with heads
                            has_heads = True  # force head-wise save
                        else:
                            emp_salary_type = 'monthly_fixed'

                        # WO Policy fields
                        wo_appl_raw = (get('wo_applicable') or '').lower().strip()
                        wo_type_raw = (get('wo_type') or '').lower().strip()
                        wo_day_raw = (get('wo_day') or '').lower().strip()

                        sal_wo_applicable = None
                        if wo_appl_raw in ('yes', 'y', '1', 'true'):
                            sal_wo_applicable = True
                        elif wo_appl_raw in ('no', 'n', '0', 'false'):
                            sal_wo_applicable = False

                        sal_wo_type = None
                        if wo_type_raw in ('paid',):
                            sal_wo_type = 'paid'
                        elif wo_type_raw in ('unpaid',):
                            sal_wo_type = 'unpaid'

                        sal_wo_day = None
                        valid_days = ('sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'rotational')
                        if wo_day_raw in valid_days:
                            sal_wo_day = wo_day_raw

                        sal = EmployeeSalary(
                            employee_id=emp.id,
                            effective_from=doj,
                            is_current=True,
                            salary_type=emp_salary_type,
                            wo_applicable=sal_wo_applicable,
                            wo_type=sal_wo_type,
                            wo_day=sal_wo_day,
                        )

                        if emp_salary_type == 'daily_wages' and daily_rate_val:
                            try:
                                sal.daily_rate = max(0, float(daily_rate_val))
                            except ValueError:
                                sal.daily_rate = 0

                        # ── CTC: auto-calculate head breakup ──
                        if emp_salary_type == 'monthly_package' and ctc_val:
                            try:
                                ctc = max(0, float(ctc_val))
                            except ValueError:
                                ctc = 0

                            if ctc > 0:
                                # Get payroll config for employer rates
                                config = PayrollConfig.query.filter_by(establishment_id=est_id).first()
                                epf_rate = 0
                                epf_admin = 0
                                esic_rate = 0
                                epf_ceiling = 15000
                                esic_ceiling = 21000
                                if config and config.epf_applicable:
                                    epf_rate = ((config.epf_ac01_rate or 3.67) + (config.epf_eps_rate or 8.33)) / 100
                                    epf_admin = ((config.epf_admin_rate or 0.50) + (config.epf_edli_rate or 0.50)) / 100
                                    epf_ceiling = config.epf_wage_ceiling or 15000
                                if config and config.esic_applicable:
                                    esic_rate = (config.esic_employer_rate or 3.25) / 100
                                    esic_ceiling = config.esic_wage_ceiling or 21000

                                # Iterative solver: CTC = Gross + Employer EPF + ESIC + Admin
                                gross = ctc
                                for _ in range(5):
                                    basic = round(gross * 0.40)
                                    epf_wages = min(basic, epf_ceiling)
                                    emp_epf = round(epf_wages * epf_rate)
                                    emp_admin = round(epf_wages * epf_admin)
                                    emp_esic = round(gross * esic_rate) if gross <= esic_ceiling else 0
                                    gross = ctc - emp_epf - emp_admin - emp_esic

                                # Final head split
                                basic = round(gross * 0.40)
                                hra = round(basic * 0.50)
                                oth_alw = max(0, gross - basic - hra)

                                sal.gross_salary = round(gross)
                                db.session.add(sal)
                                db.session.flush()

                                # Ensure default heads exist
                                heads = SalaryHead.query.filter_by(establishment_id=est_id).order_by(SalaryHead.display_order).all()
                                if not heads:
                                    from app.routes.payroll import _create_default_salary_heads
                                    _create_default_salary_heads(est_id)
                                    heads = SalaryHead.query.filter_by(establishment_id=est_id).order_by(SalaryHead.display_order).all()

                                head_amounts = {'BASIC': basic, 'HRA': hra, 'OTH_ALW': oth_alw}
                                for h in heads:
                                    amt = head_amounts.get(h.short_code, 0)
                                    if amt > 0:
                                        db.session.add(EmployeeSalaryHead(
                                            employee_salary_id=sal.id, salary_head_id=h.id, amount=amt
                                        ))
                            else:
                                db.session.add(sal)

                        # ── Monthly Heads: save individual head amounts ──
                        elif (emp_salary_type == 'monthly_fixed' and has_heads):
                            head_map = {
                                'BASIC': get('head_basic'),
                                'DA': get('head_da'),
                                'HRA': get('head_hra'),
                                'CONV': get('head_conv'),
                                'OTH_ALW': get('head_oth_alw'),
                                'WASH': get('head_wash'),
                            }
                            # Calculate gross from head amounts
                            total_gross = 0
                            for v in head_map.values():
                                if v:
                                    try:
                                        total_gross += max(0, float(v))
                                    except ValueError:
                                        pass
                            sal.gross_salary = round(total_gross)
                            db.session.add(sal)
                            db.session.flush()

                            # Ensure default heads exist
                            heads = SalaryHead.query.filter_by(establishment_id=est_id).order_by(SalaryHead.display_order).all()
                            if not heads:
                                from app.routes.payroll import _create_default_salary_heads
                                _create_default_salary_heads(est_id)
                                heads = SalaryHead.query.filter_by(establishment_id=est_id).order_by(SalaryHead.display_order).all()

                            for h in heads:
                                raw = head_map.get(h.short_code)
                                if raw:
                                    try:
                                        amt = max(0, float(raw))
                                    except ValueError:
                                        amt = 0
                                    if amt > 0:
                                        db.session.add(EmployeeSalaryHead(
                                            employee_salary_id=sal.id, salary_head_id=h.id, amount=amt
                                        ))

                        # ── Monthly Fixed / fallback ──
                        else:
                            if emp_salary_type != 'daily_wages' and gross_val:
                                try:
                                    sal.gross_salary = max(0, float(gross_val))
                                except ValueError:
                                    sal.gross_salary = 0
                            db.session.add(sal)

                    imported += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            if imported > 0:
                db.session.commit()

            msg = []
            if imported > 0:
                msg.append(f'{imported} employees imported')
            if skipped > 0:
                msg.append(f'{skipped} duplicates skipped')
            if errors:
                msg.append(f'{len(errors)} errors')

            if imported > 0:
                flash('. '.join(msg) + '.', 'success')
            elif skipped > 0:
                flash('. '.join(msg) + '.', 'info')
            else:
                flash('No employees imported. ' + '. '.join(msg), 'warning')

            for err in errors[:5]:
                flash(err, 'danger')

            return redirect(url_for('employee.employee_list'))

        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('employee_bulk.import_employees'))

    return render_template('employees/import.html')


# =============================================
# EPF ACTIVE MEMBER DATA IMPORT (CSV from EPFO)
# =============================================

_NAME_PREFIXES = re.compile(
    r'^(Mr\.?|Mrs\.?|Ms\.?|Smt\.?|Shri\.?|Shrimati\.?|Sri\.?|Dr\.?)\s+',
    re.IGNORECASE
)


def _clean_name(raw_name):
    """Remove Mr./Mrs./Smt./Shri prefixes and normalize"""
    if not raw_name:
        return '', ''
    raw_name = raw_name.strip()
    cleaned = _NAME_PREFIXES.sub('', raw_name).strip().upper()
    # Collapse multiple spaces
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return raw_name, cleaned


def _parse_epf_date(val):
    """Parse EPF date format: 25-AUG-2006 and various alternatives"""
    if not val or str(val).strip().upper() in ('NOT AVAILABLE', 'NA', 'N/A', ''):
        return None
    val = str(val).strip()
    # Try all common date formats from EPF portal exports
    for fmt in [
        '%d-%b-%Y',     # 25-AUG-2006 (EPF standard)
        '%d-%B-%Y',     # 25-AUGUST-2006
        '%d-%m-%Y',     # 25-08-2006
        '%d/%m/%Y',     # 25/08/2006
        '%d/%b/%Y',     # 25/AUG/2006
        '%Y-%m-%d',     # 2006-08-25 (ISO)
        '%m/%d/%Y',     # 08/25/2006 (US format — rare but possible)
        '%d-%b-%y',     # 25-AUG-06 (2-digit year)
        '%d/%m/%y',     # 25/08/06
    ]:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bank_ifsc(val):
    """Parse '6234891XXXX,IFSC:SBIN0020228' format"""
    if not val or val == 'NOT AVAILABLE':
        return None, None
    parts = val.split(',')
    account = parts[0].strip() if parts else None
    ifsc = None
    for p in parts:
        p = p.strip()
        if p.upper().startswith('IFSC:'):
            ifsc = p[5:].strip()
    return account, ifsc


def _yes_no_bool(val):
    """Convert YES/NO to boolean"""
    if not val:
        return None
    return str(val).strip().upper() == 'YES'


def _not_available(val):
    """Return None if 'NOT AVAILABLE', else stripped value"""
    if not val or str(val).strip().upper() == 'NOT AVAILABLE':
        return None
    return str(val).strip()


@employee_bulk_bp.route('/establishments/<int:est_id>/epf-import', methods=['GET', 'POST'])
def epf_import(est_id):
    """Import employees from EPF Active Member CSV download"""
    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)

    if request.method == 'POST':
        action = request.form.get('action', 'preview')

        if action == 'preview':
            return _epf_import_preview(est)
        elif action == 'confirm':
            return _epf_import_confirm(est)

    return render_template('employees/epf_import.html', est=est, step='upload')


def _epf_import_preview(est):
    """Step 2: Parse CSV and show preview"""
    file = request.files.get('csv_file')
    if not file or file.filename == '':
        flash('Please select a CSV file.', 'danger')
        return redirect(url_for('employee_bulk.epf_import', est_id=est.id))

    try:
        content = file.read().decode('utf-8-sig')  # BOM-safe
        reader = csv.reader(content.splitlines())
        rows = list(reader)

        if len(rows) < 2:
            flash('CSV file is empty or has no data rows.', 'danger')
            return redirect(url_for('employee_bulk.epf_import', est_id=est.id))

        header = [h.strip() for h in rows[0]]

        # Map columns
        col_map = {}
        for idx, h in enumerate(header):
            hl = h.upper().strip()
            if hl == 'UAN' or hl == 'UAN NUMBER':
                col_map['uan'] = idx
            elif hl == 'MEMBER ID' or hl == 'MEMBER_ID':
                col_map['member_id'] = idx
            elif hl in ('NAME', 'MEMBER NAME', 'EMPLOYEE NAME'):
                col_map['name'] = idx
            elif hl == 'GENDER' or hl == 'SEX':
                col_map['gender'] = idx
            elif hl.startswith('DOB') or hl == 'DATE OF BIRTH' or 'BIRTH' in hl:
                col_map['dob'] = idx
            elif hl.startswith('DOJ') or hl == 'DATE OF JOINING' or hl == 'JOINING DATE':
                col_map['doj'] = idx
            elif ("FATHER" in hl or "HUSBAND" in hl) and 'RELATION' not in hl:
                col_map['father'] = idx
            elif hl == 'RELATION' or hl == 'RELATIONSHIP':
                col_map['relation'] = idx
            elif 'MARITAL' in hl:
                col_map['marital'] = idx
            elif hl == 'MOBILE' or hl == 'MOBILE NUMBER' or hl == 'PHONE':
                col_map['mobile'] = idx
            elif 'EMAIL' in hl:
                col_map['email'] = idx
            elif 'AADHAAR' in hl and 'VERIFIED' in hl:
                # Must check this BEFORE plain 'AADHAAR' match
                col_map['aadhaar_verified'] = idx
            elif hl in ('AADHAAR', 'AADHAAR NUMBER', 'AADHAR', 'AADHAR NUMBER'):
                col_map['aadhaar'] = idx
            elif hl == 'PAN' or hl == 'PAN NUMBER':
                col_map['pan'] = idx
            elif 'BANK' in hl:
                col_map['bank'] = idx
            elif 'NOMINATION' in hl:
                col_map['nomination'] = idx
            elif 'FACE' in hl:
                col_map['face_auth'] = idx
            elif 'ESIC' in hl or 'IP NUMBER' in hl or 'E-PEHCHAN' in hl or 'E PEHCHAN' in hl:
                col_map['esic'] = idx

        if 'uan' not in col_map:
            flash('CSV does not have a UAN column. Please check the file format.', 'danger')
            return redirect(url_for('employee_bulk.epf_import', est_id=est.id))

        # Parse all rows
        preview_data = []
        for row_num, row in enumerate(rows[1:], 2):
            if not row or all(not c.strip() for c in row):
                continue

            def get(key):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    return row[idx].strip()
                return None

            uan = get('uan')
            if not uan:
                continue

            raw_name = get('name') or ''
            epfo_name_raw, cleaned_name = _clean_name(raw_name)

            raw_father = get('father') or ''
            _, cleaned_father = _clean_name(raw_father)

            dob = _parse_epf_date(get('dob'))
            doj = _parse_epf_date(get('doj'))
            gender_raw = _not_available(get('gender'))
            gender = gender_raw.capitalize() if gender_raw else None

            relation = _not_available(get('relation'))
            if relation:
                relation = relation.upper()

            marital = _not_available(get('marital'))
            if marital:
                marital = marital.replace('-', ' ').title()  # UN-MARRIED -> Un Married

            mobile = _not_available(get('mobile'))
            email = _not_available(get('email'))
            aadhaar = _not_available(get('aadhaar'))
            pan = _not_available(get('pan'))
            esic = _not_available(get('esic'))

            bank_raw = get('bank')
            bank_account, bank_ifsc = _parse_bank_ifsc(bank_raw)

            nomination = _yes_no_bool(get('nomination'))
            aadhaar_verified = _yes_no_bool(get('aadhaar_verified'))
            face_auth = _yes_no_bool(get('face_auth'))
            member_id = _not_available(get('member_id'))

            # Check if employee already exists
            existing = Employee.query.filter_by(uan_number=uan).first()
            status = 'update' if existing else 'new'

            # Check name mismatch for existing
            name_mismatch = False
            if existing and existing.name and cleaned_name:
                name_mismatch = existing.name.strip().upper() != cleaned_name

            rec = {
                'row_num': row_num,
                'uan': uan,
                'member_id': member_id,
                'epfo_name_raw': epfo_name_raw,
                'cleaned_name': cleaned_name,
                'gender': gender,
                'dob': dob.strftime('%d-%m-%Y') if dob else '',
                'doj': doj.strftime('%d-%m-%Y') if doj else '',
                'father': cleaned_father,
                'relation': relation,
                'marital': marital,
                'mobile': mobile,
                'email': email,
                'aadhaar': aadhaar,
                'pan': pan,
                'esic': esic,
                'bank_account': bank_account,
                'bank_ifsc': bank_ifsc,
                'nomination': nomination,
                'aadhaar_verified': aadhaar_verified,
                'face_auth': face_auth,
                'status': status,
                'name_mismatch': name_mismatch,
                'existing_name': existing.name if existing else None,
                'existing_id': existing.id if existing else None,
            }
            preview_data.append(rec)

        new_count = sum(1 for r in preview_data if r['status'] == 'new')
        update_count = sum(1 for r in preview_data if r['status'] == 'update')
        mismatch_count = sum(1 for r in preview_data if r['name_mismatch'])

        # Store preview data in session for confirm step
        session['epf_import_data'] = preview_data
        session['epf_import_est_id'] = est.id

        # Get salary templates for assignment after import
        salary_templates = SalaryTemplate.query.filter_by(
            establishment_id=est.id, is_active=True
        ).order_by(SalaryTemplate.name).all()

        return render_template('employees/epf_import.html',
                               est=est, step='preview',
                               preview_data=preview_data,
                               new_count=new_count,
                               update_count=update_count,
                               mismatch_count=mismatch_count,
                               total=len(preview_data),
                               salary_templates=salary_templates)

    except Exception as e:
        flash(f'Error reading CSV: {str(e)}', 'danger')
        return redirect(url_for('employee_bulk.epf_import', est_id=est.id))


def _epf_import_confirm(est):
    """Step 3: Actually create/update employees"""
    preview_data = session.get('epf_import_data')
    stored_est_id = session.get('epf_import_est_id')

    if not preview_data or stored_est_id != est.id:
        flash('Import session expired. Please upload the CSV again.', 'warning')
        return redirect(url_for('employee_bulk.epf_import', est_id=est.id))

    # Get salary template selection (if any)
    tmpl_id = request.form.get('salary_template_id', '')
    salary_tmpl = None
    tmpl_head_values = {}
    config = None
    heads = []
    if tmpl_id:
        try:
            salary_tmpl = SalaryTemplate.query.get(int(tmpl_id))
            if salary_tmpl and salary_tmpl.establishment_id == est.id:
                config = PayrollConfig.query.filter_by(establishment_id=est.id).first()
                heads = SalaryHead.query.filter_by(
                    establishment_id=est.id, is_active=True
                ).order_by(SalaryHead.display_order).all()
                for th in salary_tmpl.head_values:
                    tmpl_head_values[th.salary_head_id] = th.amount
            else:
                salary_tmpl = None
        except (ValueError, TypeError):
            salary_tmpl = None

    created = 0
    updated = 0
    salary_assigned = 0
    errors = []

    for rec in preview_data:
        try:
            uan = rec['uan']
            existing = Employee.query.filter_by(uan_number=uan).first()

            dob = _parse_epf_date(rec['dob']) if rec.get('dob') else None
            doj = _parse_epf_date(rec['doj']) if rec.get('doj') else None

            if existing:
                # UPDATE existing employee
                existing.epfo_name = rec['cleaned_name']
                existing.member_id = rec.get('member_id')
                existing.relation = rec.get('relation')
                existing.marital_status = rec.get('marital')
                existing.nomination_filed = rec.get('nomination')
                existing.aadhaar_verified = rec.get('aadhaar_verified')
                existing.face_auth_status = rec.get('face_auth')

                # Update non-masked fields
                if rec.get('gender'):
                    existing.gender = rec['gender']
                if dob:
                    existing.date_of_birth = dob
                if doj:
                    existing.date_of_joining = doj
                if rec.get('father'):
                    existing.father_husband_name = rec['father']
                if rec.get('mobile'):
                    existing.mobile_number = rec['mobile']
                if rec.get('email'):
                    existing.email = rec['email']
                if rec.get('esic'):
                    existing.esic_ip_number = rec['esic']

                # Only update aadhaar/pan/bank if not masked (no XXXX)
                if rec.get('aadhaar') and 'X' not in rec['aadhaar']:
                    existing.aadhaar_number = rec['aadhaar']
                if rec.get('pan') and 'X' not in rec['pan']:
                    existing.pan_number = rec['pan']
                if rec.get('bank_account') and 'X' not in rec['bank_account']:
                    existing.bank_account_number = rec['bank_account']
                if rec.get('bank_ifsc'):
                    existing.bank_ifsc_code = rec['bank_ifsc']

                # Reset mismatch acceptance if name changed
                if existing.has_name_mismatch:
                    existing.name_mismatch_accepted = False

                updated += 1
            else:
                # CREATE new employee
                emp_code = Employee.generate_emp_code()
                emp = Employee(
                    emp_code=emp_code,
                    establishment_id=est.id,
                    name=rec['cleaned_name'],
                    epfo_name=rec['cleaned_name'],
                    father_husband_name=rec.get('father') or 'N/A',
                    gender=rec.get('gender') or 'Male',
                    date_of_birth=dob,
                    date_of_joining=doj,
                    uan_number=uan,
                    esic_ip_number=rec.get('esic'),
                    member_id=rec.get('member_id'),
                    relation=rec.get('relation'),
                    marital_status=rec.get('marital'),
                    mobile_number=rec.get('mobile'),
                    email=rec.get('email'),
                    nomination_filed=rec.get('nomination'),
                    aadhaar_verified=rec.get('aadhaar_verified'),
                    face_auth_status=rec.get('face_auth'),
                    is_active=True
                )

                # Only set aadhaar/pan/bank if not masked
                if rec.get('aadhaar') and 'X' not in rec['aadhaar']:
                    emp.aadhaar_number = rec['aadhaar']
                if rec.get('pan') and 'X' not in rec['pan']:
                    emp.pan_number = rec['pan']
                if rec.get('bank_account') and 'X' not in rec['bank_account']:
                    emp.bank_account_number = rec['bank_account']
                if rec.get('bank_ifsc'):
                    emp.bank_ifsc_code = rec['bank_ifsc']

                db.session.add(emp)
                db.session.flush()  # Get emp.id for salary assignment
                created += 1

                # Auto-assign salary from template (for new employees only)
                if salary_tmpl and config:
                    new_sal = EmployeeSalary(
                        employee_id=emp.id,
                        salary_template_id=salary_tmpl.id,
                        effective_from=emp.date_of_joining or date.today(),
                        is_current=True,
                        salary_type=salary_tmpl.salary_type,
                        weekly_off_policy=salary_tmpl.weekly_off_policy,
                        daily_rate=salary_tmpl.daily_rate if salary_tmpl.salary_type == 'daily_wages' else None,
                        revision_reason=f'Auto-assigned: {salary_tmpl.name}',
                    )
                    if config.salary_structure == 'gross_only':
                        new_sal.gross_salary = salary_tmpl.gross_salary
                        db.session.add(new_sal)
                    else:
                        db.session.add(new_sal)
                        db.session.flush()
                        total_gross = 0
                        for head in heads:
                            amt = tmpl_head_values.get(head.id, 0)
                            esh = EmployeeSalaryHead(
                                employee_salary_id=new_sal.id,
                                salary_head_id=head.id,
                                amount=amt
                            )
                            db.session.add(esh)
                            if head.is_in_gross and head.head_type == 'earning':
                                total_gross += amt
                        new_sal.gross_salary = total_gross
                    salary_assigned += 1

        except Exception as e:
            errors.append(f'UAN {rec.get("uan", "?")}: {str(e)}')

    if created > 0 or updated > 0:
        db.session.commit()
        log_activity('epf_import', f'EPF Import: {created} new, {updated} updated for {est.company_name}')

    # Clear session data
    session.pop('epf_import_data', None)
    session.pop('epf_import_est_id', None)

    msg_parts = []
    if created > 0:
        msg_parts.append(f'{created} new employees created')
    if salary_assigned > 0:
        msg_parts.append(f'{salary_assigned} salaries auto-assigned from template "{salary_tmpl.name}"')
    if updated > 0:
        msg_parts.append(f'{updated} existing employees updated')
    if errors:
        msg_parts.append(f'{len(errors)} errors')

    if created > 0 or updated > 0:
        flash('. '.join(msg_parts) + '.', 'success')
    else:
        flash('No employees imported. ' + '. '.join(msg_parts), 'warning')

    for err in errors[:5]:
        flash(err, 'danger')

    return redirect(url_for('employee.employee_list', establishment=est.id))


@employee_bulk_bp.route('/employees/<int:emp_id>/accept-mismatch', methods=['POST'])
def accept_name_mismatch(emp_id):
    """Accept a name mismatch for an employee"""
    emp = Employee.query.get_or_404(emp_id)
    est = emp.establishment
    verify_est_ownership(est)
    emp.name_mismatch_accepted = True
    db.session.commit()
    flash(f'Name mismatch accepted for {emp.name}.', 'info')
    return redirect(request.referrer or url_for('employee.employee_view', id=emp_id))


@employee_bulk_bp.route('/employees/<int:emp_id>/reject-mismatch', methods=['POST'])
def reject_name_mismatch(emp_id):
    """Reset mismatch acceptance — flag it as unresolved again"""
    emp = Employee.query.get_or_404(emp_id)
    est = emp.establishment
    verify_est_ownership(est)
    emp.name_mismatch_accepted = False
    db.session.commit()
    flash(f'Name mismatch flagged for {emp.name}.', 'warning')
    return redirect(request.referrer or url_for('employee.employee_view', id=emp_id))


# =============================================
# REPORT: Active Member Data Sheet (EPF-style)
# =============================================

@employee_bulk_bp.route('/establishments/<int:est_id>/active-member-sheet')
def active_member_sheet(est_id):
    """Download Active Member Data Sheet (EPF-style) as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    est = Establishment.query.get_or_404(est_id)
    verify_est_ownership(est)

    employees = Employee.query.filter_by(
        establishment_id=est_id, is_active=True
    ).order_by(Employee.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Active Members"

    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )
    red_font = Font(name='Calibri', size=10, color='DC2626', bold=True)
    orange_font = Font(name='Calibri', size=10, color='D97706')
    normal_font = Font(name='Calibri', size=10)

    # Title
    cols = ['Sr No', 'UAN', 'Member ID', 'Name (ERP)', 'EPFO Registered Name', 'Name Status',
            'Gender', 'Date of Birth', 'Date of Joining', "Father's/Husband's Name", 'Relation',
            'Marital Status', 'Mobile', 'Email', 'ESIC IP Number', 'Aadhaar',
            'PAN', 'Bank Account', 'IFSC Code', 'Aadhaar Verified', 'Nomination Filed', 'Face Auth']

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws['A1'].value = f'{est.company_name} — Active Member Data Sheet'
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color='2563EB')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    info = f'Total: {len(employees)} active employees'
    if est.pf_code:
        info += f' | PF Code: {est.pf_code}'
    if est.esic_code:
        info += f' | ESIC Code: {est.esic_code}'
    info += f' | Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ws['A2'].value = info
    ws['A2'].font = Font(size=9, color='666666')

    # Headers
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    widths = [6, 14, 28, 22, 22, 14, 8, 14, 14, 22, 10, 14, 14, 22, 20, 16, 12, 18, 14, 12, 12, 10]
    for i, w in enumerate(widths):
        letter = chr(65 + i) if i < 26 else 'A' + chr(65 + i - 26)
        ws.column_dimensions[letter].width = w

    # Data
    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    mismatch_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    accepted_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')

    for row_idx, emp in enumerate(employees, 4):
        status = emp.name_status
        row_data = [
            row_idx - 3,
            emp.uan_number or '',
            emp.member_id or '',
            emp.name,
            emp.epfo_name or '',
            {'match': 'OK', 'mismatch': 'MISMATCH', 'accepted': 'ACCEPTED', 'no_epfo': 'NO EPFO DATA'}.get(status, ''),
            emp.gender,
            emp.date_of_birth.strftime('%d-%b-%Y') if emp.date_of_birth else '',
            emp.date_of_joining.strftime('%d-%b-%Y') if emp.date_of_joining else '',
            emp.father_husband_name or '',
            emp.relation or '',
            emp.marital_status or '',
            emp.mobile_number or '',
            emp.email or '',
            emp.esic_ip_number or '',
            emp.aadhaar_number or '',
            emp.pan_number or '',
            emp.bank_account_number or '',
            emp.bank_ifsc_code or '',
            'YES' if emp.aadhaar_verified else ('NO' if emp.aadhaar_verified is False else ''),
            'YES' if emp.nomination_filed else ('NO' if emp.nomination_filed is False else ''),
            'YES' if emp.face_auth_status else ('NO' if emp.face_auth_status is False else ''),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Name status coloring
            if status == 'mismatch':
                cell.fill = mismatch_fill
                if col_idx in (4, 5, 6):  # Name columns
                    cell.font = red_font
                else:
                    cell.font = normal_font
            elif status == 'accepted':
                cell.fill = accepted_fill
                if col_idx in (4, 5, 6):
                    cell.font = orange_font
                else:
                    cell.font = normal_font
            else:
                cell.font = normal_font
                if row_idx % 2 == 0:
                    cell.fill = even_fill

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'ActiveMembers_{est.company_name.replace(" ", "_")}_{datetime.now().strftime("%d%m%Y")}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# =============================================
# REPORT: Employee Profile Card (Individual)
# =============================================

@employee_bulk_bp.route('/employees/<int:emp_id>/profile-card')
def employee_profile_card(emp_id):
    """Download a professional Employee Profile Card as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    emp = Employee.query.get_or_404(emp_id)
    est = emp.establishment
    verify_est_ownership(est)

    from app.models.payroll import EmployeeSalary
    salary = EmployeeSalary.query.filter_by(employee_id=emp.id, is_current=True).first()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Profile"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Styles
    title_font = Font(name='Calibri', bold=True, size=16, color='2563EB')
    section_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    section_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    label_font = Font(name='Calibri', bold=True, size=10, color='475569')
    value_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0')
    )
    mismatch_font = Font(name='Calibri', size=11, color='DC2626', bold=True)

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 28

    row = 1

    # Header
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value=est.company_name).font = title_font
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    sub_info = ''
    if est.pf_code:
        sub_info += f'PF: {est.pf_code}'
    if est.esic_code:
        sub_info += f' | ESIC: {est.esic_code}'
    ws.cell(row=row, column=2, value=sub_info).font = Font(size=9, color='666666')
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value='EMPLOYEE PROFILE CARD').font = Font(name='Calibri', bold=True, size=13, color='1E293B')
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    row += 2

    def add_section(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=2, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 24
        row += 1

    def add_row(label1, value1, label2=None, value2=None, font_override=None):
        nonlocal row
        ws.cell(row=row, column=2, value=label1).font = label_font
        ws.cell(row=row, column=2).border = thin_border
        c = ws.cell(row=row, column=3, value=value1 or '')
        c.font = font_override or value_font
        c.border = thin_border
        if label2:
            ws.cell(row=row, column=4, value=label2).font = label_font
            ws.cell(row=row, column=4).border = thin_border
            c2 = ws.cell(row=row, column=5, value=value2 or '')
            c2.font = font_override if font_override and label2 == label1 else value_font
            c2.border = thin_border
        else:
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=5).border = thin_border
        row += 1

    # Section: Basic Information
    add_section('BASIC INFORMATION')
    add_row('UAN Number', emp.uan_number or '', 'Status', 'Active' if emp.is_active else 'Exited')
    add_row('Employee Name', emp.name, 'Gender', emp.gender)
    add_row("Father's/Husband Name", emp.father_husband_name, 'Relation', emp.relation or '')
    add_row('Date of Birth', emp.date_of_birth.strftime('%d-%b-%Y') if emp.date_of_birth else '', 'Date of Joining', emp.date_of_joining.strftime('%d-%b-%Y') if emp.date_of_joining else '')
    add_row('Marital Status', emp.marital_status or '', 'Designation', emp.designation or '')
    add_row('Department', emp.department or '', 'Internal Code', emp.internal_emp_code or '')
    row += 1

    # Section: Statutory & Compliance
    add_section('STATUTORY & COMPLIANCE')
    add_row('Employee Code', emp.emp_code, 'EPF Member ID', emp.member_id or '')
    add_row('ESIC IP Number', emp.esic_ip_number or '', 'Aadhaar Number', emp.aadhaar_number or '')
    add_row('PAN Number', emp.pan_number or '')

    # EPFO name with mismatch highlight
    status = emp.name_status
    epfo_font = mismatch_font if status == 'mismatch' else value_font
    add_row('EPFO Registered Name', emp.epfo_name or '', 'Name Status',
            {'match': 'OK', 'mismatch': 'MISMATCH', 'accepted': 'ACCEPTED', 'no_epfo': 'No EPFO Data'}.get(status, ''),
            font_override=epfo_font if status == 'mismatch' else None)
    add_row('Aadhaar Verified', 'YES' if emp.aadhaar_verified else ('NO' if emp.aadhaar_verified is False else 'N/A'),
            'Face Auth', 'YES' if emp.face_auth_status else ('NO' if emp.face_auth_status is False else 'N/A'))
    add_row('Nomination Filed', 'YES' if emp.nomination_filed else ('NO' if emp.nomination_filed is False else 'N/A'))
    row += 1

    # Section: Contact
    add_section('CONTACT DETAILS')
    add_row('Mobile Number', emp.mobile_number or '', 'Email', emp.email or '')
    add_row('Address', emp.address or '')
    row += 1

    # Section: Bank Details
    add_section('BANK DETAILS')
    add_row('Bank Name', emp.bank_name or '', 'Account Number', emp.bank_account_number or '')
    add_row('IFSC Code', emp.bank_ifsc_code or '')
    row += 1

    # Section: Current Salary
    add_section('CURRENT SALARY')
    if salary:
        add_row('Gross Salary', f'{salary.gross_salary:,.0f}' if salary.gross_salary else '',
                'Daily Rate', f'{salary.daily_rate:,.0f}' if salary.daily_rate else '')
        add_row('Effective From', salary.effective_from.strftime('%d-%b-%Y') if salary.effective_from else '',
                'Revision Reason', salary.revision_reason or '')
    else:
        add_row('Salary', 'Not assigned yet')
    row += 1

    # Section: Nominees
    if emp.nominees:
        add_section('NOMINEES')
        for nom in emp.nominees:
            add_row(nom.name, f'{nom.relation} ({nom.share_percentage or 0}%)',
                    'DOB', nom.date_of_birth.strftime('%d-%b-%Y') if nom.date_of_birth else '')
        row += 1

    # Section: Exit Details
    if emp.date_of_exit:
        add_section('EXIT DETAILS')
        add_row('Date of Exit', emp.date_of_exit.strftime('%d-%b-%Y'), 'Reason', emp.exit_reason or '')
        row += 1

    # Footer
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value=f'Generated on {datetime.now().strftime("%d-%b-%Y %H:%M")} | Vaishnavi Consultant ERP').font = Font(size=8, color='999999', italic=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Profile_{emp.name.replace(" ", "_")}_{emp.uan_number or emp.emp_code}.xlsx'
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
