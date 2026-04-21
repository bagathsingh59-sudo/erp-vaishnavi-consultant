"""
UAN & ESIC IP Tracker — Routes
Simple enrollment tracker for recording new UAN and ESIC IP numbers.
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from app import db
from app.models.enrollment import Enrollment
from app.models.establishment import Establishment
from app.models.employee import Employee
from app.user_context import (current_user_id, is_admin, user_establishments,
                               verify_est_ownership, set_owner, log_activity)
from datetime import datetime, date
from sqlalchemy import func, extract
import calendar
import io

enrollment_bp = Blueprint('enrollment', __name__)


def _user_enrollments():
    """Get enrollments scoped to current user."""
    q = Enrollment.query
    if not is_admin():
        uid = current_user_id()
        if uid:
            q = q.filter(Enrollment.owner_id == uid)
    return q


# =============================================
# DASHBOARD + ENROLLMENT LIST
# =============================================
@enrollment_bp.route('/enrollment')
def enrollment_home():
    """Dashboard + list of all enrollments with filters."""
    est_id = request.args.get('est_id', type=int)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int, default=date.today().year)
    status_filter = request.args.get('status', '')   # completed / partial / pending / linked

    establishments = user_establishments().filter_by(is_active=True)\
        .order_by(Establishment.company_name).all()

    # Base query
    q = _user_enrollments().join(Establishment).order_by(Enrollment.created_at.desc())

    if est_id:
        q = q.filter(Enrollment.establishment_id == est_id)
    if month:
        q = q.filter(extract('month', Enrollment.created_at) == month)
    if year:
        q = q.filter(extract('year', Enrollment.created_at) == year)
    if status_filter == 'completed':
        q = q.filter(Enrollment.uan_number.isnot(None), Enrollment.esic_ip_number.isnot(None))
    elif status_filter == 'partial':
        q = q.filter(
            db.or_(
                db.and_(Enrollment.uan_number.isnot(None), Enrollment.esic_ip_number.is_(None)),
                db.and_(Enrollment.uan_number.is_(None), Enrollment.esic_ip_number.isnot(None))
            ))
    elif status_filter == 'pending':
        q = q.filter(Enrollment.uan_number.is_(None), Enrollment.esic_ip_number.is_(None))
    elif status_filter == 'linked':
        q = q.filter(Enrollment.is_linked == True)

    enrollments = q.all()

    # Dashboard stats
    base_q = _user_enrollments()
    if year:
        base_q = base_q.filter(extract('year', Enrollment.created_at) == year)

    current_month = date.today().month
    this_month_q = base_q.filter(extract('month', Enrollment.created_at) == current_month)

    stats = {
        'total_year': base_q.count(),
        'total_month': this_month_q.count(),
        'pending': base_q.filter(Enrollment.uan_number.is_(None), Enrollment.esic_ip_number.is_(None)).count(),
        'completed': base_q.filter(Enrollment.uan_number.isnot(None), Enrollment.esic_ip_number.isnot(None)).count(),
        'linked': base_q.filter(Enrollment.is_linked == True).count(),
        'not_linked': base_q.filter(Enrollment.is_linked == False,
                                     db.or_(Enrollment.uan_number.isnot(None), Enrollment.esic_ip_number.isnot(None))).count(),
    }

    return render_template('enrollment/home.html',
                           establishments=establishments,
                           enrollments=enrollments,
                           stats=stats,
                           est_id=est_id, month=month, year=year,
                           status_filter=status_filter)


# =============================================
# NEW ENROLLMENT ENTRY
# =============================================
@enrollment_bp.route('/enrollment/add', methods=['GET', 'POST'])
def enrollment_add():
    """Add new enrollment record."""
    if request.method == 'POST':
        try:
            est_id = int(request.form.get('establishment_id'))
            est = Establishment.query.get(est_id)
            verify_est_ownership(est)

            enrollment = Enrollment(
                establishment_id=est_id,
                employee_name=request.form.get('employee_name', '').strip().upper(),
                father_husband_name=request.form.get('father_husband_name', '').strip().upper(),
                gender=request.form.get('gender', 'Male'),
                date_of_birth=datetime.strptime(request.form.get('date_of_birth'), '%Y-%m-%d').date(),
                date_of_joining=datetime.strptime(request.form.get('date_of_joining'), '%Y-%m-%d').date(),
                uan_number=request.form.get('uan_number', '').strip() or None,
                esic_ip_number=request.form.get('esic_ip_number', '').strip() or None,
                aadhaar_number=request.form.get('aadhaar_number', '').strip() or None,
                mobile_number=request.form.get('mobile_number', '').strip() or None,
                designation=request.form.get('designation', '').strip() or None,
                remarks=request.form.get('remarks', '').strip() or None,
            )
            set_owner(enrollment)
            db.session.add(enrollment)
            db.session.commit()

            log_activity('created', 'enrollment', entity_id=enrollment.id,
                         entity_name=f'{enrollment.employee_name} — {est.display_name}',
                         details=f'UAN: {enrollment.uan_number or "—"}, ESIC: {enrollment.esic_ip_number or "—"}',
                         establishment_id=est_id)

            flash(f'Enrollment recorded: {enrollment.employee_name} under {est.display_name}', 'success')

            # If "Save & Add Another" was clicked
            if request.form.get('action') == 'save_add':
                return redirect(url_for('enrollment.enrollment_add', est_id=est_id))
            return redirect(url_for('enrollment.enrollment_home'))

        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')
            return redirect(url_for('enrollment.enrollment_add'))

    establishments = user_establishments().filter_by(is_active=True)\
        .order_by(Establishment.company_name).all()
    pre_est_id = request.args.get('est_id', type=int)

    return render_template('enrollment/add.html',
                           establishments=establishments,
                           pre_est_id=pre_est_id,
                           today=date.today())


# =============================================
# EDIT ENROLLMENT
# =============================================
@enrollment_bp.route('/enrollment/<int:enrollment_id>/edit', methods=['GET', 'POST'])
def enrollment_edit(enrollment_id):
    """Edit an enrollment record — update UAN/ESIC IP or other details."""
    enrollment = Enrollment.query.get_or_404(enrollment_id)

    if request.method == 'POST':
        try:
            enrollment.employee_name = request.form.get('employee_name', '').strip().upper()
            enrollment.father_husband_name = request.form.get('father_husband_name', '').strip().upper()
            enrollment.gender = request.form.get('gender', 'Male')
            enrollment.date_of_birth = datetime.strptime(request.form.get('date_of_birth'), '%Y-%m-%d').date()
            enrollment.date_of_joining = datetime.strptime(request.form.get('date_of_joining'), '%Y-%m-%d').date()
            enrollment.uan_number = request.form.get('uan_number', '').strip() or None
            enrollment.esic_ip_number = request.form.get('esic_ip_number', '').strip() or None
            enrollment.aadhaar_number = request.form.get('aadhaar_number', '').strip() or None
            enrollment.mobile_number = request.form.get('mobile_number', '').strip() or None
            enrollment.designation = request.form.get('designation', '').strip() or None
            enrollment.remarks = request.form.get('remarks', '').strip() or None

            db.session.commit()
            flash(f'Enrollment updated: {enrollment.employee_name}', 'success')
            return redirect(url_for('enrollment.enrollment_home'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {e}', 'danger')

    establishments = user_establishments().filter_by(is_active=True)\
        .order_by(Establishment.company_name).all()

    return render_template('enrollment/add.html',
                           establishments=establishments,
                           enrollment=enrollment,
                           edit_mode=True,
                           today=date.today())


# =============================================
# DELETE ENROLLMENT
# =============================================
@enrollment_bp.route('/enrollment/<int:enrollment_id>/delete', methods=['POST'])
def enrollment_delete(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    name = enrollment.employee_name
    db.session.delete(enrollment)
    db.session.commit()
    flash(f'Enrollment deleted: {name}', 'warning')
    return redirect(url_for('enrollment.enrollment_home'))


# =============================================
# QUICK LINK — Single employee to Add Employee form (pre-filled)
# =============================================
@enrollment_bp.route('/enrollment/<int:enrollment_id>/quick-link')
def quick_link(enrollment_id):
    """Redirect to Add Employee form with pre-filled data from enrollment."""
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    # Build query params to pre-fill the add employee form
    params = {
        'from_enrollment': enrollment.id,
        'est_id': enrollment.establishment_id,
        'name': enrollment.employee_name,
        'father_husband_name': enrollment.father_husband_name,
        'gender': enrollment.gender,
        'date_of_birth': enrollment.date_of_birth.strftime('%Y-%m-%d'),
        'date_of_joining': enrollment.date_of_joining.strftime('%Y-%m-%d'),
        'uan_number': enrollment.uan_number or '',
        'esic_ip_number': enrollment.esic_ip_number or '',
        'aadhaar_number': enrollment.aadhaar_number or '',
        'mobile_number': enrollment.mobile_number or '',
        'designation': enrollment.designation or '',
    }
    return redirect(url_for('employee.employee_add', **params))


# =============================================
# BULK LINK — Download pre-filled template for selected enrollments
# =============================================
@enrollment_bp.route('/enrollment/bulk-link-template', methods=['POST'])
def bulk_link_template():
    """Generate Excel template pre-filled with selected enrollment data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    ids = request.form.getlist('enrollment_ids', type=int)
    if not ids:
        flash('Please select at least one enrollment to generate template.', 'warning')
        return redirect(url_for('enrollment.enrollment_home'))

    enrollments = Enrollment.query.filter(Enrollment.id.in_(ids)).all()
    if not enrollments:
        flash('No enrollments found.', 'warning')
        return redirect(url_for('enrollment.enrollment_home'))

    # Use same column structure as employee_bulk import template
    COLUMNS = [
        'Establishment Name or PF Code*',
        'Employee Name (as per Aadhaar)*',
        "Father's / Husband Name*",
        'Gender (Male/Female/Other)*',
        'Date of Birth (DD-MM-YYYY)*',
        'Date of Joining (DD-MM-YYYY)*',
        'UAN Number',
        'ESIC IP Number',
        'Salary Type (Daily/Monthly/MonthlyHeads/CTC)',
        'Daily Rate',
        'Gross Salary (Monthly)',
        'Weekly Off (Paid/Unpaid/OT Rate)',
        'Basic', 'DA', 'HRA', 'Conveyance', 'Other Allowance', 'Washing Allowance',
        'CTC Amount (Monthly)',
        'WO Applicable (Yes/No)',
        'WO Type (Paid/Unpaid)',
        'WO Day (Sunday/Monday/etc/Rotational)',
        'Aadhaar Number',
        'PAN Number',
        'Mobile Number',
        'Email',
        'Marital Status',
        'Designation',
        'Department',
        'Address',
        'Bank Name',
        'Account Number',
        'IFSC Code',
        'Internal Emp Code',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Employee Import'

    # Title
    title_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    cell = ws.cell(row=1, column=1,
                   value=f'Employee Import Template — Pre-filled from UAN & ESIC Tracker ({len(enrollments)} employees)')
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Instructions row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    ws.cell(row=2, column=1,
            value='Columns with * are mandatory. Pre-filled columns (highlighted green) are from tracker. Fill remaining columns (salary, bank, etc.) and upload.').font = Font(name='Arial', size=9, italic=True)

    # Header row
    hdr_font = Font(name='Arial', size=9, bold=True)
    hdr_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    prefill_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # Columns that are pre-filled (indices 0-7, 22, 24, 27)
    prefill_cols = {0, 1, 2, 3, 4, 5, 6, 7, 22, 24, 27}  # est, name, father, gender, dob, doj, uan, esic, aadhaar, mobile, designation

    for ci, col_name in enumerate(COLUMNS):
        cell = ws.cell(row=3, column=ci + 1, value=col_name)
        cell.font = hdr_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = prefill_fill if ci in prefill_cols else hdr_fill

    # Data rows — pre-filled from enrollment records
    data_font = Font(name='Arial', size=9)
    green_fill = PatternFill(start_color='EAFAF1', end_color='EAFAF1', fill_type='solid')

    for ri, enr in enumerate(enrollments, 4):
        est = enr.establishment
        est_name = est.company_name if est else ''

        row_data = [
            est_name,                                                    # Establishment Name
            enr.employee_name,                                           # Name
            enr.father_husband_name,                                     # Father Name
            enr.gender,                                                  # Gender
            enr.date_of_birth.strftime('%d-%m-%Y') if enr.date_of_birth else '',   # DOB
            enr.date_of_joining.strftime('%d-%m-%Y') if enr.date_of_joining else '', # DOJ
            enr.uan_number or '',                                        # UAN
            enr.esic_ip_number or '',                                    # ESIC IP
            '', '', '', '',                                              # Salary fields (empty)
            '', '', '', '', '', '',                                      # Head-wise (empty)
            '',                                                          # CTC
            '', '', '',                                                  # WO policy
            enr.aadhaar_number or '',                                    # Aadhaar
            '',                                                          # PAN
            enr.mobile_number or '',                                     # Mobile
            '',                                                          # Email
            '',                                                          # Marital Status
            enr.designation or '',                                       # Designation
            '',                                                          # Department
            '',                                                          # Address
            '', '', '',                                                  # Bank details
            '',                                                          # Internal code
        ]

        for ci, val in enumerate(row_data):
            cell = ws.cell(row=ri, column=ci + 1, value=val)
            cell.font = data_font
            cell.border = thin_border
            if ci in prefill_cols and val:
                cell.fill = green_fill

    # Column widths
    widths = [35, 30, 30, 22, 25, 25, 18, 22, 28, 14, 18, 22,
              14, 14, 14, 16, 18, 18, 18, 18, 18, 28, 16, 14, 16,
              25, 14, 18, 18, 30, 25, 20, 14, 16]
    for i, w in enumerate(widths):
        if i < len(COLUMNS):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.row_dimensions[3].height = 35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Employee_Import_From_Tracker_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)


# =============================================
# MARK AS LINKED (after employee is added via import or quick link)
# =============================================
@enrollment_bp.route('/enrollment/<int:enrollment_id>/mark-linked', methods=['POST'])
def mark_linked(enrollment_id):
    """Mark an enrollment as linked to employee system."""
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    enrollment.is_linked = True
    db.session.commit()
    flash(f'{enrollment.employee_name} marked as linked.', 'success')
    return redirect(url_for('enrollment.enrollment_home'))


@enrollment_bp.route('/enrollment/mark-linked-bulk', methods=['POST'])
def mark_linked_bulk():
    """Mark multiple enrollments as linked."""
    ids = request.form.getlist('enrollment_ids', type=int)
    if ids:
        Enrollment.query.filter(Enrollment.id.in_(ids)).update(
            {Enrollment.is_linked: True}, synchronize_session=False)
        db.session.commit()
        flash(f'{len(ids)} enrollment(s) marked as linked.', 'success')
    return redirect(url_for('enrollment.enrollment_home'))


# =============================================
# REPORT — Printable enrollment report
# =============================================
@enrollment_bp.route('/enrollment/report')
def enrollment_report():
    """Printable enrollment report — grouped by establishment."""
    est_id = request.args.get('est_id', type=int)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int, default=date.today().year)

    q = _user_enrollments().join(Establishment)\
        .order_by(Establishment.company_name, Enrollment.date_of_joining)

    if est_id:
        q = q.filter(Enrollment.establishment_id == est_id)
    if month:
        q = q.filter(extract('month', Enrollment.created_at) == month)
    if year:
        q = q.filter(extract('year', Enrollment.created_at) == year)

    enrollments = q.all()

    # Group by establishment
    grouped = {}
    for enr in enrollments:
        est_name = enr.establishment.company_name
        if est_name not in grouped:
            grouped[est_name] = []
        grouped[est_name].append(enr)

    establishments = user_establishments().filter_by(is_active=True)\
        .order_by(Establishment.company_name).all()
    generated_on = datetime.now().strftime('%d %b %Y, %I:%M %p')

    return render_template('enrollment/report.html',
                           grouped=grouped,
                           establishments=establishments,
                           est_id=est_id, month=month, year=year,
                           total=len(enrollments),
                           generated_on=generated_on)
