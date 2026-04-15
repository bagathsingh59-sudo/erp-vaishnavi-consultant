from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, current_app
from app import db
from app.models.establishment import Establishment
from app.user_context import current_user_id, user_establishments
from datetime import datetime
import os
import io

bulk_bp = Blueprint('bulk', __name__)

# Template columns — these define the standard import format
TEMPLATE_COLUMNS = [
    'Company Name*',
    'Type of Industry',
    'Date of Registration (DD-MM-YYYY)',
    'Address',
    'Contact Person',
    'Contact Phone',
    'Contact Email',
    'PF Code Number',
    'ESIC Code Number',
    'PAN Number',
    'GST Number',
    'Fee Type (Monthly/Quarterly/Yearly)',
    'Fee Amount',
    'Service Type (With Records/Only Returns)',
    'Status (Active/Inactive)'
]

# Column widths for better readability
COLUMN_WIDTHS = [35, 22, 28, 45, 25, 18, 30, 25, 25, 15, 20, 32, 14, 35, 22]

# Sample data row for reference
SAMPLE_ROW = [
    'ABC Enterprises Pvt Ltd',
    'Manufacturing',
    '15-06-2020',
    '123 Industrial Area, Gulbarga, Karnataka - 585101',
    'Ramesh Kumar',
    '9876543210',
    'ramesh@abc.com',
    'GBGLB1234567000',
    '71000012340000606',
    'ABCDE1234F',
    '29ABCDE1234F1Z5',
    'Monthly',
    '2000',
    'With Records',
    'Active'
]


@bulk_bp.route('/establishments/download-template')
def download_template():
    """Download blank import template as .xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Establishments"

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sample_font = Font(name='Calibri', size=10, color='888888', italic=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Title row
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'Vaishnavi Consultant ERP - Establishment Import Template'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Instructions row
    ws.merge_cells('A2:O2')
    inst_cell = ws['A2']
    inst_cell.value = 'Instructions: Fill data from Row 4 onwards. Row 3 is a sample (delete it before uploading). Fields marked with * are required.'
    inst_cell.font = Font(name='Calibri', size=10, color='E53E3E', italic=True)
    inst_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    # Header row (Row 3 visually, but we use row 3 for headers)
    header_row = 3
    ws.row_dimensions[header_row].height = 30
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width
    # Fix for columns beyond Z
    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
    for i, letter in enumerate(col_letters):
        ws.column_dimensions[letter].width = COLUMN_WIDTHS[i]

    # Sample data row (Row 4)
    sample_row = 4
    sample_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    for col_idx, value in enumerate(SAMPLE_ROW, 1):
        cell = ws.cell(row=sample_row, column=col_idx, value=value)
        cell.font = sample_font
        cell.fill = sample_fill
        cell.border = thin_border

    # Add note in row 5
    ws.cell(row=5, column=1, value='<-- Delete the sample row above and start entering your data here')
    ws.cell(row=5, column=1).font = Font(name='Calibri', size=9, color='999999', italic=True)

    # Freeze header row
    ws.freeze_panes = 'A4'

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Vaishnavi_ERP_Import_Template.xlsx'
    )


@bulk_bp.route('/establishments/export')
def export_establishments():
    """Export all establishments to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Get filter params (same as list page)
    filter_status = request.args.get('status', 'all')
    filter_service = request.args.get('service', 'all')

    query = user_establishments()
    if filter_status == 'active':
        query = query.filter_by(is_active=True)
    elif filter_status == 'inactive':
        query = query.filter_by(is_active=False)
    if filter_service == 'with_records':
        query = query.filter_by(service_type='With Records')
    elif filter_service == 'only_returns':
        query = query.filter_by(service_type='Only Returns')

    establishments = query.order_by(Establishment.company_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Establishments"

    # Styles
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Export columns (slightly different from template — includes ID)
    export_columns = [
        'Sr No', 'Company Name', 'Type of Industry', 'Date of Registration',
        'Address', 'Contact Person', 'Contact Phone', 'Contact Email',
        'PF Code Number', 'ESIC Code Number', 'PAN Number', 'GST Number',
        'Fee Type', 'Fee Amount', 'Service Type', 'Status'
    ]

    col_widths = [8, 35, 22, 18, 45, 25, 18, 30, 25, 25, 15, 20, 14, 14, 18, 12]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_columns))
    title_cell = ws['A1']
    title_cell.value = f'Vaishnavi Consultant ERP - Establishments Export ({datetime.now().strftime("%d-%m-%Y")})'
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Summary row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(export_columns))
    ws['A2'].value = f'Total Records: {len(establishments)}'
    ws['A2'].font = Font(name='Calibri', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='left')

    # Headers
    header_row = 3
    ws.row_dimensions[header_row].height = 28
    for col_idx, col_name in enumerate(export_columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_letters = [chr(64 + i) if i <= 26 else 'A' for i in range(1, len(export_columns) + 1)]
    actual_letters = []
    for i in range(len(export_columns)):
        if i < 26:
            actual_letters.append(chr(65 + i))
        else:
            actual_letters.append('A' + chr(65 + i - 26))
    for i, letter in enumerate(actual_letters):
        ws.column_dimensions[letter].width = col_widths[i]

    # Data rows
    even_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    data_font = Font(name='Calibri', size=10)

    for row_idx, est in enumerate(establishments, 4):
        row_data = [
            row_idx - 3,
            est.company_name,
            est.type_of_industry or '',
            est.date_of_registration.strftime('%d-%m-%Y') if est.date_of_registration else '',
            est.address or '',
            est.contact_person or '',
            est.contact_phone or '',
            est.contact_email or '',
            est.pf_code or '',
            est.esic_code or '',
            est.pan_number or '',
            est.gst_number or '',
            est.fee_type or '',
            est.fee_amount or '',
            est.service_type or '',
            'Active' if est.is_active else 'Inactive'
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill

    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Establishments_Export_{datetime.now().strftime("%d%m%Y_%H%M")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bulk_bp.route('/establishments/import', methods=['GET', 'POST'])
def import_establishments():
    """Import establishments from Excel file"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Please select a file to upload.', 'danger')
            return redirect(url_for('bulk.import_establishments'))

        filename = file.filename.lower()

        try:
            if filename.endswith('.xlsx'):
                return _process_xlsx(file)
            elif filename.endswith('.xls'):
                return _process_xls(file)
            else:
                flash('Please upload an .xlsx or .xls file.', 'danger')
                return redirect(url_for('bulk.import_establishments'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'danger')
            return redirect(url_for('bulk.import_establishments'))

    return render_template('establishments/import.html')


def _parse_date(value):
    """Parse date from various formats including EPF portal (DD-MMM-YYYY)"""
    if not value:
        return None
    if isinstance(value, (int, float)):
        # Excel serial date
        try:
            from datetime import date, timedelta
            # Excel date serial: days since 1899-12-30
            base = date(1899, 12, 30)
            return base + timedelta(days=int(value))
        except Exception:
            return None
    # Handle datetime objects from openpyxl
    if hasattr(value, 'date'):
        return value.date() if callable(getattr(value, 'date')) else value
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    value = str(value).strip()
    if not value or value.upper() in ('NOT AVAILABLE', 'NA', 'N/A'):
        return None
    for fmt in [
        '%d-%b-%Y',     # 25-AUG-2006 (EPF standard)
        '%d-%B-%Y',     # 25-AUGUST-2006
        '%d-%m-%Y',     # 25-08-2006
        '%d/%m/%Y',     # 25/08/2006
        '%d/%b/%Y',     # 25/AUG/2006
        '%Y-%m-%d',     # 2006-08-25 (ISO)
        '%d-%m-%y',     # 25-08-06
        '%d/%m/%y',     # 25/08/06
        '%d-%b-%y',     # 25-AUG-06
    ]:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_fee(value):
    """Parse fee amount"""
    if not value:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _parse_status(value):
    """Parse status to boolean"""
    if not value:
        return True
    value = str(value).strip().upper()
    if value in ['INACTIVE', 'CLOSED', 'NO', 'FALSE', '0', 'IREGULAR', 'IRREGULAR']:
        return False
    return True


def _parse_service_type(value):
    """Parse service type"""
    if not value:
        return None
    value = str(value).strip().lower()
    if 'record' in value:
        return 'With Records'
    elif 'return' in value:
        return 'Only Returns'
    return None


def _parse_fee_type(value):
    """Parse fee type"""
    if not value:
        return None
    value = str(value).strip().lower()
    if 'month' in value:
        return 'Monthly'
    elif 'quarter' in value:
        return 'Quarterly'
    elif 'year' in value or 'annual' in value:
        return 'Yearly'
    return None


def _process_xlsx(file):
    """Process .xlsx file"""
    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash('The file appears to be empty.', 'danger')
        return redirect(url_for('bulk.import_establishments'))

    # Find header row (look for 'Company Name' in first few rows)
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows[:5]):
        row_str = [str(c).strip().lower() if c else '' for c in row]
        if any('company name' in cell or 'company_name' in cell or 'companyname' in cell for cell in row_str):
            header_row_idx = idx
            headers = [str(c).strip() if c else '' for c in row]
            break

    if header_row_idx is None:
        flash('Could not find header row. Make sure "Company Name" is in the header.', 'danger')
        return redirect(url_for('bulk.import_establishments'))

    return _import_rows(rows[header_row_idx + 1:], headers)


def _process_xls(file):
    """Process .xls file"""
    import xlrd

    # Save to temp file
    temp_path = os.path.join(current_app.root_path, '..', 'data', 'temp_import.xls')
    file.save(temp_path)

    try:
        wb = xlrd.open_workbook(temp_path)
        ws = wb.sheet_by_index(0)

        # Find header row
        header_row_idx = None
        headers = []
        for idx in range(min(5, ws.nrows)):
            row = [str(ws.cell_value(idx, c)).strip().lower() for c in range(ws.ncols)]
            if any('company name' in cell or 'company_name' in cell or 'companyname' in cell for cell in row):
                header_row_idx = idx
                headers = [str(ws.cell_value(idx, c)).strip() for c in range(ws.ncols)]
                break

        if header_row_idx is None:
            flash('Could not find header row. Make sure "Company Name" is in the header.', 'danger')
            return redirect(url_for('bulk.import_establishments'))

        rows = []
        for r in range(header_row_idx + 1, ws.nrows):
            row = [ws.cell_value(r, c) for c in range(ws.ncols)]
            rows.append(row)

        return _import_rows(rows, headers)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def _map_column(headers, possible_names):
    """Find column index by trying multiple possible header names"""
    for idx, header in enumerate(headers):
        h = header.lower().strip().replace('_', '').replace(' ', '')
        for name in possible_names:
            if name.lower().replace('_', '').replace(' ', '') in h:
                return idx
    return None


def _import_rows(rows, headers):
    """Import rows with smart column mapping"""
    # Map columns (supports both template format and user's existing Excel format)
    col_map = {
        'company_name': _map_column(headers, ['company name', 'companyname', 'name', 'establishment']),
        'type_of_industry': _map_column(headers, ['industry', 'type of industry', 'nature of business', 'natureofbusiness']),
        'date_of_registration': _map_column(headers, ['date of registration', 'dateofcommencement', 'date of commencement', 'registration date']),
        'address': _map_column(headers, ['address', 'company address', 'companyaddress']),
        'contact_person': _map_column(headers, ['contact person', 'contactperson', 'authorized person', 'authorizedpersonname']),
        'contact_phone': _map_column(headers, ['phone', 'contact phone', 'contactnumber', 'authorizedcontactnumber', 'mobile']),
        'contact_email': _map_column(headers, ['email', 'contact email', 'companyemailid', 'emailid']),
        'pf_code': _map_column(headers, ['pf code', 'pfcode', 'epfcodenumber', 'epf code']),
        'esic_code': _map_column(headers, ['esic code', 'esiccode', 'esiccodenumber']),
        'pan_number': _map_column(headers, ['pan', 'pannumber', 'pan number']),
        'gst_number': _map_column(headers, ['gst', 'gstnumber', 'gst number']),
        'fee_type': _map_column(headers, ['fee type', 'feetype', 'fee cycle']),
        'fee_amount': _map_column(headers, ['fee amount', 'feeamount', 'professional fees', 'professionalfees']),
        'service_type': _map_column(headers, ['service type', 'servicetype', 'service provide', 'serviceprovide']),
        'status': _map_column(headers, ['status', 'active']),
    }

    if col_map['company_name'] is None:
        flash('Could not find "Company Name" column in the file.', 'danger')
        return redirect(url_for('bulk.import_establishments'))

    imported = 0
    skipped = 0
    errors = []

    for row_num, row in enumerate(rows, 1):
        try:
            # Get company name
            name_idx = col_map['company_name']
            if name_idx >= len(row):
                continue
            company_name = str(row[name_idx]).strip() if row[name_idx] else ''
            if not company_name or company_name == '0' or company_name == '0.0':
                continue

            # Check for duplicate
            existing = Establishment.query.filter(
                db.func.lower(Establishment.company_name) == company_name.lower()
            ).first()
            if existing:
                skipped += 1
                continue

            def get_val(field):
                idx = col_map.get(field)
                if idx is not None and idx < len(row) and row[idx]:
                    val = str(row[idx]).strip()
                    # Clean phone numbers (remove .0 from float conversion)
                    if field == 'contact_phone' and val.endswith('.0'):
                        val = val[:-2]
                    return val if val and val != '0' and val != '0.0' else None
                return None

            est = Establishment(
                company_name=company_name,
                type_of_industry=get_val('type_of_industry'),
                date_of_registration=_parse_date(row[col_map['date_of_registration']] if col_map['date_of_registration'] is not None and col_map['date_of_registration'] < len(row) else None),
                address=get_val('address'),
                contact_person=get_val('contact_person'),
                contact_phone=get_val('contact_phone'),
                contact_email=get_val('contact_email'),
                pf_code=get_val('pf_code'),
                esic_code=get_val('esic_code'),
                pan_number=get_val('pan_number').upper() if get_val('pan_number') else None,
                gst_number=get_val('gst_number').upper() if get_val('gst_number') else None,
                fee_type=_parse_fee_type(get_val('fee_type')),
                fee_amount=_parse_fee(row[col_map['fee_amount']] if col_map['fee_amount'] is not None and col_map['fee_amount'] < len(row) else None),
                service_type=_parse_service_type(get_val('service_type')),
                is_active=_parse_status(get_val('status')),
                owner_id=current_user_id(),
            )

            db.session.add(est)
            imported += 1

        except Exception as e:
            errors.append(f'Row {row_num}: {str(e)}')
            continue

    if imported > 0:
        db.session.commit()

    # Flash results
    msg_parts = []
    if imported > 0:
        msg_parts.append(f'{imported} establishments imported successfully')
    if skipped > 0:
        msg_parts.append(f'{skipped} duplicates skipped')
    if errors:
        msg_parts.append(f'{len(errors)} rows had errors')

    if imported > 0:
        flash('. '.join(msg_parts) + '.', 'success')
    elif skipped > 0:
        flash('. '.join(msg_parts) + '.', 'info')
    else:
        flash('No records were imported. ' + '. '.join(msg_parts), 'warning')

    if errors and len(errors) <= 5:
        for err in errors:
            flash(err, 'danger')

    return redirect(url_for('establishment.establishment_list'))
